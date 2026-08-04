# =============================================================================
# DockAnalyzer — Data export and serialization
# Section 1 — Header, imports, aliases and public interface
# =============================================================================

"""Serialization and file export utilities for DockAnalyzer.

This module converts DockAnalyzer results into portable representations and
writes them to JSON, JSON Lines, CSV, TSV, Excel and plain-text files.

The export layer consumes existing analysis and scoring results. It does not
redetect molecular interactions, recalculate scores or generate scientific
reports. Specialized modules remain responsible for their own chemistry,
geometry and scoring logic.

The implementation is designed to work:

1. inside UCSF ChimeraX with native molecular objects;
2. outside ChimeraX with ordinary Python objects;
3. with dictionaries loaded from previous exports;
4. with synthetic objects used by DockAnalyzer self-tests.

Optional dependencies are loaded defensively. JSON, JSON Lines, CSV, TSV and
text export remain available without pandas or openpyxl.
"""

from __future__ import annotations

# -----------------------------------------------------------------------------
# 1.1. Standard-library imports
# -----------------------------------------------------------------------------

from collections import defaultdict
from collections.abc import Iterable, Iterator, Mapping, MutableMapping, Sequence
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field, fields, is_dataclass, replace
from datetime import date, datetime, time, timedelta, timezone
from enum import Enum
from hashlib import sha256
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile, TemporaryDirectory
from types import MappingProxyType
from typing import (
    Any,
    BinaryIO,
    Callable,
    ClassVar,
    Dict,
    Final,
    FrozenSet,
    Hashable,
    List,
    Literal,
    Optional,
    Protocol,
    Set,
    TextIO,
    Tuple,
    TypeAlias,
    TypeVar,
    Union,
    runtime_checkable,
)
import csv
import json
import math
import os
import platform
import re
import shutil
import sys

# -----------------------------------------------------------------------------
# 1.2. Optional NumPy support
# -----------------------------------------------------------------------------

try:
    import numpy as np
    from numpy.typing import NDArray

    NUMPY_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - environment dependent
    np = None  # type: ignore[assignment]
    NDArray = Any  # type: ignore[misc,assignment]
    NUMPY_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.3. Optional pandas support
# -----------------------------------------------------------------------------

try:
    import pandas as pd

    PANDAS_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - environment dependent
    pd = None  # type: ignore[assignment]
    PANDAS_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.4. Optional openpyxl support
# -----------------------------------------------------------------------------

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - environment dependent
    openpyxl = None  # type: ignore[assignment]
    Workbook = Any  # type: ignore[misc,assignment]
    Alignment = Any  # type: ignore[misc,assignment]
    Font = Any  # type: ignore[misc,assignment]
    PatternFill = Any  # type: ignore[misc,assignment]
    Table = Any  # type: ignore[misc,assignment]
    TableStyleInfo = Any  # type: ignore[misc,assignment]
    get_column_letter = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.5. Optional ChimeraX support
# -----------------------------------------------------------------------------

try:
    from chimerax.atomic import Atom as ChimeraXAtom
    from chimerax.atomic import AtomicStructure as ChimeraXAtomicStructure
    from chimerax.atomic import Residue as ChimeraXResidue

    CHIMERAX_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - expected outside ChimeraX
    ChimeraXAtom = Any  # type: ignore[misc,assignment]
    ChimeraXAtomicStructure = Any  # type: ignore[misc,assignment]
    ChimeraXResidue = Any  # type: ignore[misc,assignment]
    CHIMERAX_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.6. Internal DockAnalyzer imports
# -----------------------------------------------------------------------------

if __package__:
    from . import config
    from .utils import DockLogger, DockModel
else:
    import config
    from utils import DockLogger, DockModel

# -----------------------------------------------------------------------------
# 1.7. Module metadata
# -----------------------------------------------------------------------------

__author__: Final[str] = "Leonardo Bastos and DockAnalyzer contributors"
__version__: Final[str] = "0.1.0"
__license__: Final[str] = "MIT"
__status__: Final[str] = "Development"

_MODULE_NAME: Final[str] = "export"
_MODULE_DESCRIPTION: Final[str] = (
    "Serialization and multi-format export utilities for DockAnalyzer."
)

_LOGGER: Final[DockLogger] = DockLogger(_MODULE_NAME)

# -----------------------------------------------------------------------------
# 1.8. Public interface
# -----------------------------------------------------------------------------

# Public names are registered after their defining sections.
__all__: List[str] = []

# -----------------------------------------------------------------------------
# 1.9. Generic type variables
# -----------------------------------------------------------------------------

T = TypeVar("T")
K = TypeVar("K", bound=Hashable)
V = TypeVar("V")
RecordT = TypeVar("RecordT", bound=Mapping[str, Any])
ExportObjectT = TypeVar("ExportObjectT")

# -----------------------------------------------------------------------------
# 1.10. General aliases
# -----------------------------------------------------------------------------

PathLike: TypeAlias = Union[str, os.PathLike[str], Path]
FileLike: TypeAlias = Union[TextIO, BinaryIO]

Number: TypeAlias = Union[int, float]
if NUMPY_AVAILABLE:
    Number = Union[int, float, np.integer, np.floating]

JSONPrimitive: TypeAlias = Union[str, int, float, bool, None]
JSONValue: TypeAlias = Union[
    JSONPrimitive,
    List["JSONValue"],
    Dict[str, "JSONValue"],
]
JSONObject: TypeAlias = Dict[str, JSONValue]
JSONArray: TypeAlias = List[JSONValue]

Record: TypeAlias = Dict[str, Any]
RecordSequence: TypeAlias = Sequence[Mapping[str, Any]]
RecordList: TypeAlias = List[Record]
TableMapping: TypeAlias = Mapping[str, RecordSequence]
MutableTableMapping: TypeAlias = MutableMapping[str, RecordList]

Metadata: TypeAlias = Mapping[str, Any]
MutableMetadata: TypeAlias = MutableMapping[str, Any]

ColumnName: TypeAlias = str
SheetName: TypeAlias = str
TableName: TypeAlias = str
FormatName: TypeAlias = str

# -----------------------------------------------------------------------------
# 1.11. Molecular and DockAnalyzer aliases
# -----------------------------------------------------------------------------

AtomLike: TypeAlias = Any
ResidueLike: TypeAlias = Any
StructureLike: TypeAlias = Any
InteractionLike: TypeAlias = Any
AnalysisResultLike: TypeAlias = Any
ScoringResultLike: TypeAlias = Any
DockModelLike: TypeAlias = Union[DockModel, Any]

AtomCollection: TypeAlias = Iterable[AtomLike]
ResidueCollection: TypeAlias = Iterable[ResidueLike]
InteractionCollection: TypeAlias = Iterable[InteractionLike]
DockModelCollection: TypeAlias = Iterable[DockModelLike]

# -----------------------------------------------------------------------------
# 1.12. Callable aliases
# -----------------------------------------------------------------------------

Serializer: TypeAlias = Callable[[Any], Any]
ObjectSerializer: TypeAlias = Callable[[Any, "ExportContextLike"], Any]
RecordBuilder: TypeAlias = Callable[[Any], Record]
PathResolver: TypeAlias = Callable[[PathLike], Path]
ErrorHandler: TypeAlias = Callable[[Exception], None]

# -----------------------------------------------------------------------------
# 1.13. Structural protocols
# -----------------------------------------------------------------------------

@runtime_checkable
class SupportsToDict(Protocol):
    """Object exposing ``to_dict``."""

    def to_dict(self) -> Mapping[str, Any]: ...


@runtime_checkable
class SupportsAsDict(Protocol):
    """Object exposing ``as_dict``."""

    def as_dict(self) -> Mapping[str, Any]: ...


@runtime_checkable
class SupportsJSON(Protocol):
    """Object exposing ``to_json``."""

    def to_json(self, *args: Any, **kwargs: Any) -> str: ...


@runtime_checkable
class SupportsFileRecord(Protocol):
    """Object exposing a file-record mapping."""

    files: MutableMapping[str, Any]


@runtime_checkable
class ExportContextLike(Protocol):
    """Minimum serializer-context interface."""

    include_metadata: bool
    include_private: bool
    strict: bool


# -----------------------------------------------------------------------------
# 1.14. Section registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "NUMPY_AVAILABLE",
        "PANDAS_AVAILABLE",
        "OPENPYXL_AVAILABLE",
        "CHIMERAX_AVAILABLE",
        "PathLike",
        "FileLike",
        "Number",
        "JSONPrimitive",
        "JSONValue",
        "JSONObject",
        "JSONArray",
        "Record",
        "RecordSequence",
        "RecordList",
        "TableMapping",
        "MutableTableMapping",
        "Metadata",
        "MutableMetadata",
        "ColumnName",
        "SheetName",
        "TableName",
        "FormatName",
        "AtomLike",
        "ResidueLike",
        "StructureLike",
        "InteractionLike",
        "AnalysisResultLike",
        "ScoringResultLike",
        "DockModelLike",
        "AtomCollection",
        "ResidueCollection",
        "InteractionCollection",
        "DockModelCollection",
        "Serializer",
        "ObjectSerializer",
        "RecordBuilder",
        "PathResolver",
        "ErrorHandler",
        "SupportsToDict",
        "SupportsAsDict",
        "SupportsJSON",
        "SupportsFileRecord",
        "ExportContextLike",
    ]
)

# =============================================================================
# End of Section 1
# =============================================================================
# Section 2 — Constants, formats and export conventions
# =============================================================================

# -----------------------------------------------------------------------------
# 2.1. Schema and encoding
# -----------------------------------------------------------------------------

EXPORT_SCHEMA_NAME: Final[str] = "dockanalyzer-export"
EXPORT_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_ENCODING: Final[str] = "utf-8"
DEFAULT_NEWLINE: Final[str] = "\n"
DEFAULT_FLOAT_PRECISION: Final[int] = 6
DEFAULT_JSON_INDENT: Final[int] = 2
DEFAULT_CSV_DELIMITER: Final[str] = ","
DEFAULT_TSV_DELIMITER: Final[str] = "\t"
DEFAULT_NULL_TEXT: Final[str] = ""
DEFAULT_BOOL_TRUE: Final[str] = "true"
DEFAULT_BOOL_FALSE: Final[str] = "false"

MIN_FLOAT_PRECISION: Final[int] = 0
MAX_FLOAT_PRECISION: Final[int] = 15
MAX_JSON_DEPTH: Final[int] = 128
MAX_EXCEL_ROWS: Final[int] = 1_048_576
MAX_EXCEL_COLUMNS: Final[int] = 16_384
MAX_EXCEL_SHEET_NAME: Final[int] = 31

# -----------------------------------------------------------------------------
# 2.2. Export formats
# -----------------------------------------------------------------------------

class ExportFormat(str, Enum):
    """Supported output formats."""

    JSON = "json"
    JSONL = "jsonl"
    CSV = "csv"
    TSV = "tsv"
    EXCEL = "xlsx"
    TEXT = "txt"


EXPORT_FORMAT_JSON: Final[str] = ExportFormat.JSON.value
EXPORT_FORMAT_JSONL: Final[str] = ExportFormat.JSONL.value
EXPORT_FORMAT_CSV: Final[str] = ExportFormat.CSV.value
EXPORT_FORMAT_TSV: Final[str] = ExportFormat.TSV.value
EXPORT_FORMAT_EXCEL: Final[str] = ExportFormat.EXCEL.value
EXPORT_FORMAT_TEXT: Final[str] = ExportFormat.TEXT.value

SUPPORTED_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    item.value for item in ExportFormat
)
TABULAR_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    {EXPORT_FORMAT_CSV, EXPORT_FORMAT_TSV, EXPORT_FORMAT_EXCEL}
)
TEXT_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    {
        EXPORT_FORMAT_JSON,
        EXPORT_FORMAT_JSONL,
        EXPORT_FORMAT_CSV,
        EXPORT_FORMAT_TSV,
        EXPORT_FORMAT_TEXT,
    }
)
BINARY_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    {EXPORT_FORMAT_EXCEL}
)

# -----------------------------------------------------------------------------
# 2.3. Extensions and aliases
# -----------------------------------------------------------------------------

FORMAT_EXTENSIONS: Final[Mapping[str, str]] = MappingProxyType(
    {
        EXPORT_FORMAT_JSON: ".json",
        EXPORT_FORMAT_JSONL: ".jsonl",
        EXPORT_FORMAT_CSV: ".csv",
        EXPORT_FORMAT_TSV: ".tsv",
        EXPORT_FORMAT_EXCEL: ".xlsx",
        EXPORT_FORMAT_TEXT: ".txt",
    }
)

EXTENSION_FORMATS: Final[Mapping[str, str]] = MappingProxyType(
    {extension: name for name, extension in FORMAT_EXTENSIONS.items()}
)

FORMAT_ALIASES: Final[Mapping[str, str]] = MappingProxyType(
    {
        "json": EXPORT_FORMAT_JSON,
        ".json": EXPORT_FORMAT_JSON,
        "application/json": EXPORT_FORMAT_JSON,
        "jsonl": EXPORT_FORMAT_JSONL,
        ".jsonl": EXPORT_FORMAT_JSONL,
        "ndjson": EXPORT_FORMAT_JSONL,
        ".ndjson": EXPORT_FORMAT_JSONL,
        "application/x-ndjson": EXPORT_FORMAT_JSONL,
        "csv": EXPORT_FORMAT_CSV,
        ".csv": EXPORT_FORMAT_CSV,
        "text/csv": EXPORT_FORMAT_CSV,
        "tsv": EXPORT_FORMAT_TSV,
        ".tsv": EXPORT_FORMAT_TSV,
        "tab": EXPORT_FORMAT_TSV,
        "text/tab-separated-values": EXPORT_FORMAT_TSV,
        "xlsx": EXPORT_FORMAT_EXCEL,
        ".xlsx": EXPORT_FORMAT_EXCEL,
        "excel": EXPORT_FORMAT_EXCEL,
        "spreadsheet": EXPORT_FORMAT_EXCEL,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": (
            EXPORT_FORMAT_EXCEL
        ),
        "txt": EXPORT_FORMAT_TEXT,
        ".txt": EXPORT_FORMAT_TEXT,
        "text": EXPORT_FORMAT_TEXT,
        "plain": EXPORT_FORMAT_TEXT,
        "text/plain": EXPORT_FORMAT_TEXT,
    }
)

# -----------------------------------------------------------------------------
# 2.4. Output modes
# -----------------------------------------------------------------------------

class ExportLayout(str, Enum):
    """File layout for multi-table exports."""

    SINGLE_FILE = "single_file"
    MULTIPLE_FILES = "multiple_files"
    PER_POSE = "per_pose"
    CONSOLIDATED = "consolidated"


class ExportDetail(str, Enum):
    """Payload detail level."""

    MINIMAL = "minimal"
    COMPACT = "compact"
    STANDARD = "standard"
    COMPLETE = "complete"


class OverwriteMode(str, Enum):
    """Existing-file behavior."""

    ERROR = "error"
    OVERWRITE = "overwrite"
    UNIQUE = "unique"
    BACKUP = "backup"


class ErrorMode(str, Enum):
    """Export error behavior."""

    RAISE = "raise"
    WARN = "warn"
    COLLECT = "collect"
    SKIP = "skip"


DEFAULT_EXPORT_LAYOUT: Final[str] = ExportLayout.SINGLE_FILE.value
DEFAULT_EXPORT_DETAIL: Final[str] = ExportDetail.STANDARD.value
DEFAULT_OVERWRITE_MODE: Final[str] = OverwriteMode.UNIQUE.value
DEFAULT_ERROR_MODE: Final[str] = ErrorMode.RAISE.value

SUPPORTED_EXPORT_LAYOUTS: Final[FrozenSet[str]] = frozenset(
    item.value for item in ExportLayout
)
SUPPORTED_EXPORT_DETAILS: Final[FrozenSet[str]] = frozenset(
    item.value for item in ExportDetail
)
SUPPORTED_OVERWRITE_MODES: Final[FrozenSet[str]] = frozenset(
    item.value for item in OverwriteMode
)
SUPPORTED_ERROR_MODES: Final[FrozenSet[str]] = frozenset(
    item.value for item in ErrorMode
)

# -----------------------------------------------------------------------------
# 2.5. Standard table names
# -----------------------------------------------------------------------------

TABLE_SUMMARY: Final[str] = "summary"
TABLE_POSES: Final[str] = "poses"
TABLE_INTERACTIONS: Final[str] = "interactions"
TABLE_CONTACTS: Final[str] = "contacts"
TABLE_HBONDS: Final[str] = "hydrogen_bonds"
TABLE_HYDROPHOBIC: Final[str] = "hydrophobic"
TABLE_PI: Final[str] = "pi_interactions"
TABLE_SALT_BRIDGES: Final[str] = "salt_bridges"
TABLE_RESIDUES: Final[str] = "residues"
TABLE_SCORES: Final[str] = "scores"
TABLE_SCORE_COMPONENTS: Final[str] = "score_components"
TABLE_RANKING: Final[str] = "ranking"
TABLE_CONSENSUS: Final[str] = "consensus"
TABLE_PERSISTENCE: Final[str] = "persistence"
TABLE_EXTERNAL_SCORES: Final[str] = "external_scores"
TABLE_EXPLAINABILITY: Final[str] = "explainability"
TABLE_METADATA: Final[str] = "metadata"
TABLE_FILES: Final[str] = "files"
TABLE_WARNINGS: Final[str] = "warnings"
TABLE_ERRORS: Final[str] = "errors"

STANDARD_TABLE_NAMES: Final[Tuple[str, ...]] = (
    TABLE_SUMMARY,
    TABLE_POSES,
    TABLE_INTERACTIONS,
    TABLE_CONTACTS,
    TABLE_HBONDS,
    TABLE_HYDROPHOBIC,
    TABLE_PI,
    TABLE_SALT_BRIDGES,
    TABLE_RESIDUES,
    TABLE_SCORES,
    TABLE_SCORE_COMPONENTS,
    TABLE_RANKING,
    TABLE_CONSENSUS,
    TABLE_PERSISTENCE,
    TABLE_EXTERNAL_SCORES,
    TABLE_EXPLAINABILITY,
    TABLE_METADATA,
    TABLE_FILES,
    TABLE_WARNINGS,
    TABLE_ERRORS,
)

# -----------------------------------------------------------------------------
# 2.6. Excel sheet names
# -----------------------------------------------------------------------------

EXCEL_SHEET_NAMES: Final[Mapping[str, str]] = MappingProxyType(
    {
        TABLE_SUMMARY: "Summary",
        TABLE_POSES: "Poses",
        TABLE_INTERACTIONS: "Interactions",
        TABLE_CONTACTS: "Contacts",
        TABLE_HBONDS: "Hydrogen Bonds",
        TABLE_HYDROPHOBIC: "Hydrophobic",
        TABLE_PI: "Pi Interactions",
        TABLE_SALT_BRIDGES: "Salt Bridges",
        TABLE_RESIDUES: "Residues",
        TABLE_SCORES: "Scores",
        TABLE_SCORE_COMPONENTS: "Score Components",
        TABLE_RANKING: "Ranking",
        TABLE_CONSENSUS: "Consensus",
        TABLE_PERSISTENCE: "Persistence",
        TABLE_EXTERNAL_SCORES: "External Scores",
        TABLE_EXPLAINABILITY: "Explainability",
        TABLE_METADATA: "Metadata",
        TABLE_FILES: "Files",
        TABLE_WARNINGS: "Warnings",
        TABLE_ERRORS: "Errors",
    }
)

DEFAULT_EXCEL_TABLE_STYLE: Final[str] = "TableStyleMedium2"
DEFAULT_EXCEL_FREEZE_PANES: Final[str] = "A2"
DEFAULT_EXCEL_HEADER_ROW: Final[int] = 1
DEFAULT_EXCEL_COLUMN_WIDTH: Final[float] = 12.0
MAX_AUTO_COLUMN_WIDTH: Final[float] = 60.0

# -----------------------------------------------------------------------------
# 2.7. Schema keys
# -----------------------------------------------------------------------------

KEY_SCHEMA_NAME: Final[str] = "schema_name"
KEY_SCHEMA_VERSION: Final[str] = "schema_version"
KEY_DOCKANALYZER_VERSION: Final[str] = "dockanalyzer_version"
KEY_GENERATED_AT: Final[str] = "generated_at"
KEY_PROVENANCE: Final[str] = "provenance"
KEY_METADATA: Final[str] = "metadata"
KEY_FILES: Final[str] = "files"
KEY_WARNINGS: Final[str] = "warnings"
KEY_ERRORS: Final[str] = "errors"
KEY_MODELS: Final[str] = "models"
KEY_INTERACTIONS: Final[str] = "interactions"
KEY_STATISTICS: Final[str] = "statistics"
KEY_SCORING: Final[str] = "scoring"
KEY_TABLES: Final[str] = "tables"
KEY_MANIFEST: Final[str] = "manifest"

REQUIRED_SCHEMA_KEYS: Final[Tuple[str, ...]] = (
    KEY_SCHEMA_NAME,
    KEY_SCHEMA_VERSION,
    KEY_GENERATED_AT,
)

# -----------------------------------------------------------------------------
# 2.8. Interaction families and aliases
# -----------------------------------------------------------------------------

INTERACTION_CONTACT: Final[str] = "contact"
INTERACTION_HBOND: Final[str] = "hbond"
INTERACTION_HYDROPHOBIC: Final[str] = "hydrophobic"
INTERACTION_PI: Final[str] = "pi"
INTERACTION_SALT_BRIDGE: Final[str] = "salt_bridge"
INTERACTION_CLASH: Final[str] = "clash"
INTERACTION_UNKNOWN: Final[str] = "unknown"

INTERACTION_FAMILIES: Final[FrozenSet[str]] = frozenset(
    {
        INTERACTION_CONTACT,
        INTERACTION_HBOND,
        INTERACTION_HYDROPHOBIC,
        INTERACTION_PI,
        INTERACTION_SALT_BRIDGE,
        INTERACTION_CLASH,
        INTERACTION_UNKNOWN,
    }
)

INTERACTION_FAMILY_ALIASES: Final[Mapping[str, str]] = MappingProxyType({
    "contact": "contact",
    "contacts": "contact",
    "generic_contact": "contact",
    "vdw": "contact",
    "van_der_waals": "contact",
    "hbond": "hbond",
    "hbonds": "hbond",
    "hydrogen_bond": "hbond",
    "hydrogen-bond": "hbond",
    "hydrophobic": "hydrophobic",
    "hydrophobe": "hydrophobic",
    "pi": "pi",
    "pi_interaction": "pi",
    "pi-pi": "pi",
    "cation-pi": "pi",
    "anion-pi": "pi",
    "amide-pi": "pi",
    "saltbridge": "saltbridge",
    "salt_bridge": "saltbridge",
    "salt-bridge": "saltbridge",
    "ionic": "saltbridge",
})

# -----------------------------------------------------------------------------
# 2.9. File naming
# -----------------------------------------------------------------------------

DEFAULT_EXPORT_BASENAME: Final[str] = "dockanalyzer_export"
DEFAULT_MANIFEST_BASENAME: Final[str] = "manifest"
DEFAULT_SUMMARY_BASENAME: Final[str] = "summary"
DEFAULT_POSE_PREFIX: Final[str] = "pose"
DEFAULT_BACKUP_SUFFIX: Final[str] = ".bak"
DEFAULT_TEMP_SUFFIX: Final[str] = ".tmp"

INVALID_FILENAME_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"[<>:\"/\\|?*\x00-\x1F]"
)
INVALID_SHEET_NAME_PATTERN: Final[re.Pattern[str]] = re.compile(r"[\\/*?:\[\]]")
WHITESPACE_PATTERN: Final[re.Pattern[str]] = re.compile(r"\s+")
MULTIPLE_UNDERSCORES_PATTERN: Final[re.Pattern[str]] = re.compile(r"_+")
WINDOWS_RESERVED_NAMES: Final[FrozenSet[str]] = frozenset(
    {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        *(f"COM{index}" for index in range(1, 10)),
        *(f"LPT{index}" for index in range(1, 10)),
    }
)

# -----------------------------------------------------------------------------
# 2.10. Delimited-file conventions
# -----------------------------------------------------------------------------

DELIMITER_BY_FORMAT: Final[Mapping[str, str]] = MappingProxyType(
    {
        EXPORT_FORMAT_CSV: DEFAULT_CSV_DELIMITER,
        EXPORT_FORMAT_TSV: DEFAULT_TSV_DELIMITER,
    }
)

DEFAULT_CSV_QUOTING: Final[int] = csv.QUOTE_MINIMAL
DEFAULT_CSV_QUOTECHAR: Final[str] = '"'
DEFAULT_CSV_ESCAPECHAR: Final[Optional[str]] = None
DEFAULT_CSV_DOUBLEQUOTE: Final[bool] = True
DEFAULT_CSV_LINETERMINATOR: Final[str] = DEFAULT_NEWLINE

# -----------------------------------------------------------------------------
# 2.11. JSON conventions
# -----------------------------------------------------------------------------

DEFAULT_JSON_ENSURE_ASCII: Final[bool] = False
DEFAULT_JSON_SORT_KEYS: Final[bool] = False
DEFAULT_JSON_ALLOW_NAN: Final[bool] = False
DEFAULT_JSON_COMPACT_SEPARATORS: Final[Tuple[str, str]] = (",", ":")
DEFAULT_JSON_PRETTY_SEPARATORS: Final[Tuple[str, str]] = (",", ": ")
DEFAULT_JSONL_INDENT: Final[None] = None

# -----------------------------------------------------------------------------
# 2.12. Compression and hashing
# -----------------------------------------------------------------------------

class CompressionFormat(str, Enum):
    """Supported stream compression."""

    NONE = "none"
    GZIP = "gzip"


DEFAULT_COMPRESSION: Final[str] = CompressionFormat.NONE.value
SUPPORTED_COMPRESSION_FORMATS: Final[FrozenSet[str]] = frozenset(
    item.value for item in CompressionFormat
)
COMPRESSED_SUFFIXES: Final[Mapping[str, str]] = MappingProxyType(
    {CompressionFormat.GZIP.value: ".gz"}
)
DEFAULT_HASH_ALGORITHM: Final[str] = "sha256"
DEFAULT_HASH_BLOCK_SIZE: Final[int] = 65_536

# -----------------------------------------------------------------------------
# 2.13. Export status
# -----------------------------------------------------------------------------

class ExportStatus(str, Enum):
    """Export operation status."""

    PENDING = "pending"
    SUCCESS = "success"
    PARTIAL = "partial"
    FAILED = "failed"
    SKIPPED = "skipped"


SUCCESS_EXPORT_STATUSES: Final[FrozenSet[str]] = frozenset(
    {ExportStatus.SUCCESS.value, ExportStatus.PARTIAL.value}
)
TERMINAL_EXPORT_STATUSES: Final[FrozenSet[str]] = frozenset(
    {
        ExportStatus.SUCCESS.value,
        ExportStatus.PARTIAL.value,
        ExportStatus.FAILED.value,
        ExportStatus.SKIPPED.value,
    }
)

# -----------------------------------------------------------------------------
# 2.14. Configuration lookup keys
# -----------------------------------------------------------------------------

CONFIG_DIRECTORY_KEYS: Final[Mapping[str, Tuple[str, ...]]] = MappingProxyType(
    {
        EXPORT_FORMAT_JSON: ("JSON_DIR", "JSON_OUTPUT_DIR", "OUTPUT_JSON_DIR"),
        EXPORT_FORMAT_JSONL: ("JSON_DIR", "JSON_OUTPUT_DIR", "OUTPUT_JSON_DIR"),
        EXPORT_FORMAT_CSV: ("CSV_DIR", "CSV_OUTPUT_DIR", "OUTPUT_CSV_DIR"),
        EXPORT_FORMAT_TSV: ("CSV_DIR", "CSV_OUTPUT_DIR", "OUTPUT_CSV_DIR"),
        EXPORT_FORMAT_EXCEL: (
            "EXCEL_DIR",
            "EXCEL_OUTPUT_DIR",
            "OUTPUT_EXCEL_DIR",
        ),
        EXPORT_FORMAT_TEXT: (
            "REPORT_DIR",
            "REPORTS_DIR",
            "TEXT_OUTPUT_DIR",
        ),
    }
)

CONFIG_ENABLE_KEYS: Final[Mapping[str, Tuple[str, ...]]] = MappingProxyType(
    {
        EXPORT_FORMAT_JSON: ("EXPORT_JSON", "ENABLE_JSON_EXPORT"),
        EXPORT_FORMAT_JSONL: ("EXPORT_JSON", "ENABLE_JSON_EXPORT"),
        EXPORT_FORMAT_CSV: ("EXPORT_CSV", "ENABLE_CSV_EXPORT"),
        EXPORT_FORMAT_TSV: ("EXPORT_CSV", "ENABLE_CSV_EXPORT"),
        EXPORT_FORMAT_EXCEL: ("EXPORT_EXCEL", "ENABLE_EXCEL_EXPORT"),
        EXPORT_FORMAT_TEXT: ("EXPORT_REPORT", "ENABLE_TEXT_EXPORT"),
    }
)

# -----------------------------------------------------------------------------
# 2.15. Default table order by detail level
# -----------------------------------------------------------------------------

TABLES_BY_DETAIL: Final[Mapping[str, Tuple[str, ...]]] = MappingProxyType(
    {
        ExportDetail.MINIMAL.value: (
            TABLE_SUMMARY,
            TABLE_POSES,
            TABLE_SCORES,
        ),
        ExportDetail.COMPACT.value: (
            TABLE_SUMMARY,
            TABLE_POSES,
            TABLE_INTERACTIONS,
            TABLE_RESIDUES,
            TABLE_SCORES,
            TABLE_RANKING,
        ),
        ExportDetail.STANDARD.value: (
            TABLE_SUMMARY,
            TABLE_POSES,
            TABLE_INTERACTIONS,
            TABLE_CONTACTS,
            TABLE_HBONDS,
            TABLE_HYDROPHOBIC,
            TABLE_PI,
            TABLE_SALT_BRIDGES,
            TABLE_RESIDUES,
            TABLE_SCORES,
            TABLE_SCORE_COMPONENTS,
            TABLE_RANKING,
            TABLE_CONSENSUS,
            TABLE_PERSISTENCE,
            TABLE_EXTERNAL_SCORES,
            TABLE_METADATA,
        ),
        ExportDetail.COMPLETE.value: STANDARD_TABLE_NAMES,
    }
)

# -----------------------------------------------------------------------------
# 2.16. Public registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "EXPORT_SCHEMA_NAME",
        "EXPORT_SCHEMA_VERSION",
        "DEFAULT_ENCODING",
        "DEFAULT_NEWLINE",
        "DEFAULT_FLOAT_PRECISION",
        "DEFAULT_JSON_INDENT",
        "ExportFormat",
        "ExportLayout",
        "ExportDetail",
        "OverwriteMode",
        "ErrorMode",
        "CompressionFormat",
        "ExportStatus",
        "SUPPORTED_EXPORT_FORMATS",
        "FORMAT_EXTENSIONS",
        "FORMAT_ALIASES",
        "STANDARD_TABLE_NAMES",
        "EXCEL_SHEET_NAMES",
        "INTERACTION_FAMILIES",
        "INTERACTION_FAMILY_ALIASES",
        "TABLES_BY_DETAIL",
    ]
)

# =============================================================================
# End of Section 2
# =============================================================================
# Section 3 — Export exceptions and warnings
# =============================================================================

# -----------------------------------------------------------------------------
# 3.1. Base exception
# -----------------------------------------------------------------------------


class ExportError(Exception):
    """Base exception for export failures."""

    default_code: ClassVar[str] = "export_error"

    def __init__(
        self,
        message: str,
        *,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        clean_message = str(message).strip() or self.__class__.__name__
        self.message = clean_message
        self.code = str(code or self.default_code)
        self.context: Dict[str, Any] = dict(context or {})
        self.cause = cause
        super().__init__(clean_message)

    def __str__(self) -> str:
        text = self.message
        if self.code:
            text = f"[{self.code}] {text}"
        if self.context:
            details = ", ".join(
                f"{key}={value!r}" for key, value in sorted(self.context.items())
            )
            text = f"{text} ({details})"
        return text

    def with_context(self, **context: Any) -> "ExportError":
        """Add context and return this exception."""
        self.context.update(context)
        return self

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable error record."""
        record: Dict[str, Any] = {
            "type": self.__class__.__name__,
            "code": self.code,
            "message": self.message,
            "context": dict(self.context),
        }
        if self.cause is not None:
            record["cause"] = {
                "type": self.cause.__class__.__name__,
                "message": str(self.cause),
            }
        return record


# -----------------------------------------------------------------------------
# 3.2. Input and configuration errors
# -----------------------------------------------------------------------------


class ExportInputError(ExportError):
    """Raised for unsupported or malformed input objects."""

    default_code = "invalid_input"


class ExportConfigurationError(ExportError):
    """Raised for invalid export options."""

    default_code = "invalid_configuration"


class ExportFormatError(ExportConfigurationError):
    """Raised for unsupported formats or format options."""

    default_code = "unsupported_format"


class ExportDependencyError(ExportConfigurationError):
    """Raised when an optional dependency is required."""

    default_code = "missing_dependency"

    def __init__(
        self,
        dependency: str,
        *,
        feature: Optional[str] = None,
        message: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        dependency_name = str(dependency).strip() or "unknown"
        details = dict(context or {})
        details.setdefault("dependency", dependency_name)
        if feature:
            details.setdefault("feature", str(feature))
        if message is None:
            message = f"Optional dependency {dependency_name!r} is required"
            if feature:
                message += f" for {feature}"
        super().__init__(
            message,
            context=details,
            cause=cause,
        )
        self.dependency = dependency_name
        self.feature = feature


# -----------------------------------------------------------------------------
# 3.3. Path and write errors
# -----------------------------------------------------------------------------


class ExportPathError(ExportError):
    """Raised for invalid or inaccessible output paths."""

    default_code = "invalid_path"

    def __init__(
        self,
        message: str,
        *,
        path: Optional[PathLike] = None,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        details = dict(context or {})
        if path is not None:
            details.setdefault("path", str(path))
        super().__init__(
            message,
            code=code,
            context=details,
            cause=cause,
        )
        self.path = Path(path) if path is not None else None


class ExportWriteError(ExportPathError):
    """Raised when writing or replacing an output file fails."""

    default_code = "write_failed"


# -----------------------------------------------------------------------------
# 3.4. Serialization and validation errors
# -----------------------------------------------------------------------------


class ExportSerializationError(ExportError):
    """Raised when an object cannot be serialized safely."""

    default_code = "serialization_failed"

    def __init__(
        self,
        message: str,
        *,
        object_type: Optional[Union[str, type]] = None,
        location: Optional[str] = None,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        details = dict(context or {})
        if object_type is not None:
            type_name = (
                object_type.__qualname__
                if isinstance(object_type, type)
                else str(object_type)
            )
            details.setdefault("object_type", type_name)
        if location:
            details.setdefault("location", str(location))
        super().__init__(
            message,
            code=code,
            context=details,
            cause=cause,
        )
        self.object_type = object_type
        self.location = location


class ExportValidationError(ExportError):
    """Raised when export data fail validation."""

    default_code = "validation_failed"

    def __init__(
        self,
        message: str,
        *,
        field_name: Optional[str] = None,
        value: Any = None,
        include_value: bool = False,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        details = dict(context or {})
        if field_name:
            details.setdefault("field", str(field_name))
        if include_value:
            details.setdefault("value", value)
        super().__init__(
            message,
            code=code,
            context=details,
            cause=cause,
        )
        self.field_name = field_name
        self.value = value


# -----------------------------------------------------------------------------
# 3.5. Batch errors
# -----------------------------------------------------------------------------


class ExportBatchError(ExportError):
    """Raised when one or more batch exports fail."""

    default_code = "batch_failed"

    def __init__(
        self,
        message: str,
        *,
        failures: Optional[Sequence[BaseException]] = None,
        completed: int = 0,
        total: Optional[int] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        self.failures: Tuple[BaseException, ...] = tuple(failures or ())
        self.completed = max(0, int(completed))
        self.total = None if total is None else max(0, int(total))
        details = dict(context or {})
        details.setdefault("completed", self.completed)
        details.setdefault("failure_count", len(self.failures))
        if self.total is not None:
            details.setdefault("total", self.total)
        super().__init__(message, context=details, cause=cause)

    def to_dict(self) -> Dict[str, Any]:
        """Return the batch error and nested failures."""
        record = super().to_dict()
        record["failures"] = [
            failure.to_dict()
            if isinstance(failure, ExportError)
            else {
                "type": failure.__class__.__name__,
                "message": str(failure),
            }
            for failure in self.failures
        ]
        return record


# -----------------------------------------------------------------------------
# 3.6. Warnings
# -----------------------------------------------------------------------------


class ExportWarning(UserWarning):
    """Base warning for recoverable export issues."""


class ExportLossyConversionWarning(ExportWarning):
    """Warn that conversion discarded or simplified data."""


class ExportDependencyWarning(ExportWarning):
    """Warn that an optional feature is unavailable."""


class ExportCompatibilityWarning(ExportWarning):
    """Warn about legacy or partially supported data."""


class ExportOverwriteWarning(ExportWarning):
    """Warn that an existing output may be replaced."""


# -----------------------------------------------------------------------------
# 3.7. Public registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "ExportError",
        "ExportInputError",
        "ExportConfigurationError",
        "ExportFormatError",
        "ExportDependencyError",
        "ExportPathError",
        "ExportWriteError",
        "ExportSerializationError",
        "ExportValidationError",
        "ExportBatchError",
        "ExportWarning",
        "ExportLossyConversionWarning",
        "ExportDependencyWarning",
        "ExportCompatibilityWarning",
        "ExportOverwriteWarning",
    ]
)

# =============================================================================
# End of Section 3
# =============================================================================
# Section 4 — Configuration and result dataclasses
# =============================================================================

# -----------------------------------------------------------------------------
# 4.1. Shared helpers
# -----------------------------------------------------------------------------


def _enum_value(value: Any) -> Any:
    """Return an enum value or the original object."""
    return value.value if isinstance(value, Enum) else value


def _string_tuple(values: Optional[Iterable[Any]]) -> Tuple[str, ...]:
    """Return unique non-empty strings preserving order."""
    if values is None:
        return ()
    output: List[str] = []
    seen: Set[str] = set()
    for value in values:
        text = str(_enum_value(value)).strip()
        if text and text not in seen:
            output.append(text)
            seen.add(text)
    return tuple(output)


def _path_or_none(value: Optional[PathLike]) -> Optional[Path]:
    """Convert a path-like value when present."""
    return None if value is None else Path(value).expanduser()


def _copy_mapping(value: Optional[Mapping[str, Any]]) -> Dict[str, Any]:
    """Return a mutable shallow mapping copy."""
    return dict(value or {})


def _copy_records(
    value: Optional[Iterable[Mapping[str, Any]]],
) -> List[Dict[str, Any]]:
    """Return mutable record copies."""
    return [dict(record) for record in (value or ())]


def _error_record(error: Any) -> Dict[str, Any]:
    """Convert an error-like object to a compact record."""
    if isinstance(error, ExportError):
        return error.to_dict()
    if isinstance(error, BaseException):
        return {
            "type": error.__class__.__name__,
            "message": str(error),
        }
    if isinstance(error, Mapping):
        return dict(error)
    return {"type": type(error).__name__, "message": str(error)}


def _warning_record(item: Any) -> Dict[str, Any]:
    """Convert a warning-like object to a compact record."""
    if isinstance(item, Mapping):
        return dict(item)
    if isinstance(item, Warning):
        return {
            "type": item.__class__.__name__,
            "message": str(item),
        }
    return {"type": "warning", "message": str(item)}


def _plain_field_get(
    value: Any,
    names: Sequence[str],
    default: Any = None,
) -> Any:
    """Read the first mapping key or non-None attribute."""
    if value is None:
        return default
    if isinstance(value, Mapping):
        for name in names:
            if name in value:
                return value[name]
        return default
    for name in names:
        try:
            candidate = getattr(value, name)
        except (AttributeError, TypeError, ValueError):
            continue
        if candidate is not None:
            return candidate
    return default


# -----------------------------------------------------------------------------
# 4.2. Format-specific options
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class JSONExportOptions:
    """Options for JSON and JSON Lines output."""

    indent: Optional[int] = DEFAULT_JSON_INDENT
    ensure_ascii: bool = DEFAULT_JSON_ENSURE_ASCII
    sort_keys: bool = DEFAULT_JSON_SORT_KEYS
    allow_nan: bool = DEFAULT_JSON_ALLOW_NAN
    compact: bool = False
    json_lines: bool = False
    append: bool = False
    separators: Optional[Tuple[str, str]] = None

    def __post_init__(self) -> None:
        if self.indent is not None and self.indent < 0:
            raise ExportConfigurationError("JSON indent cannot be negative")
        if self.json_lines:
            self.indent = DEFAULT_JSONL_INDENT
        if self.compact and self.separators is None:
            self.separators = DEFAULT_JSON_COMPACT_SEPARATORS
        elif self.separators is None:
            self.separators = DEFAULT_JSON_PRETTY_SEPARATORS
        if len(self.separators) != 2:
            raise ExportConfigurationError("JSON separators require two values")

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive JSON options."""
        return {
            "indent": self.indent,
            "ensure_ascii": self.ensure_ascii,
            "sort_keys": self.sort_keys,
            "allow_nan": self.allow_nan,
            "compact": self.compact,
            "json_lines": self.json_lines,
            "append": self.append,
            "separators": tuple(self.separators or ()),
        }


@dataclass(slots=True)
class DelimitedExportOptions:
    """Options for CSV and TSV output."""

    delimiter: str = DEFAULT_CSV_DELIMITER
    quotechar: str = DEFAULT_CSV_QUOTECHAR
    quoting: int = DEFAULT_CSV_QUOTING
    escapechar: Optional[str] = DEFAULT_CSV_ESCAPECHAR
    doublequote: bool = DEFAULT_CSV_DOUBLEQUOTE
    lineterminator: str = DEFAULT_CSV_LINETERMINATOR
    include_header: bool = True
    append: bool = False
    extras_action: Literal["raise", "ignore"] = "ignore"
    null_text: str = DEFAULT_NULL_TEXT
    true_text: str = DEFAULT_BOOL_TRUE
    false_text: str = DEFAULT_BOOL_FALSE

    def __post_init__(self) -> None:
        if len(self.delimiter) != 1:
            raise ExportConfigurationError("Delimiter must contain one character")
        if len(self.quotechar) != 1:
            raise ExportConfigurationError("Quote character must contain one character")
        if self.escapechar is not None and len(self.escapechar) != 1:
            raise ExportConfigurationError("Escape character must contain one character")
        if self.extras_action not in {"raise", "ignore"}:
            raise ExportConfigurationError("Invalid CSV extras action")
        if not self.lineterminator:
            raise ExportConfigurationError("Line terminator cannot be empty")

    @classmethod
    def for_format(cls, export_format: Union[str, ExportFormat]) -> "DelimitedExportOptions":
        """Create delimiter defaults for CSV or TSV."""
        value = str(_enum_value(export_format)).lower().lstrip(".")
        delimiter = DEFAULT_TSV_DELIMITER if value == EXPORT_FORMAT_TSV else DEFAULT_CSV_DELIMITER
        return cls(delimiter=delimiter)

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive delimited options."""
        return {
            "delimiter": self.delimiter,
            "quotechar": self.quotechar,
            "quoting": self.quoting,
            "escapechar": self.escapechar,
            "doublequote": self.doublequote,
            "lineterminator": self.lineterminator,
            "include_header": self.include_header,
            "append": self.append,
            "extras_action": self.extras_action,
            "null_text": self.null_text,
            "true_text": self.true_text,
            "false_text": self.false_text,
        }


@dataclass(slots=True)
class ExcelExportOptions:
    """Options for Excel workbook output."""

    engine: str = "openpyxl"
    include_empty_sheets: bool = False
    freeze_panes: Optional[str] = DEFAULT_EXCEL_FREEZE_PANES
    auto_filter: bool = True
    auto_width: bool = True
    default_width: float = DEFAULT_EXCEL_COLUMN_WIDTH
    max_width: float = MAX_AUTO_COLUMN_WIDTH
    table_style: Optional[str] = DEFAULT_EXCEL_TABLE_STYLE
    format_headers: bool = True
    bold_headers: bool = True
    wrap_text: bool = False
    include_index: bool = False
    create_tables: bool = True

    def __post_init__(self) -> None:
        self.engine = str(self.engine).strip().lower()
        if not self.engine:
            raise ExportConfigurationError("Excel engine cannot be empty")
        if self.default_width <= 0:
            raise ExportConfigurationError("Excel default width must be positive")
        if self.max_width < self.default_width:
            raise ExportConfigurationError(
                "Excel maximum width cannot be smaller than default width"
            )

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive Excel options."""
        return asdict(self)


# -----------------------------------------------------------------------------
# 4.3. General export options
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class ExportOptions:
    """Shared options for one export operation."""

    output_dir: Optional[PathLike] = None
    basename: str = DEFAULT_EXPORT_BASENAME
    formats: Tuple[str, ...] = (EXPORT_FORMAT_JSON,)
    layout: str = DEFAULT_EXPORT_LAYOUT
    detail: str = DEFAULT_EXPORT_DETAIL
    overwrite: str = DEFAULT_OVERWRITE_MODE
    error_mode: str = DEFAULT_ERROR_MODE
    encoding: str = DEFAULT_ENCODING
    newline: str = DEFAULT_NEWLINE
    float_precision: int = DEFAULT_FLOAT_PRECISION
    compression: str = DEFAULT_COMPRESSION
    hash_algorithm: Optional[str] = DEFAULT_HASH_ALGORITHM
    create_directories: bool = True
    atomic_write: bool = True
    backup_existing: bool = False
    include_metadata: bool = True
    include_manifest: bool = True
    include_provenance: bool = True
    include_empty_tables: bool = False
    include_raw_objects: bool = False
    update_model_files: bool = True
    preserve_previous_files: bool = True
    selected_tables: Tuple[str, ...] = ()
    excluded_tables: Tuple[str, ...] = ()
    json: JSONExportOptions = field(default_factory=JSONExportOptions)
    delimited: DelimitedExportOptions = field(default_factory=DelimitedExportOptions)
    excel: ExcelExportOptions = field(default_factory=ExcelExportOptions)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.output_dir = _path_or_none(self.output_dir)
        self.basename = str(self.basename).strip() or DEFAULT_EXPORT_BASENAME
        self.formats = _string_tuple(self.formats) or (EXPORT_FORMAT_JSON,)
        self.layout = str(_enum_value(self.layout)).strip().lower()
        self.detail = str(_enum_value(self.detail)).strip().lower()
        self.overwrite = str(_enum_value(self.overwrite)).strip().lower()
        self.error_mode = str(_enum_value(self.error_mode)).strip().lower()
        self.compression = str(_enum_value(self.compression)).strip().lower()
        self.selected_tables = _string_tuple(self.selected_tables)
        self.excluded_tables = _string_tuple(self.excluded_tables)
        self.metadata = _copy_mapping(self.metadata)
        if not self.encoding:
            raise ExportConfigurationError("Encoding cannot be empty")
        if not MIN_FLOAT_PRECISION <= self.float_precision <= MAX_FLOAT_PRECISION:
            raise ExportConfigurationError(
                f"Float precision must be between {MIN_FLOAT_PRECISION} and "
                f"{MAX_FLOAT_PRECISION}"
            )
        if self.layout not in SUPPORTED_EXPORT_LAYOUTS:
            raise ExportConfigurationError(
                f"Unsupported export layout: {self.layout!r}"
            )
        if self.detail not in SUPPORTED_EXPORT_DETAILS:
            raise ExportConfigurationError(
                f"Unsupported export detail: {self.detail!r}"
            )
        if self.overwrite not in SUPPORTED_OVERWRITE_MODES:
            raise ExportConfigurationError(
                f"Unsupported overwrite mode: {self.overwrite!r}"
            )
        if self.error_mode not in SUPPORTED_ERROR_MODES:
            raise ExportConfigurationError(
                f"Unsupported error mode: {self.error_mode!r}"
            )
        if self.compression not in SUPPORTED_COMPRESSION_FORMATS:
            raise ExportConfigurationError(
                f"Unsupported compression format: {self.compression!r}"
            )
        overlap = set(self.selected_tables) & set(self.excluded_tables)
        if overlap:
            raise ExportConfigurationError(
                f"Tables cannot be selected and excluded: {sorted(overlap)!r}"
            )

    def copy(self, **changes: Any) -> "ExportOptions":
        """Return a modified options copy."""
        return replace(self, **changes)

    def wants_format(self, export_format: Union[str, ExportFormat]) -> bool:
        """Return whether a format was requested."""
        value = str(_enum_value(export_format)).lower().lstrip(".")
        value = FORMAT_ALIASES.get(value, value)
        return value in self.formats

    def wants_table(self, table_name: str) -> bool:
        """Return whether a table should be included."""
        name = str(table_name).strip()
        if name in self.excluded_tables:
            return False
        return not self.selected_tables or name in self.selected_tables

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive general options."""
        return {
            "output_dir": str(self.output_dir) if self.output_dir else None,
            "basename": self.basename,
            "formats": list(self.formats),
            "layout": self.layout,
            "detail": self.detail,
            "overwrite": self.overwrite,
            "error_mode": self.error_mode,
            "encoding": self.encoding,
            "newline": self.newline,
            "float_precision": self.float_precision,
            "compression": self.compression,
            "hash_algorithm": self.hash_algorithm,
            "create_directories": self.create_directories,
            "atomic_write": self.atomic_write,
            "backup_existing": self.backup_existing,
            "include_metadata": self.include_metadata,
            "include_manifest": self.include_manifest,
            "include_provenance": self.include_provenance,
            "include_empty_tables": self.include_empty_tables,
            "include_raw_objects": self.include_raw_objects,
            "update_model_files": self.update_model_files,
            "preserve_previous_files": self.preserve_previous_files,
            "selected_tables": list(self.selected_tables),
            "excluded_tables": list(self.excluded_tables),
            "json": self.json.to_dict(),
            "delimited": self.delimited.to_dict(),
            "excel": self.excel.to_dict(),
            "metadata": dict(self.metadata),
        }


@dataclass(slots=True)
class BatchExportOptions:
    """Options for multi-model and batch exports."""

    export: ExportOptions = field(default_factory=ExportOptions)
    separate_pose_directories: bool = False
    consolidate_tables: bool = True
    continue_on_error: bool = True
    max_failures: Optional[int] = None
    write_batch_manifest: bool = True
    batch_name: Optional[str] = None
    pose_prefix: str = DEFAULT_POSE_PREFIX
    group_key: Optional[str] = None
    sort_key: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if self.max_failures is not None and self.max_failures < 0:
            raise ExportConfigurationError("Maximum failures cannot be negative")
        self.batch_name = (
            str(self.batch_name).strip() if self.batch_name is not None else None
        )
        self.pose_prefix = str(self.pose_prefix).strip() or DEFAULT_POSE_PREFIX
        self.group_key = str(self.group_key).strip() if self.group_key else None
        self.sort_key = str(self.sort_key).strip() if self.sort_key else None
        self.metadata = _copy_mapping(self.metadata)

    def copy(self, **changes: Any) -> "BatchExportOptions":
        """Return a modified batch options copy."""
        return replace(self, **changes)

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive batch options."""
        return {
            "export": self.export.to_dict(),
            "separate_pose_directories": self.separate_pose_directories,
            "consolidate_tables": self.consolidate_tables,
            "continue_on_error": self.continue_on_error,
            "max_failures": self.max_failures,
            "write_batch_manifest": self.write_batch_manifest,
            "batch_name": self.batch_name,
            "pose_prefix": self.pose_prefix,
            "group_key": self.group_key,
            "sort_key": self.sort_key,
            "metadata": dict(self.metadata),
        }


# -----------------------------------------------------------------------------
# 4.4. Exported file records
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class ExportedFile:
    """Metadata for one generated file."""

    path: PathLike
    format: str
    status: str = ExportStatus.SUCCESS.value
    table: Optional[str] = None
    sheet_names: Tuple[str, ...] = ()
    record_count: Optional[int] = None
    size_bytes: Optional[int] = None
    hash_value: Optional[str] = None
    hash_algorithm: Optional[str] = None
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.path = Path(self.path)
        self.format = str(_enum_value(self.format)).strip().lower().lstrip(".")
        self.format = FORMAT_ALIASES.get(self.format, self.format)
        self.status = str(_enum_value(self.status)).strip().lower()
        self.sheet_names = _string_tuple(self.sheet_names)
        self.metadata = _copy_mapping(self.metadata)
        if self.record_count is not None and self.record_count < 0:
            raise ExportValidationError(
                "Record count cannot be negative",
                field_name="record_count",
            )
        if self.size_bytes is not None and self.size_bytes < 0:
            raise ExportValidationError(
                "File size cannot be negative",
                field_name="size_bytes",
            )

    @property
    def exists(self) -> bool:
        """Return whether the file currently exists."""
        return self.path.is_file()

    @property
    def succeeded(self) -> bool:
        """Return whether the file has a success status."""
        return self.status in SUCCESS_EXPORT_STATUSES

    def refresh_size(self) -> Optional[int]:
        """Refresh the file size when available."""
        self.size_bytes = self.path.stat().st_size if self.exists else None
        return self.size_bytes

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable file record."""
        return {
            "path": str(self.path),
            "name": self.path.name,
            "format": self.format,
            "status": self.status,
            "table": self.table,
            "sheet_names": list(self.sheet_names),
            "record_count": self.record_count,
            "size_bytes": self.size_bytes,
            "hash": self.hash_value,
            "hash_algorithm": self.hash_algorithm,
            "created_at": self.created_at.isoformat(),
            "metadata": dict(self.metadata),
        }


# -----------------------------------------------------------------------------
# 4.5. Manifest and operation results
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class ExportManifest:
    """Manifest describing one export operation."""

    schema_name: str = EXPORT_SCHEMA_NAME
    schema_version: str = EXPORT_SCHEMA_VERSION
    dockanalyzer_version: str = __version__
    generated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    export_id: Optional[str] = None
    source_name: Optional[str] = None
    files: List[ExportedFile] = field(default_factory=list)
    options: Optional[ExportOptions] = None
    provenance: Dict[str, Any] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.files = list(self.files)
        self.provenance = _copy_mapping(self.provenance)
        self.metadata = _copy_mapping(self.metadata)
        self.warnings = [_warning_record(item) for item in self.warnings]
        self.errors = [_error_record(item) for item in self.errors]

    @property
    def file_count(self) -> int:
        """Return the number of registered files."""
        return len(self.files)

    @property
    def succeeded(self) -> bool:
        """Return whether the manifest contains no failed files or errors."""
        return not self.errors and all(item.succeeded for item in self.files)

    def add_file(self, exported_file: ExportedFile) -> ExportedFile:
        """Register and return an exported file."""
        self.files.append(exported_file)
        return exported_file

    def add_warning(self, item: Any) -> None:
        """Register a warning record."""
        self.warnings.append(_warning_record(item))

    def add_error(self, item: Any) -> None:
        """Register an error record."""
        self.errors.append(_error_record(item))

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable manifest."""
        return {
            KEY_SCHEMA_NAME: self.schema_name,
            KEY_SCHEMA_VERSION: self.schema_version,
            KEY_DOCKANALYZER_VERSION: self.dockanalyzer_version,
            KEY_GENERATED_AT: self.generated_at.isoformat(),
            "export_id": self.export_id,
            "source_name": self.source_name,
            KEY_FILES: [item.to_dict() for item in self.files],
            "options": self.options.to_dict() if self.options else None,
            KEY_PROVENANCE: dict(self.provenance),
            KEY_METADATA: dict(self.metadata),
            KEY_WARNINGS: list(self.warnings),
            KEY_ERRORS: list(self.errors),
        }


@dataclass(slots=True)
class ExportResult:
    """Result of exporting one payload or model."""

    status: str = ExportStatus.PENDING.value
    source_name: Optional[str] = None
    output_dir: Optional[PathLike] = None
    files: List[ExportedFile] = field(default_factory=list)
    manifest: Optional[ExportManifest] = None
    tables: Dict[str, List[Dict[str, Any]]] = field(default_factory=dict)
    payload: Optional[Any] = None
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)
    started_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    finished_at: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.status = str(_enum_value(self.status)).strip().lower()
        self.output_dir = _path_or_none(self.output_dir)
        self.files = list(self.files)
        self.tables = {
            str(name): _copy_records(records)
            for name, records in self.tables.items()
        }
        self.warnings = [_warning_record(item) for item in self.warnings]
        self.errors = [_error_record(item) for item in self.errors]
        self.metadata = _copy_mapping(self.metadata)

    @property
    def succeeded(self) -> bool:
        """Return whether the operation completed successfully."""
        return self.status in SUCCESS_EXPORT_STATUSES and not self.errors

    @property
    def failed(self) -> bool:
        """Return whether the operation failed."""
        return self.status == ExportStatus.FAILED.value

    @property
    def file_count(self) -> int:
        """Return the number of generated files."""
        return len(self.files)

    @property
    def duration_seconds(self) -> Optional[float]:
        """Return operation duration when finished."""
        if self.finished_at is None:
            return None
        return max(0.0, (self.finished_at - self.started_at).total_seconds())

    def add_file(self, exported_file: ExportedFile) -> ExportedFile:
        """Register and return an exported file."""
        self.files.append(exported_file)
        if self.manifest is not None:
            self.manifest.add_file(exported_file)
        return exported_file

    def add_warning(self, item: Any) -> None:
        """Register a warning."""
        record = _warning_record(item)
        self.warnings.append(record)
        if self.manifest is not None:
            self.manifest.warnings.append(dict(record))

    def add_error(self, item: Any) -> None:
        """Register an error."""
        record = _error_record(item)
        self.errors.append(record)
        if self.manifest is not None:
            self.manifest.errors.append(dict(record))

    def finish(self, status: Optional[Union[str, ExportStatus]] = None) -> "ExportResult":
        """Mark the operation as finished."""
        if status is None:
            status = (
                ExportStatus.FAILED.value
                if self.errors and not self.files
                else ExportStatus.PARTIAL.value
                if self.errors
                else ExportStatus.SUCCESS.value
            )
        self.status = str(_enum_value(status)).strip().lower()
        self.finished_at = datetime.now(timezone.utc)
        return self

    def to_dict(self, *, include_payload: bool = False) -> Dict[str, Any]:
        """Return a serializable result summary."""
        record: Dict[str, Any] = {
            "status": self.status,
            "source_name": self.source_name,
            "output_dir": str(self.output_dir) if self.output_dir else None,
            "files": [item.to_dict() for item in self.files],
            "manifest": self.manifest.to_dict() if self.manifest else None,
            "table_counts": {
                name: len(records) for name, records in self.tables.items()
            },
            "warnings": list(self.warnings),
            "errors": list(self.errors),
            "started_at": self.started_at.isoformat(),
            "finished_at": self.finished_at.isoformat() if self.finished_at else None,
            "duration_seconds": self.duration_seconds,
            "metadata": dict(self.metadata),
        }
        if include_payload:
            record["payload"] = self.payload
            record["tables"] = {
                name: [dict(item) for item in records]
                for name, records in self.tables.items()
            }
        return record


# -----------------------------------------------------------------------------
# 4.6. Batch item and batch result
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class BatchExportItem:
    """One model or payload within a batch export."""

    index: int
    source: Any
    name: Optional[str] = None
    group: Optional[str] = None
    output_dir: Optional[PathLike] = None
    result: Optional[ExportResult] = None
    status: str = ExportStatus.PENDING.value
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if self.index < 0:
            raise ExportValidationError(
                "Batch item index cannot be negative",
                field_name="index",
            )
        self.name = str(self.name).strip() if self.name is not None else None
        self.group = str(self.group).strip() if self.group is not None else None
        self.output_dir = _path_or_none(self.output_dir)
        self.status = str(_enum_value(self.status)).strip().lower()
        self.metadata = _copy_mapping(self.metadata)

    @property
    def succeeded(self) -> bool:
        """Return whether the item completed successfully."""
        return bool(self.result and self.result.succeeded)

    def attach_result(self, result: ExportResult) -> ExportResult:
        """Attach an export result and synchronize status."""
        self.result = result
        self.status = result.status
        return result

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable batch item."""
        return {
            "index": self.index,
            "name": self.name,
            "group": self.group,
            "output_dir": str(self.output_dir) if self.output_dir else None,
            "status": self.status,
            "result": self.result.to_dict() if self.result else None,
            "metadata": dict(self.metadata),
        }


@dataclass(slots=True)
class BatchExportResult:
    """Result of exporting multiple models or payloads."""

    status: str = ExportStatus.PENDING.value
    batch_name: Optional[str] = None
    output_dir: Optional[PathLike] = None
    items: List[BatchExportItem] = field(default_factory=list)
    files: List[ExportedFile] = field(default_factory=list)
    manifest: Optional[ExportManifest] = None
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)
    started_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    finished_at: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.status = str(_enum_value(self.status)).strip().lower()
        self.batch_name = (
            str(self.batch_name).strip() if self.batch_name is not None else None
        )
        self.output_dir = _path_or_none(self.output_dir)
        self.items = list(self.items)
        self.files = list(self.files)
        self.warnings = [_warning_record(item) for item in self.warnings]
        self.errors = [_error_record(item) for item in self.errors]
        self.metadata = _copy_mapping(self.metadata)

    @property
    def total_count(self) -> int:
        """Return total batch items."""
        return len(self.items)

    @property
    def success_count(self) -> int:
        """Return successful batch items."""
        return sum(item.succeeded for item in self.items)

    @property
    def failure_count(self) -> int:
        """Return failed or unsuccessful batch items."""
        return self.total_count - self.success_count

    @property
    def succeeded(self) -> bool:
        """Return whether all items completed successfully."""
        return (
            self.status in SUCCESS_EXPORT_STATUSES
            and not self.errors
            and self.failure_count == 0
        )

    @property
    def duration_seconds(self) -> Optional[float]:
        """Return batch duration when finished."""
        if self.finished_at is None:
            return None
        return max(0.0, (self.finished_at - self.started_at).total_seconds())

    def add_item(self, item: BatchExportItem) -> BatchExportItem:
        """Register and return a batch item."""
        self.items.append(item)
        return item

    def add_file(self, exported_file: ExportedFile) -> ExportedFile:
        """Register and return a batch-level file."""
        self.files.append(exported_file)
        if self.manifest is not None:
            self.manifest.add_file(exported_file)
        return exported_file

    def add_warning(self, item: Any) -> None:
        """Register a warning."""
        self.warnings.append(_warning_record(item))

    def add_error(self, item: Any) -> None:
        """Register an error."""
        self.errors.append(_error_record(item))

    def finish(self) -> "BatchExportResult":
        """Finalize aggregate status and timing."""
        failed = self.failure_count
        if self.errors and not self.items:
            self.status = ExportStatus.FAILED.value
        elif failed == 0 and not self.errors:
            self.status = ExportStatus.SUCCESS.value
        elif self.success_count > 0:
            self.status = ExportStatus.PARTIAL.value
        else:
            self.status = ExportStatus.FAILED.value
        self.finished_at = datetime.now(timezone.utc)
        return self

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable batch summary."""
        return {
            "status": self.status,
            "batch_name": self.batch_name,
            "output_dir": str(self.output_dir) if self.output_dir else None,
            "counts": {
                "total": self.total_count,
                "success": self.success_count,
                "failure": self.failure_count,
            },
            "items": [item.to_dict() for item in self.items],
            "files": [item.to_dict() for item in self.files],
            "manifest": self.manifest.to_dict() if self.manifest else None,
            "warnings": list(self.warnings),
            "errors": list(self.errors),
            "started_at": self.started_at.isoformat(),
            "finished_at": self.finished_at.isoformat() if self.finished_at else None,
            "duration_seconds": self.duration_seconds,
            "metadata": dict(self.metadata),
        }


# -----------------------------------------------------------------------------
# 4.7. Public registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "JSONExportOptions",
        "DelimitedExportOptions",
        "ExcelExportOptions",
        "ExportOptions",
        "BatchExportOptions",
        "ExportedFile",
        "ExportManifest",
        "ExportResult",
        "BatchExportItem",
        "BatchExportResult",
    ]
)

# =============================================================================
# End of Section 4
# =============================================================================
# Section 5 — Paths and file names
# =============================================================================


def normalize_export_format(value: Any, *, strict: bool = True) -> str:
    """Return a canonical export format."""
    if isinstance(value, ExportFormat):
        return value.value
    text = str(value or "").strip().lower()
    if not text:
        if strict:
            raise ExportFormatError("Export format cannot be empty.")
        return ""
    text = text.removeprefix(".")
    normalized = FORMAT_ALIASES.get(text, text)
    if normalized not in SUPPORTED_EXPORT_FORMATS:
        if strict:
            raise ExportFormatError(
                f"Unsupported export format: {value!r}.",
                context={"supported": sorted(SUPPORTED_EXPORT_FORMATS)},
            )
        return normalized
    return normalized


def normalize_compression(value: Any, *, strict: bool = True) -> str:
    """Return a canonical compression format."""
    if isinstance(value, CompressionFormat):
        return value.value
    text = str(value or DEFAULT_COMPRESSION).strip().lower().removeprefix(".")
    aliases = {"": "none", "off": "none", "false": "none", "gzip": "gz"}
    text = aliases.get(text, text)
    if text not in SUPPORTED_COMPRESSION_FORMATS and strict:
        raise ExportFormatError(
            f"Unsupported compression format: {value!r}.",
            context={"supported": sorted(SUPPORTED_COMPRESSION_FORMATS)},
        )
    return text


def canonical_extension(format_name: Any, *, include_dot: bool = True) -> str:
    """Return the canonical extension for a format."""
    normalized = normalize_export_format(format_name)
    extension = FORMAT_EXTENSIONS[normalized]
    return extension if include_dot else extension.lstrip(".")


def split_compression_suffix(path: PathLike) -> Tuple[Path, str]:
    """Split a known compression suffix from a path."""
    value = Path(path)
    suffix = value.suffix.lower()
    reverse = {item: key for key, item in COMPRESSED_SUFFIXES.items() if item}
    if suffix in reverse:
        return value.with_suffix(""), reverse[suffix]
    return value, CompressionFormat.NONE.value


def detect_export_format(path: PathLike, *, strict: bool = True) -> str:
    """Infer an export format from a file name."""
    base_path, _ = split_compression_suffix(path)
    suffix = base_path.suffix.lower()
    format_name = EXTENSION_FORMATS.get(suffix)
    if format_name is None and strict:
        raise ExportFormatError(
            f"Cannot infer export format from: {path!s}.",
            context={"path": str(path)},
        )
    return format_name or ""


def normalize_file_extension(
    path: PathLike,
    format_name: Any,
    *,
    compression: Any = DEFAULT_COMPRESSION,
    replace: bool = True,
) -> Path:
    """Apply canonical data and compression extensions."""
    value = Path(path)
    normalized_format = normalize_export_format(format_name)
    normalized_compression = normalize_compression(compression)
    base, existing_compression = split_compression_suffix(value)
    expected = canonical_extension(normalized_format)
    if base.suffix.lower() != expected:
        base = base.with_suffix(expected) if replace and base.suffix else Path(f"{base}{expected}")
    suffix = COMPRESSED_SUFFIXES.get(normalized_compression, "")
    if suffix:
        return Path(f"{base}{suffix}")
    if existing_compression != CompressionFormat.NONE.value:
        return base
    return base


def _clean_name_text(value: Any) -> str:
    text = str(value or "").strip()
    text = WHITESPACE_PATTERN.sub("_", text)
    text = MULTIPLE_UNDERSCORES_PATTERN.sub("_", text)
    return text


def sanitize_filename(
    value: Any,
    *,
    fallback: str = DEFAULT_EXPORT_BASENAME,
    preserve_extension: bool = True,
    max_length: int = 240,
) -> str:
    """Return a filesystem-safe file name."""
    raw = Path(str(value or "")).name
    suffixes = "".join(Path(raw).suffixes) if preserve_extension else ""
    stem = raw[: -len(suffixes)] if suffixes else raw
    stem = INVALID_FILENAME_PATTERN.sub("_", _clean_name_text(stem))
    stem = stem.strip(" ._-") or fallback
    if stem.upper() in WINDOWS_RESERVED_NAMES:
        stem = f"_{stem}"
    suffixes = INVALID_FILENAME_PATTERN.sub("", suffixes)
    allowed = max(1, max_length - len(suffixes))
    stem = stem[:allowed].rstrip(" ._-") or fallback[:allowed]
    return f"{stem}{suffixes}"


def sanitize_stem(value: Any, *, fallback: str = DEFAULT_EXPORT_BASENAME) -> str:
    """Return a safe file stem without extensions."""
    return sanitize_filename(value, fallback=fallback, preserve_extension=False)


def join_filename_parts(
    *parts: Any,
    separator: str = "_",
    fallback: str = DEFAULT_EXPORT_BASENAME,
) -> str:
    """Join non-empty values into a safe file stem."""
    cleaned = [sanitize_stem(item, fallback="") for item in parts if str(item or "").strip()]
    cleaned = [item for item in cleaned if item]
    return separator.join(cleaned) or fallback


def build_export_filename(
    basename: Any = DEFAULT_EXPORT_BASENAME,
    *,
    format_name: Any,
    table: Optional[str] = None,
    pose: Optional[Any] = None,
    suffix: Optional[Any] = None,
    compression: Any = DEFAULT_COMPRESSION,
) -> str:
    """Build a canonical export file name."""
    parts: List[Any] = [basename]
    if pose is not None:
        parts.extend((DEFAULT_POSE_PREFIX, pose))
    if table:
        parts.append(table)
    if suffix is not None:
        parts.append(suffix)
    stem = join_filename_parts(*parts)
    return normalize_file_extension(
        stem,
        format_name,
        compression=compression,
        replace=True,
    ).name


def expand_path(path: PathLike) -> Path:
    """Expand user and environment variables in a path."""
    return Path(os.path.expandvars(os.path.expanduser(str(path))))


def absolute_path(path: PathLike, *, base_dir: Optional[PathLike] = None) -> Path:
    """Return an absolute path without requiring it to exist."""
    value = expand_path(path)
    if not value.is_absolute():
        value = expand_path(base_dir or Path.cwd()) / value
    return value.resolve(strict=False)


def _read_config_value(names: Iterable[str], default: Any = None) -> Any:
    for name in names:
        if hasattr(config, name):
            value = getattr(config, name)
            if value is not None:
                return value
    return default


def configured_output_directory(format_name: Any) -> Optional[Path]:
    """Return a configured output directory for a format."""
    normalized = normalize_export_format(format_name)
    names = CONFIG_DIRECTORY_KEYS.get(normalized, ())
    value = _read_config_value(names)
    return expand_path(value) if value else None


def is_format_enabled(format_name: Any, *, default: bool = True) -> bool:
    """Return whether configuration enables an export format."""
    normalized = normalize_export_format(format_name)
    names = CONFIG_ENABLE_KEYS.get(normalized, ())
    value = _read_config_value(names, default)
    if isinstance(value, str):
        return value.strip().lower() not in {"0", "false", "no", "off"}
    return bool(value)


def resolve_output_directory(
    output_dir: Optional[PathLike] = None,
    *,
    format_name: Optional[Any] = None,
    create: bool = False,
) -> Path:
    """Resolve the directory used for export."""
    if output_dir is not None:
        result = absolute_path(output_dir)
    elif format_name is not None:
        configured = configured_output_directory(format_name)
        result = absolute_path(configured or Path.cwd())
    else:
        result = absolute_path(Path.cwd())
    if result.exists() and not result.is_dir():
        raise ExportPathError("Output directory is not a directory.", path=result)
    if create:
        ensure_output_directory(result)
    return result


def ensure_output_directory(path: PathLike, *, parents: bool = True) -> Path:
    """Create and return an output directory."""
    value = absolute_path(path)
    try:
        value.mkdir(parents=parents, exist_ok=True)
    except OSError as exc:
        raise ExportPathError("Cannot create output directory.", path=value, cause=exc) from exc
    if not value.is_dir():
        raise ExportPathError("Output path is not a directory.", path=value)
    return value


def ensure_parent_directory(path: PathLike) -> Path:
    """Create and return a file's parent directory."""
    value = absolute_path(path)
    ensure_output_directory(value.parent)
    return value.parent


def validate_output_path(
    path: PathLike,
    *,
    expect_directory: bool = False,
    require_parent: bool = False,
) -> Path:
    """Validate a prospective output path."""
    value = absolute_path(path)
    if expect_directory:
        if value.exists() and not value.is_dir():
            raise ExportPathError("Expected a directory path.", path=value)
        return value
    if value.exists() and value.is_dir():
        raise ExportPathError("Expected a file path.", path=value)
    if require_parent and not value.parent.is_dir():
        raise ExportPathError("Parent directory does not exist.", path=value.parent)
    return value


def path_is_writable(path: PathLike) -> bool:
    """Return whether a path can be written."""
    value = absolute_path(path)
    target = value if value.is_dir() else value.parent
    while not target.exists() and target != target.parent:
        target = target.parent
    return target.is_dir() and os.access(target, os.W_OK)


def resolve_unique_path(path: PathLike, *, start: int = 2) -> Path:
    """Return a non-existing path by adding a numeric suffix."""
    value = absolute_path(path)
    if not value.exists():
        return value
    base, compression = split_compression_suffix(value)
    data_suffix = base.suffix
    stem = base.stem
    compression_suffix = COMPRESSED_SUFFIXES.get(compression, "")
    index = max(1, int(start))
    while True:
        candidate = base.with_name(f"{stem}_{index}{data_suffix}")
        candidate = Path(f"{candidate}{compression_suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def backup_path(path: PathLike, *, suffix: str = DEFAULT_BACKUP_SUFFIX) -> Path:
    """Return an available backup path."""
    value = absolute_path(path)
    candidate = Path(f"{value}{suffix}")
    return resolve_unique_path(candidate)


def resolve_overwrite_path(path: PathLike, mode: Any = DEFAULT_OVERWRITE_MODE) -> Path:
    """Apply overwrite policy to a path."""
    value = absolute_path(path)
    normalized = mode.value if isinstance(mode, OverwriteMode) else str(mode).strip().lower()
    if normalized not in SUPPORTED_OVERWRITE_MODES:
        raise ExportConfigurationError(f"Unsupported overwrite mode: {mode!r}.")
    if not value.exists() or normalized == OverwriteMode.OVERWRITE.value:
        return value
    if normalized == OverwriteMode.UNIQUE.value:
        return resolve_unique_path(value)
    if normalized == OverwriteMode.ERROR.value:
        raise ExportPathError("Output file already exists.", path=value)
    if normalized == OverwriteMode.BACKUP.value:
        shutil.copy2(value, backup_path(value))
        return value
    return value


def resolve_output_path(
    path: Optional[PathLike] = None,
    *,
    output_dir: Optional[PathLike] = None,
    basename: Any = DEFAULT_EXPORT_BASENAME,
    format_name: Any,
    table: Optional[str] = None,
    pose: Optional[Any] = None,
    suffix: Optional[Any] = None,
    compression: Any = DEFAULT_COMPRESSION,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    create_directory: bool = True,
) -> Path:
    """Resolve a complete export path."""
    normalized_format = normalize_export_format(format_name)
    if path is None:
        directory = resolve_output_directory(
            output_dir,
            format_name=normalized_format,
            create=create_directory,
        )
        filename = build_export_filename(
            basename,
            format_name=normalized_format,
            table=table,
            pose=pose,
            suffix=suffix,
            compression=compression,
        )
        value = directory / filename
    else:
        value = absolute_path(path, base_dir=output_dir)
        value = normalize_file_extension(
            value,
            normalized_format,
            compression=compression,
        )
        if create_directory:
            ensure_parent_directory(value)
    validate_output_path(value)
    if not path_is_writable(value):
        raise ExportPathError("Output path is not writable.", path=value)
    return resolve_overwrite_path(value, overwrite)


def relative_export_path(path: PathLike, root: PathLike) -> str:
    """Return a portable relative path when possible."""
    value = absolute_path(path)
    base = absolute_path(root)
    try:
        return value.relative_to(base).as_posix()
    except ValueError:
        return value.as_posix()


def create_export_directory_tree(
    root: PathLike,
    *,
    formats: Iterable[Any] = SUPPORTED_EXPORT_FORMATS,
) -> Dict[str, Path]:
    """Create standard format directories."""
    root_path = ensure_output_directory(root)
    result: Dict[str, Path] = {"root": root_path}
    labels = {
        EXPORT_FORMAT_JSON: "JSON",
        EXPORT_FORMAT_JSONL: "JSON",
        EXPORT_FORMAT_CSV: "CSV",
        EXPORT_FORMAT_TSV: "CSV",
        EXPORT_FORMAT_EXCEL: "Excel",
        EXPORT_FORMAT_TEXT: "Text",
    }
    for value in formats:
        format_name = normalize_export_format(value)
        key = labels[format_name]
        result[format_name] = ensure_output_directory(root_path / key)
    return result


__all__.extend(
    [
        "normalize_export_format",
        "normalize_compression",
        "canonical_extension",
        "split_compression_suffix",
        "detect_export_format",
        "normalize_file_extension",
        "sanitize_filename",
        "sanitize_stem",
        "sanitize_sheet_name",
        "join_filename_parts",
        "build_export_filename",
        "expand_path",
        "absolute_path",
        "configured_output_directory",
        "is_format_enabled",
        "resolve_output_directory",
        "ensure_output_directory",
        "ensure_parent_directory",
        "validate_output_path",
        "path_is_writable",
        "resolve_unique_path",
        "backup_path",
        "resolve_overwrite_path",
        "resolve_output_path",
        "relative_export_path",
        "create_export_directory_tree",
    ]
)

# =============================================================================
# End of Section 5
# =============================================================================
# Section 6 — Generic conversion to serializable types
# =============================================================================


_SERIALIZATION_SKIP = object()


class UnknownObjectMode(str, Enum):
    """Fallback behavior for unsupported objects."""

    ERROR = "error"
    STRING = "string"
    REPR = "repr"
    DICT = "dict"
    TYPE = "type"
    NONE = "none"


SERIALIZATION_UNKNOWN_MODES: Final[FrozenSet[str]] = frozenset(
    item.value for item in UnknownObjectMode
)


@dataclass
class SerializationOptions:
    """Control recursive conversion to portable Python values."""

    max_depth: int = MAX_JSON_DEPTH
    float_precision: Optional[int] = DEFAULT_FLOAT_PRECISION
    allow_nan: bool = DEFAULT_JSON_ALLOW_NAN
    sort_sets: bool = True
    stringify_mapping_keys: bool = True
    omit_none: bool = False
    omit_private: bool = True
    include_type: bool = False
    include_module: bool = False
    use_to_dict: bool = True
    use_as_dict: bool = True
    use_json_method: bool = True
    unknown: str = UnknownObjectMode.STRING.value
    circular_value: Any = "<circular-reference>"
    depth_value: Any = "<maximum-depth>"
    bytes_encoding: str = DEFAULT_ENCODING
    bytes_errors: str = "replace"
    path_mode: str = "string"
    datetime_mode: str = "iso"
    timedelta_mode: str = "seconds"
    enum_mode: str = "value"
    complex_mode: str = "mapping"
    dataframe_mode: str = "records"
    ndarray_mode: str = "list"
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.max_depth = int(self.max_depth)
        if self.max_depth < 1:
            raise ExportConfigurationError("max_depth must be positive.")
        if self.float_precision is not None:
            self.float_precision = int(self.float_precision)
            if not MIN_FLOAT_PRECISION <= self.float_precision <= MAX_FLOAT_PRECISION:
                raise ExportConfigurationError("float_precision is outside supported limits.")
        self.unknown = str(self.unknown).strip().lower()
        if self.unknown not in SERIALIZATION_UNKNOWN_MODES:
            raise ExportConfigurationError(f"Unsupported unknown-object mode: {self.unknown!r}.")
        self.metadata = dict(self.metadata)


@dataclass
class SerializationState:
    """Track recursion, paths and conversion warnings."""

    options: SerializationOptions = field(default_factory=SerializationOptions)
    active_ids: Set[int] = field(default_factory=set)
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    converted: int = 0
    circular_references: int = 0
    truncated_values: int = 0

    def warn(self, message: str, *, path: str = "$", value: Any = None) -> None:
        record = {"message": message, "path": path}
        if value is not None:
            record["type"] = qualified_type_name(value)
        self.warnings.append(record)


_SIMPLE_SERIALIZABLE = (str, int, bool, type(None))


def qualified_type_name(value: Any) -> str:
    """Return an object's qualified type name."""
    cls = value if isinstance(value, type) else type(value)
    module = getattr(cls, "__module__", "")
    name = getattr(cls, "__qualname__", getattr(cls, "__name__", str(cls)))
    return f"{module}.{name}" if module and module != "builtins" else name


def is_finite_number(value: Any) -> bool:
    """Return whether a numeric value is finite."""
    if isinstance(value, bool):
        return True
    try:
        return math.isfinite(float(value))
    except (TypeError, ValueError, OverflowError):
        return False


def serialize_float(
    value: Any,
    *,
    precision: Optional[int] = DEFAULT_FLOAT_PRECISION,
    allow_nan: bool = DEFAULT_JSON_ALLOW_NAN,
) -> Optional[float]:
    """Convert a value to a finite JSON-compatible float."""
    result = float(value)
    if not math.isfinite(result):
        if allow_nan:
            return result
        return None
    if precision is not None:
        result = round(result, int(precision))
        if result == 0:
            result = 0.0
    return result


def serialize_bytes(value: Any, options: SerializationOptions) -> str:
    """Decode bytes using configured text settings."""
    raw = bytes(value)
    return raw.decode(options.bytes_encoding, errors=options.bytes_errors)


def serialize_datetime(value: Any, options: SerializationOptions) -> Any:
    """Convert date and time values."""
    mode = options.datetime_mode
    if mode == "timestamp" and isinstance(value, datetime):
        return value.timestamp()
    if mode == "string":
        return str(value)
    if isinstance(value, datetime) and value.tzinfo is None:
        return value.isoformat()
    return value.isoformat()


def serialize_timedelta(value: timedelta, options: SerializationOptions) -> Any:
    """Convert a timedelta."""
    if options.timedelta_mode == "string":
        return str(value)
    if options.timedelta_mode == "mapping":
        return {
            "days": value.days,
            "seconds": value.seconds,
            "microseconds": value.microseconds,
            "total_seconds": value.total_seconds(),
        }
    return value.total_seconds()


def serialize_enum(value: Enum, options: SerializationOptions) -> Any:
    """Convert an enum according to configured mode."""
    if options.enum_mode == "name":
        return value.name
    if options.enum_mode == "mapping":
        return {"name": value.name, "value": value.value}
    return value.value


def serialize_complex(value: complex, options: SerializationOptions) -> Any:
    """Convert a complex number."""
    if options.complex_mode == "string":
        return str(value)
    if options.complex_mode == "list":
        return [value.real, value.imag]
    return {"real": value.real, "imag": value.imag}


def serialize_path(value: PathLike, options: SerializationOptions) -> Any:
    """Convert a path."""
    path = Path(value)
    if options.path_mode == "posix":
        return path.as_posix()
    if options.path_mode == "absolute":
        return str(path.resolve(strict=False))
    if options.path_mode == "mapping":
        return {
            "path": str(path),
            "name": path.name,
            "suffix": path.suffix,
            "absolute": path.is_absolute(),
        }
    return str(path)


def _serialize_decimal(value: Any, options: SerializationOptions) -> Any:
    try:
        result = float(value)
    except (TypeError, ValueError, OverflowError):
        return str(value)
    return serialize_float(
        result,
        precision=options.float_precision,
        allow_nan=options.allow_nan,
    )


def _serialize_fraction(value: Any) -> Dict[str, int]:
    return {"numerator": int(value.numerator), "denominator": int(value.denominator)}


def _serialize_uuid(value: Any) -> str:
    return str(value)


def _serialize_range(value: range) -> Dict[str, int]:
    return {"start": value.start, "stop": value.stop, "step": value.step}


def _serialize_slice(value: slice) -> Dict[str, Optional[int]]:
    return {"start": value.start, "stop": value.stop, "step": value.step}


def _safe_method_call(value: Any, method_name: str) -> Any:
    method = getattr(value, method_name, None)
    if not callable(method):
        return _SERIALIZATION_SKIP
    try:
        return method()
    except TypeError:
        return _SERIALIZATION_SKIP
    except Exception as exc:
        raise ExportSerializationError(
            f"{method_name}() failed for {qualified_type_name(value)}.",
            cause=exc,
        ) from exc


def _iter_object_attributes(value: Any, options: SerializationOptions) -> Dict[str, Any]:
    try:
        source = vars(value)
    except TypeError:
        return {}
    result: Dict[str, Any] = {}
    for key, item in source.items():
        name = str(key)
        if options.omit_private and name.startswith("_"):
            continue
        if callable(item):
            continue
        result[name] = item
    return result


def _mapping_key(value: Any, options: SerializationOptions) -> Any:
    if isinstance(value, str):
        return value
    if isinstance(value, Enum):
        value = serialize_enum(value, options)
    elif isinstance(value, Path):
        value = serialize_path(value, options)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return str(value) if options.stringify_mapping_keys else value
    if options.stringify_mapping_keys:
        return str(value)
    raise ExportSerializationError(
        f"Mapping key is not serializable: {qualified_type_name(value)}."
    )


def _stable_sort_key(value: Any) -> Tuple[str, str]:
    return qualified_type_name(value), repr(value)


def serialize_mapping(
    value: Mapping[Any, Any],
    *,
    state: Optional[SerializationState] = None,
    path: str = "$",
    depth: int = 0,
) -> Dict[str, Any]:
    """Recursively convert a mapping."""
    active_state = state or SerializationState()
    result: Dict[str, Any] = {}
    for key, item in value.items():
        converted_key = _mapping_key(key, active_state.options)
        if active_state.options.omit_none and item is None:
            continue
        child_path = f"{path}.{converted_key}"
        result[str(converted_key)] = _to_serializable(item, active_state, child_path, depth + 1)
    return result


def serialize_sequence(
    value: Iterable[Any],
    *,
    state: Optional[SerializationState] = None,
    path: str = "$",
    depth: int = 0,
) -> List[Any]:
    """Recursively convert an iterable."""
    active_state = state or SerializationState()
    source = value
    if isinstance(value, (set, frozenset)) and active_state.options.sort_sets:
        source = sorted(value, key=_stable_sort_key)
    result: List[Any] = []
    for index, item in enumerate(source):
        if active_state.options.omit_none and item is None:
            continue
        result.append(_to_serializable(item, active_state, f"{path}[{index}]", depth + 1))
    return result


def serialize_dataclass(
    value: Any,
    *,
    state: Optional[SerializationState] = None,
    path: str = "$",
    depth: int = 0,
) -> Dict[str, Any]:
    """Convert a dataclass without deep-copying it first."""
    active_state = state or SerializationState()
    result: Dict[str, Any] = {}
    if active_state.options.include_type:
        result["__type__"] = type(value).__qualname__
    if active_state.options.include_module:
        result["__module__"] = type(value).__module__
    for item in fields(value):
        name = item.name
        if active_state.options.omit_private and name.startswith("_"):
            continue
        raw = getattr(value, name)
        if active_state.options.omit_none and raw is None:
            continue
        result[name] = _to_serializable(raw, active_state, f"{path}.{name}", depth + 1)
    return result


def serialize_numpy_value(
    value: Any,
    *,
    state: Optional[SerializationState] = None,
    path: str = "$",
    depth: int = 0,
) -> Any:
    """Convert a NumPy scalar, array or masked value."""
    active_state = state or SerializationState()
    if not NUMPY_AVAILABLE:
        return _SERIALIZATION_SKIP
    if value is np.ma.masked:
        return None
    if isinstance(value, np.generic):
        return _to_serializable(value.item(), active_state, path, depth + 1)
    if isinstance(value, np.ndarray):
        if active_state.options.ndarray_mode == "mapping":
            return {
                "dtype": str(value.dtype),
                "shape": list(value.shape),
                "data": _to_serializable(value.tolist(), active_state, f"{path}.data", depth + 1),
            }
        return _to_serializable(value.tolist(), active_state, path, depth + 1)
    return _SERIALIZATION_SKIP


def serialize_pandas_value(
    value: Any,
    *,
    state: Optional[SerializationState] = None,
    path: str = "$",
    depth: int = 0,
) -> Any:
    """Convert common pandas objects."""
    active_state = state or SerializationState()
    if not PANDAS_AVAILABLE:
        return _SERIALIZATION_SKIP
    if value is pd.NA or value is pd.NaT:
        return None
    if isinstance(value, pd.Timestamp):
        return serialize_datetime(value.to_pydatetime(), active_state.options)
    if isinstance(value, pd.Timedelta):
        return serialize_timedelta(value.to_pytimedelta(), active_state.options)
    if isinstance(value, pd.DataFrame):
        if active_state.options.dataframe_mode == "split":
            raw = value.to_dict(orient="split")
        elif active_state.options.dataframe_mode == "columns":
            raw = value.to_dict(orient="list")
        else:
            raw = value.to_dict(orient="records")
        return _to_serializable(raw, active_state, path, depth + 1)
    if isinstance(value, pd.Series):
        return _to_serializable(value.to_dict(), active_state, path, depth + 1)
    if isinstance(value, pd.Index):
        return _to_serializable(value.tolist(), active_state, path, depth + 1)
    return _SERIALIZATION_SKIP


def serialize_unknown_object(
    value: Any,
    *,
    state: Optional[SerializationState] = None,
    path: str = "$",
    depth: int = 0,
) -> Any:
    """Apply configured fallback conversion."""
    active_state = state or SerializationState()
    options = active_state.options
    mode = options.unknown
    if mode == UnknownObjectMode.ERROR.value:
        raise ExportSerializationError(
            f"Unsupported object type: {qualified_type_name(value)}.",
            context={"path": path},
        )
    if mode == UnknownObjectMode.NONE.value:
        return None
    if mode == UnknownObjectMode.TYPE.value:
        return qualified_type_name(value)
    if mode == UnknownObjectMode.REPR.value:
        return repr(value)
    if mode == UnknownObjectMode.DICT.value:
        attributes = _iter_object_attributes(value, options)
        if attributes:
            result = _to_serializable(attributes, active_state, path, depth + 1)
            if options.include_type and isinstance(result, dict):
                result.setdefault("__type__", qualified_type_name(value))
            return result
    active_state.warn("Object converted to string.", path=path, value=value)
    return str(value)


def _method_payload(value: Any, options: SerializationOptions) -> Any:
    methods: List[str] = []
    if options.use_to_dict:
        methods.append("to_dict")
    if options.use_as_dict:
        methods.append("as_dict")
    if options.use_json_method:
        methods.extend(("to_json", "__json__"))
    for name in methods:
        result = _safe_method_call(value, name)
        if result is _SERIALIZATION_SKIP or result is value:
            continue
        if name == "to_json" and isinstance(result, str):
            try:
                return json.loads(result)
            except json.JSONDecodeError:
                return result
        return result
    return _SERIALIZATION_SKIP


def _to_serializable(
    value: Any,
    state: SerializationState,
    path: str,
    depth: int,
) -> Any:
    options = state.options
    state.converted += 1
    if depth > options.max_depth:
        state.truncated_values += 1
        state.warn("Maximum serialization depth reached.", path=path, value=value)
        return options.depth_value
    if isinstance(value, _SIMPLE_SERIALIZABLE):
        return value
    if isinstance(value, float):
        return serialize_float(value, precision=options.float_precision, allow_nan=options.allow_nan)
    if isinstance(value, (bytes, bytearray, memoryview)):
        return serialize_bytes(value, options)
    if isinstance(value, Path):
        return serialize_path(value, options)
    if isinstance(value, Enum):
        return _to_serializable(serialize_enum(value, options), state, path, depth + 1)
    if isinstance(value, datetime):
        return serialize_datetime(value, options)
    if isinstance(value, (date, time)):
        return serialize_datetime(value, options)
    if isinstance(value, timedelta):
        return serialize_timedelta(value, options)
    if isinstance(value, complex):
        return _to_serializable(serialize_complex(value, options), state, path, depth + 1)
    if isinstance(value, range):
        return _serialize_range(value)
    if isinstance(value, slice):
        return _serialize_slice(value)

    module_name = type(value).__module__
    type_name = type(value).__name__
    if module_name == "decimal" and type_name == "Decimal":
        return _serialize_decimal(value, options)
    if module_name == "fractions" and type_name == "Fraction":
        return _serialize_fraction(value)
    if module_name == "uuid" and type_name == "UUID":
        return _serialize_uuid(value)

    numpy_result = serialize_numpy_value(value, state=state, path=path, depth=depth)
    if numpy_result is not _SERIALIZATION_SKIP:
        return numpy_result
    pandas_result = serialize_pandas_value(value, state=state, path=path, depth=depth)
    if pandas_result is not _SERIALIZATION_SKIP:
        return pandas_result

    identity = id(value)
    track = isinstance(value, (Mapping, list, tuple, set, frozenset)) or is_dataclass(value)
    if track and identity in state.active_ids:
        state.circular_references += 1
        state.warn("Circular reference replaced.", path=path, value=value)
        return options.circular_value
    if track:
        state.active_ids.add(identity)
    try:
        if is_dataclass(value) and not isinstance(value, type):
            return serialize_dataclass(value, state=state, path=path, depth=depth)
        if isinstance(value, Mapping):
            return serialize_mapping(value, state=state, path=path, depth=depth)
        if isinstance(value, (list, tuple, set, frozenset)):
            return serialize_sequence(value, state=state, path=path, depth=depth)
        payload = _method_payload(value, options)
        if payload is not _SERIALIZATION_SKIP:
            return _to_serializable(payload, state, path, depth + 1)
        return serialize_unknown_object(value, state=state, path=path, depth=depth)
    finally:
        if track:
            state.active_ids.discard(identity)


def to_serializable(
    value: Any,
    *,
    options: Optional[SerializationOptions] = None,
    state: Optional[SerializationState] = None,
) -> Any:
    """Convert nested values to portable Python types."""
    active_state = state or SerializationState(options or SerializationOptions())
    if options is not None and state is not None:
        active_state.options = options
    return _to_serializable(value, active_state, "$", 0)


def make_json_safe(
    value: Any,
    *,
    options: Optional[SerializationOptions] = None,
) -> JSONValue:
    """Convert a value to a JSON-compatible structure."""
    result = to_serializable(value, options=options)
    try:
        json.dumps(result, allow_nan=False)
    except (TypeError, ValueError) as exc:
        raise ExportSerializationError("Conversion did not produce JSON-safe data.", cause=exc) from exc
    return result


def is_json_safe(value: Any, *, allow_nan: bool = False) -> bool:
    """Return whether json.dumps accepts a value."""
    try:
        json.dumps(value, allow_nan=allow_nan)
        return True
    except (TypeError, ValueError, OverflowError, RecursionError):
        return False


def find_non_serializable(
    value: Any,
    *,
    path: str = "$",
    max_results: int = 100,
) -> List[Dict[str, str]]:
    """Locate values rejected by strict JSON serialization."""
    failures: List[Dict[str, str]] = []
    active: Set[int] = set()

    def visit(item: Any, item_path: str) -> None:
        if len(failures) >= max_results:
            return
        if isinstance(item, _SIMPLE_SERIALIZABLE):
            return
        if isinstance(item, float):
            if not math.isfinite(item):
                failures.append({"path": item_path, "type": "float", "reason": "non-finite"})
            return
        identity = id(item)
        if identity in active:
            failures.append({"path": item_path, "type": qualified_type_name(item), "reason": "circular"})
            return
        if isinstance(item, Mapping):
            active.add(identity)
            try:
                for key, child in item.items():
                    if not isinstance(key, str):
                        failures.append({"path": item_path, "type": qualified_type_name(key), "reason": "non-string-key"})
                    visit(child, f"{item_path}.{key}")
            finally:
                active.discard(identity)
            return
        if isinstance(item, (list, tuple)):
            active.add(identity)
            try:
                for index, child in enumerate(item):
                    visit(child, f"{item_path}[{index}]")
            finally:
                active.discard(identity)
            return
        failures.append({"path": item_path, "type": qualified_type_name(item), "reason": "unsupported"})

    visit(value, path)
    return failures


def serialization_report(
    value: Any,
    *,
    options: Optional[SerializationOptions] = None,
) -> Tuple[Any, Dict[str, Any]]:
    """Return converted data and conversion diagnostics."""
    state = SerializationState(options or SerializationOptions())
    result = to_serializable(value, state=state)
    report = {
        "converted_values": state.converted,
        "circular_references": state.circular_references,
        "truncated_values": state.truncated_values,
        "warnings": list(state.warnings),
        "json_safe": is_json_safe(result),
    }
    return result, report


__all__.extend(
    [
        "UnknownObjectMode",
        "SERIALIZATION_UNKNOWN_MODES",
        "SerializationOptions",
        "SerializationState",
        "qualified_type_name",
        "is_finite_number",
        "serialize_float",
        "serialize_bytes",
        "serialize_datetime",
        "serialize_timedelta",
        "serialize_enum",
        "serialize_complex",
        "serialize_path",
        "serialize_mapping",
        "serialize_sequence",
        "serialize_dataclass",
        "serialize_numpy_value",
        "serialize_pandas_value",
        "serialize_unknown_object",
        "to_serializable",
        "make_json_safe",
        "is_json_safe",
        "find_non_serializable",
        "serialization_report",
    ]
)

# =============================================================================
# End of Section 6
# =============================================================================
# Section 7 — Molecular objects
# =============================================================================

MOLECULAR_SCHEMA_VERSION: Final[str] = "1.0"
MOLECULAR_OBJECT_TYPES: Final[Tuple[str, ...]] = (
    "atom",
    "residue",
    "structure",
    "ligand",
    "molecular_object",
)

_ATOM_NAME_FIELDS: Final[Tuple[str, ...]] = ("name", "atom_name", "label")
_ATOM_SERIAL_FIELDS: Final[Tuple[str, ...]] = (
    "serial_number",
    "serial",
    "serialNumber",
    "idatm_serial",
)
_ATOM_INDEX_FIELDS: Final[Tuple[str, ...]] = ("index", "coord_index", "scene_coord_index")
_ATOM_COORD_FIELDS: Final[Tuple[str, ...]] = (
    "scene_coord",
    "coord",
    "coords",
    "coordinate",
    "coordinates",
    "xyz",
)
_RESIDUE_NAME_FIELDS: Final[Tuple[str, ...]] = ("name", "resname", "residue_name", "type")
_RESIDUE_NUMBER_FIELDS: Final[Tuple[str, ...]] = (
    "number",
    "resnum",
    "residue_number",
    "seq_id",
    "position",
)
_CHAIN_ID_FIELDS: Final[Tuple[str, ...]] = (
    "chain_id",
    "chain",
    "chainId",
    "asym_id",
    "auth_asym_id",
)
_INSERTION_CODE_FIELDS: Final[Tuple[str, ...]] = (
    "insertion_code",
    "insert",
    "icode",
    "insertionCode",
)
_STRUCTURE_NAME_FIELDS: Final[Tuple[str, ...]] = ("name", "model_name", "title", "filename")
_STRUCTURE_ID_FIELDS: Final[Tuple[str, ...]] = ("id_string", "model_id", "id", "identifier")


@dataclass
class MolecularExportOptions:
    """Control molecular-object record generation."""

    include_coordinates: bool = True
    include_element: bool = True
    include_residue: bool = True
    include_structure: bool = True
    include_atom_count: bool = True
    include_residue_count: bool = True
    include_atoms: bool = False
    include_residues: bool = False
    include_alt_loc: bool = True
    include_occupancy: bool = True
    include_bfactor: bool = True
    include_charge: bool = True
    include_bonds: bool = False
    include_type: bool = True
    include_schema: bool = True
    compact_references: bool = True
    coordinate_precision: Optional[int] = 4
    omit_none: bool = True
    strict: bool = False
    max_atoms: Optional[int] = None
    max_residues: Optional[int] = None

    def __post_init__(self) -> None:
        if self.coordinate_precision is not None and self.coordinate_precision < 0:
            raise ExportConfigurationError("coordinate_precision must be non-negative or None.")
        if self.max_atoms is not None and self.max_atoms < 0:
            raise ExportConfigurationError("max_atoms must be non-negative or None.")
        if self.max_residues is not None and self.max_residues < 0:
            raise ExportConfigurationError("max_residues must be non-negative or None.")

    def to_dict(self) -> Dict[str, Any]:
        """Return option values."""
        return {item.name: getattr(self, item.name) for item in fields(self)}


@dataclass
class MolecularRecordContext:
    """Track molecular records and circular references."""

    options: MolecularExportOptions = field(default_factory=MolecularExportOptions)
    atom_ids: Dict[int, str] = field(default_factory=dict)
    residue_ids: Dict[int, str] = field(default_factory=dict)
    structure_ids: Dict[int, str] = field(default_factory=dict)
    active_ids: Set[int] = field(default_factory=set)
    warnings: List[str] = field(default_factory=list)

    def warn(self, message: str) -> None:
        """Store one warning."""
        self.warnings.append(message)


def _molecular_get(
    value: Any,
    names: Union[str, Sequence[str]],
    default: Any = None,
) -> Any:
    """Read the first available mapping key or attribute."""
    if value is None:
        return default
    candidates = (names,) if isinstance(names, str) else names
    for name in candidates:
        if isinstance(value, Mapping):
            if name in value:
                item = value[name]
            else:
                continue
        else:
            try:
                item = getattr(value, name)
            except (AttributeError, RuntimeError, TypeError, ValueError):
                continue
        if callable(item):
            try:
                item = item()
            except (RuntimeError, TypeError, ValueError):
                continue
        if item is not None:
            return item
    return default


def _clean_molecular_record(record: Mapping[str, Any], *, omit_none: bool = True) -> Dict[str, Any]:
    """Remove missing values from a record."""
    if not omit_none:
        return dict(record)
    return {key: value for key, value in record.items() if value is not None}


def _safe_int(value: Any) -> Optional[int]:
    """Convert a value to int when possible."""
    if value is None or isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError, OverflowError):
        return None


def _safe_float(value: Any, precision: Optional[int] = None) -> Optional[float]:
    """Convert a finite value to float."""
    if value is None or isinstance(value, bool):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if not math.isfinite(result):
        return None
    return round(result, precision) if precision is not None else result


def _safe_text(value: Any) -> Optional[str]:
    """Convert a non-empty value to text."""
    if value is None:
        return None
    try:
        result = str(value).strip()
    except (RuntimeError, TypeError, ValueError):
        return None
    return result or None


def _iter_collection(value: Any) -> Iterator[Any]:
    """Iterate over a molecular collection defensively."""
    if value is None or isinstance(value, (str, bytes, bytearray, Mapping)):
        return iter(())
    try:
        return iter(value)
    except TypeError:
        return iter(())


def _collection_length(value: Any) -> Optional[int]:
    """Read collection length without consuming iterators."""
    if value is None:
        return None
    try:
        return len(value)
    except (TypeError, RuntimeError):
        return None


def _limited_items(value: Any, limit: Optional[int]) -> List[Any]:
    """Collect up to limit items."""
    result: List[Any] = []
    for item in _iter_collection(value):
        if limit is not None and len(result) >= limit:
            break
        result.append(item)
    return result


def element_symbol(value: Any) -> Optional[str]:
    """Return a normalized element symbol."""
    element = _molecular_get(value, ("element", "element_name", "symbol", "atomic_symbol"))
    if element is not None and element is not value:
        symbol = _molecular_get(element, ("symbol", "name"), element)
    else:
        symbol = element
    text = _safe_text(symbol)
    if text:
        return text[:1].upper() + text[1:].lower()
    atomic_number = _safe_int(_molecular_get(value, ("atomic_number", "atomicNumber")))
    if atomic_number is None:
        return None
    symbols = {
        1: "H", 6: "C", 7: "N", 8: "O", 9: "F", 15: "P", 16: "S",
        17: "Cl", 35: "Br", 53: "I", 11: "Na", 12: "Mg", 19: "K",
        20: "Ca", 25: "Mn", 26: "Fe", 27: "Co", 28: "Ni", 29: "Cu", 30: "Zn",
    }
    return symbols.get(atomic_number)


def coordinate_to_list(
    value: Any,
    *,
    precision: Optional[int] = 4,
) -> Optional[List[float]]:
    """Convert one 3D coordinate to a numeric list."""
    if value is None:
        return None
    raw = value
    if hasattr(raw, "data") and not isinstance(raw, (list, tuple)):
        try:
            raw = raw.data()
        except (RuntimeError, TypeError, ValueError):
            pass
    if NUMPY_AVAILABLE and isinstance(raw, np.ndarray):
        raw = raw.reshape(-1).tolist()
    if isinstance(raw, Mapping):
        raw = [raw.get("x"), raw.get("y"), raw.get("z")]
    elif all(hasattr(raw, axis) for axis in ("x", "y", "z")):
        raw = [getattr(raw, "x"), getattr(raw, "y"), getattr(raw, "z")]
    try:
        items = list(raw)
    except (TypeError, RuntimeError):
        return None
    if len(items) < 3:
        return None
    converted = [_safe_float(item, precision) for item in items[:3]]
    if any(item is None for item in converted):
        return None
    return [float(item) for item in converted if item is not None]


def atom_coordinates(
    atom: Any,
    *,
    precision: Optional[int] = 4,
) -> Optional[List[float]]:
    """Extract atom coordinates."""
    return coordinate_to_list(_molecular_get(atom, _ATOM_COORD_FIELDS), precision=precision)


def chain_identifier(value: Any) -> Optional[str]:
    """Extract a chain identifier."""
    chain = _molecular_get(value, _CHAIN_ID_FIELDS)
    if chain is not None and not isinstance(chain, (str, int)):
        chain = _molecular_get(chain, ("chain_id", "id", "name"), chain)
    return _safe_text(chain)


def residue_identifier(residue: Any) -> Optional[str]:
    """Build a stable residue identifier."""
    if residue is None:
        return None
    explicit = _safe_text(_molecular_get(residue, ("atomspec", "spec", "identifier", "id_string")))
    if explicit:
        return explicit
    name = _safe_text(_molecular_get(residue, _RESIDUE_NAME_FIELDS)) or "UNK"
    number = _molecular_get(residue, _RESIDUE_NUMBER_FIELDS)
    chain = chain_identifier(residue)
    insertion = _safe_text(_molecular_get(residue, _INSERTION_CODE_FIELDS))
    number_text = _safe_text(number) or "?"
    suffix = insertion or ""
    return f"{chain}:{name}{number_text}{suffix}" if chain else f"{name}{number_text}{suffix}"


def structure_identifier(structure: Any) -> Optional[str]:
    """Return a stable structure identifier."""
    if structure is None:
        return None
    explicit = _molecular_get(structure, _STRUCTURE_ID_FIELDS)
    if isinstance(explicit, (tuple, list)):
        explicit = ".".join(str(item) for item in explicit)
    text = _safe_text(explicit)
    if text:
        return text
    return _safe_text(_molecular_get(structure, _STRUCTURE_NAME_FIELDS))


def atom_identifier(atom: Any) -> Optional[str]:
    """Build a stable atom identifier."""
    if atom is None:
        return None
    explicit = _safe_text(_molecular_get(atom, ("atomspec", "spec", "identifier", "id_string")))
    if explicit:
        return explicit
    residue = _molecular_get(atom, ("residue", "parent_residue"))
    residue_id = residue_identifier(residue)
    name = _safe_text(_molecular_get(atom, _ATOM_NAME_FIELDS)) or "?"
    serial = _safe_int(_molecular_get(atom, _ATOM_SERIAL_FIELDS))
    if residue_id:
        return f"{residue_id}@{name}"
    if serial is not None:
        return f"atom:{serial}:{name}"
    return f"atom:{name}"


def is_atom_like(value: Any) -> bool:
    """Return whether an object resembles an atom."""
    if value is None:
        return False
    if CHIMERAX_AVAILABLE and isinstance(value, ChimeraXAtom):
        return True
    name = _molecular_get(value, _ATOM_NAME_FIELDS)
    residue = _molecular_get(value, ("residue", "parent_residue"))
    coordinates = _molecular_get(value, _ATOM_COORD_FIELDS)
    element = _molecular_get(value, ("element", "atomic_number", "symbol"))
    return name is not None and (residue is not None or coordinates is not None or element is not None)


def is_residue_like(value: Any) -> bool:
    """Return whether an object resembles a residue."""
    if value is None:
        return False
    if CHIMERAX_AVAILABLE and isinstance(value, ChimeraXResidue):
        return True
    name = _molecular_get(value, _RESIDUE_NAME_FIELDS)
    number = _molecular_get(value, _RESIDUE_NUMBER_FIELDS)
    atoms = _molecular_get(value, ("atoms", "atom_list"))
    return name is not None and (number is not None or atoms is not None)


def is_structure_like(value: Any) -> bool:
    """Return whether an object resembles a molecular structure."""
    if value is None:
        return False
    if CHIMERAX_AVAILABLE and isinstance(value, ChimeraXAtomicStructure):
        return True
    atoms = _molecular_get(value, ("atoms", "atom_list"))
    residues = _molecular_get(value, ("residues", "residue_list"))
    return atoms is not None and residues is not None


def molecular_object_kind(value: Any) -> Optional[str]:
    """Classify a molecular object."""
    if is_atom_like(value):
        return "atom"
    if is_structure_like(value):
        return "structure"
    if is_residue_like(value):
        return "residue"
    return None


def residue_to_record(
    residue: Any,
    *,
    options: Optional[MolecularExportOptions] = None,
    context: Optional[MolecularRecordContext] = None,
) -> Dict[str, Any]:
    """Convert a residue to a stable record."""
    active = context or MolecularRecordContext(options or MolecularExportOptions())
    if options is not None:
        active.options = options
    opts = active.options
    if residue is None:
        return {}
    identity = id(residue)
    residue_id = active.residue_ids.setdefault(identity, residue_identifier(residue) or f"residue:{identity}")
    record: Dict[str, Any] = {
        "object_type": "residue" if opts.include_type else None,
        "schema_version": MOLECULAR_SCHEMA_VERSION if opts.include_schema else None,
        "id": residue_id,
        "name": _safe_text(_molecular_get(residue, _RESIDUE_NAME_FIELDS)),
        "number": _safe_int(_molecular_get(residue, _RESIDUE_NUMBER_FIELDS)),
        "chain_id": chain_identifier(residue),
        "insertion_code": _safe_text(_molecular_get(residue, _INSERTION_CODE_FIELDS)),
        "polymer_type": _safe_text(_molecular_get(residue, ("polymer_type", "category", "residue_type"))),
        "is_hetero": _molecular_get(residue, ("is_hetero", "hetero", "het")),
    }
    atoms = _molecular_get(residue, ("atoms", "atom_list"))
    if opts.include_atom_count:
        record["atom_count"] = _collection_length(atoms)
    structure = _molecular_get(residue, ("structure", "model", "parent"))
    if opts.include_structure:
        record["structure_id"] = structure_identifier(structure)
    if opts.include_atoms and identity not in active.active_ids:
        active.active_ids.add(identity)
        try:
            atom_items = _limited_items(atoms, opts.max_atoms)
            record["atoms"] = [
                atom_to_record(item, options=replace(opts, include_residue=False), context=active)
                for item in atom_items
            ]
            total = _collection_length(atoms)
            if total is not None and len(atom_items) < total:
                record["atoms_truncated"] = total - len(atom_items)
        finally:
            active.active_ids.discard(identity)
    return _clean_molecular_record(record, omit_none=opts.omit_none)


def atom_to_record(
    atom: Any,
    *,
    options: Optional[MolecularExportOptions] = None,
    context: Optional[MolecularRecordContext] = None,
) -> Dict[str, Any]:
    """Convert an atom to a stable record."""
    active = context or MolecularRecordContext(options or MolecularExportOptions())
    if options is not None:
        active.options = options
    opts = active.options
    if atom is None:
        return {}
    identity = id(atom)
    atom_id = active.atom_ids.setdefault(identity, atom_identifier(atom) or f"atom:{identity}")
    residue = _molecular_get(atom, ("residue", "parent_residue"))
    structure = _molecular_get(atom, ("structure", "model"))
    if structure is None and residue is not None:
        structure = _molecular_get(residue, ("structure", "model", "parent"))
    record: Dict[str, Any] = {
        "object_type": "atom" if opts.include_type else None,
        "schema_version": MOLECULAR_SCHEMA_VERSION if opts.include_schema else None,
        "id": atom_id,
        "name": _safe_text(_molecular_get(atom, _ATOM_NAME_FIELDS)),
        "serial_number": _safe_int(_molecular_get(atom, _ATOM_SERIAL_FIELDS)),
        "index": _safe_int(_molecular_get(atom, _ATOM_INDEX_FIELDS)),
        "element": element_symbol(atom) if opts.include_element else None,
        "atomic_number": _safe_int(_molecular_get(atom, ("atomic_number", "atomicNumber"))),
        "alt_loc": _safe_text(_molecular_get(atom, ("alt_loc", "altloc", "alternate_location")))
        if opts.include_alt_loc else None,
        "occupancy": _safe_float(_molecular_get(atom, ("occupancy",)), 4)
        if opts.include_occupancy else None,
        "bfactor": _safe_float(_molecular_get(atom, ("bfactor", "b_factor", "temp_factor")), 4)
        if opts.include_bfactor else None,
        "formal_charge": _safe_float(_molecular_get(atom, ("formal_charge", "charge")), 4)
        if opts.include_charge else None,
        "idatm_type": _safe_text(_molecular_get(atom, ("idatm_type", "atom_type", "type"))),
        "coordinates": atom_coordinates(atom, precision=opts.coordinate_precision)
        if opts.include_coordinates else None,
    }
    if opts.include_residue:
        if opts.compact_references:
            record["residue_id"] = residue_identifier(residue)
        else:
            record["residue"] = residue_to_record(
                residue,
                options=replace(opts, include_atoms=False, include_structure=False),
                context=active,
            ) if residue is not None else None
    if opts.include_structure:
        record["structure_id"] = structure_identifier(structure)
    if opts.include_bonds:
        neighbors = _molecular_get(atom, ("neighbors", "bonded_atoms"))
        record["bonded_atom_ids"] = [atom_identifier(item) for item in _iter_collection(neighbors)]
    return _clean_molecular_record(record, omit_none=opts.omit_none)


def structure_to_record(
    structure: Any,
    *,
    options: Optional[MolecularExportOptions] = None,
    context: Optional[MolecularRecordContext] = None,
) -> Dict[str, Any]:
    """Convert a molecular structure to a stable record."""
    active = context or MolecularRecordContext(options or MolecularExportOptions())
    if options is not None:
        active.options = options
    opts = active.options
    if structure is None:
        return {}
    identity = id(structure)
    structure_id = active.structure_ids.setdefault(
        identity,
        structure_identifier(structure) or f"structure:{identity}",
    )
    atoms = _molecular_get(structure, ("atoms", "atom_list"))
    residues = _molecular_get(structure, ("residues", "residue_list"))
    record: Dict[str, Any] = {
        "object_type": "structure" if opts.include_type else None,
        "schema_version": MOLECULAR_SCHEMA_VERSION if opts.include_schema else None,
        "id": structure_id,
        "name": _safe_text(_molecular_get(structure, _STRUCTURE_NAME_FIELDS)),
        "model_id": _safe_text(_molecular_get(structure, _STRUCTURE_ID_FIELDS)),
        "filename": _safe_text(_molecular_get(structure, ("filename", "file_name", "path"))),
        "displayed": _molecular_get(structure, ("display", "displayed", "visible")),
    }
    if opts.include_atom_count:
        record["atom_count"] = _collection_length(atoms)
    if opts.include_residue_count:
        record["residue_count"] = _collection_length(residues)
    if identity not in active.active_ids:
        active.active_ids.add(identity)
        try:
            if opts.include_atoms:
                atom_items = _limited_items(atoms, opts.max_atoms)
                child_opts = replace(opts, include_structure=False, include_residue=True)
                record["atoms"] = [atom_to_record(item, options=child_opts, context=active) for item in atom_items]
                total_atoms = _collection_length(atoms)
                if total_atoms is not None and len(atom_items) < total_atoms:
                    record["atoms_truncated"] = total_atoms - len(atom_items)
            if opts.include_residues:
                residue_items = _limited_items(residues, opts.max_residues)
                child_opts = replace(opts, include_structure=False, include_atoms=False)
                record["residues"] = [
                    residue_to_record(item, options=child_opts, context=active)
                    for item in residue_items
                ]
                total_residues = _collection_length(residues)
                if total_residues is not None and len(residue_items) < total_residues:
                    record["residues_truncated"] = total_residues - len(residue_items)
        finally:
            active.active_ids.discard(identity)
    return _clean_molecular_record(record, omit_none=opts.omit_none)


def ligand_to_record(
    ligand: Any,
    *,
    options: Optional[MolecularExportOptions] = None,
    context: Optional[MolecularRecordContext] = None,
) -> Dict[str, Any]:
    """Convert a ligand residue or structure to a ligand record."""
    active = context or MolecularRecordContext(options or MolecularExportOptions())
    if options is not None:
        active.options = options
    opts = active.options
    if is_residue_like(ligand):
        record = residue_to_record(ligand, options=opts, context=active)
    elif is_structure_like(ligand):
        record = structure_to_record(ligand, options=opts, context=active)
    else:
        record = molecular_object_to_record(ligand, options=opts, context=active)
    record["object_type"] = "ligand"
    record.setdefault("ligand_name", record.get("name"))
    formula = _safe_text(_molecular_get(ligand, ("formula", "molecular_formula")))
    molecular_weight = _safe_float(_molecular_get(ligand, ("molecular_weight", "mol_weight", "mass")), 6)
    if formula is not None:
        record["formula"] = formula
    if molecular_weight is not None:
        record["molecular_weight"] = molecular_weight
    return _clean_molecular_record(record, omit_none=opts.omit_none)


def molecular_object_to_record(
    value: Any,
    *,
    options: Optional[MolecularExportOptions] = None,
    context: Optional[MolecularRecordContext] = None,
) -> Dict[str, Any]:
    """Dispatch a molecular object to its record builder."""
    active = context or MolecularRecordContext(options or MolecularExportOptions())
    if options is not None:
        active.options = options
    kind = molecular_object_kind(value)
    if kind == "atom":
        return atom_to_record(value, options=active.options, context=active)
    if kind == "residue":
        return residue_to_record(value, options=active.options, context=active)
    if kind == "structure":
        return structure_to_record(value, options=active.options, context=active)
    if value is None:
        return {}
    if active.options.strict:
        raise ExportSerializationError(
            f"Unsupported molecular object: {qualified_type_name(value)}."
        )
    payload = to_serializable(
        value,
        options=SerializationOptions(
            unknown=UnknownObjectMode.DICT.value,
            omit_none=active.options.omit_none,
            include_type=active.options.include_type,
        ),
    )
    if isinstance(payload, Mapping):
        record = dict(payload)
    else:
        record = {"value": payload}
    record.setdefault("object_type", "molecular_object")
    if active.options.include_schema:
        record.setdefault("schema_version", MOLECULAR_SCHEMA_VERSION)
    return _clean_molecular_record(record, omit_none=active.options.omit_none)


def atoms_to_records(
    atoms: Iterable[Any],
    *,
    options: Optional[MolecularExportOptions] = None,
) -> List[Dict[str, Any]]:
    """Convert atoms using one shared context."""
    context = MolecularRecordContext(options or MolecularExportOptions())
    return [atom_to_record(atom, context=context) for atom in atoms]


def residues_to_records(
    residues: Iterable[Any],
    *,
    options: Optional[MolecularExportOptions] = None,
) -> List[Dict[str, Any]]:
    """Convert residues using one shared context."""
    context = MolecularRecordContext(options or MolecularExportOptions())
    return [residue_to_record(residue, context=context) for residue in residues]


def molecular_objects_to_records(
    values: Iterable[Any],
    *,
    options: Optional[MolecularExportOptions] = None,
) -> List[Dict[str, Any]]:
    """Convert mixed molecular objects."""
    context = MolecularRecordContext(options or MolecularExportOptions())
    return [molecular_object_to_record(value, context=context) for value in values]


def atom_reference(atom: Any) -> Dict[str, Any]:
    """Return a compact atom reference."""
    return _clean_molecular_record(
        {
            "id": atom_identifier(atom),
            "name": _safe_text(_molecular_get(atom, _ATOM_NAME_FIELDS)),
            "element": element_symbol(atom),
            "residue_id": residue_identifier(_molecular_get(atom, ("residue", "parent_residue"))),
        }
    )


def residue_reference(residue: Any) -> Dict[str, Any]:
    """Return a compact residue reference."""
    return _clean_molecular_record(
        {
            "id": residue_identifier(residue),
            "name": _safe_text(_molecular_get(residue, _RESIDUE_NAME_FIELDS)),
            "number": _safe_int(_molecular_get(residue, _RESIDUE_NUMBER_FIELDS)),
            "chain_id": chain_identifier(residue),
        }
    )


def structure_reference(structure: Any) -> Dict[str, Any]:
    """Return a compact structure reference."""
    return _clean_molecular_record(
        {
            "id": structure_identifier(structure),
            "name": _safe_text(_molecular_get(structure, _STRUCTURE_NAME_FIELDS)),
        }
    )


def molecular_identifier(value: Any) -> Optional[str]:
    """Return a stable identifier for a molecular object."""
    kind = molecular_object_kind(value)
    if kind == "atom":
        return atom_identifier(value)
    if kind == "residue":
        return residue_identifier(value)
    if kind == "structure":
        return structure_identifier(value)
    return _safe_text(_molecular_get(value, ("identifier", "id_string", "id", "name")))


def molecular_reference(
    value: Any,
    *,
    options: Optional[MolecularExportOptions] = None,
    context: Optional[MolecularRecordContext] = None,
) -> Dict[str, Any]:
    """Return a compact reference for a molecular object."""
    kind = molecular_object_kind(value)
    if kind == "atom":
        return atom_reference(value)
    if kind == "residue":
        return residue_reference(value)
    if kind == "structure":
        return structure_reference(value)
    active = context or MolecularRecordContext(options or MolecularExportOptions())
    if options is not None:
        active.options = options
    record = molecular_object_to_record(value, options=active.options, context=active)
    identifier = molecular_identifier(value)
    result: Dict[str, Any] = {
        "id": identifier,
        "object_type": record.get("object_type", "molecular_object"),
    }
    label = record.get("name") or record.get("label")
    if label is not None:
        result["name"] = label
    if identifier is None:
        result["value"] = record
    return _clean_molecular_record(result, omit_none=active.options.omit_none)


__all__.extend(
    [
        "MOLECULAR_SCHEMA_VERSION",
        "MOLECULAR_OBJECT_TYPES",
        "MolecularExportOptions",
        "MolecularRecordContext",
        "element_symbol",
        "coordinate_to_list",
        "atom_coordinates",
        "chain_identifier",
        "atom_identifier",
        "residue_identifier",
        "structure_identifier",
        "molecular_identifier",
        "molecular_reference",
        "is_atom_like",
        "is_residue_like",
        "is_structure_like",
        "molecular_object_kind",
        "atom_to_record",
        "residue_to_record",
        "structure_to_record",
        "ligand_to_record",
        "molecular_object_to_record",
        "atoms_to_records",
        "residues_to_records",
        "molecular_objects_to_records",
        "atom_reference",
        "residue_reference",
        "structure_reference",
    ]
)

# =============================================================================
# End of Section 7
# =============================================================================
# Section 8 — Serializer registry
# =============================================================================

SERIALIZER_REGISTRY_VERSION: Final[str] = "1.0"
DEFAULT_SERIALIZER_PRIORITY: Final[int] = 100
FALLBACK_SERIALIZER_PRIORITY: Final[int] = -1000

SerializerFunction: TypeAlias = Callable[..., Any]
SerializerPredicate: TypeAlias = Callable[[Any], bool]


class SerializerMatchMode(str, Enum):
    """Define how a serializer entry matches values."""

    EXACT = "exact"
    INSTANCE = "instance"
    PREDICATE = "predicate"
    FALLBACK = "fallback"


@dataclass(frozen=True)
class SerializerEntry:
    """Describe one registered serializer."""

    name: str
    serializer: SerializerFunction
    target_type: Optional[type] = None
    predicate: Optional[SerializerPredicate] = None
    priority: int = DEFAULT_SERIALIZER_PRIORITY
    match_mode: SerializerMatchMode = SerializerMatchMode.INSTANCE
    aliases: Tuple[str, ...] = ()
    enabled: bool = True
    builtin: bool = False
    description: Optional[str] = None
    metadata: Mapping[str, Any] = field(default_factory=dict)
    registration_order: int = 0

    def __post_init__(self) -> None:
        normalized_name = _normalize_serializer_name(self.name)
        object.__setattr__(self, "name", normalized_name)
        object.__setattr__(
            self,
            "aliases",
            tuple(dict.fromkeys(_normalize_serializer_name(alias) for alias in self.aliases)),
        )
        object.__setattr__(self, "metadata", MappingProxyType(dict(self.metadata)))

        if not callable(self.serializer):
            raise ExportConfigurationError("serializer must be callable.")
        if self.target_type is not None and not isinstance(self.target_type, type):
            raise ExportConfigurationError("target_type must be a type or None.")
        if self.predicate is not None and not callable(self.predicate):
            raise ExportConfigurationError("predicate must be callable or None.")
        if self.match_mode is SerializerMatchMode.PREDICATE and self.predicate is None:
            raise ExportConfigurationError("Predicate serializers require a predicate.")
        if self.match_mode in {SerializerMatchMode.EXACT, SerializerMatchMode.INSTANCE}:
            if self.target_type is None:
                raise ExportConfigurationError("Type serializers require target_type.")
        if self.match_mode is SerializerMatchMode.FALLBACK:
            if self.target_type is not None or self.predicate is not None:
                raise ExportConfigurationError("Fallback serializers cannot define match targets.")
        if self.name in self.aliases:
            raise ExportConfigurationError("Serializer aliases cannot repeat the canonical name.")

    def matches(self, value: Any) -> bool:
        """Return whether this entry accepts a value."""
        if not self.enabled:
            return False
        if self.match_mode is SerializerMatchMode.FALLBACK:
            return True
        if self.match_mode is SerializerMatchMode.EXACT:
            return type(value) is self.target_type
        if self.match_mode is SerializerMatchMode.INSTANCE:
            return isinstance(value, self.target_type)  # type: ignore[arg-type]
        if self.match_mode is SerializerMatchMode.PREDICATE:
            try:
                return bool(self.predicate(value))  # type: ignore[misc]
            except (AttributeError, KeyError, RuntimeError, TypeError, ValueError):
                return False
        return False

    def to_dict(self) -> Dict[str, Any]:
        """Return serializable entry metadata."""
        return {
            "name": self.name,
            "target_type": qualified_type_name(self.target_type) if self.target_type else None,
            "priority": self.priority,
            "match_mode": self.match_mode.value,
            "aliases": list(self.aliases),
            "enabled": self.enabled,
            "builtin": self.builtin,
            "description": self.description,
            "metadata": dict(self.metadata),
            "registration_order": self.registration_order,
        }


@dataclass(frozen=True)
class SerializerResolution:
    """Store one registry resolution result."""

    entry: SerializerEntry
    value_type: str
    matched_by: str

    @property
    def name(self) -> str:
        """Return the serializer name."""
        return self.entry.name

    def to_dict(self) -> Dict[str, Any]:
        """Return resolution metadata."""
        return {
            "serializer": self.entry.name,
            "value_type": self.value_type,
            "matched_by": self.matched_by,
            "priority": self.entry.priority,
        }


@dataclass
class SerializerCallContext:
    """Provide optional state to serializer functions."""

    registry: "SerializerRegistry"
    serialization_options: Optional[SerializationOptions] = None
    molecular_options: Optional[MolecularExportOptions] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    path: Tuple[Union[str, int], ...] = ()
    warnings: List[str] = field(default_factory=list)

    def child(self, component: Union[str, int]) -> "SerializerCallContext":
        """Return a child context."""
        return SerializerCallContext(
            registry=self.registry,
            serialization_options=self.serialization_options,
            molecular_options=self.molecular_options,
            metadata=self.metadata,
            path=(*self.path, component),
            warnings=self.warnings,
        )

    def warn(self, message: str) -> None:
        """Store one warning."""
        self.warnings.append(message)


class SerializerRegistry:
    """Register and resolve object serializers."""

    def __init__(self, *, name: str = "default", allow_fallback: bool = True) -> None:
        self.name = _normalize_serializer_name(name)
        self.allow_fallback = bool(allow_fallback)
        self._entries: Dict[str, SerializerEntry] = {}
        self._aliases: Dict[str, str] = {}
        self._order = 0
        self._frozen = False
        self._resolution_cache: Dict[type, Optional[str]] = {}

    def __len__(self) -> int:
        return len(self._entries)

    def __iter__(self) -> Iterator[SerializerEntry]:
        return iter(self.entries())

    def __contains__(self, name: object) -> bool:
        if not isinstance(name, str):
            return False
        try:
            return self.canonical_name(name) in self._entries
        except KeyError:
            return False

    @property
    def frozen(self) -> bool:
        """Return whether mutations are disabled."""
        return self._frozen

    def freeze(self) -> "SerializerRegistry":
        """Disable registry mutations."""
        self._frozen = True
        return self

    def thaw(self) -> "SerializerRegistry":
        """Enable registry mutations."""
        self._frozen = False
        return self

    def _ensure_mutable(self) -> None:
        if self._frozen:
            raise ExportConfigurationError(f"Serializer registry {self.name!r} is frozen.")

    def _invalidate_cache(self) -> None:
        self._resolution_cache.clear()

    def canonical_name(self, name: str) -> str:
        """Resolve a canonical serializer name."""
        normalized = _normalize_serializer_name(name)
        if normalized in self._entries:
            return normalized
        if normalized in self._aliases:
            return self._aliases[normalized]
        raise KeyError(f"Unknown serializer: {name!r}.")

    def register(
        self,
        name: str,
        serializer: SerializerFunction,
        *,
        target_type: Optional[type] = None,
        predicate: Optional[SerializerPredicate] = None,
        priority: int = DEFAULT_SERIALIZER_PRIORITY,
        match_mode: Optional[Union[SerializerMatchMode, str]] = None,
        aliases: Sequence[str] = (),
        enabled: bool = True,
        builtin: bool = False,
        description: Optional[str] = None,
        metadata: Optional[Mapping[str, Any]] = None,
        replace_existing: bool = False,
    ) -> SerializerEntry:
        """Register one serializer."""
        self._ensure_mutable()
        canonical = _normalize_serializer_name(name)
        if match_mode is None:
            if predicate is not None:
                resolved_mode = SerializerMatchMode.PREDICATE
            elif target_type is not None:
                resolved_mode = SerializerMatchMode.INSTANCE
            else:
                resolved_mode = SerializerMatchMode.FALLBACK
        else:
            resolved_mode = SerializerMatchMode(match_mode)

        if canonical in self._entries and not replace_existing:
            raise ExportConfigurationError(f"Serializer {canonical!r} is already registered.")

        normalized_aliases = tuple(_normalize_serializer_name(alias) for alias in aliases)
        collisions = [
            alias for alias in normalized_aliases
            if alias in self._entries or (alias in self._aliases and self._aliases[alias] != canonical)
        ]
        if collisions:
            raise ExportConfigurationError(
                f"Serializer alias collision: {', '.join(sorted(collisions))}."
            )

        if canonical in self._entries:
            self.unregister(canonical, missing_ok=False, allow_builtin=True)

        self._order += 1
        entry = SerializerEntry(
            name=canonical,
            serializer=serializer,
            target_type=target_type,
            predicate=predicate,
            priority=int(priority),
            match_mode=resolved_mode,
            aliases=normalized_aliases,
            enabled=bool(enabled),
            builtin=bool(builtin),
            description=description,
            metadata=dict(metadata or {}),
            registration_order=self._order,
        )
        self._entries[canonical] = entry
        for alias in entry.aliases:
            self._aliases[alias] = canonical
        self._invalidate_cache()
        return entry

    def register_type(
        self,
        target_type: type,
        serializer: SerializerFunction,
        *,
        name: Optional[str] = None,
        exact: bool = False,
        **kwargs: Any,
    ) -> SerializerEntry:
        """Register a type-based serializer."""
        serializer_name = name or target_type.__name__
        mode = SerializerMatchMode.EXACT if exact else SerializerMatchMode.INSTANCE
        return self.register(
            serializer_name,
            serializer,
            target_type=target_type,
            match_mode=mode,
            **kwargs,
        )

    def register_predicate(
        self,
        name: str,
        predicate: SerializerPredicate,
        serializer: SerializerFunction,
        **kwargs: Any,
    ) -> SerializerEntry:
        """Register a predicate-based serializer."""
        return self.register(
            name,
            serializer,
            predicate=predicate,
            match_mode=SerializerMatchMode.PREDICATE,
            **kwargs,
        )

    def register_fallback(
        self,
        name: str,
        serializer: SerializerFunction,
        *,
        priority: int = FALLBACK_SERIALIZER_PRIORITY,
        **kwargs: Any,
    ) -> SerializerEntry:
        """Register a fallback serializer."""
        return self.register(
            name,
            serializer,
            priority=priority,
            match_mode=SerializerMatchMode.FALLBACK,
            **kwargs,
        )

    def unregister(
        self,
        name: str,
        *,
        missing_ok: bool = False,
        allow_builtin: bool = False,
    ) -> Optional[SerializerEntry]:
        """Remove a serializer."""
        self._ensure_mutable()
        try:
            canonical = self.canonical_name(name)
        except KeyError:
            if missing_ok:
                return None
            raise
        entry = self._entries[canonical]
        if entry.builtin and not allow_builtin:
            raise ExportConfigurationError(
                f"Built-in serializer {canonical!r} cannot be removed without allow_builtin=True."
            )
        del self._entries[canonical]
        for alias in entry.aliases:
            self._aliases.pop(alias, None)
        self._invalidate_cache()
        return entry

    def clear(self, *, include_builtin: bool = False) -> None:
        """Remove registered serializers."""
        self._ensure_mutable()
        names = [
            name for name, entry in self._entries.items()
            if include_builtin or not entry.builtin
        ]
        for name in names:
            self.unregister(name, allow_builtin=True)

    def get(self, name: str) -> SerializerEntry:
        """Return one serializer entry."""
        return self._entries[self.canonical_name(name)]

    def entries(
        self,
        *,
        enabled_only: bool = False,
        include_fallback: bool = True,
    ) -> Tuple[SerializerEntry, ...]:
        """Return entries in resolution order."""
        values = [
            entry for entry in self._entries.values()
            if (not enabled_only or entry.enabled)
            and (include_fallback or entry.match_mode is not SerializerMatchMode.FALLBACK)
        ]
        values.sort(key=_serializer_sort_key)
        return tuple(values)

    def names(self, *, enabled_only: bool = False) -> Tuple[str, ...]:
        """Return canonical serializer names."""
        return tuple(entry.name for entry in self.entries(enabled_only=enabled_only))

    def aliases(self) -> Mapping[str, str]:
        """Return registered aliases."""
        return MappingProxyType(dict(self._aliases))

    def set_enabled(self, name: str, enabled: bool) -> SerializerEntry:
        """Enable or disable one serializer."""
        self._ensure_mutable()
        canonical = self.canonical_name(name)
        current = self._entries[canonical]
        updated = replace(current, enabled=bool(enabled))
        self._entries[canonical] = updated
        self._invalidate_cache()
        return updated

    def resolve(self, value: Any, *, allow_fallback: Optional[bool] = None) -> SerializerResolution:
        """Resolve the best serializer for a value."""
        fallback_allowed = self.allow_fallback if allow_fallback is None else allow_fallback
        value_type = type(value)

        cached_name = self._resolution_cache.get(value_type, "__missing__")
        if cached_name != "__missing__":
            if cached_name is None:
                raise ExportSerializationError(
                    f"No serializer registered for {qualified_type_name(value_type)}."
                )
            entry = self._entries.get(cached_name)
            if entry is not None and entry.matches(value):
                return SerializerResolution(
                    entry=entry,
                    value_type=qualified_type_name(value_type),
                    matched_by=entry.match_mode.value,
                )

        fallback: Optional[SerializerEntry] = None
        cacheable = True
        for entry in self.entries(enabled_only=True):
            if entry.match_mode is SerializerMatchMode.PREDICATE:
                cacheable = False
            if entry.match_mode is SerializerMatchMode.FALLBACK:
                if fallback is None:
                    fallback = entry
                continue
            if entry.matches(value):
                if cacheable:
                    self._resolution_cache[value_type] = entry.name
                return SerializerResolution(
                    entry=entry,
                    value_type=qualified_type_name(value_type),
                    matched_by=entry.match_mode.value,
                )

        if fallback_allowed and fallback is not None:
            return SerializerResolution(
                entry=fallback,
                value_type=qualified_type_name(value_type),
                matched_by=SerializerMatchMode.FALLBACK.value,
            )

        if cacheable:
            self._resolution_cache[value_type] = None
        raise ExportSerializationError(
            f"No serializer registered for {qualified_type_name(value_type)}."
        )

    def serialize(
        self,
        value: Any,
        *,
        serializer: Optional[str] = None,
        context: Optional[SerializerCallContext] = None,
        allow_fallback: Optional[bool] = None,
        **kwargs: Any,
    ) -> Any:
        """Serialize a value using a resolved or named serializer."""
        active_context = context or SerializerCallContext(registry=self)
        if active_context.registry is not self:
            active_context = replace(active_context, registry=self)

        if serializer is None:
            resolution = self.resolve(value, allow_fallback=allow_fallback)
            entry = resolution.entry
        else:
            entry = self.get(serializer)
            if not entry.enabled:
                raise ExportSerializationError(f"Serializer {entry.name!r} is disabled.")

        try:
            return _invoke_serializer(entry.serializer, value, active_context, kwargs)
        except ExportError:
            raise
        except Exception as exc:
            raise ExportSerializationError(
                f"Serializer {entry.name!r} failed for {qualified_type_name(value)}: {exc}"
            ) from exc

    def clone(
        self,
        *,
        name: Optional[str] = None,
        frozen: Optional[bool] = None,
    ) -> "SerializerRegistry":
        """Return an independent registry copy."""
        clone = SerializerRegistry(
            name=name or self.name,
            allow_fallback=self.allow_fallback,
        )
        clone._entries = dict(self._entries)
        clone._aliases = dict(self._aliases)
        clone._order = self._order
        clone._frozen = self._frozen if frozen is None else bool(frozen)
        return clone

    def describe(self) -> Dict[str, Any]:
        """Return registry metadata."""
        return {
            "name": self.name,
            "version": SERIALIZER_REGISTRY_VERSION,
            "frozen": self.frozen,
            "allow_fallback": self.allow_fallback,
            "serializer_count": len(self),
            "serializers": [entry.to_dict() for entry in self.entries()],
            "aliases": dict(self._aliases),
        }

    @contextmanager
    def temporary_registration(self, *args: Any, **kwargs: Any) -> Iterator[SerializerEntry]:
        """Register a serializer for one context block."""
        entry = self.register(*args, **kwargs)
        try:
            yield entry
        finally:
            self.unregister(entry.name, missing_ok=True, allow_builtin=True)


def _normalize_serializer_name(value: str) -> str:
    """Normalize a serializer name."""
    text = str(value).strip().lower().replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        raise ExportConfigurationError("Serializer names cannot be empty.")
    if not re.fullmatch(r"[a-z][a-z0-9_]*", text):
        raise ExportConfigurationError(f"Invalid serializer name: {value!r}.")
    return text


def _serializer_sort_key(entry: SerializerEntry) -> Tuple[int, int, int]:
    """Return deterministic serializer resolution order."""
    mode_rank = {
        SerializerMatchMode.EXACT: 0,
        SerializerMatchMode.INSTANCE: 1,
        SerializerMatchMode.PREDICATE: 2,
        SerializerMatchMode.FALLBACK: 3,
    }[entry.match_mode]
    return (-entry.priority, mode_rank, entry.registration_order)


def _invoke_serializer(
    serializer: SerializerFunction,
    value: Any,
    context: SerializerCallContext,
    kwargs: Mapping[str, Any],
) -> Any:
    """Invoke serializers with compatible call signatures."""
    attempts = (
        lambda: serializer(value, context=context, **kwargs),
        lambda: serializer(value, context, **kwargs),
        lambda: serializer(value, **kwargs),
        lambda: serializer(value),
    )
    last_error: Optional[TypeError] = None
    for attempt in attempts:
        try:
            return attempt()
        except TypeError as exc:
            last_error = exc
    if last_error is None:
        raise ExportSerializationError(
            "Serializer invocation failed without a captured TypeError."
        )
    raise last_error


def serializer(
    registry: SerializerRegistry,
    *,
    name: Optional[str] = None,
    target_type: Optional[type] = None,
    predicate: Optional[SerializerPredicate] = None,
    priority: int = DEFAULT_SERIALIZER_PRIORITY,
    exact: bool = False,
    aliases: Sequence[str] = (),
    description: Optional[str] = None,
) -> Callable[[SerializerFunction], SerializerFunction]:
    """Register a serializer through a decorator."""
    def decorator(function: SerializerFunction) -> SerializerFunction:
        serializer_name = name or function.__name__
        if target_type is not None:
            registry.register_type(
                target_type,
                function,
                name=serializer_name,
                exact=exact,
                priority=priority,
                aliases=aliases,
                description=description,
            )
        elif predicate is not None:
            registry.register_predicate(
                serializer_name,
                predicate,
                function,
                priority=priority,
                aliases=aliases,
                description=description,
            )
        else:
            registry.register_fallback(
                serializer_name,
                function,
                priority=priority,
                aliases=aliases,
                description=description,
            )
        return function

    return decorator


def _serialize_atom_default(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize an atom-like object."""
    options = context.molecular_options if context else None
    return atom_to_record(value, options=options)


def _serialize_residue_default(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize a residue-like object."""
    options = context.molecular_options if context else None
    return residue_to_record(value, options=options)


def _serialize_structure_default(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize a structure-like object."""
    options = context.molecular_options if context else None
    return structure_to_record(value, options=options)


def _serialize_dataclass_default(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Any:
    """Serialize a dataclass instance."""
    options = context.serialization_options if context else None
    return to_serializable(value, options=options)


def _serialize_mapping_default(
    value: Mapping[Any, Any],
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Any:
    """Serialize a mapping."""
    options = context.serialization_options if context else None
    return to_serializable(value, options=options)


def _serialize_sequence_default(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Any:
    """Serialize a sequence."""
    options = context.serialization_options if context else None
    return to_serializable(value, options=options)


def _serialize_generic_default(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Any:
    """Serialize an unsupported object generically."""
    options = context.serialization_options if context else None
    return to_serializable(value, options=options)


def create_default_serializer_registry(
    *,
    name: str = "dockanalyzer",
    include_fallback: bool = True,
    frozen: bool = False,
) -> SerializerRegistry:
    """Create the standard DockAnalyzer registry."""
    registry = SerializerRegistry(name=name, allow_fallback=include_fallback)

    registry.register_predicate(
        "atom",
        is_atom_like,
        _serialize_atom_default,
        priority=500,
        aliases=("atomic", "chimerax_atom"),
        builtin=True,
        description="Atom-like molecular objects.",
    )
    registry.register_predicate(
        "residue",
        is_residue_like,
        _serialize_residue_default,
        priority=490,
        aliases=("amino_acid", "chimerax_residue"),
        builtin=True,
        description="Residue-like molecular objects.",
    )
    registry.register_predicate(
        "structure",
        is_structure_like,
        _serialize_structure_default,
        priority=480,
        aliases=("model", "molecule", "chimerax_structure"),
        builtin=True,
        description="Structure-like molecular objects.",
    )
    registry.register_predicate(
        "dataclass",
        lambda value: is_dataclass(value) and not isinstance(value, type),
        _serialize_dataclass_default,
        priority=300,
        builtin=True,
        description="Dataclass instances.",
    )
    registry.register_type(
        Mapping,
        _serialize_mapping_default,
        name="mapping",
        priority=200,
        aliases=("dict",),
        builtin=True,
        description="Mapping objects.",
    )
    registry.register_predicate(
        "sequence",
        lambda value: isinstance(value, (list, tuple, set, frozenset, range)),
        _serialize_sequence_default,
        priority=190,
        aliases=("collection", "iterable"),
        builtin=True,
        description="Common finite collections.",
    )
    registry.register_predicate(
        "scalar",
        lambda value: value is None or isinstance(
            value,
            (str, bytes, bytearray, bool, int, float, complex, date, time, timedelta, Enum, Path),
        ),
        _serialize_generic_default,
        priority=180,
        builtin=True,
        description="Common scalar and standard-library values.",
    )

    if NUMPY_AVAILABLE:
        registry.register_predicate(
            "numpy",
            lambda value: isinstance(value, (np.ndarray, np.generic)),
            _serialize_generic_default,
            priority=250,
            aliases=("ndarray", "numpy_scalar"),
            builtin=True,
            description="NumPy arrays and scalar values.",
        )
    if PANDAS_AVAILABLE:
        registry.register_predicate(
            "pandas",
            lambda value: isinstance(value, (pd.DataFrame, pd.Series, pd.Index)),
            _serialize_generic_default,
            priority=240,
            aliases=("dataframe", "series"),
            builtin=True,
            description="pandas tabular objects.",
        )
    if include_fallback:
        registry.register_fallback(
            "generic",
            _serialize_generic_default,
            priority=FALLBACK_SERIALIZER_PRIORITY,
            aliases=("fallback", "default"),
            builtin=True,
            description="Generic recursive serialization fallback.",
        )
    if frozen:
        registry.freeze()
    return registry


DEFAULT_SERIALIZER_REGISTRY: Final[SerializerRegistry] = create_default_serializer_registry()


def get_serializer_registry(*, clone: bool = False) -> SerializerRegistry:
    """Return the module-level serializer registry."""
    return DEFAULT_SERIALIZER_REGISTRY.clone() if clone else DEFAULT_SERIALIZER_REGISTRY


def register_serializer(
    name: str,
    serializer_function: SerializerFunction,
    **kwargs: Any,
) -> SerializerEntry:
    """Register a serializer in the module registry."""
    return DEFAULT_SERIALIZER_REGISTRY.register(name, serializer_function, **kwargs)


def unregister_serializer(
    name: str,
    *,
    missing_ok: bool = False,
    allow_builtin: bool = False,
) -> Optional[SerializerEntry]:
    """Remove a serializer from the module registry."""
    return DEFAULT_SERIALIZER_REGISTRY.unregister(
        name,
        missing_ok=missing_ok,
        allow_builtin=allow_builtin,
    )


def resolve_serializer(
    value: Any,
    *,
    registry: Optional[SerializerRegistry] = None,
    allow_fallback: Optional[bool] = None,
) -> SerializerResolution:
    """Resolve a serializer from a registry."""
    active = registry or DEFAULT_SERIALIZER_REGISTRY
    return active.resolve(value, allow_fallback=allow_fallback)


def serialize_registered(
    value: Any,
    *,
    registry: Optional[SerializerRegistry] = None,
    serializer_name: Optional[str] = None,
    context: Optional[SerializerCallContext] = None,
    allow_fallback: Optional[bool] = None,
    **kwargs: Any,
) -> Any:
    """Serialize a value using the registry layer."""
    active = registry or DEFAULT_SERIALIZER_REGISTRY
    return active.serialize(
        value,
        serializer=serializer_name,
        context=context,
        allow_fallback=allow_fallback,
        **kwargs,
    )


__all__.extend(
    [
        "SERIALIZER_REGISTRY_VERSION",
        "DEFAULT_SERIALIZER_PRIORITY",
        "FALLBACK_SERIALIZER_PRIORITY",
        "SerializerFunction",
        "SerializerPredicate",
        "SerializerMatchMode",
        "SerializerEntry",
        "SerializerResolution",
        "SerializerCallContext",
        "SerializerRegistry",
        "serializer",
        "create_default_serializer_registry",
        "DEFAULT_SERIALIZER_REGISTRY",
        "get_serializer_registry",
        "register_serializer",
        "unregister_serializer",
        "resolve_serializer",
        "serialize_registered",
    ]
)

# =============================================================================
# End of Section 8
# =============================================================================
# Section 9 — Molecular interactions
# =============================================================================

INTERACTION_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_INTERACTION_DISTANCE_PRECISION: Final[int] = 4
DEFAULT_INTERACTION_ANGLE_PRECISION: Final[int] = 3
DEFAULT_INTERACTION_SCORE_PRECISION: Final[int] = 6

_INTERACTION_FAMILY_FIELDS: Final[Tuple[str, ...]] = (
    "family", "interaction_family", "category", "module", "source_module",
)
_INTERACTION_TYPE_FIELDS: Final[Tuple[str, ...]] = (
    "interaction_type", "type", "kind", "subtype", "classification", "geometry_type",
)
_INTERACTION_ID_FIELDS: Final[Tuple[str, ...]] = (
    "interaction_id", "id", "uid", "identifier", "key",
)
_INTERACTION_DISTANCE_FIELDS: Final[Tuple[str, ...]] = (
    "distance", "distance_a", "distance_angstrom", "centroid_distance", "min_distance",
)
_INTERACTION_ANGLE_FIELDS: Final[Tuple[str, ...]] = (
    "angle", "angle_deg", "dha_angle", "theta", "normal_angle", "plane_angle",
)
_INTERACTION_SCORE_FIELDS: Final[Tuple[str, ...]] = (
    "score", "interaction_score", "geometric_score", "strength_score", "weight",
)
_INTERACTION_STRENGTH_FIELDS: Final[Tuple[str, ...]] = (
    "strength", "strength_class", "quality", "grade",
)
_INTERACTION_POSE_FIELDS: Final[Tuple[str, ...]] = (
    "pose", "pose_id", "pose_name", "model", "model_id",
)
_INTERACTION_ATOM1_FIELDS: Final[Tuple[str, ...]] = (
    "atom1", "atom_a", "source_atom", "donor_atom", "ligand_atom", "positive_atom",
)
_INTERACTION_ATOM2_FIELDS: Final[Tuple[str, ...]] = (
    "atom2", "atom_b", "target_atom", "acceptor_atom", "receptor_atom", "negative_atom",
)
_INTERACTION_RESIDUE1_FIELDS: Final[Tuple[str, ...]] = (
    "residue1", "residue_a", "source_residue", "donor_residue", "ligand_residue",
)
_INTERACTION_RESIDUE2_FIELDS: Final[Tuple[str, ...]] = (
    "residue2", "residue_b", "target_residue", "acceptor_residue", "receptor_residue",
)
_INTERACTION_ATOMS_FIELDS: Final[Tuple[str, ...]] = (
    "atoms", "atom_pair", "atom_pairs", "participants", "members",
)
_INTERACTION_RESIDUES_FIELDS: Final[Tuple[str, ...]] = (
    "residues", "residue_pair", "residue_pairs",
)
_INTERACTION_METADATA_FIELDS: Final[Tuple[str, ...]] = (
    "metadata", "details", "extra", "properties", "annotations",
)

_INTERACTION_CLASSIFICATION_KEYS: Final[FrozenSet[str]] = frozenset(
    _INTERACTION_FAMILY_FIELDS + _INTERACTION_TYPE_FIELDS
)
_INTERACTION_PARTICIPANT_KEYS: Final[FrozenSet[str]] = frozenset(
    _INTERACTION_ATOM1_FIELDS
    + _INTERACTION_ATOM2_FIELDS
    + _INTERACTION_RESIDUE1_FIELDS
    + _INTERACTION_RESIDUE2_FIELDS
    + _INTERACTION_ATOMS_FIELDS
    + _INTERACTION_RESIDUES_FIELDS
)
_INTERACTION_GEOMETRY_KEYS: Final[FrozenSet[str]] = frozenset(
    _INTERACTION_DISTANCE_FIELDS + _INTERACTION_ANGLE_FIELDS
)
_INTERACTION_KNOWN_FIELDS: Final[FrozenSet[str]] = frozenset(
    _INTERACTION_FAMILY_FIELDS
    + _INTERACTION_TYPE_FIELDS
    + _INTERACTION_ID_FIELDS
    + _INTERACTION_DISTANCE_FIELDS
    + _INTERACTION_ANGLE_FIELDS
    + _INTERACTION_SCORE_FIELDS
    + _INTERACTION_STRENGTH_FIELDS
    + _INTERACTION_POSE_FIELDS
    + _INTERACTION_ATOM1_FIELDS
    + _INTERACTION_ATOM2_FIELDS
    + _INTERACTION_RESIDUE1_FIELDS
    + _INTERACTION_RESIDUE2_FIELDS
    + _INTERACTION_ATOMS_FIELDS
    + _INTERACTION_RESIDUES_FIELDS
    + _INTERACTION_METADATA_FIELDS
)
_INTERACTION_ROLE_PAIRS: Final[Mapping[str, Tuple[str, str]]] = MappingProxyType(
    {
        "hbond": ("donor", "acceptor"),
        "saltbridge": ("positive", "negative"),
        "pi": ("source", "target"),
        "hydrophobic": ("ligand", "receptor"),
        "contact": ("source", "target"),
    }
)


class InteractionOrientation(str, Enum):
    """Describe participant ordering."""

    DIRECTED = "directed"
    UNDIRECTED = "undirected"
    AUTO = "auto"


class InteractionGranularity(str, Enum):
    """Control participant detail."""

    COMPACT = "compact"
    STANDARD = "standard"
    FULL = "full"


@dataclass
class InteractionExportOptions:
    """Control interaction record generation."""

    include_schema: bool = True
    include_family: bool = True
    include_type: bool = True
    include_id: bool = True
    include_atoms: bool = True
    include_residues: bool = True
    include_pose: bool = True
    include_geometry: bool = True
    include_score: bool = True
    include_strength: bool = True
    include_metadata: bool = True
    include_raw_fields: bool = False
    include_participant_roles: bool = True
    omit_none: bool = True
    strict: bool = False
    orientation: InteractionOrientation = InteractionOrientation.AUTO
    granularity: InteractionGranularity = InteractionGranularity.STANDARD
    distance_precision: Optional[int] = DEFAULT_INTERACTION_DISTANCE_PRECISION
    angle_precision: Optional[int] = DEFAULT_INTERACTION_ANGLE_PRECISION
    score_precision: Optional[int] = DEFAULT_INTERACTION_SCORE_PRECISION
    molecular_options: MolecularExportOptions = field(default_factory=MolecularExportOptions)

    def __post_init__(self) -> None:
        if isinstance(self.orientation, str):
            self.orientation = InteractionOrientation(self.orientation)
        if isinstance(self.granularity, str):
            self.granularity = InteractionGranularity(self.granularity)
        for name in ("distance_precision", "angle_precision", "score_precision"):
            value = getattr(self, name)
            if value is not None and value < 0:
                raise ExportConfigurationError(f"{name} must be non-negative or None.")

    def to_dict(self) -> Dict[str, Any]:
        """Return option values."""
        result = {item.name: getattr(self, item.name) for item in fields(self)}
        result["orientation"] = self.orientation.value
        result["granularity"] = self.granularity.value
        result["molecular_options"] = self.molecular_options.to_dict()
        return result


@dataclass(frozen=True)
class InteractionParticipant:
    """Store one interaction participant."""

    role: str
    atom: Any = None
    residue: Any = None
    structure: Any = None
    group: Any = None
    label: Optional[str] = None
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        object.__setattr__(self, "role", _normalize_interaction_token(self.role, "participant"))
        object.__setattr__(self, "metadata", MappingProxyType(dict(self.metadata)))


@dataclass
class InteractionRecordContext:
    """Track interaction conversion state."""

    options: InteractionExportOptions = field(default_factory=InteractionExportOptions)
    molecular_context: MolecularRecordContext = field(default_factory=MolecularRecordContext)
    active_ids: Set[int] = field(default_factory=set)
    seen_ids: Set[str] = field(default_factory=set)
    warnings: List[str] = field(default_factory=list)
    _atom_cache: Dict[Tuple[str, int], Dict[str, Any]] = field(
        default_factory=dict,
        init=False,
        repr=False,
    )
    _residue_cache: Dict[int, Dict[str, Any]] = field(
        default_factory=dict,
        init=False,
        repr=False,
    )
    _structure_cache: Dict[int, Dict[str, Any]] = field(
        default_factory=dict,
        init=False,
        repr=False,
    )
    _label_cache: Dict[Tuple[str, int], str] = field(
        default_factory=dict,
        init=False,
        repr=False,
    )

    def __post_init__(self) -> None:
        self.molecular_context.options = self.options.molecular_options

    def warn(self, message: str) -> None:
        """Store one warning."""
        self.warnings.append(message)


@dataclass(frozen=True)
class InteractionCollectionSummary:
    """Summarize serialized interactions."""

    total: int
    by_family: Mapping[str, int]
    by_type: Mapping[str, int]
    by_strength: Mapping[str, int]
    residues: Tuple[str, ...]
    poses: Tuple[str, ...]
    distance_min: Optional[float] = None
    distance_max: Optional[float] = None
    distance_mean: Optional[float] = None
    score_total: Optional[float] = None

    def to_dict(self) -> Dict[str, Any]:
        """Return serializable summary data."""
        return {
            "total": self.total,
            "by_family": dict(self.by_family),
            "by_type": dict(self.by_type),
            "by_strength": dict(self.by_strength),
            "residues": list(self.residues),
            "poses": list(self.poses),
            "distance_min": self.distance_min,
            "distance_max": self.distance_max,
            "distance_mean": self.distance_mean,
            "score_total": self.score_total,
        }


def _normalize_interaction_token(value: Any, default: str = "unknown") -> str:
    """Normalize an interaction token."""
    if value is None:
        return default
    text = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")
    return text or default


def normalize_interaction_family(value: Any, default: str = "unknown") -> str:
    """Return a canonical interaction family."""
    token = _normalize_interaction_token(value, default)
    return INTERACTION_FAMILY_ALIASES.get(token, token)


def normalize_interaction_type(value: Any, default: str = "unknown") -> str:
    """Return a normalized interaction type."""
    return _normalize_interaction_token(value, default)


def _interaction_get(value: Any, names: Sequence[str], default: Any = None) -> Any:
    """Read the first available field."""
    return _molecular_get(value, names, default)


def _interaction_mapping(value: Any) -> Mapping[str, Any]:
    """Return visible fields as a mapping."""
    if isinstance(value, Mapping):
        return value
    if is_dataclass(value) and not isinstance(value, type):
        return {item.name: getattr(value, item.name) for item in fields(value)}
    try:
        namespace = vars(value)
    except (TypeError, AttributeError):
        return {}
    return {key: item for key, item in namespace.items() if not key.startswith("_")}


def _interaction_float(value: Any, precision: Optional[int]) -> Optional[float]:
    """Convert one interaction number."""
    return _safe_float(value, precision)


def _interaction_bool(value: Any) -> Optional[bool]:
    """Convert one optional boolean."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "yes", "1", "y"}:
        return True
    if text in {"false", "no", "0", "n"}:
        return False
    return None


def _clean_interaction_record(record: Mapping[str, Any], *, omit_none: bool) -> Dict[str, Any]:
    """Remove missing record fields."""
    if not omit_none:
        return dict(record)
    return {key: value for key, value in record.items() if value is not None}


def infer_interaction_family(value: Any) -> str:
    """Infer the interaction family from fields and type names."""
    explicit = _interaction_get(value, _INTERACTION_FAMILY_FIELDS)
    if explicit is not None:
        return normalize_interaction_family(explicit)

    type_value = _interaction_get(value, _INTERACTION_TYPE_FIELDS)
    token = normalize_interaction_type(type_value, "")
    if token:
        for alias, family in INTERACTION_FAMILY_ALIASES.items():
            if alias in token:
                return family
        if any(part in token for part in ("parallel", "t_shaped", "cation_pi", "anion_pi", "amide_pi")):
            return "pi"

    type_name = qualified_type_name(type(value)).lower()
    for alias, family in INTERACTION_FAMILY_ALIASES.items():
        if alias.replace("_", "") in type_name.replace("_", ""):
            return family

    fields_map = _interaction_mapping(value)
    names = set(fields_map)
    if {"donor", "acceptor"} & names or {"dha_angle", "hydrogen"} & names:
        return "hbond"
    if {"ring1", "ring2", "centroid1", "centroid2", "normal_angle"} & names:
        return "pi"
    if {"positive_group", "negative_group"} & names:
        return "saltbridge"
    if {"hydrophobic_atoms", "contact_area"} & names:
        return "hydrophobic"
    return "contact" if is_interaction_like(value) else "unknown"


def infer_interaction_type(value: Any, family: Optional[str] = None) -> str:
    """Infer a normalized interaction type."""
    explicit = _interaction_get(value, _INTERACTION_TYPE_FIELDS)
    if explicit is not None:
        return normalize_interaction_type(explicit)
    family = family or infer_interaction_family(value)
    type_name = type(value).__name__
    normalized = normalize_interaction_type(type_name, family)
    suffixes = ("_interaction", "_result", "_record", "interaction", "result", "record")
    for suffix in suffixes:
        if normalized.endswith(suffix):
            normalized = normalized[: -len(suffix)].rstrip("_")
            break
    return normalized or family


def is_interaction_like(value: Any) -> bool:
    """Return whether a value resembles one interaction."""
    if value is None or isinstance(value, (str, bytes, bytearray, int, float, bool)):
        return False
    keys = set(_interaction_mapping(value))
    if keys & _INTERACTION_CLASSIFICATION_KEYS:
        return True
    if keys & _INTERACTION_PARTICIPANT_KEYS and keys & _INTERACTION_GEOMETRY_KEYS:
        return True
    name = type(value).__name__.lower()
    return any(
        token in name
        for token in ("interaction", "contact", "hbond", "hydrophobic", "saltbridge")
    )


def _interaction_identifier_from_parts(
    value: Any,
    family: str,
    interaction_type: str,
    participant_labels: Sequence[str],
) -> str:
    """Build an identifier from precomputed interaction parts."""
    explicit = _interaction_get(value, _INTERACTION_ID_FIELDS)
    if explicit is not None:
        return str(explicit)
    labels = sorted(filter(None, participant_labels))
    distance = _interaction_get(value, _INTERACTION_DISTANCE_FIELDS)
    payload = "|".join((family, interaction_type, *labels, str(distance or "")))
    digest = sha256(payload.encode("utf-8", errors="replace")).hexdigest()[:16]
    return f"{family}:{digest}"


def interaction_identifier(value: Any, *, family: Optional[str] = None) -> str:
    """Return a stable interaction identifier."""
    resolved_family = family or infer_interaction_family(value)
    interaction_type = infer_interaction_type(value, resolved_family)
    participants = _extract_interaction_participants(value, resolved_family)
    labels = tuple(participant_label(item) for item in participants)
    return _interaction_identifier_from_parts(
        value,
        resolved_family,
        interaction_type,
        labels,
    )


def participant_label(participant: InteractionParticipant) -> str:
    """Return a compact participant label."""
    if participant.label:
        return participant.label
    if participant.atom is not None:
        return atom_identifier(participant.atom)
    if participant.residue is not None:
        return residue_identifier(participant.residue)
    if participant.group is not None:
        name = _interaction_get(participant.group, ("name", "label", "type", "id"))
        if name is not None:
            return str(name)
    return participant.role


def _participant_from_value(value: Any, role: str) -> InteractionParticipant:
    """Create a participant from an arbitrary value."""
    if isinstance(value, InteractionParticipant):
        return value
    if value is None:
        return InteractionParticipant(role=role)
    if is_atom_like(value):
        return InteractionParticipant(
            role=role,
            atom=value,
            residue=_interaction_get(value, ("residue",)),
            structure=_interaction_get(value, ("structure", "model")),
        )
    if is_residue_like(value):
        return InteractionParticipant(
            role=role,
            residue=value,
            structure=_interaction_get(value, ("structure", "model")),
        )
    if isinstance(value, Mapping):
        return InteractionParticipant(
            role=str(value.get("role", role)),
            atom=value.get("atom"),
            residue=value.get("residue"),
            structure=value.get("structure"),
            group=value.get("group"),
            label=value.get("label"),
            metadata=value.get("metadata", {}),
        )
    return InteractionParticipant(role=role, group=value)


def _sequence_without_text(value: Any) -> List[Any]:
    """Return a safe materialized sequence."""
    if value is None or isinstance(value, (str, bytes, bytearray, Mapping)):
        return []
    try:
        return list(value)
    except (TypeError, RuntimeError, ValueError):
        return []


def _extract_interaction_participants(
    value: Any,
    family: str,
) -> List[InteractionParticipant]:
    """Extract participants using a precomputed family."""
    items = _sequence_without_text(_interaction_get(value, ("participants",)))
    if items:
        return [
            _participant_from_value(item, f"participant_{index + 1}")
            for index, item in enumerate(items)
        ]

    role1, role2 = _INTERACTION_ROLE_PAIRS.get(family, ("source", "target"))
    atom1 = _interaction_get(value, _INTERACTION_ATOM1_FIELDS)
    atom2 = _interaction_get(value, _INTERACTION_ATOM2_FIELDS)
    residue1 = _interaction_get(value, _INTERACTION_RESIDUE1_FIELDS)
    residue2 = _interaction_get(value, _INTERACTION_RESIDUE2_FIELDS)
    result: List[InteractionParticipant] = []
    if atom1 is not None or residue1 is not None:
        result.append(InteractionParticipant(role=role1, atom=atom1, residue=residue1))
    if atom2 is not None or residue2 is not None:
        result.append(InteractionParticipant(role=role2, atom=atom2, residue=residue2))
    if result:
        return result

    atoms = _sequence_without_text(_interaction_get(value, _INTERACTION_ATOMS_FIELDS))
    residues = _sequence_without_text(_interaction_get(value, _INTERACTION_RESIDUES_FIELDS))
    for index in range(max(len(atoms), len(residues))):
        role = role1 if index == 0 else role2 if index == 1 else f"participant_{index + 1}"
        result.append(
            InteractionParticipant(
                role=role,
                atom=atoms[index] if index < len(atoms) else None,
                residue=residues[index] if index < len(residues) else None,
            )
        )
    return result


def extract_interaction_participants(value: Any) -> List[InteractionParticipant]:
    """Extract ordered interaction participants."""
    return _extract_interaction_participants(value, infer_interaction_family(value))


def _clone_cached_record(record: Mapping[str, Any]) -> Dict[str, Any]:
    """Clone flat molecular records stored in operation-local caches."""
    return {
        key: list(value)
        if isinstance(value, list)
        else dict(value)
        if isinstance(value, Mapping)
        else value
        for key, value in record.items()
    }


def _cached_molecular_record(
    cache: Dict[Any, Dict[str, Any]],
    key: Any,
    builder: Callable[[], Dict[str, Any]],
) -> Dict[str, Any]:
    """Return a cloned operation-local molecular record."""
    cached = cache.get(key)
    if cached is None:
        cached = builder()
        cache[key] = cached
    return _clone_cached_record(cached)


def _cached_participant_label(
    participant: InteractionParticipant,
    context: InteractionRecordContext,
) -> str:
    """Return a participant label cached by molecular object identity."""
    if participant.label:
        return participant.label
    if participant.atom is not None:
        key = ("atom", id(participant.atom))
    elif participant.residue is not None:
        key = ("residue", id(participant.residue))
    elif participant.group is not None:
        key = ("group", id(participant.group))
    else:
        return participant.role
    label = context._label_cache.get(key)
    if label is None:
        label = participant_label(participant)
        context._label_cache[key] = label
    return label


def _interaction_participant_to_record(
    participant: InteractionParticipant,
    label: str,
    options: InteractionExportOptions,
    context: InteractionRecordContext,
) -> Dict[str, Any]:
    """Convert a participant using its precomputed label."""
    molecular = context.molecular_context
    record: Dict[str, Any] = {"label": label}
    if options.include_participant_roles:
        record["role"] = participant.role
    if options.include_atoms and participant.atom is not None:
        atom_key = (options.granularity.value, id(participant.atom))
        record["atom"] = _cached_molecular_record(
            context._atom_cache,
            atom_key,
            lambda: molecular_reference(participant.atom, context=molecular)
            if options.granularity is InteractionGranularity.COMPACT
            else atom_to_record(participant.atom, context=molecular),
        )
    if options.include_residues and participant.residue is not None:
        if options.granularity is InteractionGranularity.FULL:
            record["residue"] = residue_to_record(participant.residue, context=molecular)
        else:
            record["residue"] = _cached_molecular_record(
                context._residue_cache,
                id(participant.residue),
                lambda: molecular_reference(participant.residue, context=molecular),
            )
    if options.granularity is InteractionGranularity.FULL and participant.structure is not None:
        record["structure"] = _cached_molecular_record(
            context._structure_cache,
            id(participant.structure),
            lambda: molecular_reference(participant.structure, context=molecular),
        )
    if participant.group is not None:
        record["group"] = to_serializable(participant.group)
    if options.include_metadata and participant.metadata:
        record["metadata"] = to_serializable(participant.metadata)
    return _clean_interaction_record(record, omit_none=options.omit_none)


def interaction_participant_to_record(
    participant: InteractionParticipant,
    *,
    options: Optional[InteractionExportOptions] = None,
    context: Optional[InteractionRecordContext] = None,
) -> Dict[str, Any]:
    """Convert one interaction participant."""
    resolved_options = options or InteractionExportOptions()
    resolved_context = context or InteractionRecordContext(options=resolved_options)
    return _interaction_participant_to_record(
        participant,
        _cached_participant_label(participant, resolved_context),
        resolved_options,
        resolved_context,
    )


def extract_interaction_geometry(
    value: Any,
    *,
    options: Optional[InteractionExportOptions] = None,
) -> Dict[str, Any]:
    """Extract common interaction geometry."""
    options = options or InteractionExportOptions()
    geometry: Dict[str, Any] = {}

    distance = _interaction_get(value, _INTERACTION_DISTANCE_FIELDS)
    angle = _interaction_get(value, _INTERACTION_ANGLE_FIELDS)
    geometry["distance"] = _interaction_float(distance, options.distance_precision)
    geometry["angle"] = _interaction_float(angle, options.angle_precision)

    aliases: Mapping[str, Tuple[str, ...]] = {
        "donor_acceptor_distance": ("donor_acceptor_distance", "da_distance"),
        "hydrogen_acceptor_distance": ("hydrogen_acceptor_distance", "ha_distance"),
        "centroid_distance": ("centroid_distance", "center_distance"),
        "offset": ("offset", "lateral_offset", "slip"),
        "planarity": ("planarity", "planarity_score"),
        "overlap": ("overlap", "overlap_fraction", "overlap_score"),
        "contact_area": ("contact_area", "surface_area"),
        "normal_angle": ("normal_angle", "interplanar_angle"),
        "theta": ("theta",),
        "phi": ("phi",),
    }
    for key, names in aliases.items():
        raw = _interaction_get(value, names)
        precision = options.angle_precision if "angle" in key or key in {"theta", "phi"} else options.distance_precision
        geometry[key] = _interaction_float(raw, precision)

    boolean_aliases: Mapping[str, Tuple[str, ...]] = {
        "within_distance": ("within_distance", "distance_valid"),
        "within_angle": ("within_angle", "angle_valid"),
        "valid_geometry": ("valid_geometry", "geometry_valid", "is_valid"),
    }
    for key, names in boolean_aliases.items():
        geometry[key] = _interaction_bool(_interaction_get(value, names))

    vectors = {
        "centroid1": ("centroid1", "centroid_a", "source_centroid"),
        "centroid2": ("centroid2", "centroid_b", "target_centroid"),
        "normal1": ("normal1", "normal_a", "source_normal"),
        "normal2": ("normal2", "normal_b", "target_normal"),
    }
    for key, names in vectors.items():
        vector = _interaction_get(value, names)
        if vector is not None:
            geometry[key] = coordinate_to_list(vector, precision=options.distance_precision)

    return _clean_interaction_record(geometry, omit_none=options.omit_none)


def extract_interaction_metadata(value: Any) -> Dict[str, Any]:
    """Extract explicit interaction metadata."""
    metadata = _interaction_get(value, _INTERACTION_METADATA_FIELDS, {})
    if isinstance(metadata, Mapping):
        return dict(metadata)
    return {"value": metadata} if metadata is not None else {}


def interaction_to_record(
    value: Any,
    *,
    options: Optional[InteractionExportOptions] = None,
    context: Optional[InteractionRecordContext] = None,
) -> Dict[str, Any]:
    """Convert one molecular interaction to a record."""
    options = options or InteractionExportOptions()
    context = context or InteractionRecordContext(options=options)
    object_id = id(value)
    if object_id in context.active_ids:
        if options.strict:
            raise ExportSerializationError("Circular interaction reference detected.")
        context.warn("Circular interaction reference replaced by an identifier.")
        return {"interaction_id": interaction_identifier(value)}

    context.active_ids.add(object_id)
    try:
        family = infer_interaction_family(value)
        interaction_type = infer_interaction_type(value, family)
        participants = _extract_interaction_participants(value, family)
        labels = tuple(
            _cached_participant_label(item, context) for item in participants
        )
        identifier = _interaction_identifier_from_parts(
            value,
            family,
            interaction_type,
            labels,
        )
        record: Dict[str, Any] = {}

        if options.include_schema:
            record["schema_version"] = INTERACTION_SCHEMA_VERSION
            record["object_type"] = "interaction"
        if options.include_id:
            record["interaction_id"] = identifier
        if options.include_family:
            record["family"] = family
        if options.include_type:
            record["interaction_type"] = interaction_type
        if options.include_pose:
            pose = _interaction_get(value, _INTERACTION_POSE_FIELDS)
            record["pose"] = to_serializable(pose) if pose is not None else None
        if participants:
            record["participants"] = [
                _interaction_participant_to_record(item, label, options, context)
                for item, label in zip(participants, labels)
            ]
        if options.include_geometry:
            geometry = extract_interaction_geometry(value, options=options)
            record["geometry"] = geometry or None
        if options.include_score:
            score = _interaction_get(value, _INTERACTION_SCORE_FIELDS)
            record["score"] = _interaction_float(score, options.score_precision)
        if options.include_strength:
            strength = _interaction_get(value, _INTERACTION_STRENGTH_FIELDS)
            record["strength"] = _normalize_interaction_token(strength, "") or None
        if options.include_metadata:
            metadata = extract_interaction_metadata(value)
            record["metadata"] = to_serializable(metadata) if metadata else None
        if options.include_raw_fields:
            raw = {
                key: item
                for key, item in _interaction_mapping(value).items()
                if key not in _INTERACTION_KNOWN_FIELDS
            }
            record["raw"] = to_serializable(raw) if raw else None

        context.seen_ids.add(identifier)
        return _clean_interaction_record(record, omit_none=options.omit_none)
    except ExportError:
        raise
    except Exception as exc:
        if options.strict:
            raise ExportSerializationError(
                f"Failed to serialize interaction {type(value).__name__}: {exc}"
            ) from exc
        context.warn(f"Interaction serialization fallback used for {type(value).__name__}.")
        return {
            "schema_version": INTERACTION_SCHEMA_VERSION,
            "object_type": "interaction",
            "interaction_id": interaction_identifier(value),
            "family": infer_interaction_family(value),
            "interaction_type": infer_interaction_type(value),
            "value": to_serializable(value),
        }
    finally:
        context.active_ids.discard(object_id)


def interaction_reference(value: Any) -> Dict[str, Any]:
    """Return a compact interaction reference."""
    family = infer_interaction_family(value)
    interaction_type = infer_interaction_type(value, family)
    participants = _extract_interaction_participants(value, family)
    labels = tuple(participant_label(item) for item in participants)
    return {
        "interaction_id": _interaction_identifier_from_parts(
            value,
            family,
            interaction_type,
            labels,
        ),
        "family": family,
        "interaction_type": interaction_type,
    }


def interaction_deduplication_key(
    value: Any,
    *,
    directed: Optional[bool] = None,
    distance_precision: int = 3,
) -> Tuple[Any, ...]:
    """Return a deterministic interaction key."""
    family = infer_interaction_family(value)
    interaction_type = infer_interaction_type(value, family)
    participants = tuple(
        filter(
            None,
            (
                participant_label(item)
                for item in _extract_interaction_participants(value, family)
            ),
        )
    )
    if directed is None:
        directed = family in {"hbond", "saltbridge"}
    ordered = participants if directed else tuple(sorted(participants))
    distance = _interaction_float(_interaction_get(value, _INTERACTION_DISTANCE_FIELDS), distance_precision)
    pose = _interaction_get(value, _INTERACTION_POSE_FIELDS)
    return family, interaction_type, ordered, distance, str(pose) if pose is not None else None


def deduplicate_interactions(
    interactions: Iterable[Any],
    *,
    directed: Optional[bool] = None,
    keep: Literal["first", "last", "highest_score", "shortest_distance"] = "first",
) -> List[Any]:
    """Deduplicate interactions while preserving useful records."""
    selected: Dict[Tuple[Any, ...], Any] = {}
    order: List[Tuple[Any, ...]] = []
    for interaction in interactions:
        key = interaction_deduplication_key(interaction, directed=directed)
        if key not in selected:
            selected[key] = interaction
            order.append(key)
            continue
        if keep == "last":
            selected[key] = interaction
        elif keep == "highest_score":
            current = _interaction_float(_interaction_get(selected[key], _INTERACTION_SCORE_FIELDS), None)
            candidate = _interaction_float(_interaction_get(interaction, _INTERACTION_SCORE_FIELDS), None)
            if candidate is not None and (current is None or candidate > current):
                selected[key] = interaction
        elif keep == "shortest_distance":
            current = _interaction_float(_interaction_get(selected[key], _INTERACTION_DISTANCE_FIELDS), None)
            candidate = _interaction_float(_interaction_get(interaction, _INTERACTION_DISTANCE_FIELDS), None)
            if candidate is not None and (current is None or candidate < current):
                selected[key] = interaction
        elif keep != "first":
            raise ExportConfigurationError(f"Unsupported deduplication mode: {keep!r}.")
    return [selected[key] for key in order]


def _interaction_residue_labels(record: Mapping[str, Any]) -> Set[str]:
    """Collect residue labels from one serialized record."""
    labels: Set[str] = set()
    for participant in record.get("participants", ()):
        if not isinstance(participant, Mapping):
            continue
        residue = participant.get("residue")
        if isinstance(residue, Mapping):
            label = residue.get("identifier") or residue.get("residue_id") or residue.get("label")
            if label:
                labels.add(str(label))
    return labels


def summarize_interaction_records(records: Sequence[Mapping[str, Any]]) -> InteractionCollectionSummary:
    """Summarize serialized interaction records."""
    by_family: Dict[str, int] = defaultdict(int)
    by_type: Dict[str, int] = defaultdict(int)
    by_strength: Dict[str, int] = defaultdict(int)
    residues: Set[str] = set()
    poses: Set[str] = set()
    distances: List[float] = []
    scores: List[float] = []

    for record in records:
        by_family[str(record.get("family", "unknown"))] += 1
        by_type[str(record.get("interaction_type", "unknown"))] += 1
        if record.get("strength") is not None:
            by_strength[str(record["strength"])] += 1
        residues.update(_interaction_residue_labels(record))
        if record.get("pose") is not None:
            poses.add(str(record["pose"]))
        geometry = record.get("geometry")
        if isinstance(geometry, Mapping):
            distance = _interaction_float(geometry.get("distance"), None)
            if distance is not None:
                distances.append(distance)
        score = _interaction_float(record.get("score"), None)
        if score is not None:
            scores.append(score)

    return InteractionCollectionSummary(
        total=len(records),
        by_family=dict(sorted(by_family.items())),
        by_type=dict(sorted(by_type.items())),
        by_strength=dict(sorted(by_strength.items())),
        residues=tuple(sorted(residues)),
        poses=tuple(sorted(poses)),
        distance_min=min(distances) if distances else None,
        distance_max=max(distances) if distances else None,
        distance_mean=(sum(distances) / len(distances)) if distances else None,
        score_total=sum(scores) if scores else None,
    )


def interactions_to_records(
    interactions: Iterable[Any],
    *,
    options: Optional[InteractionExportOptions] = None,
    deduplicate: bool = False,
    deduplication_keep: Literal["first", "last", "highest_score", "shortest_distance"] = "first",
    include_summary: bool = False,
) -> Union[List[Dict[str, Any]], Dict[str, Any]]:
    """Convert an interaction collection."""
    options = options or InteractionExportOptions()
    items = list(interactions)
    if deduplicate:
        items = deduplicate_interactions(items, keep=deduplication_keep)
    context = InteractionRecordContext(options=options)
    records = [interaction_to_record(item, options=options, context=context) for item in items]
    if not include_summary:
        return records
    return {
        "schema_version": INTERACTION_SCHEMA_VERSION,
        "object_type": "interaction_collection",
        "interactions": records,
        "summary": summarize_interaction_records(records).to_dict(),
        "warnings": list(context.warnings),
    }


def group_interactions_by_family(interactions: Iterable[Any]) -> Dict[str, List[Any]]:
    """Group interactions by canonical family."""
    grouped: Dict[str, List[Any]] = defaultdict(list)
    for interaction in interactions:
        grouped[infer_interaction_family(interaction)].append(interaction)
    return dict(sorted(grouped.items()))


def group_interactions_by_residue(interactions: Iterable[Any]) -> Dict[str, List[Any]]:
    """Group interactions by participant residue."""
    grouped: Dict[str, List[Any]] = defaultdict(list)
    for interaction in interactions:
        labels: Set[str] = set()
        for participant in extract_interaction_participants(interaction):
            if participant.residue is not None:
                labels.add(residue_identifier(participant.residue))
            elif participant.atom is not None:
                residue = _interaction_get(participant.atom, ("residue",))
                if residue is not None:
                    labels.add(residue_identifier(residue))
        for label in labels:
            grouped[label].append(interaction)
    return dict(sorted(grouped.items()))


def interaction_serializer(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize one registry-resolved interaction."""
    options = None
    if context is not None:
        candidate = context.metadata.get("interaction_options")
        if isinstance(candidate, InteractionExportOptions):
            options = candidate
    return interaction_to_record(value, options=options)


def register_interaction_serializer(
    registry: Optional[SerializerRegistry] = None,
    *,
    priority: int = 850,
    replace_existing: bool = True,
) -> SerializerRegistry:
    """Register the generic interaction serializer."""
    target = registry or DEFAULT_SERIALIZER_REGISTRY
    target.register(
        "molecular_interaction",
        interaction_serializer,
        predicate=is_interaction_like,
        match_mode=SerializerMatchMode.PREDICATE,
        priority=priority,
        aliases=("interaction", "contact_interaction"),
        builtin=True,
        description="Serialize DockAnalyzer molecular interactions.",
        replace_existing=replace_existing,
    )
    return target


register_interaction_serializer()

__all__.extend([
    "INTERACTION_SCHEMA_VERSION",
    "InteractionOrientation",
    "InteractionGranularity",
    "InteractionExportOptions",
    "InteractionParticipant",
    "InteractionRecordContext",
    "InteractionCollectionSummary",
    "normalize_interaction_family",
    "normalize_interaction_type",
    "infer_interaction_family",
    "infer_interaction_type",
    "is_interaction_like",
    "interaction_identifier",
    "participant_label",
    "extract_interaction_participants",
    "interaction_participant_to_record",
    "extract_interaction_geometry",
    "extract_interaction_metadata",
    "interaction_to_record",
    "interaction_reference",
    "interaction_deduplication_key",
    "deduplicate_interactions",
    "summarize_interaction_records",
    "interactions_to_records",
    "group_interactions_by_family",
    "group_interactions_by_residue",
    "interaction_serializer",
    "register_interaction_serializer",
])
# =============================================================================
# Section 10 — Analysis results
# =============================================================================

ANALYSIS_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_ANALYSIS_FLOAT_PRECISION: Final[int] = 6

_ANALYSIS_ID_FIELDS: Final[Tuple[str, ...]] = (
    "analysis_id", "result_id", "id", "uid", "identifier", "key",
)
_ANALYSIS_NAME_FIELDS: Final[Tuple[str, ...]] = (
    "analysis_name", "name", "title", "label",
)
_ANALYSIS_TYPE_FIELDS: Final[Tuple[str, ...]] = (
    "analysis_type", "result_type", "type", "kind", "module", "category",
)
_ANALYSIS_POSE_FIELDS: Final[Tuple[str, ...]] = (
    "pose", "pose_id", "pose_name", "model", "dock_model", "structure",
)
_ANALYSIS_INTERACTION_FIELDS: Final[Tuple[str, ...]] = (
    "interactions", "contacts", "hbonds", "hydrogen_bonds", "hydrophobic",
    "pi", "pi_interactions", "saltbridge", "salt_bridges",
)
_ANALYSIS_STATISTIC_FIELDS: Final[Tuple[str, ...]] = (
    "statistics", "stats", "summary", "metrics", "counts",
)
_ANALYSIS_SCORE_FIELDS: Final[Tuple[str, ...]] = (
    "score", "total_score", "interaction_score", "normalized_score",
)
_ANALYSIS_METADATA_FIELDS: Final[Tuple[str, ...]] = (
    "metadata", "details", "extra", "properties", "annotations",
)
_ANALYSIS_WARNING_FIELDS: Final[Tuple[str, ...]] = (
    "warnings", "warning_messages", "issues",
)
_ANALYSIS_ERROR_FIELDS: Final[Tuple[str, ...]] = (
    "errors", "error_messages", "failures",
)
_ANALYSIS_TIMESTAMP_FIELDS: Final[Tuple[str, ...]] = (
    "created_at", "timestamp", "analyzed_at", "completed_at",
)


class AnalysisGranularity(str, Enum):
    """Control analysis detail."""

    COMPACT = "compact"
    STANDARD = "standard"
    FULL = "full"


class AnalysisInteractionLayout(str, Enum):
    """Control interaction grouping."""

    FLAT = "flat"
    BY_FAMILY = "by_family"
    PRESERVE = "preserve"


@dataclass
class AnalysisExportOptions:
    """Control analysis record generation."""

    include_schema: bool = True
    include_id: bool = True
    include_name: bool = True
    include_type: bool = True
    include_pose: bool = True
    include_interactions: bool = True
    include_statistics: bool = True
    include_scores: bool = True
    include_metadata: bool = True
    include_warnings: bool = True
    include_errors: bool = True
    include_timestamp: bool = True
    include_raw_fields: bool = False
    include_empty: bool = False
    omit_none: bool = True
    strict: bool = False
    granularity: AnalysisGranularity = AnalysisGranularity.STANDARD
    interaction_layout: AnalysisInteractionLayout = AnalysisInteractionLayout.BY_FAMILY
    float_precision: Optional[int] = DEFAULT_ANALYSIS_FLOAT_PRECISION
    molecular_options: MolecularExportOptions = field(default_factory=MolecularExportOptions)
    interaction_options: InteractionExportOptions = field(default_factory=InteractionExportOptions)

    def __post_init__(self) -> None:
        if isinstance(self.granularity, str):
            self.granularity = AnalysisGranularity(self.granularity)
        if isinstance(self.interaction_layout, str):
            self.interaction_layout = AnalysisInteractionLayout(self.interaction_layout)
        if self.float_precision is not None and self.float_precision < 0:
            raise ExportConfigurationError("float_precision must be non-negative or None.")
        self.interaction_options.molecular_options = self.molecular_options

    def to_dict(self) -> Dict[str, Any]:
        """Return option values."""
        result = {item.name: getattr(self, item.name) for item in fields(self)}
        result["granularity"] = self.granularity.value
        result["interaction_layout"] = self.interaction_layout.value
        result["molecular_options"] = self.molecular_options.to_dict()
        result["interaction_options"] = self.interaction_options.to_dict()
        return result


@dataclass
class AnalysisRecordContext:
    """Track analysis conversion state."""

    options: AnalysisExportOptions = field(default_factory=AnalysisExportOptions)
    active_ids: Set[int] = field(default_factory=set)
    warnings: List[str] = field(default_factory=list)

    def warn(self, message: str) -> None:
        """Store one warning."""
        self.warnings.append(message)


@dataclass(frozen=True)
class AnalysisCollectionSummary:
    """Summarize analysis records."""

    total: int
    successful: int
    failed: int
    with_warnings: int
    by_type: Mapping[str, int]
    interaction_total: int
    score_min: Optional[float] = None
    score_max: Optional[float] = None
    score_mean: Optional[float] = None

    def to_dict(self) -> Dict[str, Any]:
        """Return serializable summary data."""
        return {
            "total": self.total,
            "successful": self.successful,
            "failed": self.failed,
            "with_warnings": self.with_warnings,
            "by_type": dict(self.by_type),
            "interaction_total": self.interaction_total,
            "score_min": self.score_min,
            "score_max": self.score_max,
            "score_mean": self.score_mean,
        }


@dataclass(frozen=True)
class AnalysisBundle:
    """Store a normalized analysis result."""

    analysis_type: str
    interactions: Tuple[Any, ...] = ()
    statistics: Mapping[str, Any] = field(default_factory=dict)
    scores: Mapping[str, Any] = field(default_factory=dict)
    metadata: Mapping[str, Any] = field(default_factory=dict)
    pose: Any = None
    analysis_id: Optional[str] = None
    name: Optional[str] = None
    warnings: Tuple[str, ...] = ()
    errors: Tuple[str, ...] = ()
    created_at: Optional[Union[str, datetime]] = None

    def __post_init__(self) -> None:
        object.__setattr__(self, "analysis_type", normalize_analysis_type(self.analysis_type))
        object.__setattr__(self, "interactions", tuple(self.interactions))
        object.__setattr__(self, "statistics", MappingProxyType(dict(self.statistics)))
        object.__setattr__(self, "scores", MappingProxyType(dict(self.scores)))
        object.__setattr__(self, "metadata", MappingProxyType(dict(self.metadata)))
        object.__setattr__(self, "warnings", tuple(str(item) for item in self.warnings))
        object.__setattr__(self, "errors", tuple(str(item) for item in self.errors))


def _analysis_get(value: Any, names: Sequence[str], default: Any = None) -> Any:
    """Read the first available field."""
    return _plain_field_get(value, names, default)


def _analysis_mapping(value: Any) -> Dict[str, Any]:
    """Return visible object fields."""
    if isinstance(value, Mapping):
        return dict(value)
    if is_dataclass(value) and not isinstance(value, type):
        return {item.name: getattr(value, item.name) for item in fields(value)}
    try:
        return dict(vars(value))
    except (TypeError, ValueError):
        return {}


def _analysis_token(value: Any, default: str = "analysis") -> str:
    """Normalize a token."""
    if value is None:
        return default
    text = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")
    return text or default


def normalize_analysis_type(value: Any, default: str = "analysis") -> str:
    """Return a canonical analysis type."""
    token = _analysis_token(value, default)
    aliases = {
        "contact": "contacts",
        "hbond": "hbonds",
        "hydrogen_bond": "hbonds",
        "hydrogen_bonds": "hbonds",
        "salt_bridge": "saltbridge",
        "salt_bridges": "saltbridge",
        "pi_interaction": "pi",
        "pi_interactions": "pi",
        "multi_pose": "multipose",
        "multi_pose_analysis": "multipose",
        "pose_analysis": "pose",
    }
    return aliases.get(token, token)


def infer_analysis_type(value: Any) -> str:
    """Infer an analysis type."""
    explicit = _analysis_get(value, _ANALYSIS_TYPE_FIELDS)
    if explicit is not None:
        return normalize_analysis_type(explicit)
    name = value.__class__.__name__.lower()
    for token in ("saltbridge", "hydrophobic", "hbond", "contact", "pi", "multipose", "pose"):
        if token in name:
            return normalize_analysis_type(token)
    fields_map = _analysis_mapping(value)
    present = [name for name in _ANALYSIS_INTERACTION_FIELDS if name in fields_map]
    if len(present) == 1:
        return normalize_analysis_type(present[0])
    if present:
        return "combined"
    return "analysis"


def is_analysis_result_like(value: Any) -> bool:
    """Return whether a value resembles an analysis result."""
    if value is None or isinstance(value, (str, bytes, int, float, bool, Path)):
        return False
    if is_interaction_like(value) or is_atom_like(value) or is_residue_like(value):
        return False
    fields_map = _analysis_mapping(value)
    names = set(fields_map)
    signals = (
        set(_ANALYSIS_INTERACTION_FIELDS)
        | set(_ANALYSIS_STATISTIC_FIELDS)
        | set(_ANALYSIS_SCORE_FIELDS)
        | set(_ANALYSIS_ERROR_FIELDS)
    )
    if names.intersection(signals):
        return True
    class_name = value.__class__.__name__.lower()
    return "analysis" in class_name or class_name.endswith("result")


def analysis_identifier(value: Any) -> str:
    """Return a stable analysis identifier."""
    explicit = _analysis_get(value, _ANALYSIS_ID_FIELDS)
    if explicit not in (None, ""):
        return str(explicit)
    analysis_type = infer_analysis_type(value)
    pose = _analysis_get(value, _ANALYSIS_POSE_FIELDS)
    pose_id = structure_identifier(pose) if pose is not None and is_structure_like(pose) else str(pose or "")
    name = _analysis_get(value, _ANALYSIS_NAME_FIELDS, "")
    payload = f"{analysis_type}|{pose_id}|{name}"
    return f"analysis:{sha256(payload.encode('utf-8')).hexdigest()[:16]}"


def _analysis_float(value: Any, precision: Optional[int]) -> Optional[float]:
    """Return a finite float."""
    try:
        number = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if not math.isfinite(number):
        return None
    return round(number, precision) if precision is not None else number


def _analysis_sequence(value: Any) -> List[Any]:
    """Return a non-string sequence as a list."""
    if value is None:
        return []
    if isinstance(value, Mapping):
        return list(value.values())
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return list(value)
    if isinstance(value, Iterable) and not isinstance(value, (str, bytes, bytearray)):
        try:
            return list(value)
        except TypeError:
            pass
    return [value]


def extract_analysis_interactions(value: Any) -> Dict[str, List[Any]]:
    """Extract interactions grouped by family."""
    grouped: Dict[str, List[Any]] = defaultdict(list)
    fields_map = _analysis_mapping(value)
    generic = fields_map.get("interactions")
    if generic is not None:
        if isinstance(generic, Mapping):
            for family, items in generic.items():
                canonical = normalize_interaction_family(family)
                grouped[canonical].extend(_analysis_sequence(items))
        else:
            for item in _analysis_sequence(generic):
                grouped[infer_interaction_family(item)].append(item)
    aliases = {
        "contacts": "contact",
        "hbonds": "hbond",
        "hydrogen_bonds": "hbond",
        "hydrophobic": "hydrophobic",
        "pi": "pi",
        "pi_interactions": "pi",
        "saltbridge": "saltbridge",
        "salt_bridges": "saltbridge",
    }
    for field_name, family in aliases.items():
        if field_name in fields_map:
            grouped[family].extend(_analysis_sequence(fields_map[field_name]))
    return {family: items for family, items in sorted(grouped.items()) if items}


def extract_analysis_statistics(value: Any) -> Dict[str, Any]:
    """Extract statistics and metrics."""
    result: Dict[str, Any] = {}
    fields_map = _analysis_mapping(value)
    for field_name in _ANALYSIS_STATISTIC_FIELDS:
        candidate = fields_map.get(field_name)
        if isinstance(candidate, Mapping):
            result.update(candidate)
    for name, candidate in fields_map.items():
        token = _analysis_token(name)
        if token.startswith(("count_", "total_", "mean_", "median_", "min_", "max_")):
            result.setdefault(name, candidate)
    return result


def extract_analysis_scores(value: Any) -> Dict[str, Any]:
    """Extract score fields."""
    result: Dict[str, Any] = {}
    fields_map = _analysis_mapping(value)
    nested = fields_map.get("scores")
    if isinstance(nested, Mapping):
        result.update(nested)
    for field_name in _ANALYSIS_SCORE_FIELDS:
        if field_name in fields_map:
            result.setdefault(field_name, fields_map[field_name])
    for name, candidate in fields_map.items():
        if "score" in _analysis_token(name):
            result.setdefault(name, candidate)
    return result


def extract_analysis_messages(value: Any, fields_: Sequence[str]) -> List[str]:
    """Extract warning or error messages."""
    messages: List[str] = []
    for candidate in (_analysis_get(value, (name,)) for name in fields_):
        if candidate is None:
            continue
        for item in _analysis_sequence(candidate):
            if isinstance(item, BaseException):
                messages.append(str(item))
            elif isinstance(item, Mapping):
                messages.append(str(item.get("message", item.get("error", item))))
            else:
                messages.append(str(item))
    return list(dict.fromkeys(message for message in messages if message))


def extract_analysis_metadata(value: Any) -> Dict[str, Any]:
    """Extract metadata mappings."""
    result: Dict[str, Any] = {}
    for field_name in _ANALYSIS_METADATA_FIELDS:
        candidate = _analysis_get(value, (field_name,))
        if isinstance(candidate, Mapping):
            result.update(candidate)
    return result


def _serialize_analysis_interactions(
    grouped: Mapping[str, Sequence[Any]],
    options: AnalysisExportOptions,
) -> Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]]:
    """Serialize grouped interactions."""
    by_family = {
        family: [interaction_to_record(item, options=options.interaction_options) for item in items]
        for family, items in grouped.items()
    }
    if options.interaction_layout == AnalysisInteractionLayout.BY_FAMILY:
        return by_family
    flat = [record for records in by_family.values() for record in records]
    if options.interaction_layout == AnalysisInteractionLayout.FLAT:
        return flat
    return by_family if len(by_family) > 1 else next(iter(by_family.values()), [])


def _analysis_serialize_mapping(value: Mapping[str, Any], options: AnalysisExportOptions) -> Dict[str, Any]:
    """Serialize one mapping."""
    converted = make_json_safe(value)
    if isinstance(converted, Mapping):
        return dict(converted)
    return {"value": converted}


def analysis_to_record(
    value: Any,
    *,
    options: Optional[AnalysisExportOptions] = None,
    context: Optional[AnalysisRecordContext] = None,
) -> Dict[str, Any]:
    """Convert one analysis result to a record."""
    options = options or AnalysisExportOptions()
    context = context or AnalysisRecordContext(options=options)
    object_id = id(value)
    if object_id in context.active_ids:
        raise ExportSerializationError("Circular analysis result reference detected.")
    context.active_ids.add(object_id)
    try:
        record: Dict[str, Any] = {}
        if options.include_schema:
            record["schema_version"] = ANALYSIS_SCHEMA_VERSION
            record["object_type"] = "analysis_result"
        if options.include_id:
            record["analysis_id"] = analysis_identifier(value)
        if options.include_name:
            name = _analysis_get(value, _ANALYSIS_NAME_FIELDS)
            if name is not None:
                record["name"] = str(name)
        if options.include_type:
            record["analysis_type"] = infer_analysis_type(value)
        if options.include_pose:
            pose = _analysis_get(value, _ANALYSIS_POSE_FIELDS)
            if pose is not None:
                record["pose"] = (
                    structure_to_record(pose, options=options.molecular_options)
                    if is_structure_like(pose)
                    else make_json_safe(pose)
                )
        grouped = extract_analysis_interactions(value)
        if options.include_interactions and (grouped or options.include_empty):
            record["interactions"] = _serialize_analysis_interactions(grouped, options)
            record["interaction_count"] = sum(len(items) for items in grouped.values())
        statistics = extract_analysis_statistics(value)
        if options.include_statistics and (statistics or options.include_empty):
            record["statistics"] = _analysis_serialize_mapping(statistics, options)
        scores = extract_analysis_scores(value)
        if options.include_scores and (scores or options.include_empty):
            serialized_scores: Dict[str, Any] = {}
            for name, candidate in scores.items():
                number = _analysis_float(candidate, options.float_precision)
                serialized_scores[name] = number if number is not None else make_json_safe(candidate)
            record["scores"] = serialized_scores
        metadata = extract_analysis_metadata(value)
        if options.include_metadata and (metadata or options.include_empty):
            record["metadata"] = _analysis_serialize_mapping(metadata, options)
        warnings_ = extract_analysis_messages(value, _ANALYSIS_WARNING_FIELDS)
        errors = extract_analysis_messages(value, _ANALYSIS_ERROR_FIELDS)
        if options.include_warnings and (warnings_ or options.include_empty):
            record["warnings"] = warnings_
        if options.include_errors and (errors or options.include_empty):
            record["errors"] = errors
        if options.include_timestamp:
            timestamp = _analysis_get(value, _ANALYSIS_TIMESTAMP_FIELDS)
            if timestamp is not None:
                record["timestamp"] = make_json_safe(timestamp)
        record["status"] = "failed" if errors else "success"
        if options.include_raw_fields:
            excluded = set(
                _ANALYSIS_ID_FIELDS + _ANALYSIS_NAME_FIELDS + _ANALYSIS_TYPE_FIELDS
                + _ANALYSIS_POSE_FIELDS + _ANALYSIS_INTERACTION_FIELDS
                + _ANALYSIS_STATISTIC_FIELDS + _ANALYSIS_SCORE_FIELDS
                + _ANALYSIS_METADATA_FIELDS + _ANALYSIS_WARNING_FIELDS
                + _ANALYSIS_ERROR_FIELDS + _ANALYSIS_TIMESTAMP_FIELDS
            )
            raw = {key: item for key, item in _analysis_mapping(value).items() if key not in excluded}
            if raw:
                record["raw_fields"] = make_json_safe(raw)
        if options.omit_none:
            record = {key: item for key, item in record.items() if item is not None}
        if not options.include_empty:
            record = {
                key: item for key, item in record.items()
                if item not in ({}, [], (), "") or key in {"status", "analysis_type", "analysis_id"}
            }
        return record
    except ExportError:
        raise
    except Exception as exc:
        if options.strict:
            raise ExportSerializationError("Failed to serialize analysis result.") from exc
        context.warn(str(exc))
        return {
            "schema_version": ANALYSIS_SCHEMA_VERSION,
            "object_type": "analysis_result",
            "analysis_id": analysis_identifier(value),
            "analysis_type": infer_analysis_type(value),
            "status": "failed",
            "errors": [str(exc)],
        }
    finally:
        context.active_ids.discard(object_id)


def analysis_reference(value: Any) -> Dict[str, Any]:
    """Return a compact analysis reference."""
    return {
        "analysis_id": analysis_identifier(value),
        "analysis_type": infer_analysis_type(value),
        "name": _analysis_get(value, _ANALYSIS_NAME_FIELDS),
    }


def summarize_analysis_records(records: Iterable[Mapping[str, Any]]) -> AnalysisCollectionSummary:
    """Summarize serialized analysis records."""
    items = list(records)
    by_type: Dict[str, int] = defaultdict(int)
    scores: List[float] = []
    successful = 0
    failed = 0
    with_warnings = 0
    interaction_total = 0
    for record in items:
        by_type[str(record.get("analysis_type", "analysis"))] += 1
        if record.get("status") == "failed" or record.get("errors"):
            failed += 1
        else:
            successful += 1
        if record.get("warnings"):
            with_warnings += 1
        try:
            interaction_total += int(record.get("interaction_count", 0) or 0)
        except (TypeError, ValueError):
            pass
        score_map = record.get("scores")
        if isinstance(score_map, Mapping):
            preferred = score_map.get("total_score", score_map.get("score"))
            number = _analysis_float(preferred, None)
            if number is not None:
                scores.append(number)
    return AnalysisCollectionSummary(
        total=len(items),
        successful=successful,
        failed=failed,
        with_warnings=with_warnings,
        by_type=dict(sorted(by_type.items())),
        interaction_total=interaction_total,
        score_min=min(scores) if scores else None,
        score_max=max(scores) if scores else None,
        score_mean=(sum(scores) / len(scores)) if scores else None,
    )


def analyses_to_records(
    analyses: Iterable[Any],
    *,
    options: Optional[AnalysisExportOptions] = None,
    include_summary: bool = False,
) -> Union[List[Dict[str, Any]], Dict[str, Any]]:
    """Convert an analysis collection."""
    options = options or AnalysisExportOptions()
    context = AnalysisRecordContext(options=options)
    records = [analysis_to_record(item, options=options, context=context) for item in analyses]
    if not include_summary:
        return records
    return {
        "schema_version": ANALYSIS_SCHEMA_VERSION,
        "object_type": "analysis_collection",
        "analyses": records,
        "summary": summarize_analysis_records(records).to_dict(),
        "warnings": list(context.warnings),
    }


def group_analyses_by_type(analyses: Iterable[Any]) -> Dict[str, List[Any]]:
    """Group analyses by canonical type."""
    grouped: Dict[str, List[Any]] = defaultdict(list)
    for analysis in analyses:
        grouped[infer_analysis_type(analysis)].append(analysis)
    return dict(sorted(grouped.items()))


def analysis_serializer(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize one registry-resolved analysis result."""
    options = None
    if context is not None:
        candidate = context.metadata.get("analysis_options")
        if isinstance(candidate, AnalysisExportOptions):
            options = candidate
    return analysis_to_record(value, options=options)


def register_analysis_serializer(
    registry: Optional[SerializerRegistry] = None,
    *,
    priority: int = 800,
    replace_existing: bool = True,
) -> SerializerRegistry:
    """Register the generic analysis serializer."""
    target = registry or DEFAULT_SERIALIZER_REGISTRY
    target.register(
        "analysis_result",
        analysis_serializer,
        predicate=is_analysis_result_like,
        match_mode=SerializerMatchMode.PREDICATE,
        priority=priority,
        aliases=("analysis", "result"),
        builtin=True,
        description="Serialize DockAnalyzer analysis results.",
        replace_existing=replace_existing,
    )
    return target


register_analysis_serializer()

__all__.extend([
    "ANALYSIS_SCHEMA_VERSION",
    "AnalysisGranularity",
    "AnalysisInteractionLayout",
    "AnalysisExportOptions",
    "AnalysisRecordContext",
    "AnalysisCollectionSummary",
    "AnalysisBundle",
    "normalize_analysis_type",
    "infer_analysis_type",
    "is_analysis_result_like",
    "analysis_identifier",
    "extract_analysis_interactions",
    "extract_analysis_statistics",
    "extract_analysis_scores",
    "extract_analysis_messages",
    "extract_analysis_metadata",
    "analysis_to_record",
    "analysis_reference",
    "summarize_analysis_records",
    "analyses_to_records",
    "group_analyses_by_type",
    "analysis_serializer",
    "register_analysis_serializer",
])
# =============================================================================
# DockAnalyzer — Data export and serialization
# Section 11 — Scoring
# =============================================================================

"""Serialization helpers for DockAnalyzer scoring results."""

SCORING_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_SCORE_PRECISION: Final[int] = 6
DEFAULT_SCORE_SERIALIZER_PRIORITY: Final[int] = 850

_SCORE_TOTAL_FIELDS: Final[Tuple[str, ...]] = (
    "total_score",
    "score_total",
    "final_score",
    "combined_score",
    "consensus_score",
    "score",
)
_SCORE_COMPONENT_FIELDS: Final[Tuple[str, ...]] = (
    "components",
    "score_components",
    "component_scores",
    "terms",
    "contributions",
)
_SCORE_EXTERNAL_FIELDS: Final[Tuple[str, ...]] = (
    "external_scores",
    "docking_scores",
    "affinities",
    "external",
)
_SCORE_NORMALIZATION_FIELDS: Final[Tuple[str, ...]] = (
    "normalization",
    "normalization_info",
    "normalized",
)
_SCORE_EXPLANATION_FIELDS: Final[Tuple[str, ...]] = (
    "explanation",
    "explainability",
    "rationale",
    "details",
)
_SCORE_RANK_FIELDS: Final[Tuple[str, ...]] = (
    "rank",
    "ranking",
    "position",
)
_SCORE_POSE_FIELDS: Final[Tuple[str, ...]] = (
    "pose_id",
    "pose_name",
    "model_id",
    "name",
)


class ScoreDirection(str, Enum):
    """Define whether larger or smaller scores are better."""

    HIGHER_IS_BETTER = "higher_is_better"
    LOWER_IS_BETTER = "lower_is_better"
    UNKNOWN = "unknown"


class ScoreLayout(str, Enum):
    """Define score component layout."""

    MAPPING = "mapping"
    RECORDS = "records"
    BOTH = "both"


@dataclass(frozen=True)
class ScoringExportOptions:
    """Configure scoring serialization."""

    precision: Optional[int] = DEFAULT_SCORE_PRECISION
    component_layout: ScoreLayout = ScoreLayout.RECORDS
    include_components: bool = True
    include_external_scores: bool = True
    include_normalization: bool = True
    include_explanation: bool = True
    include_metadata: bool = True
    include_raw_values: bool = False
    include_rank: bool = True
    include_pose_reference: bool = True
    omit_none: bool = True
    stable_identifier: bool = True

    def __post_init__(self) -> None:
        if self.precision is not None and self.precision < 0:
            raise ExportConfigurationError("Score precision must be non-negative.")
        object.__setattr__(self, "component_layout", ScoreLayout(self.component_layout))


@dataclass(frozen=True)
class ScoreComponentRecord:
    """Store one score component."""

    name: str
    value: Optional[float]
    weight: Optional[float] = None
    weighted_value: Optional[float] = None
    normalized_value: Optional[float] = None
    direction: ScoreDirection = ScoreDirection.UNKNOWN
    family: Optional[str] = None
    source: Optional[str] = None
    description: Optional[str] = None
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def to_dict(self, *, precision: Optional[int] = DEFAULT_SCORE_PRECISION) -> Dict[str, Any]:
        """Return a serializable component record."""
        record = {
            "name": self.name,
            "value": _score_round(self.value, precision),
            "weight": _score_round(self.weight, precision),
            "weighted_value": _score_round(self.weighted_value, precision),
            "normalized_value": _score_round(self.normalized_value, precision),
            "direction": self.direction.value,
            "family": self.family,
            "source": self.source,
            "description": self.description,
            "metadata": to_serializable(dict(self.metadata)),
        }
        return {key: value for key, value in record.items() if value is not None}


@dataclass(frozen=True)
class ExternalScoreRecord:
    """Store one external docking score."""

    name: str
    value: Optional[float]
    unit: Optional[str] = None
    direction: ScoreDirection = ScoreDirection.LOWER_IS_BETTER
    source: Optional[str] = None
    normalized_value: Optional[float] = None
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def to_dict(self, *, precision: Optional[int] = DEFAULT_SCORE_PRECISION) -> Dict[str, Any]:
        """Return a serializable external score."""
        record = {
            "name": self.name,
            "value": _score_round(self.value, precision),
            "unit": self.unit,
            "direction": self.direction.value,
            "source": self.source,
            "normalized_value": _score_round(self.normalized_value, precision),
            "metadata": to_serializable(dict(self.metadata)),
        }
        return {key: value for key, value in record.items() if value is not None}


@dataclass
class ScoringCollectionSummary:
    """Summarize a score collection."""

    count: int = 0
    valid_count: int = 0
    missing_count: int = 0
    minimum: Optional[float] = None
    maximum: Optional[float] = None
    mean: Optional[float] = None
    median: Optional[float] = None
    best_identifier: Optional[str] = None
    best_score: Optional[float] = None
    direction: ScoreDirection = ScoreDirection.UNKNOWN

    def to_dict(self, *, precision: Optional[int] = DEFAULT_SCORE_PRECISION) -> Dict[str, Any]:
        """Return serializable summary data."""
        return {
            "count": self.count,
            "valid_count": self.valid_count,
            "missing_count": self.missing_count,
            "minimum": _score_round(self.minimum, precision),
            "maximum": _score_round(self.maximum, precision),
            "mean": _score_round(self.mean, precision),
            "median": _score_round(self.median, precision),
            "best_identifier": self.best_identifier,
            "best_score": _score_round(self.best_score, precision),
            "direction": self.direction.value,
        }


def _score_get(value: Any, names: Sequence[str], default: Any = None) -> Any:
    """Return the first available score field."""
    return _plain_field_get(value, names, default)


def _score_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    """Convert a value to a finite float."""
    if value is None or isinstance(value, bool):
        return default
    try:
        number = float(value)
    except (TypeError, ValueError, OverflowError):
        return default
    if not math.isfinite(number):
        return default
    return number


def _score_round(value: Optional[float], precision: Optional[int]) -> Optional[float]:
    """Round one score value."""
    if value is None:
        return None
    return round(float(value), precision) if precision is not None else float(value)


def normalize_score_direction(value: Any) -> ScoreDirection:
    """Normalize score direction labels."""
    if isinstance(value, ScoreDirection):
        return value
    text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if text in {"higher", "higher_is_better", "maximize", "max", "descending"}:
        return ScoreDirection.HIGHER_IS_BETTER
    if text in {"lower", "lower_is_better", "minimize", "min", "ascending"}:
        return ScoreDirection.LOWER_IS_BETTER
    return ScoreDirection.UNKNOWN


def infer_score_direction(value: Any) -> ScoreDirection:
    """Infer score direction from explicit fields or names."""
    explicit = _score_get(value, ("direction", "score_direction", "optimization"))
    direction = normalize_score_direction(explicit)
    if direction is not ScoreDirection.UNKNOWN:
        return direction
    name = str(_score_get(value, ("name", "score_name", "metric"), "")).lower()
    if any(token in name for token in ("affinity", "energy", "rmsd", "penalty", "loss")):
        return ScoreDirection.LOWER_IS_BETTER
    if any(token in name for token in ("confidence", "quality", "coverage", "score")):
        return ScoreDirection.HIGHER_IS_BETTER
    return ScoreDirection.UNKNOWN


def is_scoring_result_like(value: Any) -> bool:
    """Return whether an object resembles a scoring result."""
    if value is None or isinstance(value, (str, bytes, int, float, bool)):
        return False
    if isinstance(value, Mapping):
        keys = {str(key).lower() for key in value}
    else:
        keys = {name.lower() for name in dir(value)}
    score_keys = set(_SCORE_TOTAL_FIELDS + _SCORE_COMPONENT_FIELDS + _SCORE_EXTERNAL_FIELDS)
    if keys & score_keys:
        return True
    class_name = type(value).__name__.lower()
    return "score" in class_name or "scoring" in class_name


def scoring_identifier(value: Any, *, stable: bool = True) -> str:
    """Return a stable scoring identifier."""
    explicit = _score_get(value, ("scoring_id", "score_id", "id", "identifier"))
    if explicit not in (None, ""):
        return str(explicit)
    pose = _score_get(value, _SCORE_POSE_FIELDS)
    label = str(pose or type(value).__name__)
    if not stable:
        return sanitize_filename(label, default="scoring")
    payload = {
        "pose": pose,
        "score": _score_get(value, _SCORE_TOTAL_FIELDS),
        "components": _score_get(value, _SCORE_COMPONENT_FIELDS),
    }
    digest = sha256(json.dumps(to_serializable(payload), sort_keys=True).encode("utf-8")).hexdigest()[:12]
    return f"score_{digest}"


def _component_name(value: Any, fallback: str) -> str:
    """Return a component name."""
    candidate = _score_get(value, ("name", "component", "term", "metric", "type"))
    return str(candidate or fallback)


def score_component_to_record(
    value: Any,
    *,
    name: Optional[str] = None,
    precision: Optional[int] = DEFAULT_SCORE_PRECISION,
) -> Dict[str, Any]:
    """Convert one score component to a record."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        component = ScoreComponentRecord(name=name or "score", value=_score_float(value))
        return component.to_dict(precision=precision)

    raw_value = _score_get(value, ("value", "score", "raw_value", "component_score"))
    weight = _score_float(_score_get(value, ("weight", "coefficient", "factor")))
    weighted = _score_float(_score_get(value, ("weighted_value", "weighted_score", "contribution")))
    numeric = _score_float(raw_value)
    if weighted is None and numeric is not None and weight is not None:
        weighted = numeric * weight
    component = ScoreComponentRecord(
        name=name or _component_name(value, "score"),
        value=numeric,
        weight=weight,
        weighted_value=weighted,
        normalized_value=_score_float(_score_get(value, ("normalized_value", "normalized_score"))),
        direction=infer_score_direction(value),
        family=_score_get(value, ("family", "category", "group")),
        source=_score_get(value, ("source", "module", "origin")),
        description=_score_get(value, ("description", "label")),
        metadata=to_serializable(_score_get(value, ("metadata", "meta"), {})),
    )
    return component.to_dict(precision=precision)


def extract_score_components(
    value: Any,
    *,
    precision: Optional[int] = DEFAULT_SCORE_PRECISION,
) -> List[Dict[str, Any]]:
    """Extract score component records."""
    components = _score_get(value, _SCORE_COMPONENT_FIELDS, {})
    records: List[Dict[str, Any]] = []
    if isinstance(components, Mapping):
        for name, component in components.items():
            records.append(score_component_to_record(component, name=str(name), precision=precision))
    elif isinstance(components, Iterable) and not isinstance(components, (str, bytes)):
        for index, component in enumerate(components):
            records.append(score_component_to_record(component, name=None, precision=precision))
            if not records[-1].get("name"):
                records[-1]["name"] = f"component_{index + 1}"
    return records


def external_score_to_record(
    value: Any,
    *,
    name: Optional[str] = None,
    precision: Optional[int] = DEFAULT_SCORE_PRECISION,
) -> Dict[str, Any]:
    """Convert one external score to a record."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        record = ExternalScoreRecord(name=name or "external_score", value=_score_float(value))
        return record.to_dict(precision=precision)
    record = ExternalScoreRecord(
        name=name or _component_name(value, "external_score"),
        value=_score_float(_score_get(value, ("value", "score", "affinity", "energy"))),
        unit=_score_get(value, ("unit", "units")),
        direction=infer_score_direction(value),
        source=_score_get(value, ("source", "engine", "software", "origin")),
        normalized_value=_score_float(_score_get(value, ("normalized_value", "normalized_score"))),
        metadata=to_serializable(_score_get(value, ("metadata", "meta"), {})),
    )
    return record.to_dict(precision=precision)


def extract_external_scores(
    value: Any,
    *,
    precision: Optional[int] = DEFAULT_SCORE_PRECISION,
) -> List[Dict[str, Any]]:
    """Extract external score records."""
    external = _score_get(value, _SCORE_EXTERNAL_FIELDS, {})
    records: List[Dict[str, Any]] = []
    if isinstance(external, Mapping):
        for name, score in external.items():
            records.append(external_score_to_record(score, name=str(name), precision=precision))
    elif isinstance(external, Iterable) and not isinstance(external, (str, bytes)):
        for score in external:
            records.append(external_score_to_record(score, precision=precision))
    return records


def extract_score_metadata(value: Any) -> Dict[str, Any]:
    """Extract scoring metadata."""
    metadata = _score_get(value, ("metadata", "meta", "context"), {})
    return dict(to_serializable(metadata)) if isinstance(metadata, Mapping) else {"value": to_serializable(metadata)}


def scoring_to_record(
    value: Any,
    *,
    options: Optional[ScoringExportOptions] = None,
) -> Dict[str, Any]:
    """Convert one scoring result to a portable record."""
    opts = options or ScoringExportOptions()
    total = _score_float(_score_get(value, _SCORE_TOTAL_FIELDS))
    direction = infer_score_direction(value)
    components = extract_score_components(value, precision=opts.precision) if opts.include_components else []
    external = extract_external_scores(value, precision=opts.precision) if opts.include_external_scores else []

    component_mapping = {item["name"]: item.get("value") for item in components}
    record: Dict[str, Any] = {
        "schema_version": SCORING_SCHEMA_VERSION,
        "object_type": "scoring_result",
        "scoring_id": scoring_identifier(value, stable=opts.stable_identifier),
        "pose_id": _score_get(value, _SCORE_POSE_FIELDS) if opts.include_pose_reference else None,
        "total_score": _score_round(total, opts.precision),
        "direction": direction.value,
        "rank": _score_get(value, _SCORE_RANK_FIELDS) if opts.include_rank else None,
    }
    if opts.include_components:
        if opts.component_layout in {ScoreLayout.RECORDS, ScoreLayout.BOTH}:
            record["components"] = components
        if opts.component_layout in {ScoreLayout.MAPPING, ScoreLayout.BOTH}:
            record["component_values"] = component_mapping
    if opts.include_external_scores:
        record["external_scores"] = external
    if opts.include_normalization:
        record["normalization"] = to_serializable(_score_get(value, _SCORE_NORMALIZATION_FIELDS))
    if opts.include_explanation:
        record["explanation"] = to_serializable(_score_get(value, _SCORE_EXPLANATION_FIELDS))
    if opts.include_metadata:
        record["metadata"] = extract_score_metadata(value)
    if opts.include_raw_values:
        record["raw"] = to_serializable(value)
    if opts.omit_none:
        record = {key: item for key, item in record.items() if item is not None}
    return record


def scoring_reference(value: Any) -> Dict[str, Any]:
    """Return a compact scoring reference."""
    return {
        "scoring_id": scoring_identifier(value),
        "pose_id": _score_get(value, _SCORE_POSE_FIELDS),
        "total_score": _score_float(_score_get(value, _SCORE_TOTAL_FIELDS)),
        "direction": infer_score_direction(value).value,
    }


def scorings_to_records(
    values: Iterable[Any],
    *,
    options: Optional[ScoringExportOptions] = None,
) -> List[Dict[str, Any]]:
    """Convert multiple scoring results."""
    return [scoring_to_record(value, options=options) for value in values]


def summarize_scoring_records(
    records: Iterable[Mapping[str, Any]],
    *,
    direction: Union[ScoreDirection, str] = ScoreDirection.UNKNOWN,
) -> ScoringCollectionSummary:
    """Summarize serialized scoring records."""
    items = list(records)
    resolved_direction = normalize_score_direction(direction)
    if resolved_direction is ScoreDirection.UNKNOWN:
        for item in items:
            candidate = normalize_score_direction(item.get("direction"))
            if candidate is not ScoreDirection.UNKNOWN:
                resolved_direction = candidate
                break
    pairs: List[Tuple[str, float]] = []
    for index, item in enumerate(items):
        score = _score_float(item.get("total_score"))
        if score is not None:
            pairs.append((str(item.get("scoring_id") or index), score))
    values = sorted(score for _, score in pairs)
    summary = ScoringCollectionSummary(
        count=len(items),
        valid_count=len(values),
        missing_count=len(items) - len(values),
        direction=resolved_direction,
    )
    if not values:
        return summary
    summary.minimum = values[0]
    summary.maximum = values[-1]
    summary.mean = sum(values) / len(values)
    midpoint = len(values) // 2
    summary.median = values[midpoint] if len(values) % 2 else (values[midpoint - 1] + values[midpoint]) / 2
    reverse = resolved_direction is ScoreDirection.HIGHER_IS_BETTER
    best_id, best_score = sorted(pairs, key=lambda pair: pair[1], reverse=reverse)[0]
    summary.best_identifier = best_id
    summary.best_score = best_score
    return summary


def rank_scoring_records(
    records: Iterable[Mapping[str, Any]],
    *,
    direction: Union[ScoreDirection, str] = ScoreDirection.UNKNOWN,
) -> List[Dict[str, Any]]:
    """Rank serialized scoring records."""
    items = [dict(item) for item in records]
    resolved = normalize_score_direction(direction)
    if resolved is ScoreDirection.UNKNOWN and items:
        resolved = normalize_score_direction(items[0].get("direction"))
    reverse = resolved is ScoreDirection.HIGHER_IS_BETTER
    def ranking_key(item: Mapping[str, Any]) -> Tuple[int, float]:
        score = _score_float(item.get("total_score"))
        if score is None or not math.isfinite(score):
            return (1, 0.0)
        return (0, -score if reverse else score)

    items.sort(key=ranking_key)
    for rank, item in enumerate(items, start=1):
        item["rank"] = rank
    return items


def scoring_serializer(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize one registry-resolved scoring result."""
    options = None
    if context is not None:
        candidate = context.metadata.get("scoring_options")
        if isinstance(candidate, ScoringExportOptions):
            options = candidate
    return scoring_to_record(value, options=options)


def register_scoring_serializer(
    registry: Optional[SerializerRegistry] = None,
    *,
    priority: int = DEFAULT_SCORE_SERIALIZER_PRIORITY,
    replace_existing: bool = True,
) -> SerializerRegistry:
    """Register the generic scoring serializer."""
    target = registry or DEFAULT_SERIALIZER_REGISTRY
    target.register(
        "scoring_result",
        scoring_serializer,
        predicate=is_scoring_result_like,
        match_mode=SerializerMatchMode.PREDICATE,
        priority=priority,
        aliases=("scoring", "score_result"),
        builtin=True,
        description="Serialize DockAnalyzer scoring results.",
        replace_existing=replace_existing,
    )
    return target


register_scoring_serializer()

__all__.extend([
    "SCORING_SCHEMA_VERSION",
    "DEFAULT_SCORE_PRECISION",
    "DEFAULT_SCORE_SERIALIZER_PRIORITY",
    "ScoreDirection",
    "ScoreLayout",
    "ScoringExportOptions",
    "ScoreComponentRecord",
    "ExternalScoreRecord",
    "ScoringCollectionSummary",
    "normalize_score_direction",
    "infer_score_direction",
    "is_scoring_result_like",
    "scoring_identifier",
    "score_component_to_record",
    "extract_score_components",
    "external_score_to_record",
    "extract_external_scores",
    "extract_score_metadata",
    "scoring_to_record",
    "scoring_reference",
    "scorings_to_records",
    "summarize_scoring_records",
    "rank_scoring_records",
    "scoring_serializer",
    "register_scoring_serializer",
])
# =============================================================================
# Section 12 — DockModel
# =============================================================================

DOCK_MODEL_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_DOCK_MODEL_SERIALIZER_PRIORITY: Final[int] = 1200

_DOCK_MODEL_ID_FIELDS: Final[Tuple[str, ...]] = (
    "dock_model_id", "model_id", "pose_id", "id", "uid", "identifier",
)
_DOCK_MODEL_NAME_FIELDS: Final[Tuple[str, ...]] = (
    "name", "model_name", "pose_name", "title", "label",
)
_DOCK_MODEL_STRUCTURE_FIELDS: Final[Tuple[str, ...]] = (
    "structure", "model", "atomic_structure", "chimerax_model", "pose",
)
_DOCK_MODEL_LIGAND_FIELDS: Final[Tuple[str, ...]] = (
    "ligand", "ligand_model", "ligand_structure", "compound",
)
_DOCK_MODEL_RECEPTOR_FIELDS: Final[Tuple[str, ...]] = (
    "receptor", "target", "protein", "receptor_model",
)
_DOCK_MODEL_ANALYSIS_FIELDS: Final[Tuple[str, ...]] = (
    "analysis", "analyses", "analysis_results", "results",
)
_DOCK_MODEL_SCORING_FIELDS: Final[Tuple[str, ...]] = (
    "scoring", "score_result", "scoring_result", "scores",
)
_DOCK_MODEL_METADATA_FIELDS: Final[Tuple[str, ...]] = (
    "metadata", "properties", "details", "annotations", "extra",
)
_DOCK_MODEL_FILE_FIELDS: Final[Tuple[str, ...]] = (
    "files", "file_paths", "paths", "artifacts", "outputs",
)
_DOCK_MODEL_STATUS_FIELDS: Final[Tuple[str, ...]] = (
    "status", "state", "analysis_status",
)
_DOCK_MODEL_RANK_FIELDS: Final[Tuple[str, ...]] = (
    "rank", "pose_rank", "ranking",
)
_DOCK_MODEL_SCORE_FIELDS: Final[Tuple[str, ...]] = (
    "total_score", "score", "normalized_score", "docking_score", "affinity",
)
_DOCK_MODEL_INTERACTION_FIELDS: Final[Tuple[str, ...]] = (
    "contacts",
    "hbonds",
    "hydrogen_bonds",
    "hydrophobic",
    "pi",
    "pi_interactions",
    "saltbridge",
    "saltbridges",
    "salt_bridges",
    "clash",
    "clashes",
    "interactions",
)


class DockModelLayout(str, Enum):
    """Control DockModel representation."""

    COMPACT = "compact"
    STANDARD = "standard"
    FULL = "full"


class DockModelAttachmentMode(str, Enum):
    """Control result attachment behavior."""

    PRESERVE = "preserve"
    REPLACE = "replace"
    MERGE = "merge"


@dataclass
class DockModelExportOptions:
    """Control DockModel record generation."""

    include_schema: bool = True
    include_id: bool = True
    include_name: bool = True
    include_structure: bool = True
    include_ligand: bool = True
    include_receptor: bool = True
    include_interactions: bool = True
    include_analysis: bool = True
    include_scoring: bool = True
    include_metadata: bool = True
    include_files: bool = True
    include_status: bool = True
    include_rank: bool = True
    include_raw_fields: bool = False
    include_empty: bool = False
    omit_none: bool = True
    strict: bool = False
    layout: DockModelLayout = DockModelLayout.STANDARD
    molecular_options: Optional[MolecularExportOptions] = None
    interaction_options: Optional[InteractionExportOptions] = None
    analysis_options: Optional[AnalysisExportOptions] = None
    scoring_options: Optional[ScoringExportOptions] = None

    def __post_init__(self) -> None:
        if isinstance(self.layout, str):
            self.layout = DockModelLayout(self.layout)


@dataclass
class DockModelRecordContext:
    """Track DockModel serialization state."""

    options: DockModelExportOptions = field(default_factory=DockModelExportOptions)
    registry: Optional[SerializerRegistry] = None
    seen: Set[int] = field(default_factory=set)
    warnings: List[str] = field(default_factory=list)

    def child(self) -> "DockModelRecordContext":
        """Create a child context."""
        return DockModelRecordContext(
            options=self.options,
            registry=self.registry,
            seen=self.seen,
            warnings=self.warnings,
        )


@dataclass
class DockModelCollectionSummary:
    """Summarize DockModel records."""

    count: int = 0
    successful: int = 0
    failed: int = 0
    scored: int = 0
    ranked: int = 0
    interaction_count: int = 0
    best_identifier: Optional[str] = None
    best_score: Optional[float] = None
    score_direction: ScoreDirection = ScoreDirection.UNKNOWN

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable summary."""
        return to_serializable(self)


def _dock_model_get(value: Any, names: Sequence[str], default: Any = None) -> Any:
    """Read the first available DockModel field."""
    return _plain_field_get(value, names, default)


def _dock_model_iter_fields(value: Any) -> Iterable[Tuple[str, Any]]:
    """Iterate public DockModel fields."""
    if isinstance(value, Mapping):
        yield from value.items()
        return
    if is_dataclass(value) and not isinstance(value, type):
        for item in fields(value):
            try:
                yield item.name, getattr(value, item.name)
            except (AttributeError, TypeError, ValueError):
                continue
        return
    namespace = getattr(value, "__dict__", None)
    if isinstance(namespace, Mapping):
        for key, item in namespace.items():
            if not str(key).startswith("_"):
                yield str(key), item


def is_dock_model_like(value: Any) -> bool:
    """Return whether a value resembles DockModel."""
    if value is None or isinstance(value, (str, bytes, bytearray)):
        return False
    class_name = value.__class__.__name__.lower()
    if class_name in {"dockmodel", "dock_model"} or "dockmodel" in class_name:
        return True
    names = {str(name).lower() for name, _ in _dock_model_iter_fields(value)}
    interaction_hits = len(names.intersection(_DOCK_MODEL_INTERACTION_FIELDS))
    identity_hit = bool(names.intersection(_DOCK_MODEL_ID_FIELDS + _DOCK_MODEL_NAME_FIELDS))
    structure_hit = bool(names.intersection(_DOCK_MODEL_STRUCTURE_FIELDS))
    return interaction_hits >= 2 and (identity_hit or structure_hit)


def dock_model_identifier(value: Any, fallback: Optional[str] = None) -> str:
    """Return a stable DockModel identifier."""
    candidate = _dock_model_get(value, _DOCK_MODEL_ID_FIELDS)
    if candidate not in (None, ""):
        return str(candidate)
    name = _dock_model_get(value, _DOCK_MODEL_NAME_FIELDS)
    if name not in (None, ""):
        return str(name)
    structure = _dock_model_get(value, _DOCK_MODEL_STRUCTURE_FIELDS)
    if structure is not None:
        try:
            return molecular_identifier(structure)
        except Exception:
            pass
    if fallback:
        return fallback
    return f"dock_model_{id(value):x}"


def _dock_model_status(value: Any) -> str:
    """Infer DockModel status."""
    candidate = _dock_model_get(value, _DOCK_MODEL_STATUS_FIELDS)
    if isinstance(candidate, Enum):
        candidate = candidate.value
    if candidate not in (None, ""):
        return str(candidate).strip().lower()
    errors = _dock_model_get(value, ("errors", "error", "failures"))
    if errors:
        return "failed"
    analyses = _dock_model_get(value, _DOCK_MODEL_ANALYSIS_FIELDS)
    interactions = any(_dock_model_get(value, (name,)) for name in _DOCK_MODEL_INTERACTION_FIELDS)
    if analyses or interactions:
        return "completed"
    return "pending"


def _dock_model_rank(value: Any) -> Optional[int]:
    """Extract a DockModel rank."""
    candidate = _dock_model_get(value, _DOCK_MODEL_RANK_FIELDS)
    try:
        return int(candidate) if candidate is not None else None
    except (TypeError, ValueError, OverflowError):
        return None


def _dock_model_total_score(value: Any) -> Optional[float]:
    """Extract the preferred DockModel score."""
    candidate = _dock_model_get(value, _DOCK_MODEL_SCORE_FIELDS)
    if candidate is None:
        scoring = _dock_model_get(value, _DOCK_MODEL_SCORING_FIELDS)
        candidate = _score_get(scoring, _DOCK_MODEL_SCORE_FIELDS) if scoring is not None else None
    try:
        return float(candidate) if candidate is not None else None
    except (TypeError, ValueError, OverflowError):
        return None


def _serialize_dock_structure(
    value: Any,
    context: DockModelRecordContext,
    *,
    ligand: bool = False,
) -> Any:
    """Serialize a molecular DockModel member."""
    if value is None:
        return None
    options = context.options.molecular_options or MolecularExportOptions()
    if context.options.layout is DockModelLayout.COMPACT:
        return molecular_reference(value, options=options)
    if ligand:
        return ligand_to_record(value, options=options)
    return molecular_object_to_record(value, options=options)


def _normalize_interaction_source(value: Any) -> List[Any]:
    """Normalize an interaction source."""
    if value is None:
        return []
    if isinstance(value, Mapping):
        if is_interaction_like(value):
            return [value]
        output: List[Any] = []
        for item in value.values():
            output.extend(_normalize_interaction_source(item))
        return output
    if isinstance(value, Iterable) and not isinstance(value, (str, bytes, bytearray)):
        return list(value)
    return [value]


def extract_dock_model_interactions(value: Any) -> Dict[str, List[Any]]:
    """Extract unique DockModel interactions grouped by canonical family."""
    output: Dict[str, List[Any]] = {}
    seen_objects: Set[int] = set()
    seen_identifiers: Set[str] = set()

    def add(item: Any, family_hint: Optional[str] = None) -> None:
        object_marker = id(item)
        explicit = _interaction_get(item, _INTERACTION_ID_FIELDS)
        identifier_marker = str(explicit) if explicit not in (None, "") else ""
        if object_marker in seen_objects or (identifier_marker and identifier_marker in seen_identifiers):
            return
        family = normalize_interaction_family(
            family_hint or infer_interaction_family(item)
        )
        output.setdefault(family, []).append(item)
        seen_objects.add(object_marker)
        if identifier_marker:
            seen_identifiers.add(identifier_marker)

    aliases = (
        ("contacts", "contact"),
        ("hbonds", "hbond"),
        ("hydrogen_bonds", "hbond"),
        ("hydrophobic", "hydrophobic"),
        ("pi", "pi"),
        ("pi_interactions", "pi"),
        ("saltbridge", "saltbridge"),
        ("saltbridges", "saltbridge"),
        ("salt_bridges", "saltbridge"),
        ("clash", "clash"),
        ("clashes", "clash"),
    )
    for field_name, family in aliases:
        for item in _normalize_interaction_source(
            _dock_model_get(value, (field_name,))
        ):
            add(item, family)

    for item in _normalize_interaction_source(
        _dock_model_get(value, ("interactions",))
    ):
        add(item)

    return output


def dock_model_interactions_to_record(
    value: Any,
    *,
    options: Optional[InteractionExportOptions] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    """Serialize all DockModel interaction families."""
    output: Dict[str, List[Dict[str, Any]]] = {}
    for family, items in extract_dock_model_interactions(value).items():
        records = interactions_to_records(items, options=options)
        if records:
            output[family] = records
    return output


def _dock_model_analysis_records(
    value: Any,
    context: DockModelRecordContext,
) -> Any:
    """Serialize attached analysis results."""
    source = _dock_model_get(value, _DOCK_MODEL_ANALYSIS_FIELDS)
    if source is None:
        return None
    options = context.options.analysis_options or AnalysisExportOptions()
    if isinstance(source, Mapping) and not is_analysis_result_like(source):
        return {
            str(key): analysis_to_record(item, options=options)
            for key, item in source.items()
        }
    if isinstance(source, Iterable) and not isinstance(source, (str, bytes, bytearray, Mapping)):
        return analyses_to_records(source, options=options)
    return analysis_to_record(source, options=options)


def _dock_model_scoring_record(
    value: Any,
    context: DockModelRecordContext,
) -> Any:
    """Serialize attached scoring results."""
    source = _dock_model_get(value, _DOCK_MODEL_SCORING_FIELDS)
    if source is None:
        total = _dock_model_total_score(value)
        return {"total_score": total} if total is not None else None
    options = context.options.scoring_options or ScoringExportOptions()
    if isinstance(source, Mapping) and not is_scoring_result_like(source):
        return {
            str(key): scoring_to_record(item, options=options)
            for key, item in source.items()
        }
    if isinstance(source, Iterable) and not isinstance(source, (str, bytes, bytearray, Mapping)):
        return scorings_to_records(source, options=options)
    return scoring_to_record(source, options=options)


def _dock_model_files_record(value: Any) -> Any:
    """Serialize DockModel files."""
    source = _dock_model_get(value, _DOCK_MODEL_FILE_FIELDS)
    if source is None:
        return None
    if isinstance(source, Mapping):
        return {str(key): str(item) if isinstance(item, (Path, os.PathLike)) else to_serializable(item) for key, item in source.items()}
    if isinstance(source, Iterable) and not isinstance(source, (str, bytes, bytearray)):
        return [str(item) if isinstance(item, (Path, os.PathLike)) else to_serializable(item) for item in source]
    return str(source) if isinstance(source, (Path, os.PathLike)) else to_serializable(source)


def _clean_dock_model_record(record: Dict[str, Any], options: DockModelExportOptions) -> Dict[str, Any]:
    """Remove omitted DockModel values."""
    if options.omit_none:
        record = {key: value for key, value in record.items() if value is not None}
    if not options.include_empty:
        record = {key: value for key, value in record.items() if value not in ({}, [], (), "")}
    return record


def dock_model_to_record(
    value: Any,
    *,
    options: Optional[DockModelExportOptions] = None,
    context: Optional[DockModelRecordContext] = None,
) -> Dict[str, Any]:
    """Convert one DockModel to a serializable record."""
    if value is None:
        raise ValueError("DockModel cannot be None.")
    resolved_options = options or (context.options if context else DockModelExportOptions())
    resolved_context = context or DockModelRecordContext(options=resolved_options)
    object_id = id(value)
    if object_id in resolved_context.seen:
        return {"dock_model_id": dock_model_identifier(value), "$ref": True}
    resolved_context.seen.add(object_id)
    try:
        record: Dict[str, Any] = {}
        if resolved_options.include_schema:
            record["schema_version"] = DOCK_MODEL_SCHEMA_VERSION
            record["record_type"] = "dock_model"
        if resolved_options.include_id:
            record["dock_model_id"] = dock_model_identifier(value)
        if resolved_options.include_name:
            record["name"] = _dock_model_get(value, _DOCK_MODEL_NAME_FIELDS)
        if resolved_options.include_status:
            record["status"] = _dock_model_status(value)
        if resolved_options.include_rank:
            record["rank"] = _dock_model_rank(value)
        record["total_score"] = _dock_model_total_score(value)
        if resolved_options.include_structure:
            record["structure"] = _serialize_dock_structure(
                _dock_model_get(value, _DOCK_MODEL_STRUCTURE_FIELDS),
                resolved_context,
            )
        if resolved_options.include_ligand:
            record["ligand"] = _serialize_dock_structure(
                _dock_model_get(value, _DOCK_MODEL_LIGAND_FIELDS),
                resolved_context,
                ligand=True,
            )
        if resolved_options.include_receptor:
            record["receptor"] = _serialize_dock_structure(
                _dock_model_get(value, _DOCK_MODEL_RECEPTOR_FIELDS),
                resolved_context,
            )
        if resolved_options.include_interactions:
            record["interactions"] = dock_model_interactions_to_record(
                value,
                options=resolved_options.interaction_options,
            )
        if resolved_options.include_analysis:
            record["analysis"] = _dock_model_analysis_records(value, resolved_context)
        if resolved_options.include_scoring:
            record["scoring"] = _dock_model_scoring_record(value, resolved_context)
        if resolved_options.include_metadata:
            record["metadata"] = to_serializable(
                _dock_model_get(value, _DOCK_MODEL_METADATA_FIELDS)
            )
        if resolved_options.include_files:
            record["files"] = _dock_model_files_record(value)
        if resolved_options.include_raw_fields and resolved_options.layout is DockModelLayout.FULL:
            excluded = set(
                _DOCK_MODEL_ID_FIELDS
                + _DOCK_MODEL_NAME_FIELDS
                + _DOCK_MODEL_STRUCTURE_FIELDS
                + _DOCK_MODEL_LIGAND_FIELDS
                + _DOCK_MODEL_RECEPTOR_FIELDS
                + _DOCK_MODEL_ANALYSIS_FIELDS
                + _DOCK_MODEL_SCORING_FIELDS
                + _DOCK_MODEL_METADATA_FIELDS
                + _DOCK_MODEL_FILE_FIELDS
                + _DOCK_MODEL_STATUS_FIELDS
                + _DOCK_MODEL_RANK_FIELDS
                + _DOCK_MODEL_INTERACTION_FIELDS
            )
            record["raw_fields"] = {
                name: to_serializable(item)
                for name, item in _dock_model_iter_fields(value)
                if name not in excluded
            }
        return _clean_dock_model_record(record, resolved_options)
    except Exception as exc:
        if resolved_options.strict:
            raise
        resolved_context.warnings.append(str(exc))
        return {
            "schema_version": DOCK_MODEL_SCHEMA_VERSION,
            "record_type": "dock_model",
            "dock_model_id": dock_model_identifier(value),
            "status": "serialization_failed",
            "error": str(exc),
        }
    finally:
        resolved_context.seen.discard(object_id)


def dock_model_reference(value: Any) -> Dict[str, Any]:
    """Return a compact DockModel reference."""
    record: Dict[str, Any] = {
        "dock_model_id": dock_model_identifier(value),
        "name": _dock_model_get(value, _DOCK_MODEL_NAME_FIELDS),
        "status": _dock_model_status(value),
        "rank": _dock_model_rank(value),
        "total_score": _dock_model_total_score(value),
    }
    return {key: item for key, item in record.items() if item is not None}


def dock_models_to_records(
    values: Iterable[Any],
    *,
    options: Optional[DockModelExportOptions] = None,
) -> List[Dict[str, Any]]:
    """Serialize multiple DockModels."""
    resolved = options or DockModelExportOptions()
    context = DockModelRecordContext(options=resolved)
    return [dock_model_to_record(value, context=context) for value in values]


def summarize_dock_model_records(
    records: Iterable[Mapping[str, Any]],
    *,
    direction: Union[ScoreDirection, str] = ScoreDirection.UNKNOWN,
) -> DockModelCollectionSummary:
    """Summarize serialized DockModels."""
    items = [dict(item) for item in records]
    resolved_direction = normalize_score_direction(direction)
    scores: List[Tuple[str, float]] = []
    summary = DockModelCollectionSummary(count=len(items), score_direction=resolved_direction)
    for index, item in enumerate(items):
        status = str(item.get("status", "")).lower()
        if status in {"failed", "error", "serialization_failed"}:
            summary.failed += 1
        else:
            summary.successful += 1
        if item.get("rank") is not None:
            summary.ranked += 1
        score = _score_float(item.get("total_score"))
        if score is not None:
            summary.scored += 1
            scores.append((str(item.get("dock_model_id") or index), score))
        interactions = item.get("interactions")
        if isinstance(interactions, Mapping):
            summary.interaction_count += sum(
                len(value) for value in interactions.values() if isinstance(value, Sequence)
            )
    if scores:
        reverse = resolved_direction is ScoreDirection.HIGHER_IS_BETTER
        best_id, best_score = sorted(scores, key=lambda pair: pair[1], reverse=reverse)[0]
        summary.best_identifier = best_id
        summary.best_score = best_score
    return summary


def rank_dock_model_records(
    records: Iterable[Mapping[str, Any]],
    *,
    direction: Union[ScoreDirection, str] = ScoreDirection.UNKNOWN,
) -> List[Dict[str, Any]]:
    """Rank serialized DockModels."""
    items = [dict(item) for item in records]
    resolved = normalize_score_direction(direction)
    reverse = resolved is ScoreDirection.HIGHER_IS_BETTER
    def ranking_key(item: Mapping[str, Any]) -> Tuple[int, float]:
        score = _score_float(item.get("total_score"))
        if score is None or not math.isfinite(score):
            return (1, 0.0)
        return (0, -score if reverse else score)

    items.sort(key=ranking_key)
    for rank, item in enumerate(items, start=1):
        item["rank"] = rank
    return items


def rank_dock_models(
    values: Iterable[Any],
    *,
    direction: Union[ScoreDirection, str] = ScoreDirection.UNKNOWN,
) -> List[Any]:
    """Return DockModel-like objects ordered by their preferred score."""
    items = list(values)
    resolved = normalize_score_direction(direction)
    if resolved is ScoreDirection.UNKNOWN:
        for item in items:
            scoring = _dock_model_get(item, _DOCK_MODEL_SCORING_FIELDS)
            inferred = infer_score_direction(scoring if scoring is not None else item)
            if inferred is not ScoreDirection.UNKNOWN:
                resolved = inferred
                break
    scored: List[Tuple[float, int, Any]] = []
    missing: List[Tuple[int, Any]] = []
    for index, item in enumerate(items):
        score = _dock_model_total_score(item)
        if score is None or not math.isfinite(score):
            missing.append((index, item))
        else:
            scored.append((score, index, item))
    scored.sort(
        key=lambda entry: (entry[0], entry[1]),
        reverse=resolved is ScoreDirection.HIGHER_IS_BETTER,
    )
    return [item for _, _, item in scored] + [item for _, item in missing]


def _set_dock_model_field(target: Any, name: str, value: Any) -> None:
    """Set a DockModel field."""
    if isinstance(target, MutableMapping):
        target[name] = value
        return
    setattr(target, name, value)


def _merge_dock_model_value(existing: Any, incoming: Any) -> Any:
    """Merge one attached DockModel value."""
    if existing is None:
        return incoming
    if incoming is None:
        return existing
    if isinstance(existing, MutableMapping) and isinstance(incoming, Mapping):
        merged = dict(existing)
        merged.update(incoming)
        return merged
    if isinstance(existing, list):
        merged_list = list(existing)
        if isinstance(incoming, Iterable) and not isinstance(incoming, (str, bytes, bytearray, Mapping)):
            merged_list.extend(incoming)
        else:
            merged_list.append(incoming)
        return merged_list
    return incoming


def attach_dock_model_result(
    target: Any,
    field_name: str,
    result: Any,
    *,
    mode: Union[DockModelAttachmentMode, str] = DockModelAttachmentMode.PRESERVE,
) -> Any:
    """Attach one result to a DockModel."""
    resolved = DockModelAttachmentMode(mode)
    existing = _dock_model_get(target, (field_name,))
    if resolved is DockModelAttachmentMode.PRESERVE and existing not in (None, [], {}, ()):
        return target
    value = result
    if resolved is DockModelAttachmentMode.MERGE:
        value = _merge_dock_model_value(existing, result)
    _set_dock_model_field(target, field_name, value)
    return target


def attach_dock_model_results(
    target: Any,
    results: Mapping[str, Any],
    *,
    mode: Union[DockModelAttachmentMode, str] = DockModelAttachmentMode.PRESERVE,
) -> Any:
    """Attach multiple results to a DockModel."""
    for field_name, result in results.items():
        attach_dock_model_result(target, str(field_name), result, mode=mode)
    return target


def dock_model_serializer(
    value: Any,
    *,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize one registry-resolved DockModel."""
    options = None
    if context is not None:
        candidate = context.metadata.get("dock_model_options")
        if isinstance(candidate, DockModelExportOptions):
            options = candidate
    return dock_model_to_record(value, options=options)


def register_dock_model_serializer(
    registry: Optional[SerializerRegistry] = None,
    *,
    priority: int = DEFAULT_DOCK_MODEL_SERIALIZER_PRIORITY,
    replace_existing: bool = True,
) -> SerializerRegistry:
    """Register the DockModel serializer."""
    target = registry or DEFAULT_SERIALIZER_REGISTRY
    target.register(
        "dock_model",
        dock_model_serializer,
        predicate=is_dock_model_like,
        match_mode=SerializerMatchMode.PREDICATE,
        priority=priority,
        aliases=("dockmodel", "pose_model"),
        builtin=True,
        description="Serialize DockAnalyzer DockModel objects.",
        replace_existing=replace_existing,
    )
    return target


register_dock_model_serializer()

__all__.extend([
    "DOCK_MODEL_SCHEMA_VERSION",
    "DEFAULT_DOCK_MODEL_SERIALIZER_PRIORITY",
    "DockModelLayout",
    "DockModelAttachmentMode",
    "DockModelExportOptions",
    "DockModelRecordContext",
    "DockModelCollectionSummary",
    "is_dock_model_like",
    "dock_model_identifier",
    "extract_dock_model_interactions",
    "dock_model_interactions_to_record",
    "dock_model_to_record",
    "dock_model_reference",
    "dock_models_to_records",
    "summarize_dock_model_records",
    "rank_dock_model_records",
    "rank_dock_models",
    "attach_dock_model_result",
    "attach_dock_model_results",
    "dock_model_serializer",
    "register_dock_model_serializer",
])
# =============================================================================
# Section 13 — Table construction
# =============================================================================

TABLE_SCHEMA_VERSION = "1.0"
DEFAULT_TABLE_SEPARATOR = "."
DEFAULT_TABLE_LIST_SEPARATOR = "; "
DEFAULT_MAX_TABLE_DEPTH = 8
DEFAULT_MAX_CELL_LENGTH = 32767


class TableOrientation(str, Enum):
    """Table record orientation."""

    RECORDS = "records"
    COLUMNS = "columns"


class TableValueMode(str, Enum):
    """Nested value handling."""

    FLATTEN = "flatten"
    JSON = "json"
    TEXT = "text"
    PRESERVE = "preserve"


class TableColumnOrder(str, Enum):
    """Column ordering mode."""

    DISCOVERY = "discovery"
    ALPHABETICAL = "alphabetical"
    SCHEMA = "schema"


@dataclass(frozen=True)
class TableColumn:
    """Column definition."""

    name: str
    label: Optional[str] = None
    dtype: Optional[str] = None
    nullable: bool = True
    default: Any = None
    source: Optional[str] = None
    position: Optional[int] = None
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        normalized = str(self.name).strip()
        if not normalized:
            raise ValueError("Table column name cannot be empty.")
        object.__setattr__(self, "name", normalized)
        object.__setattr__(self, "metadata", dict(self.metadata))

    @property
    def display_name(self) -> str:
        return self.label or self.name

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "label": self.label,
            "dtype": self.dtype,
            "nullable": self.nullable,
            "default": make_json_safe(self.default),
            "source": self.source,
            "position": self.position,
            "metadata": make_json_safe(dict(self.metadata)),
        }


@dataclass(frozen=True)
class TableSchema:
    """Ordered table schema."""

    name: str
    columns: Tuple[TableColumn, ...] = ()
    primary_key: Tuple[str, ...] = ()
    description: Optional[str] = None
    version: str = TABLE_SCHEMA_VERSION
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        normalized = str(self.name).strip()
        if not normalized:
            raise ValueError("Table schema name cannot be empty.")
        columns = tuple(self.columns)
        names = [column.name for column in columns]
        if len(names) != len(set(names)):
            raise ValueError("Table schema contains duplicate columns.")
        missing = [key for key in self.primary_key if key not in names]
        if missing:
            raise ValueError(f"Primary-key columns are missing: {missing!r}.")
        object.__setattr__(self, "name", normalized)
        object.__setattr__(self, "columns", columns)
        object.__setattr__(self, "primary_key", tuple(self.primary_key))
        object.__setattr__(self, "metadata", dict(self.metadata))

    @property
    def column_names(self) -> Tuple[str, ...]:
        return tuple(column.name for column in self.columns)

    def column(self, name: str) -> Optional[TableColumn]:
        return next((column for column in self.columns if column.name == name), None)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "columns": [column.to_dict() for column in self.columns],
            "primary_key": list(self.primary_key),
            "description": self.description,
            "version": self.version,
            "metadata": make_json_safe(dict(self.metadata)),
        }


@dataclass(frozen=True)
class TableBuildOptions:
    """Table construction options."""

    value_mode: TableValueMode = TableValueMode.FLATTEN
    column_order: TableColumnOrder = TableColumnOrder.DISCOVERY
    separator: str = DEFAULT_TABLE_SEPARATOR
    list_separator: str = DEFAULT_TABLE_LIST_SEPARATOR
    max_depth: int = DEFAULT_MAX_TABLE_DEPTH
    max_cell_length: Optional[int] = DEFAULT_MAX_CELL_LENGTH
    include_none: bool = True
    include_empty: bool = True
    include_index: bool = False
    index_name: str = "index"
    stringify_unknown: bool = True
    sort_mapping_keys: bool = False
    preserve_sequences: bool = False
    schema_strict: bool = False
    extra_columns: bool = True

    def __post_init__(self) -> None:
        if not self.separator:
            raise ValueError("Table separator cannot be empty.")
        if self.max_depth < 0:
            raise ValueError("max_depth cannot be negative.")
        if self.max_cell_length is not None and self.max_cell_length < 1:
            raise ValueError("max_cell_length must be positive.")
        object.__setattr__(self, "value_mode", TableValueMode(self.value_mode))
        object.__setattr__(self, "column_order", TableColumnOrder(self.column_order))


@dataclass
class TableData:
    """Portable table representation."""

    name: str
    columns: List[str] = field(default_factory=list)
    rows: List[Dict[str, Any]] = field(default_factory=list)
    schema: Optional[TableSchema] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.name = str(self.name).strip() or "table"
        self.columns = list(dict.fromkeys(str(column) for column in self.columns))
        self.rows = [dict(row) for row in self.rows]
        self.metadata = dict(self.metadata)
        self.warnings = [str(item) for item in self.warnings]
        if not self.columns:
            self.columns = discover_columns(self.rows)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.columns)

    @property
    def empty(self) -> bool:
        return not self.rows

    def normalized_rows(self, fill: Any = None) -> List[Dict[str, Any]]:
        return [
            {column: row.get(column, fill) for column in self.columns}
            for row in self.rows
        ]

    def column_values(self, name: str) -> List[Any]:
        if name not in self.columns:
            raise KeyError(name)
        return [row.get(name) for row in self.rows]

    def to_records(self) -> List[Dict[str, Any]]:
        return self.normalized_rows()

    def to_columns(self) -> Dict[str, List[Any]]:
        return {column: self.column_values(column) for column in self.columns}

    def to_dict(self, orientation: TableOrientation = TableOrientation.RECORDS) -> Dict[str, Any]:
        orientation = TableOrientation(orientation)
        data: Any = self.to_records() if orientation is TableOrientation.RECORDS else self.to_columns()
        return {
            "name": self.name,
            "columns": list(self.columns),
            "data": make_json_safe(data),
            "row_count": self.row_count,
            "column_count": self.column_count,
            "schema": self.schema.to_dict() if self.schema else None,
            "metadata": make_json_safe(self.metadata),
            "warnings": list(self.warnings),
        }

    def to_dataframe(self) -> Any:
        if not PANDAS_AVAILABLE:
            raise ExportDependencyError("pandas", feature="DataFrame conversion")
        return pd.DataFrame(self.normalized_rows(), columns=self.columns)


@dataclass
class TableCollection:
    """Named table collection."""

    tables: Dict[str, TableData] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def add(self, table: TableData, *, replace: bool = False) -> TableData:
        if table.name in self.tables and not replace:
            raise ValueError(f"Table already exists: {table.name!r}.")
        self.tables[table.name] = table
        return table

    def get(self, name: str) -> Optional[TableData]:
        return self.tables.get(name)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "tables": {name: table.to_dict() for name, table in self.tables.items()},
            "metadata": make_json_safe(self.metadata),
        }


def _table_is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _table_sequence(value: Any) -> bool:
    return isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray))


def _table_json(value: Any) -> str:
    return json.dumps(
        make_json_safe(value),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _truncate_cell(value: Any, limit: Optional[int]) -> Any:
    if limit is None or not isinstance(value, str) or len(value) <= limit:
        return value
    if limit <= 1:
        return value[:limit]
    return value[: limit - 1] + "…"


def normalize_table_cell(value: Any, options: Optional[TableBuildOptions] = None) -> Any:
    """Normalize a value for tabular storage."""
    options = options or TableBuildOptions()
    if value is None:
        return None
    if isinstance(value, Enum):
        value = value.value
    if isinstance(value, Path):
        value = str(value)
    if isinstance(value, (datetime, date, time)):
        value = value.isoformat()
    if isinstance(value, timedelta):
        value = value.total_seconds()
    if NUMPY_AVAILABLE:
        if isinstance(value, np.generic):
            value = value.item()
        elif isinstance(value, np.ndarray):
            value = value.tolist()
    if _table_is_scalar(value):
        return _truncate_cell(value, options.max_cell_length)
    if isinstance(value, Mapping):
        if options.value_mode is TableValueMode.PRESERVE:
            return dict(value)
        if options.value_mode is TableValueMode.TEXT:
            value = str(dict(value))
        else:
            value = _table_json(value)
        return _truncate_cell(value, options.max_cell_length)
    if _table_sequence(value) or isinstance(value, (set, frozenset)):
        items = list(value)
        if options.preserve_sequences or options.value_mode is TableValueMode.PRESERVE:
            return [make_json_safe(item) for item in items]
        if options.value_mode is TableValueMode.JSON:
            value = _table_json(items)
        elif all(_table_is_scalar(item) for item in items):
            value = options.list_separator.join("" if item is None else str(item) for item in items)
        else:
            value = _table_json(items)
        return _truncate_cell(value, options.max_cell_length)
    serialized = make_json_safe(value)
    if _table_is_scalar(serialized):
        return _truncate_cell(serialized, options.max_cell_length)
    if options.stringify_unknown:
        return _truncate_cell(_table_json(serialized), options.max_cell_length)
    return serialized


def flatten_table_record(
    value: Any,
    *,
    prefix: str = "",
    options: Optional[TableBuildOptions] = None,
    _depth: int = 0,
) -> Dict[str, Any]:
    """Flatten one object into a table row."""
    options = options or TableBuildOptions()
    if _depth > options.max_depth:
        key = prefix or "value"
        return {key: normalize_table_cell(value, options)}
    if not isinstance(value, Mapping):
        value = serialize_registered(value)
    if not isinstance(value, Mapping):
        return {prefix or "value": normalize_table_cell(value, options)}
    row: Dict[str, Any] = {}
    items = value.items()
    if options.sort_mapping_keys:
        items = sorted(items, key=lambda item: str(item[0]))
    for raw_key, item in items:
        key = str(raw_key)
        full_key = f"{prefix}{options.separator}{key}" if prefix else key
        if item is None and not options.include_none:
            continue
        if isinstance(item, Mapping) and options.value_mode is TableValueMode.FLATTEN:
            nested = flatten_table_record(
                item,
                prefix=full_key,
                options=options,
                _depth=_depth + 1,
            )
            row.update(nested)
            continue
        if _table_sequence(item) and not item and not options.include_empty:
            continue
        if isinstance(item, Mapping) and not item and not options.include_empty:
            continue
        row[full_key] = normalize_table_cell(item, options)
    return row


def record_to_table_row(
    value: Any,
    *,
    options: Optional[TableBuildOptions] = None,
    index: Optional[int] = None,
) -> Dict[str, Any]:
    """Convert one value into a table row."""
    options = options or TableBuildOptions()
    row = flatten_table_record(value, options=options)
    if options.include_index:
        row = {options.index_name: index, **row}
    return row


def records_to_table_rows(
    values: Iterable[Any],
    *,
    options: Optional[TableBuildOptions] = None,
) -> List[Dict[str, Any]]:
    """Convert values into table rows."""
    options = options or TableBuildOptions()
    return [record_to_table_row(value, options=options, index=index) for index, value in enumerate(values)]


def discover_columns(rows: Iterable[Mapping[str, Any]]) -> List[str]:
    """Discover columns in first-seen order."""
    columns: List[str] = []
    seen: Set[str] = set()
    for row in rows:
        for column in row:
            name = str(column)
            if name not in seen:
                seen.add(name)
                columns.append(name)
    return columns


def infer_table_dtype(values: Iterable[Any]) -> str:
    """Infer a portable column type."""
    kinds: Set[str] = set()
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool):
            kinds.add("boolean")
        elif isinstance(value, int):
            kinds.add("integer")
        elif isinstance(value, float):
            kinds.add("number")
        elif isinstance(value, str):
            kinds.add("string")
        elif isinstance(value, Mapping):
            kinds.add("object")
        elif _table_sequence(value):
            kinds.add("array")
        else:
            kinds.add("unknown")
    if not kinds:
        return "null"
    if kinds <= {"integer", "number"}:
        return "number" if "number" in kinds else "integer"
    return next(iter(kinds)) if len(kinds) == 1 else "mixed"


def infer_table_schema(
    name: str,
    rows: Sequence[Mapping[str, Any]],
    *,
    columns: Optional[Sequence[str]] = None,
    primary_key: Sequence[str] = (),
) -> TableSchema:
    """Infer a schema from rows."""
    names = list(columns) if columns is not None else discover_columns(rows)
    definitions = []
    for position, column in enumerate(names):
        values = [row.get(column) for row in rows]
        definitions.append(
            TableColumn(
                name=column,
                dtype=infer_table_dtype(values),
                nullable=any(value is None for value in values),
                position=position,
            )
        )
    return TableSchema(name=name, columns=tuple(definitions), primary_key=tuple(primary_key))


def order_table_columns(
    columns: Sequence[str],
    *,
    options: Optional[TableBuildOptions] = None,
    schema: Optional[TableSchema] = None,
) -> List[str]:
    """Order columns according to options and schema."""
    options = options or TableBuildOptions()
    unique = list(dict.fromkeys(str(column) for column in columns))
    if options.column_order is TableColumnOrder.ALPHABETICAL:
        return sorted(unique)
    if options.column_order is TableColumnOrder.SCHEMA and schema is not None:
        ordered = [name for name in schema.column_names if name in unique]
        if options.extra_columns:
            ordered.extend(name for name in unique if name not in ordered)
        return ordered
    return unique


def apply_table_schema(
    rows: Iterable[Mapping[str, Any]],
    schema: TableSchema,
    *,
    strict: bool = False,
    include_extra: bool = True,
) -> List[Dict[str, Any]]:
    """Apply defaults and column ordering from a schema."""
    names = schema.column_names
    output: List[Dict[str, Any]] = []
    for source in rows:
        extras = [name for name in source if name not in names]
        if strict and extras:
            raise ValueError(f"Unexpected columns for {schema.name!r}: {extras!r}.")
        row: Dict[str, Any] = {}
        for column in schema.columns:
            value = source.get(column.name, column.default)
            if value is None and not column.nullable:
                raise ValueError(f"Column {column.name!r} cannot be null.")
            row[column.name] = value
        if include_extra:
            row.update({name: source[name] for name in extras})
        output.append(row)
    return output


def build_table(
    values: Iterable[Any],
    *,
    name: str = "table",
    options: Optional[TableBuildOptions] = None,
    schema: Optional[TableSchema] = None,
    primary_key: Sequence[str] = (),
    metadata: Optional[Mapping[str, Any]] = None,
) -> TableData:
    """Build a portable table."""
    options = options or TableBuildOptions()
    rows = records_to_table_rows(values, options=options)
    columns = discover_columns(rows)
    if schema is not None:
        rows = apply_table_schema(
            rows,
            schema,
            strict=options.schema_strict,
            include_extra=options.extra_columns,
        )
        columns = discover_columns(rows)
    columns = order_table_columns(columns, options=options, schema=schema)
    if schema is None:
        schema = infer_table_schema(name, rows, columns=columns, primary_key=primary_key)
    return TableData(
        name=name,
        columns=columns,
        rows=rows,
        schema=schema,
        metadata=dict(metadata or {}),
    )


def build_single_record_table(
    value: Any,
    *,
    name: str = "summary",
    options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Build a one-row table."""
    return build_table([value], name=name, options=options)


def interactions_table(
    interactions: Iterable[Any],
    *,
    name: str = "interactions",
    interaction_options: Optional[InteractionExportOptions] = None,
    table_options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Build an interaction table."""
    records = interactions_to_records(interactions, options=interaction_options)
    return build_table(records, name=name, options=table_options)


def analyses_table(
    analyses: Iterable[Any],
    *,
    name: str = "analyses",
    analysis_options: Optional[AnalysisExportOptions] = None,
    table_options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Build an analysis-results table."""
    records = analyses_to_records(analyses, options=analysis_options)
    return build_table(records, name=name, options=table_options)


def scoring_table(
    scorings: Iterable[Any],
    *,
    name: str = "scoring",
    scoring_options: Optional[ScoringExportOptions] = None,
    table_options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Build a scoring table."""
    records = scorings_to_records(scorings, options=scoring_options)
    return build_table(records, name=name, options=table_options)


def dock_models_table(
    models: Iterable[Any],
    *,
    name: str = "dock_models",
    dock_options: Optional[DockModelExportOptions] = None,
    table_options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Build a DockModel table."""
    records = dock_models_to_records(models, options=dock_options)
    return build_table(records, name=name, options=table_options)


def table_from_mapping(
    value: Mapping[str, Any],
    *,
    name: str = "mapping",
    key_column: str = "key",
    value_column: str = "value",
    options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Build a key-value table."""
    records = [{key_column: key, value_column: item} for key, item in value.items()]
    return build_table(records, name=name, options=options)


def transpose_table(table: TableData, *, name: Optional[str] = None) -> TableData:
    """Transpose a table using the first column as row labels."""
    if not table.columns:
        return TableData(name=name or f"{table.name}_transposed")
    key_column = table.columns[0]
    output_rows: List[Dict[str, Any]] = []
    for column in table.columns[1:]:
        row: Dict[str, Any] = {key_column: column}
        for index, source in enumerate(table.rows):
            label = source.get(key_column)
            label = str(label) if label not in (None, "") else str(index)
            row[label] = source.get(column)
        output_rows.append(row)
    return TableData(name=name or f"{table.name}_transposed", rows=output_rows)


def merge_tables(
    tables: Iterable[TableData],
    *,
    name: str = "merged",
    source_column: Optional[str] = "source_table",
) -> TableData:
    """Append compatible tables."""
    rows: List[Dict[str, Any]] = []
    metadata: Dict[str, Any] = {"sources": []}
    for table in tables:
        metadata["sources"].append(table.name)
        for source in table.rows:
            row = dict(source)
            if source_column:
                row = {source_column: table.name, **row}
            rows.append(row)
    return TableData(name=name, rows=rows, metadata=metadata)


def split_table(
    table: TableData,
    column: str,
    *,
    name_prefix: Optional[str] = None,
) -> TableCollection:
    """Split a table by a column value."""
    if column not in table.columns:
        raise KeyError(column)
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in table.rows:
        key = "none" if row.get(column) is None else str(row.get(column))
        groups[key].append(dict(row))
    collection = TableCollection(metadata={"source": table.name, "split_column": column})
    for key, rows in groups.items():
        table_name = f"{name_prefix or table.name}_{sanitize_filename(key)}"
        collection.add(TableData(name=table_name, rows=rows))
    return collection


def table_summary(table: TableData) -> Dict[str, Any]:
    """Return compact table statistics."""
    return {
        "name": table.name,
        "rows": table.row_count,
        "columns": table.column_count,
        "column_names": list(table.columns),
        "empty": table.empty,
        "null_counts": {
            column: sum(row.get(column) is None for row in table.rows)
            for column in table.columns
        },
    }


def table_serializer(
    value: Any,
    context: Optional[SerializerCallContext] = None,
    **_: Any,
) -> Dict[str, Any]:
    """Serialize table objects."""
    if isinstance(value, TableData):
        return value.to_dict()
    if isinstance(value, TableSchema):
        return value.to_dict()
    if isinstance(value, TableColumn):
        return value.to_dict()
    if isinstance(value, TableCollection):
        return value.to_dict()
    raise TypeError(f"Unsupported table object: {type(value).__name__}.")


def register_table_serializers(
    registry: Optional[SerializerRegistry] = None,
    *,
    priority: int = 720,
    replace_existing: bool = True,
) -> SerializerRegistry:
    """Register table serializers."""
    target = registry or DEFAULT_SERIALIZER_REGISTRY
    for name, target_type in (
        ("table_data", TableData),
        ("table_schema", TableSchema),
        ("table_column", TableColumn),
        ("table_collection", TableCollection),
    ):
        target.register(
            name,
            table_serializer,
            target_type=target_type,
            match_mode=SerializerMatchMode.EXACT,
            priority=priority,
            builtin=True,
            description=f"Serialize {target_type.__name__} objects.",
            replace_existing=replace_existing,
        )
    return target


register_table_serializers()

__all__.extend([
    "TABLE_SCHEMA_VERSION",
    "DEFAULT_TABLE_SEPARATOR",
    "DEFAULT_TABLE_LIST_SEPARATOR",
    "DEFAULT_MAX_TABLE_DEPTH",
    "DEFAULT_MAX_CELL_LENGTH",
    "TableOrientation",
    "TableValueMode",
    "TableColumnOrder",
    "TableColumn",
    "TableSchema",
    "TableBuildOptions",
    "TableData",
    "TableCollection",
    "normalize_table_cell",
    "flatten_table_record",
    "record_to_table_row",
    "records_to_table_rows",
    "discover_columns",
    "infer_table_dtype",
    "infer_table_schema",
    "order_table_columns",
    "apply_table_schema",
    "build_table",
    "build_single_record_table",
    "interactions_table",
    "analyses_table",
    "scoring_table",
    "dock_models_table",
    "table_from_mapping",
    "transpose_table",
    "merge_tables",
    "split_table",
    "table_summary",
    "table_serializer",
    "register_table_serializers",
])


# =============================================================================
# Section 14 — JSON and JSON Lines export
# =============================================================================

JSON_SCHEMA_VERSION: Final[str] = "1.0"


def _json_options(options: Optional[JSONExportOptions]) -> JSONExportOptions:
    return options if options is not None else JSONExportOptions()


def _json_payload(
    value: Any,
    *,
    registry: Optional[SerializerRegistry] = None,
    serialization_options: Optional[SerializationOptions] = None,
) -> Any:
    """Convert a value into a JSON-safe payload without reclassifying records."""
    if isinstance(value, (Mapping, list, tuple, set, frozenset)):
        return make_json_safe(value, options=serialization_options)
    registry = registry or DEFAULT_SERIALIZER_REGISTRY
    try:
        converted = serialize_registered(value, registry=registry)
    except Exception:
        converted = value
    return make_json_safe(converted, options=serialization_options)


def json_dumps(
    value: Any,
    *,
    options: Optional[JSONExportOptions] = None,
    registry: Optional[SerializerRegistry] = None,
    serialization_options: Optional[SerializationOptions] = None,
) -> str:
    """Serialize one value to JSON text."""
    opts = _json_options(options)
    payload = _json_payload(
        value,
        registry=registry,
        serialization_options=serialization_options,
    )
    try:
        return json.dumps(
            payload,
            indent=opts.indent,
            ensure_ascii=opts.ensure_ascii,
            sort_keys=opts.sort_keys,
            allow_nan=opts.allow_nan,
            separators=opts.separators,
        )
    except (TypeError, ValueError) as exc:
        raise ExportSerializationError("Unable to encode JSON payload") from exc


def json_lines_dumps(
    values: Iterable[Any],
    *,
    options: Optional[JSONExportOptions] = None,
    registry: Optional[SerializerRegistry] = None,
    serialization_options: Optional[SerializationOptions] = None,
) -> str:
    """Serialize values as JSON Lines text."""
    opts = _json_options(options)
    lines: List[str] = []
    for value in values:
        payload = _json_payload(
            value,
            registry=registry,
            serialization_options=serialization_options,
        )
        lines.append(json.dumps(
            payload,
            ensure_ascii=opts.ensure_ascii,
            sort_keys=opts.sort_keys,
            allow_nan=opts.allow_nan,
            separators=opts.separators,
        ))
    return "\n".join(lines) + ("\n" if lines else "")


def _write_text_atomic(path: Path, text: str, *, encoding: str, append: bool) -> None:
    """Write text atomically unless append mode is active."""
    ensure_parent_directory(path)
    if append:
        with path.open("a", encoding=encoding, newline="") as handle:
            handle.write(text)
        return
    with NamedTemporaryFile(
        "w",
        encoding=encoding,
        newline="",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
        handle.write(text)
    temp_path.replace(path)


def write_json(
    value: Any,
    path: Optional[PathLike] = None,
    *,
    output_dir: Optional[PathLike] = None,
    basename: str = DEFAULT_EXPORT_BASENAME,
    options: Optional[JSONExportOptions] = None,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    encoding: str = DEFAULT_ENCODING,
    registry: Optional[SerializerRegistry] = None,
    serialization_options: Optional[SerializationOptions] = None,
) -> ExportedFile:
    """Write one JSON document."""
    opts = _json_options(options)
    target = resolve_output_path(
        path,
        output_dir=output_dir,
        basename=basename,
        format_name=EXPORT_FORMAT_JSON,
        overwrite=overwrite,
    )
    text = json_dumps(
        value,
        options=opts,
        registry=registry,
        serialization_options=serialization_options,
    )
    if not text.endswith("\n"):
        text += "\n"
    _write_text_atomic(target, text, encoding=encoding, append=opts.append)
    exported = ExportedFile(path=target, format=EXPORT_FORMAT_JSON, record_count=1)
    exported.refresh_size()
    return exported


def write_json_lines(
    values: Iterable[Any],
    path: Optional[PathLike] = None,
    *,
    output_dir: Optional[PathLike] = None,
    basename: str = DEFAULT_EXPORT_BASENAME,
    options: Optional[JSONExportOptions] = None,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    encoding: str = DEFAULT_ENCODING,
    registry: Optional[SerializerRegistry] = None,
    serialization_options: Optional[SerializationOptions] = None,
) -> ExportedFile:
    """Write iterable values as JSON Lines."""
    opts = _json_options(options)
    opts.json_lines = True
    records = list(values)
    target = resolve_output_path(
        path,
        output_dir=output_dir,
        basename=basename,
        format_name=EXPORT_FORMAT_JSONL,
        overwrite=overwrite,
    )
    text = json_lines_dumps(
        records,
        options=opts,
        registry=registry,
        serialization_options=serialization_options,
    )
    _write_text_atomic(target, text, encoding=encoding, append=opts.append)
    exported = ExportedFile(
        path=target,
        format=EXPORT_FORMAT_JSONL,
        record_count=len(records),
    )
    exported.refresh_size()
    return exported


def read_json(path: PathLike, *, encoding: str = DEFAULT_ENCODING) -> Any:
    """Read a JSON document."""
    with Path(path).open("r", encoding=encoding) as handle:
        return json.load(handle)


def read_json_lines(path: PathLike, *, encoding: str = DEFAULT_ENCODING) -> List[Any]:
    """Read a JSON Lines document."""
    records: List[Any] = []
    with Path(path).open("r", encoding=encoding) as handle:
        for line_number, line in enumerate(handle, start=1):
            text = line.strip()
            if not text:
                continue
            try:
                records.append(json.loads(text))
            except json.JSONDecodeError as exc:
                raise ExportSerializationError(
                    f"Invalid JSON Lines record at line {line_number}"
                ) from exc
    return records


__all__.extend([
    "JSON_SCHEMA_VERSION",
    "json_dumps",
    "json_lines_dumps",
    "write_json",
    "write_json_lines",
    "read_json",
    "read_json_lines",
])


# =============================================================================
# Section 15 — CSV and TSV export
# =============================================================================

DELIMITED_SCHEMA_VERSION: Final[str] = "1.0"


def _delimited_options(
    format_name: Any,
    options: Optional[DelimitedExportOptions],
) -> DelimitedExportOptions:
    return options if options is not None else DelimitedExportOptions.for_format(format_name)


def _table_input(
    value: Any,
    *,
    name: str = "data",
    build_options: Optional[TableBuildOptions] = None,
) -> TableData:
    """Normalize input into TableData."""
    if isinstance(value, TableData):
        return value
    if isinstance(value, TableCollection):
        if len(value.tables) != 1:
            raise ExportValidationError("Delimited export requires one table")
        return next(iter(value.tables.values()))
    if PANDAS_AVAILABLE and isinstance(value, pd.DataFrame):
        return TableData(name=name, columns=list(value.columns), rows=value.to_dict("records"))
    return build_table(value, name=name, options=build_options)


def _delimited_cell(value: Any, options: DelimitedExportOptions) -> Any:
    if value is None:
        return options.null_text
    if isinstance(value, bool):
        return options.true_text if value else options.false_text
    if isinstance(value, (dict, list, tuple, set, frozenset)):
        return json.dumps(make_json_safe(value), ensure_ascii=False, separators=(",", ":"))
    return value


def delimited_dumps(
    value: Any,
    *,
    format_name: Any = EXPORT_FORMAT_CSV,
    options: Optional[DelimitedExportOptions] = None,
    table_name: str = "data",
    build_options: Optional[TableBuildOptions] = None,
) -> str:
    """Serialize one table to CSV or TSV text."""
    normalized_format = normalize_export_format(format_name)
    if normalized_format not in {EXPORT_FORMAT_CSV, EXPORT_FORMAT_TSV}:
        raise ExportFormatError(normalized_format)
    opts = _delimited_options(normalized_format, options)
    table = _table_input(value, name=table_name, build_options=build_options)
    buffer = StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=table.columns,
        delimiter=opts.delimiter,
        quotechar=opts.quotechar,
        quoting=opts.quoting,
        escapechar=opts.escapechar,
        doublequote=opts.doublequote,
        lineterminator=opts.lineterminator,
        extrasaction=opts.extras_action,
    )
    if opts.include_header:
        writer.writeheader()
    for row in table.normalized_rows():
        writer.writerow({key: _delimited_cell(value, opts) for key, value in row.items()})
    return buffer.getvalue()


def write_delimited(
    value: Any,
    path: Optional[PathLike] = None,
    *,
    format_name: Any = EXPORT_FORMAT_CSV,
    output_dir: Optional[PathLike] = None,
    basename: str = DEFAULT_EXPORT_BASENAME,
    options: Optional[DelimitedExportOptions] = None,
    table_name: str = "data",
    build_options: Optional[TableBuildOptions] = None,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    encoding: str = DEFAULT_ENCODING,
) -> ExportedFile:
    """Write one CSV or TSV table."""
    normalized_format = normalize_export_format(format_name)
    opts = _delimited_options(normalized_format, options)
    table = _table_input(value, name=table_name, build_options=build_options)
    target = resolve_output_path(
        path,
        output_dir=output_dir,
        basename=basename,
        format_name=normalized_format,
        table=table.name,
        overwrite=overwrite,
    )
    text = delimited_dumps(table, format_name=normalized_format, options=opts)
    append_header = opts.include_header
    if opts.append and target.exists() and target.stat().st_size > 0:
        opts.include_header = False
    try:
        _write_text_atomic(target, text if opts.include_header == append_header else delimited_dumps(
            table, format_name=normalized_format, options=opts
        ), encoding=encoding, append=opts.append)
    finally:
        opts.include_header = append_header
    exported = ExportedFile(
        path=target,
        format=normalized_format,
        table=table.name,
        record_count=table.row_count,
    )
    exported.refresh_size()
    return exported


def write_csv(value: Any, path: Optional[PathLike] = None, **kwargs: Any) -> ExportedFile:
    """Write CSV output."""
    return write_delimited(value, path, format_name=EXPORT_FORMAT_CSV, **kwargs)


def write_tsv(value: Any, path: Optional[PathLike] = None, **kwargs: Any) -> ExportedFile:
    """Write TSV output."""
    return write_delimited(value, path, format_name=EXPORT_FORMAT_TSV, **kwargs)


def read_delimited(
    path: PathLike,
    *,
    format_name: Optional[Any] = None,
    options: Optional[DelimitedExportOptions] = None,
    encoding: str = DEFAULT_ENCODING,
    table_name: Optional[str] = None,
) -> TableData:
    """Read CSV or TSV into TableData."""
    source = Path(path)
    inferred = format_name or source.suffix.lower().lstrip(".")
    normalized_format = normalize_export_format(inferred)
    opts = _delimited_options(normalized_format, options)
    with source.open("r", encoding=encoding, newline="") as handle:
        reader = csv.DictReader(
            handle,
            delimiter=opts.delimiter,
            quotechar=opts.quotechar,
            escapechar=opts.escapechar,
            doublequote=opts.doublequote,
        )
        rows = [dict(row) for row in reader]
        columns = list(reader.fieldnames or [])
    return TableData(name=table_name or source.stem, columns=columns, rows=rows)


__all__.extend([
    "DELIMITED_SCHEMA_VERSION",
    "delimited_dumps",
    "write_delimited",
    "write_csv",
    "write_tsv",
    "read_delimited",
])


# =============================================================================
# Section 16 — Excel export
# =============================================================================

EXCEL_SCHEMA_VERSION: Final[str] = "1.0"


def sanitize_sheet_name(name: Any, *, fallback: str = "Sheet") -> str:
    """Return an Excel-safe sheet name."""
    value = INVALID_SHEET_NAME_PATTERN.sub("_", str(name).strip()).strip("'")
    return (value or fallback)[:MAX_EXCEL_SHEET_NAME]


def unique_sheet_name(name: Any, used: Set[str]) -> str:
    """Return a unique Excel sheet name."""
    base = sanitize_sheet_name(name)
    candidate = base
    index = 2
    while candidate.lower() in used:
        suffix = f"_{index}"
        candidate = f"{base[:MAX_EXCEL_SHEET_NAME - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate.lower())
    return candidate


def _excel_tables(
    value: Any,
    *,
    table_name: str = "data",
    build_options: Optional[TableBuildOptions] = None,
) -> TableCollection:
    if isinstance(value, TableCollection):
        return value
    if isinstance(value, TableData):
        return TableCollection(tables={value.name: value})
    if isinstance(value, Mapping) and value and all(
        isinstance(item, TableData) for item in value.values()
    ):
        return TableCollection(tables={str(key): item for key, item in value.items()})
    return TableCollection(tables={table_name: _table_input(
        value,
        name=table_name,
        build_options=build_options,
    )})


def _excel_cell(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool, date, datetime, time)):
        return value
    return json.dumps(make_json_safe(value), ensure_ascii=False, separators=(",", ":"))


def _style_excel_sheet(
    worksheet: Any,
    table: TableData,
    options: ExcelExportOptions,
    *,
    table_index: int,
) -> None:
    """Apply workbook formatting."""
    if options.freeze_panes:
        worksheet.freeze_panes = options.freeze_panes
    if options.auto_filter and table.columns:
        worksheet.auto_filter.ref = worksheet.dimensions
    if options.format_headers and table.columns:
        for cell in worksheet[1]:
            if options.bold_headers:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    if options.wrap_text:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True,
                )
    if options.auto_width:
        for index, column in enumerate(table.columns, start=1):
            values = [column] + [row.get(column) for row in table.rows]
            width = max((len(str(value)) for value in values if value is not None), default=0) + 2
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                max(options.default_width, width),
                options.max_width,
            )
    if options.create_tables and table.rows and table.columns:
        ref = f"A1:{get_column_letter(len(table.columns))}{len(table.rows) + 1}"
        excel_table = Table(displayName=f"DockAnalyzerTable{table_index}", ref=ref)
        if options.table_style:
            excel_table.tableStyleInfo = TableStyleInfo(
                name=options.table_style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        worksheet.add_table(excel_table)


def build_excel_workbook(
    value: Any,
    *,
    options: Optional[ExcelExportOptions] = None,
    table_name: str = "data",
    build_options: Optional[TableBuildOptions] = None,
) -> Any:
    """Build an openpyxl workbook."""
    if not OPENPYXL_AVAILABLE:
        raise ExportDependencyError("openpyxl", feature="Excel export")
    opts = options or ExcelExportOptions()
    collection = _excel_tables(value, table_name=table_name, build_options=build_options)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used: Set[str] = set()
    written = 0
    for source_name, table in collection.tables.items():
        if table.empty and not opts.include_empty_sheets:
            continue
        if table.row_count + 1 > MAX_EXCEL_ROWS:
            raise ExportValidationError(f"Excel row limit exceeded for {source_name!r}")
        if table.column_count > MAX_EXCEL_COLUMNS:
            raise ExportValidationError(f"Excel column limit exceeded for {source_name!r}")
        sheet_name = unique_sheet_name(source_name, used)
        worksheet = workbook.create_sheet(sheet_name)
        if table.columns:
            worksheet.append(list(table.columns))
            for row in table.normalized_rows():
                worksheet.append([_excel_cell(row.get(column)) for column in table.columns])
        _style_excel_sheet(worksheet, table, opts, table_index=written + 1)
        written += 1
    if not workbook.sheetnames:
        workbook.create_sheet("Summary")
    workbook.properties.creator = __author__
    workbook.properties.title = "DockAnalyzer export"
    return workbook


def write_excel(
    value: Any,
    path: Optional[PathLike] = None,
    *,
    output_dir: Optional[PathLike] = None,
    basename: str = DEFAULT_EXPORT_BASENAME,
    options: Optional[ExcelExportOptions] = None,
    table_name: str = "data",
    build_options: Optional[TableBuildOptions] = None,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
) -> ExportedFile:
    """Write one Excel workbook."""
    target = resolve_output_path(
        path,
        output_dir=output_dir,
        basename=basename,
        format_name=EXPORT_FORMAT_EXCEL,
        overwrite=overwrite,
    )
    workbook = build_excel_workbook(
        value,
        options=options,
        table_name=table_name,
        build_options=build_options,
    )
    ensure_parent_directory(target)
    with NamedTemporaryFile(
        "wb",
        dir=target.parent,
        prefix=f".{target.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
    try:
        workbook.save(temp_path)
        temp_path.replace(target)
    except Exception as exc:
        temp_path.unlink(missing_ok=True)
        raise ExportWriteError("Unable to write Excel workbook", path=target) from exc
    collection = _excel_tables(value, table_name=table_name, build_options=build_options)
    exported = ExportedFile(
        path=target,
        format=EXPORT_FORMAT_EXCEL,
        sheet_names=tuple(workbook.sheetnames),
        record_count=sum(table.row_count for table in collection.tables.values()),
    )
    exported.refresh_size()
    return exported


def read_excel(
    path: PathLike,
    *,
    data_only: bool = True,
    read_only: bool = True,
) -> TableCollection:
    """Read an Excel workbook into table objects."""
    if not OPENPYXL_AVAILABLE:
        raise ExportDependencyError("openpyxl", feature="Excel import")
    workbook = openpyxl.load_workbook(path, data_only=data_only, read_only=read_only)
    collection = TableCollection()
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            collection.add(TableData(name=worksheet.title), replace=True)
            continue
        columns = [str(value) if value is not None else f"column_{index}" for index, value in enumerate(rows[0], start=1)]
        records = [dict(zip(columns, row)) for row in rows[1:]]
        collection.add(TableData(name=worksheet.title, columns=columns, rows=records), replace=True)
    workbook.close()
    return collection


__all__.extend([
    "EXCEL_SCHEMA_VERSION",
    "MAX_EXCEL_SHEET_NAME",
    "MAX_EXCEL_ROWS",
    "MAX_EXCEL_COLUMNS",
    "unique_sheet_name",
    "build_excel_workbook",
    "write_excel",
    "read_excel",
])
# =============================================================================
# Section 17 — Text and summaries
# =============================================================================

TEXT_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_TEXT_WIDTH: Final[int] = 100
DEFAULT_TEXT_INDENT: Final[int] = 2


class TextStyle(str, Enum):
    """Plain-text rendering style."""

    PLAIN = "plain"
    COMPACT = "compact"
    OUTLINE = "outline"
    KEY_VALUE = "key_value"


@dataclass(slots=True)
class TextExportOptions:
    """Options for text rendering."""

    style: TextStyle = TextStyle.PLAIN
    title: Optional[str] = None
    width: int = DEFAULT_TEXT_WIDTH
    indent: int = DEFAULT_TEXT_INDENT
    sort_keys: bool = False
    include_empty: bool = False
    include_metadata: bool = True
    include_header: bool = True
    line_prefix: str = ""
    newline: str = "\n"

    def __post_init__(self) -> None:
        if not isinstance(self.style, TextStyle):
            self.style = TextStyle(str(self.style).strip().lower())
        self.width = max(20, int(self.width))
        self.indent = max(0, int(self.indent))


@dataclass(slots=True)
class TextSummary:
    """Rendered summary and its metadata."""

    text: str
    title: Optional[str] = None
    line_count: int = 0
    character_count: int = 0
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.line_count = len(self.text.splitlines())
        self.character_count = len(self.text)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable summary."""
        return {
            "title": self.title,
            "text": self.text,
            "line_count": self.line_count,
            "character_count": self.character_count,
            "metadata": dict(self.metadata),
        }


def _text_scalar(value: Any) -> str:
    """Convert one scalar to readable text."""
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        if math.isinf(value):
            return "Infinity" if value > 0 else "-Infinity"
        return f"{value:.6g}"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    return str(value)


def _text_lines(
    value: Any,
    *,
    level: int,
    options: TextExportOptions,
) -> List[str]:
    """Render a value recursively."""
    pad = " " * (level * options.indent)
    if isinstance(value, Mapping):
        items = list(value.items())
        if options.sort_keys:
            items.sort(key=lambda item: str(item[0]))
        lines: List[str] = []
        for key, item in items:
            if not options.include_empty and item in (None, "", [], {}, ()):
                continue
            label = str(key).replace("_", " ")
            if isinstance(item, (Mapping, list, tuple, set)):
                lines.append(f"{pad}{label}:")
                lines.extend(_text_lines(item, level=level + 1, options=options))
            else:
                lines.append(f"{pad}{label}: {_text_scalar(item)}")
        return lines
    if isinstance(value, (list, tuple, set)):
        lines = []
        for item in value:
            if isinstance(item, (Mapping, list, tuple, set)):
                lines.append(f"{pad}-")
                lines.extend(_text_lines(item, level=level + 1, options=options))
            else:
                lines.append(f"{pad}- {_text_scalar(item)}")
        return lines
    return [f"{pad}{_text_scalar(value)}"]


def text_dumps(
    value: Any,
    *,
    options: Optional[TextExportOptions] = None,
    registry: Optional[SerializerRegistry] = None,
) -> str:
    """Render any supported object as plain text."""
    opts = options or TextExportOptions()
    payload = to_serializable(value)
    lines: List[str] = []
    if opts.include_header and opts.title:
        lines.extend([opts.title, "=" * min(len(opts.title), opts.width), ""])
    lines.extend(_text_lines(payload, level=0, options=opts))
    if opts.line_prefix:
        lines = [f"{opts.line_prefix}{line}" for line in lines]
    return opts.newline.join(lines).rstrip() + opts.newline


def summarize_value(
    value: Any,
    *,
    title: Optional[str] = None,
    options: Optional[TextExportOptions] = None,
    registry: Optional[SerializerRegistry] = None,
) -> TextSummary:
    """Build a reusable text summary."""
    opts = options or TextExportOptions(title=title)
    if title is not None:
        opts = replace(opts, title=title)
    text = text_dumps(value, options=opts, registry=registry)
    return TextSummary(text=text, title=opts.title)


def write_text(
    value: Any,
    path: Optional[PathLike] = None,
    *,
    output_dir: Optional[PathLike] = None,
    basename: str = DEFAULT_EXPORT_BASENAME,
    options: Optional[TextExportOptions] = None,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    encoding: str = DEFAULT_ENCODING,
    registry: Optional[SerializerRegistry] = None,
) -> ExportedFile:
    """Write a plain-text export."""
    target = resolve_output_path(
        path,
        output_dir=output_dir,
        basename=basename,
        format_name=EXPORT_FORMAT_TEXT,
        overwrite=overwrite,
    )
    text = text_dumps(value, options=options, registry=registry)
    ensure_parent_directory(target)
    with NamedTemporaryFile("w", encoding=encoding, newline="", delete=False, dir=target.parent) as handle:
        handle.write(text)
        temporary = Path(handle.name)
    os.replace(temporary, target)
    exported = ExportedFile(path=target, format=EXPORT_FORMAT_TEXT)
    exported.refresh_size()
    return exported


def read_text(path: PathLike, *, encoding: str = DEFAULT_ENCODING) -> str:
    """Read a text export."""
    return Path(path).read_text(encoding=encoding)


__all__.extend([
    "TEXT_SCHEMA_VERSION",
    "TextStyle",
    "TextExportOptions",
    "TextSummary",
    "text_dumps",
    "summarize_value",
    "write_text",
    "read_text",
])
# =============================================================================
# Section 18 — Manifest and provenance
# =============================================================================

MANIFEST_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_MANIFEST_NAME: Final[str] = "manifest"


@dataclass(slots=True)
class ProvenanceRecord:
    """Runtime and source provenance."""

    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    application: str = "DockAnalyzer"
    application_version: str = __version__
    python_version: str = field(default_factory=lambda: platform.python_version())
    platform_name: str = field(default_factory=platform.platform)
    hostname: Optional[str] = None
    command: Optional[str] = None
    working_directory: Optional[str] = None
    source_files: List[Dict[str, Any]] = field(default_factory=list)
    parameters: Dict[str, Any] = field(default_factory=dict)
    environment: Dict[str, Any] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive provenance data."""
        return {
            "created_at": self.created_at.isoformat(),
            "application": self.application,
            "application_version": self.application_version,
            "python_version": self.python_version,
            "platform": self.platform_name,
            "hostname": self.hostname,
            "command": self.command,
            "working_directory": self.working_directory,
            "source_files": list(self.source_files),
            "parameters": dict(self.parameters),
            "environment": dict(self.environment),
            "metadata": dict(self.metadata),
        }


def compute_file_hash(
    path: PathLike,
    *,
    algorithm: str = DEFAULT_HASH_ALGORITHM,
    chunk_size: int = 1024 * 1024,
) -> str:
    """Compute a cryptographic file hash."""
    import hashlib

    try:
        digest = hashlib.new(algorithm)
    except ValueError as exc:
        raise ExportConfigurationError(f"Unsupported hash algorithm: {algorithm!r}.") from exc
    with Path(path).open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def file_provenance(
    path: PathLike,
    *,
    root: Optional[PathLike] = None,
    include_hash: bool = True,
    hash_algorithm: str = DEFAULT_HASH_ALGORITHM,
) -> Dict[str, Any]:
    """Describe one source or generated file."""
    value = Path(path).resolve()
    record: Dict[str, Any] = {
        "path": relative_export_path(value, root) if root else str(value),
        "name": value.name,
        "exists": value.exists(),
    }
    if value.exists():
        stat = value.stat()
        record.update({
            "size_bytes": stat.st_size,
            "modified_at": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        })
        if include_hash and value.is_file():
            record["hash"] = compute_file_hash(value, algorithm=hash_algorithm)
            record["hash_algorithm"] = hash_algorithm
    return record


def collect_provenance(
    *,
    source_files: Optional[Iterable[PathLike]] = None,
    parameters: Optional[Mapping[str, Any]] = None,
    metadata: Optional[Mapping[str, Any]] = None,
    include_environment: bool = False,
    include_hashes: bool = True,
    root: Optional[PathLike] = None,
) -> ProvenanceRecord:
    """Collect portable execution provenance."""
    import socket

    environment: Dict[str, Any] = {}
    if include_environment:
        allowed = ("CONDA_DEFAULT_ENV", "VIRTUAL_ENV", "CHIMERAX_VERSION")
        environment = {key: os.environ[key] for key in allowed if key in os.environ}
    return ProvenanceRecord(
        hostname=socket.gethostname(),
        command=" ".join(sys.argv) if sys.argv else None,
        working_directory=str(Path.cwd()),
        source_files=[
            file_provenance(path, root=root, include_hash=include_hashes)
            for path in (source_files or ())
        ],
        parameters=dict(parameters or {}),
        environment=environment,
        metadata=dict(metadata or {}),
    )


def enrich_exported_file(
    exported_file: ExportedFile,
    *,
    include_hash: bool = True,
    hash_algorithm: str = DEFAULT_HASH_ALGORITHM,
) -> ExportedFile:
    """Refresh file metadata and optional checksum."""
    exported_file.refresh_size()
    if include_hash and exported_file.exists:
        exported_file.hash_value = compute_file_hash(
            exported_file.path,
            algorithm=hash_algorithm,
        )
        exported_file.hash_algorithm = hash_algorithm
    return exported_file


def build_manifest(
    files: Iterable[ExportedFile],
    *,
    source_name: Optional[str] = None,
    export_id: Optional[str] = None,
    options: Optional[ExportOptions] = None,
    provenance: Optional[Union[ProvenanceRecord, Mapping[str, Any]]] = None,
    metadata: Optional[Mapping[str, Any]] = None,
    include_hashes: bool = True,
    hash_algorithm: str = DEFAULT_HASH_ALGORITHM,
) -> ExportManifest:
    """Build an export manifest."""
    records = [
        enrich_exported_file(item, include_hash=include_hashes, hash_algorithm=hash_algorithm)
        for item in files
    ]
    provenance_data = (
        provenance.to_dict()
        if isinstance(provenance, ProvenanceRecord)
        else dict(provenance or {})
    )
    return ExportManifest(
        export_id=export_id,
        source_name=source_name,
        files=records,
        options=options,
        provenance=provenance_data,
        metadata=dict(metadata or {}),
    )


def write_manifest(
    manifest: ExportManifest,
    path: Optional[PathLike] = None,
    *,
    output_dir: Optional[PathLike] = None,
    basename: str = DEFAULT_MANIFEST_NAME,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    json_options: Optional[JSONExportOptions] = None,
) -> ExportedFile:
    """Write a manifest as JSON."""
    target = resolve_output_path(
        path,
        output_dir=output_dir,
        basename=basename,
        format_name=EXPORT_FORMAT_JSON,
        overwrite=overwrite,
    )
    options = json_options or JSONExportOptions()
    text = json.dumps(
        manifest.to_dict(),
        ensure_ascii=options.ensure_ascii,
        indent=options.indent,
        sort_keys=options.sort_keys,
        allow_nan=options.allow_nan,
    ) + "\n"
    atomic_write_text(target, text, overwrite=OverwriteMode.OVERWRITE.value)
    exported = ExportedFile(path=target, format=EXPORT_FORMAT_JSON, record_count=1)
    exported.refresh_size()
    return exported


def read_manifest(path: PathLike) -> ExportManifest:
    """Read a manifest JSON file."""
    data = read_json(path)
    files = [
        ExportedFile(
            path=item["path"],
            format=item.get("format", Path(item["path"]).suffix.lstrip(".")),
            status=item.get("status", ExportStatus.SUCCESS.value),
            table=item.get("table"),
            sheet_names=tuple(item.get("sheet_names", ())),
            record_count=item.get("record_count"),
            size_bytes=item.get("size_bytes"),
            hash_value=item.get("hash"),
            hash_algorithm=item.get("hash_algorithm"),
            metadata=item.get("metadata", {}),
        )
        for item in data.get(KEY_FILES, [])
    ]
    return ExportManifest(
        schema_name=data.get(KEY_SCHEMA_NAME, EXPORT_SCHEMA_NAME),
        schema_version=data.get(KEY_SCHEMA_VERSION, EXPORT_SCHEMA_VERSION),
        dockanalyzer_version=data.get(KEY_DOCKANALYZER_VERSION, __version__),
        export_id=data.get("export_id"),
        source_name=data.get("source_name"),
        files=files,
        provenance=data.get(KEY_PROVENANCE, {}),
        metadata=data.get(KEY_METADATA, {}),
        warnings=data.get(KEY_WARNINGS, []),
        errors=data.get(KEY_ERRORS, []),
    )


def verify_manifest(
    manifest: ExportManifest,
    *,
    root: Optional[PathLike] = None,
    verify_hashes: bool = True,
) -> Dict[str, Any]:
    """Verify files recorded by a manifest."""
    base = Path(root).resolve() if root else None
    results: List[Dict[str, Any]] = []
    for item in manifest.files:
        path = Path(item.path)
        if base is not None and not path.is_absolute():
            path = base / path
        exists = path.is_file()
        valid_hash: Optional[bool] = None
        if exists and verify_hashes and item.hash_value and item.hash_algorithm:
            valid_hash = compute_file_hash(path, algorithm=item.hash_algorithm) == item.hash_value
        results.append({
            "path": str(path),
            "exists": exists,
            "hash_valid": valid_hash,
            "valid": exists and valid_hash is not False,
        })
    return {
        "valid": all(item["valid"] for item in results),
        "file_count": len(results),
        "files": results,
    }


__all__.extend([
    "MANIFEST_SCHEMA_VERSION",
    "DEFAULT_MANIFEST_NAME",
    "DEFAULT_HASH_ALGORITHM",
    "ProvenanceRecord",
    "compute_file_hash",
    "file_provenance",
    "collect_provenance",
    "enrich_exported_file",
    "build_manifest",
    "write_manifest",
    "read_manifest",
    "verify_manifest",
])
# =============================================================================
# Section 19 — Safe and transactional writing
# =============================================================================

TRANSACTION_SCHEMA_VERSION: Final[str] = "1.0"


class TransactionState(str, Enum):
    """Transaction lifecycle state."""

    PENDING = "pending"
    ACTIVE = "active"
    COMMITTED = "committed"
    ROLLED_BACK = "rolled_back"
    FAILED = "failed"


@dataclass(slots=True)
class TransactionEntry:
    """One staged output file."""

    target: Path
    staged: Path
    backup: Optional[Path] = None
    committed: bool = False

    def to_dict(self) -> Dict[str, Any]:
        """Return primitive entry data."""
        return {
            "target": str(self.target),
            "staged": str(self.staged),
            "backup": str(self.backup) if self.backup else None,
            "committed": self.committed,
        }


@dataclass
class ExportTransaction:
    """Stage and atomically commit multiple files."""

    output_dir: PathLike
    overwrite: Any = DEFAULT_OVERWRITE_MODE
    keep_backups: bool = False
    state: TransactionState = TransactionState.PENDING
    entries: List[TransactionEntry] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    _temporary_directory: Optional[TemporaryDirectory] = field(default=None, init=False, repr=False)

    def __enter__(self) -> "ExportTransaction":
        self.begin()
        return self

    def __exit__(self, exc_type: Any, exc: Any, traceback: Any) -> bool:
        if exc_type is None:
            self.commit()
        else:
            self.rollback()
        return False

    @property
    def staging_dir(self) -> Path:
        """Return the active staging directory."""
        if self._temporary_directory is None:
            raise ExportConfigurationError("Transaction has not been started.")
        return Path(self._temporary_directory.name)

    def begin(self) -> None:
        """Start the transaction."""
        if self.state not in {TransactionState.PENDING, TransactionState.ROLLED_BACK}:
            raise ExportConfigurationError(f"Cannot begin transaction in state {self.state.value!r}.")
        output = Path(self.output_dir).resolve()
        output.mkdir(parents=True, exist_ok=True)
        self.output_dir = output
        self._temporary_directory = TemporaryDirectory(prefix=".dockanalyzer-", dir=output)
        self.entries.clear()
        self.state = TransactionState.ACTIVE

    def stage_path(self, target: PathLike) -> Path:
        """Create and register a staged path."""
        if self.state != TransactionState.ACTIVE:
            raise ExportConfigurationError("Transaction is not active.")
        resolved = Path(target)
        if not resolved.is_absolute():
            resolved = Path(self.output_dir) / resolved
        resolved = resolved.resolve()
        try:
            relative = resolved.relative_to(Path(self.output_dir))
        except ValueError as exc:
            raise ExportPathError("Transactional targets must be inside output_dir.", path=resolved) from exc
        staged = self.staging_dir / relative
        staged.parent.mkdir(parents=True, exist_ok=True)
        self.entries.append(TransactionEntry(target=resolved, staged=staged))
        return staged

    def register(self, staged: PathLike, target: PathLike) -> TransactionEntry:
        """Register an externally written staged file."""
        staged_path = Path(staged).resolve()
        target_path = Path(target)
        if not target_path.is_absolute():
            target_path = Path(self.output_dir) / target_path
        entry = TransactionEntry(target=target_path.resolve(), staged=staged_path)
        self.entries.append(entry)
        return entry

    def commit(self) -> List[Path]:
        """Commit all staged files."""
        if self.state != TransactionState.ACTIVE:
            raise ExportConfigurationError("Transaction is not active.")
        committed: List[TransactionEntry] = []
        try:
            for entry in self.entries:
                if not entry.staged.is_file():
                    raise ExportWriteError("Staged file does not exist.", path=entry.staged)
                entry.target.parent.mkdir(parents=True, exist_ok=True)
                if entry.target.exists():
                    mode = self.overwrite.value if isinstance(self.overwrite, OverwriteMode) else str(self.overwrite)
                    if mode == OverwriteMode.ERROR.value:
                        raise ExportPathError("Output file already exists.", path=entry.target)
                    if mode == OverwriteMode.UNIQUE.value:
                        entry.target = resolve_unique_path(entry.target)
                    elif mode == OverwriteMode.BACKUP.value:
                        entry.backup = backup_path(entry.target)
                        os.replace(entry.target, entry.backup)
                    elif mode == OverwriteMode.OVERWRITE.value:
                        entry.backup = backup_path(entry.target)
                        os.replace(entry.target, entry.backup)
                os.replace(entry.staged, entry.target)
                entry.committed = True
                committed.append(entry)
            if not self.keep_backups:
                for entry in committed:
                    if entry.backup and entry.backup.exists():
                        entry.backup.unlink()
                        entry.backup = None
            self.state = TransactionState.COMMITTED
            self._cleanup()
            return [entry.target for entry in committed]
        except Exception:
            self.state = TransactionState.FAILED
            for entry in reversed(committed):
                if entry.target.exists():
                    entry.target.unlink()
                if entry.backup and entry.backup.exists():
                    os.replace(entry.backup, entry.target)
                entry.committed = False
            self._cleanup()
            raise

    def rollback(self) -> None:
        """Discard staged files and restore backups."""
        for entry in reversed(self.entries):
            if entry.committed and entry.target.exists():
                entry.target.unlink()
            if entry.backup and entry.backup.exists():
                os.replace(entry.backup, entry.target)
            entry.committed = False
        self.state = TransactionState.ROLLED_BACK
        self._cleanup()

    def _cleanup(self) -> None:
        if self._temporary_directory is not None:
            self._temporary_directory.cleanup()
            self._temporary_directory = None

    def to_dict(self) -> Dict[str, Any]:
        """Return transaction metadata."""
        return {
            "state": self.state.value,
            "output_dir": str(self.output_dir),
            "overwrite": self.overwrite.value if isinstance(self.overwrite, OverwriteMode) else str(self.overwrite),
            "entries": [entry.to_dict() for entry in self.entries],
            "metadata": dict(self.metadata),
        }


def atomic_write_bytes(
    path: PathLike,
    data: bytes,
    *,
    overwrite: Any = OverwriteMode.OVERWRITE.value,
) -> Path:
    """Atomically write binary data."""
    target = resolve_overwrite_path(path, overwrite)
    ensure_parent_directory(target)
    with NamedTemporaryFile("wb", delete=False, dir=target.parent) as handle:
        handle.write(data)
        handle.flush()
        os.fsync(handle.fileno())
        temporary = Path(handle.name)
    try:
        os.replace(temporary, target)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return target


def atomic_write_text(
    path: PathLike,
    text: str,
    *,
    encoding: str = DEFAULT_ENCODING,
    overwrite: Any = OverwriteMode.OVERWRITE.value,
) -> Path:
    """Atomically write text data."""
    return atomic_write_bytes(path, text.encode(encoding), overwrite=overwrite)


@contextmanager
def transactional_export(
    output_dir: PathLike,
    *,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
    keep_backups: bool = False,
) -> Iterator[ExportTransaction]:
    """Yield an active export transaction."""
    transaction = ExportTransaction(
        output_dir=output_dir,
        overwrite=overwrite,
        keep_backups=keep_backups,
    )
    with transaction:
        yield transaction


def transactional_write(
    writers: Mapping[PathLike, Callable[[Path], Any]],
    *,
    output_dir: PathLike,
    overwrite: Any = DEFAULT_OVERWRITE_MODE,
) -> List[Path]:
    """Run multiple writers as one transaction."""
    with transactional_export(output_dir, overwrite=overwrite) as transaction:
        for target, writer in writers.items():
            staged = transaction.stage_path(target)
            writer(staged)
    return [entry.target for entry in transaction.entries]


__all__.extend([
    "TRANSACTION_SCHEMA_VERSION",
    "TransactionState",
    "TransactionEntry",
    "ExportTransaction",
    "atomic_write_bytes",
    "atomic_write_text",
    "transactional_export",
    "transactional_write",
])
# =============================================================================
# Section 20 — Single-pose export
# =============================================================================

SINGLE_POSE_SCHEMA_VERSION: Final[str] = "1.0"


@dataclass(slots=True)
class SinglePoseExportContext:
    """State for one pose export."""

    pose: Any
    name: str
    output_dir: Path
    options: ExportOptions
    record: Dict[str, Any] = field(default_factory=dict)
    tables: TableCollection = field(default_factory=TableCollection)
    metadata: Dict[str, Any] = field(default_factory=dict)


def _export_format_values(formats: Iterable[Any]) -> Tuple[str, ...]:
    values: List[str] = []
    for item in formats:
        value = normalize_export_format(item)
        if value not in values:
            values.append(value)
    return tuple(values)


def _pose_export_name(pose: Any, fallback: str = DEFAULT_EXPORT_BASENAME) -> str:
    for key in ("name", "pose_name", "pose_id", "id", "identifier"):
        value = _molecular_get(pose, key, default=None)
        if value not in (None, ""):
            return sanitize_filename(value, fallback=fallback)
    return sanitize_filename(fallback, fallback=DEFAULT_EXPORT_BASENAME)


def _pose_record(pose: Any, options: ExportOptions) -> Dict[str, Any]:
    if is_dock_model_like(pose):
        return dock_model_to_record(
            pose,
            options=DockModelExportOptions(
                layout=DockModelLayout.STANDARD,
                include_metadata=options.include_metadata,
                include_files=True,
            ),
        )
    value = serialize_registered(pose)
    if isinstance(value, Mapping):
        return dict(value)
    return {"value": make_json_safe(value)}


def _pose_tables(
    pose: Any,
    record: Mapping[str, Any],
    options: ExportOptions,
) -> TableCollection:
    collection = TableCollection()
    pose_table = build_single_record_table(record, name="pose")
    if options.wants_table("pose"):
        collection.add(pose_table)
    if is_dock_model_like(pose):
        try:
            table = dock_models_table([pose], name="poses")
            if options.wants_table("poses"):
                collection.add(table)
        except Exception:
            pass
    interactions = record.get("interactions")
    if interactions:
        if isinstance(interactions, Mapping):
            flattened: List[Any] = []
            for family, items in interactions.items():
                if isinstance(items, Sequence) and not isinstance(items, (str, bytes)):
                    for item in items:
                        if isinstance(item, Mapping):
                            merged = dict(item)
                            merged.setdefault("family", family)
                            flattened.append(merged)
                        else:
                            flattened.append(item)
            interactions = flattened
        try:
            table = interactions_table(interactions, name="interactions")
            if not table.empty or options.include_empty_tables:
                if options.wants_table("interactions"):
                    collection.add(table)
        except Exception:
            pass
    scoring = record.get("scoring") or record.get("score")
    if scoring:
        try:
            values = scoring if isinstance(scoring, list) else [scoring]
            table = scoring_table(values, name="scoring")
            if not table.empty or options.include_empty_tables:
                if options.wants_table("scoring"):
                    collection.add(table)
        except Exception:
            pass
    return collection


def prepare_single_pose_export(
    pose: Any,
    *,
    options: Optional[ExportOptions] = None,
    output_dir: Optional[PathLike] = None,
    basename: Optional[str] = None,
    metadata: Optional[Mapping[str, Any]] = None,
) -> SinglePoseExportContext:
    """Prepare one pose for export."""
    if pose is None:
        raise ExportInputError("Export source cannot be None.")
    opts = options or ExportOptions()
    if output_dir is not None:
        opts = opts.copy(output_dir=output_dir)
    name = sanitize_filename(basename or opts.basename or _pose_export_name(pose))
    directory = Path(opts.output_dir or Path.cwd())
    if opts.create_directories:
        directory.mkdir(parents=True, exist_ok=True)
    record = _pose_record(pose, opts)
    tables = _pose_tables(pose, record, opts)
    return SinglePoseExportContext(
        pose=pose,
        name=name,
        output_dir=directory,
        options=opts,
        record=record,
        tables=tables,
        metadata=dict(metadata or {}),
    )


def _export_pose_format(
    context: SinglePoseExportContext,
    format_name: str,
) -> List[ExportedFile]:
    opts = context.options
    name = context.name
    directory = context.output_dir
    if format_name == EXPORT_FORMAT_JSON:
        return [write_json(
            context.record,
            output_dir=directory,
            basename=name,
            options=opts.json,
            overwrite=opts.overwrite,
        )]
    if format_name == EXPORT_FORMAT_JSONL:
        return [write_json_lines(
            [context.record],
            output_dir=directory,
            basename=name,
            options=opts.json,
            overwrite=opts.overwrite,
        )]
    if format_name in {EXPORT_FORMAT_CSV, EXPORT_FORMAT_TSV}:
        files: List[ExportedFile] = []
        writer = write_csv if format_name == EXPORT_FORMAT_CSV else write_tsv
        for table_name, table in context.tables.tables.items():
            if table.empty and not opts.include_empty_tables:
                continue
            files.append(writer(
                table,
                output_dir=directory,
                basename=f"{name}_{sanitize_filename(table_name)}",
                table_name=table_name,
                options=opts.delimited,
                overwrite=opts.overwrite,
            ))
        return files
    if format_name == EXPORT_FORMAT_EXCEL:
        return [write_excel(
            context.tables,
            output_dir=directory,
            basename=name,
            options=opts.excel,
            overwrite=opts.overwrite,
        )]
    if format_name in {EXPORT_FORMAT_TEXT, "txt"}:
        return [write_text(
            context.record,
            output_dir=directory,
            basename=name,
            overwrite=opts.overwrite,
        )]
    raise ExportFormatError(f"Unsupported single-pose format: {format_name!r}")


def export_single_pose(
    pose: Any,
    *,
    options: Optional[ExportOptions] = None,
    output_dir: Optional[PathLike] = None,
    basename: Optional[str] = None,
    metadata: Optional[Mapping[str, Any]] = None,
) -> ExportResult:
    """Export one pose in all requested formats."""
    context = prepare_single_pose_export(
        pose,
        options=options,
        output_dir=output_dir,
        basename=basename,
        metadata=metadata,
    )
    result = ExportResult(
        source_name=context.name,
        output_dir=context.output_dir,
        payload=context.record,
        tables={name: table.normalized_rows() for name, table in context.tables.tables.items()},
        metadata={**context.metadata, "schema_version": SINGLE_POSE_SCHEMA_VERSION},
    )
    formats = _export_format_values(context.options.formats)
    for format_name in formats:
        try:
            for exported in _export_pose_format(context, format_name):
                result.add_file(exported)
        except Exception as exc:
            result.add_error(exc)
            if context.options.error_mode == ErrorMode.RAISE.value:
                raise
    if context.options.include_manifest:
        try:
            provenance = None
            if context.options.include_provenance:
                provenance = collect_provenance(
                    parameters=context.options.to_dict(),
                    metadata=context.metadata,
                    root=context.output_dir,
                )
            result.manifest = build_manifest(
                result.files,
                source_name=context.name,
                options=context.options,
                provenance=provenance,
                metadata=context.metadata,
                include_hashes=bool(context.options.hash_algorithm),
                hash_algorithm=context.options.hash_algorithm or DEFAULT_HASH_ALGORITHM,
            )
            manifest_file = write_manifest(
                result.manifest,
                output_dir=context.output_dir,
                basename=f"{context.name}_manifest",
                overwrite=context.options.overwrite,
                json_options=context.options.json,
            )
            result.add_file(manifest_file)
        except Exception as exc:
            result.add_error(exc)
            if context.options.error_mode == ErrorMode.RAISE.value:
                raise
    if context.options.update_model_files and (
        hasattr(pose, "files")
        or (isinstance(pose, Mapping) and "files" in pose)
    ):
        try:
            update_dock_model_files_from_result(
                pose,
                result,
                options=ModelFilesUpdateOptions(
                    preserve_existing=context.options.preserve_previous_files,
                ),
            )
        except Exception as exc:
            result.add_warning(
                {
                    "type": exc.__class__.__name__,
                    "message": "Unable to update DockModel.files",
                }
            )
    return result.finish()


__all__.extend([
    "SINGLE_POSE_SCHEMA_VERSION",
    "SinglePoseExportContext",
    "prepare_single_pose_export",
    "export_single_pose",
])
# =============================================================================
# Section 21 — Multipose export
# =============================================================================

MULTIPOSE_SCHEMA_VERSION: Final[str] = "1.0"


@dataclass(slots=True)
class MultiposeExportOptions:
    """Options specific to multipose export."""

    export: ExportOptions = field(default_factory=ExportOptions)
    include_individual_poses: bool = False
    separate_pose_directories: bool = False
    include_combined_payload: bool = True
    include_ranking: bool = True
    pose_prefix: str = DEFAULT_POSE_PREFIX
    sort_poses: bool = False
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.pose_prefix = str(self.pose_prefix).strip() or DEFAULT_POSE_PREFIX
        self.metadata = dict(self.metadata)


@dataclass(slots=True)
class MultiposeExportContext:
    """Prepared multipose payload."""

    poses: List[Any]
    records: List[Dict[str, Any]]
    tables: TableCollection
    name: str
    output_dir: Path
    options: MultiposeExportOptions


def prepare_multipose_export(
    poses: Iterable[Any],
    *,
    options: Optional[MultiposeExportOptions] = None,
    output_dir: Optional[PathLike] = None,
    basename: Optional[str] = None,
) -> MultiposeExportContext:
    """Prepare multiple poses for export."""
    if poses is None:
        raise ExportInputError("Pose collection cannot be None.")
    opts = options or MultiposeExportOptions()
    export_opts = opts.export
    if output_dir is not None:
        export_opts = export_opts.copy(output_dir=output_dir)
        opts = replace(opts, export=export_opts)
    values = list(poses)
    if opts.sort_poses:
        try:
            values = rank_dock_models(values)
        except Exception:
            pass
    name = sanitize_filename(basename or export_opts.basename or "multipose")
    directory = Path(export_opts.output_dir or Path.cwd())
    directory.mkdir(parents=True, exist_ok=True)
    records = [_pose_record(pose, export_opts) for pose in values]
    tables = TableCollection()
    try:
        tables.add(dock_models_table(values, name="poses"))
    except Exception:
        tables.add(build_table(records, name="poses"))
    interactions: List[Dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        source = record.get("interactions")
        if isinstance(source, Mapping):
            for family, items in source.items():
                if isinstance(items, Sequence) and not isinstance(items, (str, bytes)):
                    for item in items:
                        if isinstance(item, Mapping):
                            row = dict(item)
                            row.setdefault("family", family)
                            row.setdefault("pose_index", index)
                            interactions.append(row)
        elif isinstance(source, Sequence) and not isinstance(source, (str, bytes)):
            for item in source:
                if isinstance(item, Mapping):
                    row = dict(item)
                    row.setdefault("pose_index", index)
                    interactions.append(row)
    if interactions:
        tables.add(interactions_table(interactions, name="interactions"))
    if opts.include_ranking:
        ranking_rows: List[Dict[str, Any]] = []
        for index, record in enumerate(records, start=1):
            ranking_rows.append({
                "pose_index": index,
                "pose_id": record.get("id") or record.get("pose_id") or record.get("name"),
                "rank": record.get("rank"),
                "total_score": (
                    record.get("total_score")
                    if record.get("total_score") is not None
                    else record.get("score")
                ),
            })
        tables.add(build_table(ranking_rows, name="ranking"))
    return MultiposeExportContext(values, records, tables, name, directory, opts)


def export_multiple_poses(
    poses: Iterable[Any],
    *,
    options: Optional[MultiposeExportOptions] = None,
    output_dir: Optional[PathLike] = None,
    basename: Optional[str] = None,
) -> ExportResult:
    """Export a multipose collection."""
    context = prepare_multipose_export(
        poses,
        options=options,
        output_dir=output_dir,
        basename=basename,
    )
    export_opts = context.options.export
    payload: Dict[str, Any] = {
        "schema_version": MULTIPOSE_SCHEMA_VERSION,
        "name": context.name,
        "pose_count": len(context.records),
        "poses": context.records,
        "metadata": context.options.metadata,
    }
    result = ExportResult(
        source_name=context.name,
        output_dir=context.output_dir,
        payload=payload,
        tables={name: table.normalized_rows() for name, table in context.tables.tables.items()},
        metadata={"pose_count": len(context.records), **context.options.metadata},
    )
    for format_name in _export_format_values(export_opts.formats):
        try:
            if format_name == EXPORT_FORMAT_JSON:
                result.add_file(write_json(payload, output_dir=context.output_dir, basename=context.name, options=export_opts.json, overwrite=export_opts.overwrite))
            elif format_name == EXPORT_FORMAT_JSONL:
                result.add_file(write_json_lines(context.records, output_dir=context.output_dir, basename=context.name, options=export_opts.json, overwrite=export_opts.overwrite))
            elif format_name in {EXPORT_FORMAT_CSV, EXPORT_FORMAT_TSV}:
                writer = write_csv if format_name == EXPORT_FORMAT_CSV else write_tsv
                for table_name, table in context.tables.tables.items():
                    if table.empty and not export_opts.include_empty_tables:
                        continue
                    result.add_file(writer(table, output_dir=context.output_dir, basename=f"{context.name}_{sanitize_filename(table_name)}", table_name=table_name, options=export_opts.delimited, overwrite=export_opts.overwrite))
            elif format_name == EXPORT_FORMAT_EXCEL:
                result.add_file(write_excel(context.tables, output_dir=context.output_dir, basename=context.name, options=export_opts.excel, overwrite=export_opts.overwrite))
            elif format_name in {EXPORT_FORMAT_TEXT, "txt"}:
                result.add_file(write_text(payload, output_dir=context.output_dir, basename=context.name, overwrite=export_opts.overwrite))
            else:
                raise ExportFormatError(f"Unsupported multipose format: {format_name!r}")
        except Exception as exc:
            result.add_error(exc)
            if export_opts.error_mode == ErrorMode.RAISE.value:
                raise
    if context.options.include_individual_poses:
        pose_root = context.output_dir / "poses" if context.options.separate_pose_directories else context.output_dir
        for index, pose in enumerate(context.poses, start=1):
            pose_name = f"{context.options.pose_prefix}_{index:04d}"
            pose_dir = pose_root / pose_name if context.options.separate_pose_directories else pose_root
            child_options = export_opts.copy(
                output_dir=pose_dir,
                basename=pose_name,
                include_manifest=False,
            )
            child = export_single_pose(pose, options=child_options)
            result.files.extend(child.files)
            result.warnings.extend(child.warnings)
            result.errors.extend(child.errors)
    if export_opts.include_manifest:
        provenance = collect_provenance(parameters=export_opts.to_dict(), metadata=context.options.metadata, root=context.output_dir) if export_opts.include_provenance else None
        result.manifest = build_manifest(result.files, source_name=context.name, options=export_opts, provenance=provenance, metadata=result.metadata, include_hashes=bool(export_opts.hash_algorithm), hash_algorithm=export_opts.hash_algorithm or DEFAULT_HASH_ALGORITHM)
        try:
            result.add_file(write_manifest(result.manifest, output_dir=context.output_dir, basename=f"{context.name}_manifest", overwrite=export_opts.overwrite, json_options=export_opts.json))
        except Exception as exc:
            result.add_error(exc)
    if export_opts.update_model_files:
        for pose in context.poses:
            if not (
                hasattr(pose, "files")
                or (isinstance(pose, Mapping) and "files" in pose)
            ):
                continue
            try:
                update_dock_model_files_from_result(
                    pose,
                    result,
                    options=ModelFilesUpdateOptions(
                        preserve_existing=export_opts.preserve_previous_files,
                    ),
                )
            except Exception as exc:
                result.add_warning(
                    {
                        "type": exc.__class__.__name__,
                        "message": "Unable to update DockModel.files",
                    }
                )
    return result.finish()


__all__.extend([
    "MULTIPOSE_SCHEMA_VERSION",
    "MultiposeExportOptions",
    "MultiposeExportContext",
    "prepare_multipose_export",
    "export_multiple_poses",
])
# =============================================================================
# Section 22 — Batch export
# =============================================================================

BATCH_SCHEMA_VERSION: Final[str] = "1.0"


def _batch_item_name(source: Any, index: int, prefix: str) -> str:
    return _pose_export_name(source, fallback=f"{prefix}_{index:04d}")


def _batch_group(source: Any, key: Optional[str]) -> Optional[str]:
    if not key:
        return None
    value = _molecular_get(source, key, default=None)
    return None if value in (None, "") else sanitize_filename(value)


def prepare_batch_items(
    sources: Iterable[Any],
    *,
    options: Optional[BatchExportOptions] = None,
) -> List[BatchExportItem]:
    """Prepare normalized batch items."""
    if sources is None:
        raise ExportInputError("Batch source collection cannot be None.")
    opts = options or BatchExportOptions()
    items = [
        BatchExportItem(
            index=index,
            source=source,
            name=_batch_item_name(source, index, opts.pose_prefix),
            group=_batch_group(source, opts.group_key),
        )
        for index, source in enumerate(sources, start=1)
    ]
    if opts.sort_key:
        items.sort(key=lambda item: _molecular_get(item.source, opts.sort_key, default=None))
        for index, item in enumerate(items, start=1):
            item.index = index
    return items


def _write_batch_summary(
    result: BatchExportResult,
    *,
    options: BatchExportOptions,
) -> List[ExportedFile]:
    rows = []
    for item in result.items:
        rows.append({
            "index": item.index,
            "name": item.name,
            "group": item.group,
            "status": item.status,
            "file_count": item.result.file_count if item.result else 0,
            "error_count": len(item.result.errors) if item.result else 0,
            "output_dir": str(item.output_dir) if item.output_dir else None,
        })
    table = build_table(rows, name="batch_summary")
    files: List[ExportedFile] = []
    export_opts = options.export
    formats = _export_format_values(export_opts.formats)
    if EXPORT_FORMAT_JSON in formats:
        files.append(write_json({
            "schema_version": BATCH_SCHEMA_VERSION,
            "batch_name": result.batch_name,
            "items": [item.to_dict() for item in result.items],
        }, output_dir=result.output_dir, basename=f"{result.batch_name}_summary", options=export_opts.json, overwrite=export_opts.overwrite))
    if EXPORT_FORMAT_CSV in formats:
        files.append(write_csv(table, output_dir=result.output_dir, basename=f"{result.batch_name}_summary", options=export_opts.delimited, overwrite=export_opts.overwrite))
    if EXPORT_FORMAT_TSV in formats:
        files.append(write_tsv(table, output_dir=result.output_dir, basename=f"{result.batch_name}_summary", options=export_opts.delimited, overwrite=export_opts.overwrite))
    if EXPORT_FORMAT_EXCEL in formats:
        files.append(write_excel(table, output_dir=result.output_dir, basename=f"{result.batch_name}_summary", options=export_opts.excel, overwrite=export_opts.overwrite))
    return files


def export_batch(
    sources: Iterable[Any],
    *,
    options: Optional[BatchExportOptions] = None,
    output_dir: Optional[PathLike] = None,
) -> BatchExportResult:
    """Export many independent poses or payloads."""
    opts = options or BatchExportOptions()
    export_opts = opts.export
    if output_dir is not None:
        export_opts = export_opts.copy(output_dir=output_dir)
        opts = opts.copy(export=export_opts)
    directory = Path(export_opts.output_dir or Path.cwd())
    directory.mkdir(parents=True, exist_ok=True)
    batch_name = sanitize_filename(opts.batch_name or export_opts.basename or "batch")
    result = BatchExportResult(
        batch_name=batch_name,
        output_dir=directory,
        metadata={"schema_version": BATCH_SCHEMA_VERSION, **opts.metadata},
    )
    items = prepare_batch_items(sources, options=opts)
    failures = 0
    for item in items:
        group_dir = directory / item.group if item.group else directory
        item_dir = group_dir / item.name if opts.separate_pose_directories else group_dir
        item.output_dir = item_dir
        result.add_item(item)
        try:
            child_options = export_opts.copy(
                output_dir=item_dir,
                basename=item.name or f"{opts.pose_prefix}_{item.index:04d}",
                include_manifest=True,
            )
            child = export_single_pose(item.source, options=child_options, metadata=item.metadata)
            item.attach_result(child)
        except Exception as exc:
            failed = ExportResult(source_name=item.name, output_dir=item_dir)
            failed.add_error(exc)
            failed.finish(ExportStatus.FAILED.value)
            item.attach_result(failed)
        if not item.succeeded:
            failures += 1
            if not opts.continue_on_error:
                break
            if opts.max_failures is not None and failures >= opts.max_failures:
                break
    if opts.consolidate_tables:
        try:
            for exported in _write_batch_summary(result, options=opts):
                result.add_file(exported)
        except Exception as exc:
            result.add_error(exc)
    if opts.write_batch_manifest:
        try:
            all_files = list(result.files)
            for item in result.items:
                if item.result:
                    all_files.extend(item.result.files)
            provenance = collect_provenance(parameters=opts.to_dict(), metadata=opts.metadata, root=directory) if export_opts.include_provenance else None
            result.manifest = build_manifest(all_files, source_name=batch_name, options=export_opts, provenance=provenance, metadata=result.metadata, include_hashes=bool(export_opts.hash_algorithm), hash_algorithm=export_opts.hash_algorithm or DEFAULT_HASH_ALGORITHM)
            result.add_file(write_manifest(result.manifest, output_dir=directory, basename=f"{batch_name}_manifest", overwrite=export_opts.overwrite, json_options=export_opts.json))
        except Exception as exc:
            result.add_error(exc)
    return result.finish()


__all__.extend([
    "BATCH_SCHEMA_VERSION",
    "prepare_batch_items",
    "export_batch",
])
# =============================================================================
# Section 23 — DockModel.files update
# =============================================================================

DOCK_MODEL_FILES_SCHEMA_VERSION: Final[str] = "1.0"


class ModelFilesLayout(str, Enum):
    """Supported DockModel.files layouts."""

    AUTO = "auto"
    LIST = "list"
    MAPPING = "mapping"


@dataclass(slots=True)
class ModelFilesUpdateOptions:
    """Options for updating DockModel.files."""

    layout: str = ModelFilesLayout.AUTO.value
    preserve_existing: bool = True
    replace_matching_formats: bool = False
    remove_missing: bool = False
    relative_to: Optional[PathLike] = None
    store_metadata: bool = False
    deduplicate: bool = True
    sort_entries: bool = False

    def __post_init__(self) -> None:
        self.layout = str(_enum_value(self.layout)).strip().lower()
        if self.layout not in {item.value for item in ModelFilesLayout}:
            raise ExportConfigurationError(
                f"Unsupported DockModel.files layout: {self.layout!r}"
            )
        self.relative_to = _path_or_none(self.relative_to)


@dataclass(slots=True)
class ModelFilesUpdateReport:
    """Result of one DockModel.files update."""

    layout: str
    previous_count: int
    current_count: int
    added: List[str] = field(default_factory=list)
    removed: List[str] = field(default_factory=list)
    skipped: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def changed(self) -> bool:
        """Return whether the collection changed."""
        return bool(self.added or self.removed)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable update report."""
        return asdict(self)


def _model_files_value(model: Any) -> Any:
    """Read DockModel.files without assuming object type."""
    if isinstance(model, Mapping):
        return model.get("files")
    return getattr(model, "files", None)


def _set_model_files_value(model: Any, value: Any) -> None:
    """Write DockModel.files."""
    if isinstance(model, MutableMapping):
        model["files"] = value
        return
    setattr(model, "files", value)


def _detect_model_files_layout(value: Any) -> str:
    """Detect the current files layout."""
    if isinstance(value, Mapping):
        return ModelFilesLayout.MAPPING.value
    return ModelFilesLayout.LIST.value


def _stored_file_path(path: PathLike, relative_to: Optional[Path]) -> str:
    """Normalize one stored file path."""
    resolved = Path(path).expanduser()
    if relative_to is not None:
        try:
            return str(resolved.resolve().relative_to(relative_to.resolve()))
        except (OSError, ValueError):
            pass
    return str(resolved)


def _exported_file_key(item: ExportedFile) -> str:
    """Return a stable mapping key."""
    stem = sanitize_filename(item.table or item.path.stem)
    format_name = normalize_export_format(item.format)
    return f"{stem}_{format_name}" if stem else format_name


def _exported_file_entry(
    item: ExportedFile,
    *,
    options: ModelFilesUpdateOptions,
) -> Any:
    """Build one DockModel.files entry."""
    path = _stored_file_path(item.path, options.relative_to)
    if not options.store_metadata:
        return path
    record = item.to_dict()
    record["path"] = path
    return record


def _path_from_model_file_entry(value: Any) -> Optional[str]:
    """Extract a path from a stored entry."""
    if isinstance(value, (str, os.PathLike)):
        return str(value)
    if isinstance(value, Mapping) and value.get("path") is not None:
        return str(value["path"])
    return None


def _model_file_exists(value: Any, root: Optional[Path]) -> bool:
    """Return whether one stored path exists."""
    path_value = _path_from_model_file_entry(value)
    if path_value is None:
        return True
    path = Path(path_value)
    if not path.is_absolute() and root is not None:
        path = root / path
    return path.exists()


def _deduplicate_file_list(values: Iterable[Any]) -> List[Any]:
    """Deduplicate file entries while preserving order."""
    output: List[Any] = []
    seen: Set[str] = set()
    for value in values:
        path = _path_from_model_file_entry(value)
        marker = path or repr(make_json_safe(value))
        if marker in seen:
            continue
        seen.add(marker)
        output.append(value)
    return output


def update_dock_model_files(
    model: Any,
    files: Iterable[ExportedFile],
    *,
    options: Optional[ModelFilesUpdateOptions] = None,
) -> ModelFilesUpdateReport:
    """Update DockModel.files from generated files."""
    if model is None:
        raise ExportValidationError("DockModel cannot be None")
    opts = options or ModelFilesUpdateOptions()
    incoming = list(files)
    current = _model_files_value(model)
    layout = (
        _detect_model_files_layout(current)
        if opts.layout == ModelFilesLayout.AUTO.value
        else opts.layout
    )
    previous_count = len(current) if isinstance(current, (Mapping, Sequence)) else 0
    report = ModelFilesUpdateReport(
        layout=layout,
        previous_count=previous_count,
        current_count=previous_count,
    )

    if layout == ModelFilesLayout.MAPPING.value:
        output: Dict[str, Any] = dict(current) if opts.preserve_existing and isinstance(current, Mapping) else {}
        if opts.remove_missing:
            for key, value in list(output.items()):
                if not _model_file_exists(value, opts.relative_to):
                    report.removed.append(str(_path_from_model_file_entry(value) or key))
                    del output[key]
        if opts.replace_matching_formats:
            formats = {normalize_export_format(item.format) for item in incoming}
            for key, value in list(output.items()):
                path = _path_from_model_file_entry(value)
                suffix = Path(path).suffix.lower().lstrip(".") if path else ""
                if FORMAT_ALIASES.get(suffix, suffix) in formats:
                    report.removed.append(str(path or key))
                    del output[key]
        for item in incoming:
            base_key = _exported_file_key(item)
            key = base_key
            index = 2
            entry = _exported_file_entry(item, options=opts)
            path = _path_from_model_file_entry(entry) or str(item.path)
            while key in output and _path_from_model_file_entry(output[key]) != path:
                key = f"{base_key}_{index}"
                index += 1
            if key in output and _path_from_model_file_entry(output[key]) == path:
                report.skipped.append(path)
                continue
            output[key] = entry
            report.added.append(path)
        if opts.sort_entries:
            output = dict(sorted(output.items(), key=lambda pair: pair[0]))
        _set_model_files_value(model, output)
        report.current_count = len(output)
        return report

    output_list: List[Any] = list(current) if opts.preserve_existing and isinstance(current, Sequence) and not isinstance(current, (str, bytes, bytearray)) else []
    if opts.remove_missing:
        kept: List[Any] = []
        for value in output_list:
            if _model_file_exists(value, opts.relative_to):
                kept.append(value)
            else:
                report.removed.append(str(_path_from_model_file_entry(value) or value))
        output_list = kept
    if opts.replace_matching_formats:
        formats = {normalize_export_format(item.format) for item in incoming}
        kept = []
        for value in output_list:
            path = _path_from_model_file_entry(value)
            suffix = Path(path).suffix.lower().lstrip(".") if path else ""
            if FORMAT_ALIASES.get(suffix, suffix) in formats:
                report.removed.append(str(path or value))
            else:
                kept.append(value)
        output_list = kept
    existing_paths = {
        path for path in (_path_from_model_file_entry(value) for value in output_list) if path
    }
    for item in incoming:
        entry = _exported_file_entry(item, options=opts)
        path = _path_from_model_file_entry(entry) or str(item.path)
        if opts.deduplicate and path in existing_paths:
            report.skipped.append(path)
            continue
        output_list.append(entry)
        existing_paths.add(path)
        report.added.append(path)
    if opts.deduplicate:
        output_list = _deduplicate_file_list(output_list)
    if opts.sort_entries:
        output_list.sort(key=lambda value: _path_from_model_file_entry(value) or repr(value))
    _set_model_files_value(model, output_list)
    report.current_count = len(output_list)
    return report


def update_dock_model_files_from_result(
    model: Any,
    result: ExportResult,
    *,
    options: Optional[ModelFilesUpdateOptions] = None,
) -> ModelFilesUpdateReport:
    """Update DockModel.files from one export result."""
    return update_dock_model_files(model, result.files, options=options)


def clear_dock_model_files(model: Any) -> ModelFilesUpdateReport:
    """Clear DockModel.files."""
    current = _model_files_value(model)
    layout = _detect_model_files_layout(current)
    count = len(current) if isinstance(current, (Mapping, Sequence)) else 0
    _set_model_files_value(model, {} if layout == ModelFilesLayout.MAPPING.value else [])
    return ModelFilesUpdateReport(
        layout=layout,
        previous_count=count,
        current_count=0,
        removed=[str(_path_from_model_file_entry(item) or item) for item in (current.values() if isinstance(current, Mapping) else current or [])],
    )


__all__.extend([
    "DOCK_MODEL_FILES_SCHEMA_VERSION",
    "ModelFilesLayout",
    "ModelFilesUpdateOptions",
    "ModelFilesUpdateReport",
    "update_dock_model_files",
    "update_dock_model_files_from_result",
    "clear_dock_model_files",
])
# =============================================================================
# Section 24 — Validation
# =============================================================================

VALIDATION_SCHEMA_VERSION: Final[str] = "1.0"


class ValidationSeverity(str, Enum):
    """Validation issue severity."""

    INFO = "info"
    WARNING = "warning"
    ERROR = "error"


@dataclass(slots=True)
class ValidationIssue:
    """One validation issue."""

    code: str
    message: str
    severity: str = ValidationSeverity.ERROR.value
    field_name: Optional[str] = None
    path: Optional[str] = None
    value: Any = None
    context: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.code = str(self.code).strip() or "validation_issue"
        self.message = str(self.message)
        self.severity = str(_enum_value(self.severity)).strip().lower()
        if self.severity not in {item.value for item in ValidationSeverity}:
            raise ExportConfigurationError(
                f"Unsupported validation severity: {self.severity!r}"
            )
        self.context = _copy_mapping(self.context)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable issue."""
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
            "field_name": self.field_name,
            "path": self.path,
            "value": make_json_safe(self.value),
            "context": dict(self.context),
        }


@dataclass(slots=True)
class ValidationReport:
    """Collection of validation issues."""

    subject: Optional[str] = None
    issues: List[ValidationIssue] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def valid(self) -> bool:
        """Return whether no errors were found."""
        return not any(item.severity == ValidationSeverity.ERROR.value for item in self.issues)

    @property
    def error_count(self) -> int:
        """Return the number of errors."""
        return sum(item.severity == ValidationSeverity.ERROR.value for item in self.issues)

    @property
    def warning_count(self) -> int:
        """Return the number of warnings."""
        return sum(item.severity == ValidationSeverity.WARNING.value for item in self.issues)

    def add(
        self,
        code: str,
        message: str,
        *,
        severity: Union[str, ValidationSeverity] = ValidationSeverity.ERROR,
        field_name: Optional[str] = None,
        path: Optional[PathLike] = None,
        value: Any = None,
        context: Optional[Mapping[str, Any]] = None,
    ) -> ValidationIssue:
        """Add one issue."""
        issue = ValidationIssue(
            code=code,
            message=message,
            severity=str(_enum_value(severity)),
            field_name=field_name,
            path=str(path) if path is not None else None,
            value=value,
            context=dict(context or {}),
        )
        self.issues.append(issue)
        return issue

    def merge(self, other: "ValidationReport") -> "ValidationReport":
        """Merge another report."""
        self.issues.extend(other.issues)
        self.metadata.update(other.metadata)
        return self

    def raise_for_errors(self) -> None:
        """Raise when errors are present."""
        if not self.valid:
            messages = "; ".join(item.message for item in self.issues if item.severity == ValidationSeverity.ERROR.value)
            raise ExportValidationError(messages or "Export validation failed")

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable report."""
        return {
            "schema_version": VALIDATION_SCHEMA_VERSION,
            "subject": self.subject,
            "valid": self.valid,
            "error_count": self.error_count,
            "warning_count": self.warning_count,
            "issues": [item.to_dict() for item in self.issues],
            "metadata": dict(self.metadata),
        }


@dataclass(slots=True)
class ExportValidationOptions:
    """Options for export validation."""

    require_source: bool = True
    require_output_dir: bool = True
    check_output_writable: bool = True
    check_existing_files: bool = True
    check_hashes: bool = False
    check_serializable: bool = True
    require_files_for_success: bool = True
    allow_empty_formats: bool = False
    max_issues: Optional[int] = None

    def __post_init__(self) -> None:
        if self.max_issues is not None and self.max_issues <= 0:
            raise ExportConfigurationError("Maximum validation issues must be positive")


def _validation_full(report: ValidationReport, options: ExportValidationOptions) -> bool:
    """Return whether issue collection reached its limit."""
    return options.max_issues is not None and len(report.issues) >= options.max_issues


def validate_export_options(
    options: ExportOptions,
    *,
    validation: Optional[ExportValidationOptions] = None,
) -> ValidationReport:
    """Validate general export options."""
    opts = validation or ExportValidationOptions()
    report = ValidationReport(subject="ExportOptions")
    if not options.formats and not opts.allow_empty_formats:
        report.add("formats.empty", "At least one export format is required", field_name="formats")
    for format_name in options.formats:
        try:
            normalize_export_format(format_name)
        except Exception as exc:
            report.add("format.invalid", str(exc), field_name="formats", value=format_name)
        if _validation_full(report, opts):
            return report
    if opts.require_output_dir and options.output_dir is None:
        report.add("output_dir.missing", "Output directory is required", field_name="output_dir")
    if options.output_dir is not None and opts.check_output_writable:
        path = Path(options.output_dir)
        try:
            path.mkdir(parents=True, exist_ok=True)
            if not os.access(path, os.W_OK):
                report.add("output_dir.not_writable", "Output directory is not writable", path=path)
        except OSError as exc:
            report.add("output_dir.unavailable", str(exc), path=path)
    return report


def validate_exported_file(
    exported_file: ExportedFile,
    *,
    validation: Optional[ExportValidationOptions] = None,
) -> ValidationReport:
    """Validate one exported file record."""
    opts = validation or ExportValidationOptions()
    report = ValidationReport(subject=str(exported_file.path))
    path = Path(exported_file.path)
    if opts.check_existing_files and not path.exists():
        report.add("file.missing", "Exported file does not exist", path=path)
        return report
    if path.exists() and not path.is_file():
        report.add("file.not_regular", "Exported path is not a regular file", path=path)
    if path.exists() and exported_file.size_bytes is not None:
        actual_size = path.stat().st_size
        if actual_size != exported_file.size_bytes:
            report.add(
                "file.size_mismatch",
                "Exported file size does not match metadata",
                path=path,
                context={"expected": exported_file.size_bytes, "actual": actual_size},
            )
    if opts.check_hashes and exported_file.hash_value:
        algorithm = exported_file.hash_algorithm or DEFAULT_HASH_ALGORITHM
        actual_hash = compute_file_hash(path, algorithm=algorithm)
        if actual_hash != exported_file.hash_value:
            report.add(
                "file.hash_mismatch",
                "Exported file hash does not match metadata",
                path=path,
                context={"expected": exported_file.hash_value, "actual": actual_hash},
            )
    return report


def validate_manifest(
    manifest: ExportManifest,
    *,
    validation: Optional[ExportValidationOptions] = None,
) -> ValidationReport:
    """Validate one export manifest."""
    opts = validation or ExportValidationOptions()
    report = ValidationReport(subject="ExportManifest")
    if not manifest.schema_name:
        report.add("manifest.schema_name", "Manifest schema name is missing")
    if not manifest.schema_version:
        report.add("manifest.schema_version", "Manifest schema version is missing")
    for item in manifest.files:
        report.merge(validate_exported_file(item, validation=opts))
        if _validation_full(report, opts):
            break
    return report


def validate_export_result(
    result: ExportResult,
    *,
    validation: Optional[ExportValidationOptions] = None,
) -> ValidationReport:
    """Validate one export result."""
    opts = validation or ExportValidationOptions()
    report = ValidationReport(subject=result.source_name or "ExportResult")
    if opts.require_files_for_success and result.status in SUCCESS_EXPORT_STATUSES and not result.files:
        report.add("result.files_empty", "Successful export result has no files")
    for item in result.files:
        report.merge(validate_exported_file(item, validation=opts))
        if _validation_full(report, opts):
            return report
    if result.manifest is not None:
        report.merge(validate_manifest(result.manifest, validation=opts))
    if opts.check_serializable:
        problems = find_non_serializable(result.to_dict(include_payload=True))
        for problem in problems:
            report.add(
                "result.not_serializable",
                "Export result contains a non-serializable value",
                path=problem.get("path"),
                value=problem.get("type"),
                context={"reason": problem.get("reason")},
            )
            if _validation_full(report, opts):
                break
    return report


def validate_dock_model_for_export(
    model: Any,
    *,
    validation: Optional[ExportValidationOptions] = None,
) -> ValidationReport:
    """Validate a DockModel-like source."""
    opts = validation or ExportValidationOptions()
    report = ValidationReport(subject="DockModel")
    if model is None:
        report.add("dock_model.none", "DockModel cannot be None")
        return report
    if opts.require_source and not is_dock_model_like(model):
        report.add(
            "dock_model.unrecognized",
            "Object does not match the expected DockModel interface",
            value=type(model).__name__,
        )
    try:
        identifier = dock_model_identifier(model)
        if not identifier:
            report.add("dock_model.identifier", "DockModel identifier is empty")
    except Exception as exc:
        report.add("dock_model.identifier_error", str(exc))
    if opts.check_serializable:
        try:
            record = dock_model_to_record(model)
            for problem in find_non_serializable(record):
                report.add(
                    "dock_model.not_serializable",
                    "DockModel record contains a non-serializable value",
                    path=problem.get("path"),
                    value=problem.get("type"),
                    context={"reason": problem.get("reason")},
                )
        except Exception as exc:
            report.add("dock_model.serialization_error", str(exc))
    return report


def validate_export_request(
    source: Any,
    options: ExportOptions,
    *,
    validation: Optional[ExportValidationOptions] = None,
) -> ValidationReport:
    """Run preflight validation for one export request."""
    opts = validation or ExportValidationOptions()
    report = ValidationReport(subject="export_request")
    report.merge(validate_export_options(options, validation=opts))
    if source is None and opts.require_source:
        report.add("source.missing", "Export source is required")
    elif is_dock_model_like(source):
        report.merge(validate_dock_model_for_export(source, validation=opts))
    elif opts.check_serializable:
        try:
            problems = find_non_serializable(to_serializable(source))
            for problem in problems:
                report.add(
                    "source.not_serializable",
                    "Source contains a non-serializable value",
                    path=problem.get("path"),
                    value=problem.get("type"),
                    context={"reason": problem.get("reason")},
                )
        except Exception as exc:
            report.add("source.serialization_error", str(exc))
    return report


__all__.extend([
    "VALIDATION_SCHEMA_VERSION",
    "ValidationSeverity",
    "ValidationIssue",
    "ValidationReport",
    "ExportValidationOptions",
    "validate_export_options",
    "validate_exported_file",
    "validate_manifest",
    "validate_export_result",
    "validate_dock_model_for_export",
    "validate_export_request",
])
# =============================================================================
# Section 25 — Errors and permissive mode
# =============================================================================

PERMISSIVE_SCHEMA_VERSION: Final[str] = "1.0"


@dataclass(slots=True)
class PermissiveExportOptions:
    """Policies for recoverable export failures."""

    enabled: bool = True
    validate_before_export: bool = True
    continue_after_validation_errors: bool = True
    continue_after_format_errors: bool = True
    continue_after_manifest_error: bool = True
    update_model_files_on_partial: bool = True
    create_failed_result: bool = True
    warning_on_recovery: bool = True
    fallback_formats: Tuple[str, ...] = (EXPORT_FORMAT_JSON,)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.fallback_formats = _string_tuple(self.fallback_formats)
        self.metadata = _copy_mapping(self.metadata)


@dataclass(slots=True)
class ExportErrorContext:
    """Normalized information about one export failure."""

    operation: str
    exception_type: str
    message: str
    recoverable: bool = True
    format: Optional[str] = None
    source_name: Optional[str] = None
    path: Optional[str] = None
    traceback_text: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    @classmethod
    def from_exception(
        cls,
        exception: BaseException,
        *,
        operation: str,
        recoverable: bool = True,
        format_name: Optional[str] = None,
        source_name: Optional[str] = None,
        path: Optional[PathLike] = None,
        include_traceback: bool = False,
        metadata: Optional[Mapping[str, Any]] = None,
    ) -> "ExportErrorContext":
        """Build context from an exception."""
        traceback_text = None
        if include_traceback:
            import traceback
            traceback_text = "".join(traceback.format_exception(type(exception), exception, exception.__traceback__))
        return cls(
            operation=operation,
            exception_type=type(exception).__name__,
            message=str(exception),
            recoverable=recoverable,
            format=format_name,
            source_name=source_name,
            path=str(path) if path is not None else None,
            traceback_text=traceback_text,
            metadata=dict(metadata or {}),
        )

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable error context."""
        return asdict(self)


@dataclass(slots=True)
class PermissiveExportReport:
    """Report for a permissive export operation."""

    result: ExportResult
    validation: Optional[ValidationReport] = None
    recovered_errors: List[ExportErrorContext] = field(default_factory=list)
    attempted_fallbacks: List[str] = field(default_factory=list)

    @property
    def recovered(self) -> bool:
        """Return whether at least one failure was recovered."""
        return bool(self.recovered_errors and self.result.files)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable report."""
        return {
            "schema_version": PERMISSIVE_SCHEMA_VERSION,
            "result": self.result.to_dict(include_payload=False),
            "validation": self.validation.to_dict() if self.validation else None,
            "recovered_errors": [item.to_dict() for item in self.recovered_errors],
            "attempted_fallbacks": list(self.attempted_fallbacks),
            "recovered": self.recovered,
        }


def is_recoverable_export_error(exception: BaseException) -> bool:
    """Return whether an exception can be handled permissively."""
    return isinstance(
        exception,
        (
            ExportError,
            OSError,
            ValueError,
            TypeError,
            UnicodeError,
        ),
    ) and not isinstance(exception, (KeyboardInterrupt, SystemExit, MemoryError))


def options_for_error_mode(
    options: ExportOptions,
    *,
    permissive: bool,
) -> ExportOptions:
    """Return options configured for strict or permissive behavior."""
    return options.copy(
        error_mode=(ErrorMode.COLLECT.value if permissive else ErrorMode.RAISE.value)
    )


def _validation_to_result(
    report: ValidationReport,
    result: ExportResult,
    *,
    as_errors: bool,
) -> None:
    """Attach validation issues to an export result."""
    for issue in report.issues:
        record = issue.to_dict()
        if as_errors and issue.severity == ValidationSeverity.ERROR.value:
            result.add_error(record)
        else:
            result.add_warning(record)


def export_single_pose_permissive(
    pose: Any,
    *,
    options: Optional[ExportOptions] = None,
    permissive: Optional[PermissiveExportOptions] = None,
    validation: Optional[ExportValidationOptions] = None,
    output_dir: Optional[PathLike] = None,
    basename: Optional[str] = None,
    metadata: Optional[Mapping[str, Any]] = None,
) -> PermissiveExportReport:
    """Export one pose with validation and recovery."""
    policy = permissive or PermissiveExportOptions()
    base_options = options or ExportOptions()
    if output_dir is not None:
        base_options = base_options.copy(output_dir=output_dir)
    if basename is not None:
        base_options = base_options.copy(basename=basename)
    effective = options_for_error_mode(base_options, permissive=policy.enabled)
    validation_report = None
    recovered: List[ExportErrorContext] = []

    if policy.validate_before_export:
        validation_report = validate_export_request(pose, effective, validation=validation)
        if not validation_report.valid and not policy.continue_after_validation_errors:
            validation_report.raise_for_errors()

    try:
        result = export_single_pose(
            pose,
            options=effective,
            metadata={**dict(metadata or {}), **policy.metadata},
        )
    except Exception as exc:
        if not policy.enabled or not is_recoverable_export_error(exc):
            raise
        context = ExportErrorContext.from_exception(
            exc,
            operation="export_single_pose",
            recoverable=True,
            source_name=basename or effective.basename,
            path=effective.output_dir,
        )
        recovered.append(context)
        if not policy.create_failed_result:
            raise
        result = ExportResult(
            status=ExportStatus.FAILED.value,
            source_name=basename or effective.basename,
            output_dir=effective.output_dir,
            metadata={"permissive": True, **policy.metadata},
        )
        result.add_error(context.to_dict())
        result.finish()

    if validation_report is not None:
        _validation_to_result(
            validation_report,
            result,
            as_errors=not policy.continue_after_validation_errors,
        )

    attempted: List[str] = []
    if policy.enabled and not result.files and policy.fallback_formats:
        for format_name in policy.fallback_formats:
            attempted.append(format_name)
            try:
                fallback_options = effective.copy(
                    formats=(format_name,),
                    include_manifest=False,
                    update_model_files=False,
                    error_mode=ErrorMode.COLLECT.value,
                )
                fallback = export_single_pose(
                    pose,
                    options=fallback_options,
                    metadata={**dict(metadata or {}), "fallback": True},
                )
                for item in fallback.files:
                    result.add_file(item)
                result.payload = result.payload if result.payload is not None else fallback.payload
                result.tables.update(fallback.tables)
                result.warnings.extend(fallback.warnings)
                result.errors.extend(fallback.errors)
                if fallback.files:
                    if policy.warning_on_recovery:
                        result.add_warning(f"Recovered export using {format_name}")
                    break
            except Exception as exc:
                context = ExportErrorContext.from_exception(
                    exc,
                    operation="fallback_export",
                    recoverable=is_recoverable_export_error(exc),
                    format_name=format_name,
                    source_name=result.source_name,
                    path=result.output_dir,
                )
                recovered.append(context)
                result.add_error(context.to_dict())
                if not policy.continue_after_format_errors:
                    raise

    if (
        effective.update_model_files
        and hasattr(pose, "files")
        and (result.succeeded or (policy.update_model_files_on_partial and result.files))
    ):
        try:
            update_dock_model_files_from_result(
                pose,
                result,
                options=ModelFilesUpdateOptions(
                    preserve_existing=effective.preserve_previous_files,
                ),
            )
        except Exception as exc:
            context = ExportErrorContext.from_exception(
                exc,
                operation="update_dock_model_files",
                recoverable=True,
                source_name=result.source_name,
            )
            recovered.append(context)
            result.add_warning(context.to_dict())

    result.finish()
    return PermissiveExportReport(
        result=result,
        validation=validation_report,
        recovered_errors=recovered,
        attempted_fallbacks=attempted,
    )


def export_multipose_permissive(
    poses: Iterable[Any],
    *,
    options: Optional[MultiposeExportOptions] = None,
    permissive: Optional[PermissiveExportOptions] = None,
    output_dir: Optional[PathLike] = None,
) -> PermissiveExportReport:
    """Export multiple poses with normalized error handling."""
    policy = permissive or PermissiveExportOptions()
    multi_options = options or MultiposeExportOptions()
    export_options = options_for_error_mode(multi_options.export, permissive=policy.enabled)
    multi_options = replace(multi_options, export=export_options)
    recovered: List[ExportErrorContext] = []
    try:
        result = export_multiple_poses(poses, options=multi_options, output_dir=output_dir)
    except Exception as exc:
        if not policy.enabled or not is_recoverable_export_error(exc):
            raise
        context = ExportErrorContext.from_exception(exc, operation="export_multiple_poses")
        recovered.append(context)
        result = ExportResult(status=ExportStatus.FAILED.value, output_dir=output_dir)
        result.add_error(context.to_dict())
        result.finish()
    return PermissiveExportReport(result=result, recovered_errors=recovered)


def export_batch_permissive(
    sources: Iterable[Any],
    *,
    options: Optional[BatchExportOptions] = None,
    permissive: Optional[PermissiveExportOptions] = None,
    output_dir: Optional[PathLike] = None,
) -> BatchExportResult:
    """Export a batch using strict or permissive policies."""
    policy = permissive or PermissiveExportOptions()
    batch_options = options or BatchExportOptions()
    export_options = options_for_error_mode(batch_options.export, permissive=policy.enabled)
    batch_options = batch_options.copy(
        export=export_options,
        continue_on_error=(True if policy.enabled else batch_options.continue_on_error),
    )
    try:
        return export_batch(sources, options=batch_options, output_dir=output_dir)
    except Exception as exc:
        if not policy.enabled or not is_recoverable_export_error(exc):
            raise
        result = BatchExportResult(
            status=ExportStatus.FAILED.value,
            batch_name=batch_options.batch_name,
            output_dir=output_dir or export_options.output_dir,
            metadata={"permissive": True, **policy.metadata},
        )
        result.add_error(
            ExportErrorContext.from_exception(
                exc,
                operation="export_batch",
                recoverable=True,
            ).to_dict()
        )
        return result.finish()


__all__.extend([
    "PERMISSIVE_SCHEMA_VERSION",
    "PermissiveExportOptions",
    "ExportErrorContext",
    "PermissiveExportReport",
    "is_recoverable_export_error",
    "options_for_error_mode",
    "export_single_pose_permissive",
    "export_multipose_permissive",
    "export_batch_permissive",
])
# =============================================================================
# 26. Convenience public interface
# =============================================================================

CONVENIENCE_SCHEMA_VERSION: Final[str] = "1.0"


def export_pose(
    pose: Any,
    output_dir: Optional[PathLike] = None,
    *,
    formats: Optional[Iterable[Union[str, ExportFormat]]] = None,
    basename: Optional[str] = None,
    permissive: bool = False,
    update_files: bool = True,
    metadata: Optional[Mapping[str, Any]] = None,
    **option_overrides: Any,
) -> Union[ExportResult, PermissiveExportReport]:
    """Export one pose with minimal configuration."""
    options = ExportOptions(
        output_dir=output_dir,
        basename=basename,
        formats=tuple(formats or (EXPORT_FORMAT_JSON,)),
        update_model_files=update_files,
        metadata=dict(metadata or {}),
        **option_overrides,
    )
    if permissive:
        return export_single_pose_permissive(
            pose,
            options=options,
            output_dir=output_dir,
            basename=basename,
            metadata=metadata,
        )
    return export_single_pose(
        pose,
        options=options,
        output_dir=output_dir,
        basename=basename,
        metadata=metadata,
    )


def export_poses(
    poses: Iterable[Any],
    output_dir: Optional[PathLike] = None,
    *,
    formats: Optional[Iterable[Union[str, ExportFormat]]] = None,
    basename: Optional[str] = None,
    individual: bool = False,
    permissive: bool = False,
    metadata: Optional[Mapping[str, Any]] = None,
    **option_overrides: Any,
) -> Union[ExportResult, PermissiveExportReport]:
    """Export a pose collection."""
    export_options = ExportOptions(
        output_dir=output_dir,
        basename=basename,
        formats=tuple(formats or (EXPORT_FORMAT_JSON, EXPORT_FORMAT_CSV)),
        metadata=dict(metadata or {}),
        **option_overrides,
    )
    options = MultiposeExportOptions(
        export=export_options,
        export_individual=individual,
        metadata=dict(metadata or {}),
    )
    if permissive:
        return export_multipose_permissive(
            poses,
            options=options,
            output_dir=output_dir,
        )
    return export_multiple_poses(
        poses,
        options=options,
        output_dir=output_dir,
        basename=basename,
    )


def export_many(
    sources: Iterable[Any],
    output_dir: Optional[PathLike] = None,
    *,
    formats: Optional[Iterable[Union[str, ExportFormat]]] = None,
    batch_name: str = "batch",
    permissive: bool = False,
    metadata: Optional[Mapping[str, Any]] = None,
    **option_overrides: Any,
) -> BatchExportResult:
    """Export independent items as a batch."""
    export_options = ExportOptions(
        output_dir=output_dir,
        formats=tuple(formats or (EXPORT_FORMAT_JSON,)),
        metadata=dict(metadata or {}),
        **option_overrides,
    )
    options = BatchExportOptions(
        export=export_options,
        batch_name=batch_name,
        metadata=dict(metadata or {}),
    )
    if permissive:
        return export_batch_permissive(
            sources,
            options=options,
            output_dir=output_dir,
        )
    return export_batch(sources, options=options, output_dir=output_dir)


def export_data(
    value: Any,
    path: PathLike,
    *,
    format: Optional[Union[str, ExportFormat]] = None,
    overwrite: Union[str, OverwriteMode] = OverwriteMode.OVERWRITE.value,
) -> ExportedFile:
    """Export arbitrary data based on a path or explicit format."""
    target = Path(path)
    format_name = normalize_export_format(format or target.suffix or EXPORT_FORMAT_JSON)
    output_dir = target.parent
    basename = target.stem
    if format_name == EXPORT_FORMAT_JSON:
        return write_json(value, output_dir=output_dir, basename=basename, overwrite=overwrite)
    if format_name == EXPORT_FORMAT_JSONL:
        rows = value if isinstance(value, Iterable) and not isinstance(value, (str, bytes, Mapping)) else [value]
        return write_json_lines(rows, output_dir=output_dir, basename=basename, overwrite=overwrite)
    if format_name == EXPORT_FORMAT_CSV:
        return write_csv(value, output_dir=output_dir, basename=basename, overwrite=overwrite)
    if format_name == EXPORT_FORMAT_TSV:
        return write_tsv(value, output_dir=output_dir, basename=basename, overwrite=overwrite)
    if format_name == EXPORT_FORMAT_EXCEL:
        return write_excel(value, output_dir=output_dir, basename=basename, overwrite=overwrite)
    if format_name == EXPORT_FORMAT_TEXT:
        return write_text(value, output_dir=output_dir, basename=basename, overwrite=overwrite)
    raise ExportFormatError(f"Unsupported format: {format_name!r}")


def load_export(path: PathLike, *, format: Optional[Union[str, ExportFormat]] = None) -> Any:
    """Load a supported export file."""
    source = Path(path)
    format_name = normalize_export_format(format or source.suffix)
    if format_name == EXPORT_FORMAT_JSON:
        return read_json(source)
    if format_name == EXPORT_FORMAT_JSONL:
        return read_json_lines(source)
    if format_name in {EXPORT_FORMAT_CSV, EXPORT_FORMAT_TSV}:
        return read_delimited(source, delimiter="," if format_name == EXPORT_FORMAT_CSV else "\t")
    if format_name == EXPORT_FORMAT_EXCEL:
        return read_excel(source)
    if format_name == EXPORT_FORMAT_TEXT:
        return read_text(source)
    raise ExportFormatError(f"Unsupported format: {format_name!r}")


def available_export_formats() -> Tuple[str, ...]:
    """Return public format names."""
    return tuple(sorted(SUPPORTED_EXPORT_FORMATS))


__all__.extend([
    "CONVENIENCE_SCHEMA_VERSION",
    "export_pose",
    "export_poses",
    "export_many",
    "export_data",
    "load_export",
    "available_export_formats",
])
# =============================================================================
# 27. ChimeraX compatibility
# =============================================================================

CHIMERAX_SCHEMA_VERSION: Final[str] = "1.0"


@dataclass(frozen=True)
class ChimeraXExportOptions:
    """Optional ChimeraX integration settings."""

    log_results: bool = True
    select_exported: bool = False
    open_exported: bool = False
    use_run_command: bool = True
    metadata: Mapping[str, Any] = field(default_factory=dict)


def is_chimerax_session(value: Any) -> bool:
    """Return whether an object resembles a ChimeraX session."""
    return value is not None and hasattr(value, "logger") and hasattr(value, "models")


def chimerax_available() -> bool:
    """Return whether ChimeraX APIs can be imported."""
    try:
        import chimerax  # type: ignore  # noqa: F401
    except ImportError:
        return False
    return True


def _chimerax_log(session: Any, level: str, message: str) -> None:
    logger = getattr(session, "logger", None)
    method = getattr(logger, level, None)
    if callable(method):
        method(message)


def chimerax_atom_spec(value: Any) -> Optional[str]:
    """Build a ChimeraX atom specification when possible."""
    for name in ("atomspec", "atom_spec", "spec"):
        candidate = getattr(value, name, None)
        if candidate:
            return str(candidate)
    residue = getattr(value, "residue", None)
    atom_name = getattr(value, "name", None)
    chain = getattr(residue, "chain_id", None) or getattr(residue, "chain", None)
    number = getattr(residue, "number", None) or getattr(residue, "resid", None)
    if number is not None:
        spec = f":{number}"
        if chain:
            spec = f"/{chain}{spec}"
        if atom_name:
            spec += f"@{atom_name}"
        return spec
    return None


def chimerax_model_spec(value: Any) -> Optional[str]:
    """Return a ChimeraX model specification."""
    for name in ("atomspec", "id_string"):
        candidate = getattr(value, name, None)
        if candidate:
            text = str(candidate)
            return text if text.startswith("#") else f"#{text}"
    model_id = getattr(value, "id", None)
    if isinstance(model_id, Sequence) and not isinstance(model_id, (str, bytes)):
        return "#" + ".".join(str(part) for part in model_id)
    if model_id is not None:
        return f"#{model_id}"
    return None


def run_chimerax_command(session: Any, command: str) -> Any:
    """Run a command without making ChimeraX mandatory."""
    if not is_chimerax_session(session):
        raise ExportDependencyError(
            "chimerax",
            feature="command execution",
            message="A valid ChimeraX session is required.",
        )
    try:
        from chimerax.core.commands import run  # type: ignore
    except ImportError as exc:
        raise ExportDependencyError(
            "chimerax.core.commands",
            feature="command execution",
            message="ChimeraX command API is unavailable.",
            cause=exc,
        ) from exc
    return run(session, command)


def chimerax_export_result_commands(result: ExportResult) -> List[str]:
    """Build commands for exported files."""
    commands: List[str] = []
    for exported in result.files:
        path = Path(exported.path)
        if path.suffix.lower() in {".pdb", ".cif", ".mol2", ".sdf"}:
            commands.append(f'open "{path}"')
    return commands


def notify_chimerax_export(
    session: Any,
    result: Union[ExportResult, BatchExportResult],
    *,
    options: Optional[ChimeraXExportOptions] = None,
) -> Union[ExportResult, BatchExportResult]:
    """Report an export to a ChimeraX session."""
    opts = options or ChimeraXExportOptions()
    if not is_chimerax_session(session):
        raise ExportDependencyError(
            "chimerax",
            feature="export notification",
            message="A valid ChimeraX session is required.",
        )
    if opts.log_results:
        count = len(getattr(result, "files", []))
        status = getattr(result, "status", "unknown")
        _chimerax_log(session, "info", f"DockAnalyzer export: {status}; {count} file(s)")
        for warning in getattr(result, "warnings", []):
            _chimerax_log(session, "warning", str(warning))
        for error in getattr(result, "errors", []):
            _chimerax_log(session, "error", str(error))
    if opts.open_exported and isinstance(result, ExportResult):
        for command in chimerax_export_result_commands(result):
            run_chimerax_command(session, command)
    return result


def export_pose_chimerax(
    session: Any,
    pose: Any,
    *,
    options: Optional[ExportOptions] = None,
    chimerax_options: Optional[ChimeraXExportOptions] = None,
    output_dir: Optional[PathLike] = None,
    basename: Optional[str] = None,
) -> ExportResult:
    """Export one pose and report through ChimeraX."""
    result = export_single_pose(
        pose,
        options=options,
        output_dir=output_dir,
        basename=basename,
        metadata={"host": "ChimeraX"},
    )
    return notify_chimerax_export(session, result, options=chimerax_options)  # type: ignore[return-value]


def export_poses_chimerax(
    session: Any,
    poses: Iterable[Any],
    *,
    options: Optional[MultiposeExportOptions] = None,
    chimerax_options: Optional[ChimeraXExportOptions] = None,
    output_dir: Optional[PathLike] = None,
) -> ExportResult:
    """Export multiple poses and report through ChimeraX."""
    result = export_multiple_poses(poses, options=options, output_dir=output_dir)
    return notify_chimerax_export(session, result, options=chimerax_options)  # type: ignore[return-value]


__all__.extend([
    "CHIMERAX_SCHEMA_VERSION",
    "ChimeraXExportOptions",
    "is_chimerax_session",
    "chimerax_available",
    "chimerax_atom_spec",
    "chimerax_model_spec",
    "run_chimerax_command",
    "chimerax_export_result_commands",
    "notify_chimerax_export",
    "export_pose_chimerax",
    "export_poses_chimerax",
])
# =============================================================================
# 28. Versioning and migration
# =============================================================================

CURRENT_EXPORT_SCHEMA_VERSION: Final[str] = "1.0"


@dataclass(frozen=True, order=True)
class SchemaVersion:
    """Comparable semantic schema version."""

    major: int
    minor: int = 0
    patch: int = 0

    @classmethod
    def parse(cls, value: Union[str, int, Sequence[int], "SchemaVersion"]) -> "SchemaVersion":
        """Parse a version value."""
        if isinstance(value, cls):
            return value
        if isinstance(value, int):
            return cls(value)
        if isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
            parts = [int(item) for item in value]
        else:
            text = str(value).strip().lstrip("vV")
            parts = [int(part) for part in text.split(".") if part != ""]
        if not parts or len(parts) > 3:
            raise ExportValidationError(f"Invalid schema version: {value!r}")
        return cls(*(parts + [0] * (3 - len(parts))))

    def __str__(self) -> str:
        return f"{self.major}.{self.minor}.{self.patch}"


@dataclass(frozen=True)
class MigrationStep:
    """One schema migration."""

    source: SchemaVersion
    target: SchemaVersion
    function: Callable[[Mapping[str, Any]], Mapping[str, Any]]
    name: str = ""
    description: str = ""


@dataclass
class MigrationReport:
    """Migration execution report."""

    source_version: str
    target_version: str
    applied: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    payload: Optional[Dict[str, Any]] = None

    @property
    def changed(self) -> bool:
        return bool(self.applied)


class MigrationRegistry:
    """Registry for ordered schema migrations."""

    def __init__(self) -> None:
        self._steps: Dict[Tuple[SchemaVersion, SchemaVersion], MigrationStep] = {}

    def register(
        self,
        source: Union[str, SchemaVersion],
        target: Union[str, SchemaVersion],
        function: Callable[[Mapping[str, Any]], Mapping[str, Any]],
        *,
        name: str = "",
        description: str = "",
    ) -> MigrationStep:
        step = MigrationStep(
            SchemaVersion.parse(source),
            SchemaVersion.parse(target),
            function,
            name or getattr(function, "__name__", "migration"),
            description,
        )
        if step.target <= step.source:
            raise ExportValidationError("Migration target must be newer than source")
        self._steps[(step.source, step.target)] = step
        return step

    def path(
        self,
        source: Union[str, SchemaVersion],
        target: Union[str, SchemaVersion],
    ) -> List[MigrationStep]:
        start = SchemaVersion.parse(source)
        end = SchemaVersion.parse(target)
        if start == end:
            return []
        frontier: List[Tuple[SchemaVersion, List[MigrationStep]]] = [(start, [])]
        visited = {start}
        while frontier:
            current, steps = frontier.pop(0)
            for (step_source, step_target), step in sorted(self._steps.items()):
                if step_source != current or step_target in visited or step_target > end:
                    continue
                next_steps = steps + [step]
                if step_target == end:
                    return next_steps
                visited.add(step_target)
                frontier.append((step_target, next_steps))
        raise ExportDependencyError(f"No migration path from {start} to {end}")

    def migrate(
        self,
        payload: Mapping[str, Any],
        *,
        target: Union[str, SchemaVersion] = CURRENT_EXPORT_SCHEMA_VERSION,
        source: Optional[Union[str, SchemaVersion]] = None,
    ) -> MigrationReport:
        source_value = source or payload.get("schema_version") or "0.0.0"
        start = SchemaVersion.parse(source_value)
        end = SchemaVersion.parse(target)
        data: Mapping[str, Any] = dict(payload)
        report = MigrationReport(str(start), str(end))
        for step in self.path(start, end):
            data = dict(step.function(data))
            data["schema_version"] = str(step.target)
            report.applied.append(step.name)
        report.payload = dict(data)
        return report


DEFAULT_MIGRATION_REGISTRY = MigrationRegistry()


def register_migration(
    source: Union[str, SchemaVersion],
    target: Union[str, SchemaVersion],
    *,
    registry: MigrationRegistry = DEFAULT_MIGRATION_REGISTRY,
    name: str = "",
    description: str = "",
) -> Callable[[Callable[[Mapping[str, Any]], Mapping[str, Any]]], Callable[[Mapping[str, Any]], Mapping[str, Any]]]:
    """Register a migration decorator."""
    def decorator(function: Callable[[Mapping[str, Any]], Mapping[str, Any]]) -> Callable[[Mapping[str, Any]], Mapping[str, Any]]:
        registry.register(source, target, function, name=name, description=description)
        return function
    return decorator


@register_migration("0.0.0", "1.0.0", name="legacy_to_v1")
def _legacy_to_v1(payload: Mapping[str, Any]) -> Mapping[str, Any]:
    data = dict(payload)
    if "score" in data and "total_score" not in data:
        data["total_score"] = data.pop("score")
    if "pose" in data and "poses" not in data:
        data["poses"] = [data.pop("pose")]
    if "file" in data and "files" not in data:
        data["files"] = [data.pop("file")]
    data.setdefault("metadata", {})
    data["schema_version"] = "1.0.0"
    return data


def detect_schema_version(value: Any) -> SchemaVersion:
    """Detect a schema version from exported data."""
    if isinstance(value, Mapping):
        return SchemaVersion.parse(value.get("schema_version", "0.0.0"))
    candidate = getattr(value, "schema_version", None)
    return SchemaVersion.parse(candidate or "0.0.0")


def migrate_export_data(
    value: Mapping[str, Any],
    *,
    target: Union[str, SchemaVersion] = CURRENT_EXPORT_SCHEMA_VERSION,
    registry: MigrationRegistry = DEFAULT_MIGRATION_REGISTRY,
) -> MigrationReport:
    """Migrate exported data to a target version."""
    return registry.migrate(value, target=target)


def migrate_export_file(
    source: PathLike,
    destination: Optional[PathLike] = None,
    *,
    target: Union[str, SchemaVersion] = CURRENT_EXPORT_SCHEMA_VERSION,
    overwrite: Union[str, OverwriteMode] = OverwriteMode.OVERWRITE.value,
    registry: MigrationRegistry = DEFAULT_MIGRATION_REGISTRY,
) -> MigrationReport:
    """Migrate a JSON export file."""
    source_path = Path(source)
    payload = read_json(source_path)
    if not isinstance(payload, Mapping):
        raise ExportValidationError("Migration requires a JSON object")
    report = registry.migrate(payload, target=target)
    target_path = Path(destination) if destination is not None else source_path
    write_json(
        report.payload or {},
        output_dir=target_path.parent,
        basename=target_path.stem,
        overwrite=overwrite,
    )
    return report


def schema_compatible(
    found: Union[str, SchemaVersion],
    expected: Union[str, SchemaVersion] = CURRENT_EXPORT_SCHEMA_VERSION,
    *,
    allow_older_minor: bool = True,
) -> bool:
    """Check schema compatibility."""
    left = SchemaVersion.parse(found)
    right = SchemaVersion.parse(expected)
    if left.major != right.major:
        return False
    return left <= right if allow_older_minor else left == right


__all__.extend([
    "CURRENT_EXPORT_SCHEMA_VERSION",
    "SchemaVersion",
    "MigrationStep",
    "MigrationReport",
    "MigrationRegistry",
    "DEFAULT_MIGRATION_REGISTRY",
    "register_migration",
    "detect_schema_version",
    "migrate_export_data",
    "migrate_export_file",
    "schema_compatible",
])
# -----------------------------------------------------------------------------
# 29. Introspection
# -----------------------------------------------------------------------------

INTROSPECTION_SCHEMA_VERSION: Final[str] = "1.0.0"


@dataclass(slots=True)
class ExportCapability:
    """Describe one available export capability."""

    name: str
    available: bool
    description: str = ""
    dependency: Optional[str] = None
    details: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable capability record."""
        return asdict(self)


@dataclass(slots=True)
class ExportIntrospectionReport:
    """Describe the active export subsystem."""

    schema_version: str = INTROSPECTION_SCHEMA_VERSION
    module: str = _MODULE_NAME
    module_version: str = __version__
    formats: Tuple[str, ...] = ()
    serializers: List[Dict[str, Any]] = field(default_factory=list)
    capabilities: List[ExportCapability] = field(default_factory=list)
    public_names: Tuple[str, ...] = ()
    optional_dependencies: Dict[str, bool] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def available_capabilities(self) -> Tuple[str, ...]:
        """Return available capability names."""
        return tuple(item.name for item in self.capabilities if item.available)

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable report."""
        payload = asdict(self)
        payload["available_capabilities"] = list(self.available_capabilities)
        return payload


def serializer_registry_snapshot(
    registry: SerializerRegistry = DEFAULT_SERIALIZER_REGISTRY,
) -> List[Dict[str, Any]]:
    """Return a stable serializer registry snapshot."""
    records: List[Dict[str, Any]] = []
    for entry in registry.entries():
        records.append({
            "name": entry.name,
            "aliases": list(entry.aliases),
            "priority": entry.priority,
            "match_mode": entry.match_mode.value,
            "enabled": entry.enabled,
            "builtin": entry.builtin,
            "target_type": (
                f"{entry.target_type.__module__}.{entry.target_type.__qualname__}"
                if entry.target_type is not None
                else None
            ),
            "description": entry.description,
            "metadata": dict(entry.metadata),
        })
    return records


def export_capabilities() -> List[ExportCapability]:
    """Return supported runtime capabilities."""
    return [
        ExportCapability("json", True, "JSON and JSON Lines export"),
        ExportCapability("delimited", True, "CSV and TSV export"),
        ExportCapability(
            "excel",
            OPENPYXL_AVAILABLE,
            "Excel workbook export",
            dependency="openpyxl",
        ),
        ExportCapability(
            "numpy",
            NUMPY_AVAILABLE,
            "NumPy-aware serialization",
            dependency="numpy",
        ),
        ExportCapability(
            "pandas",
            PANDAS_AVAILABLE,
            "DataFrame conversion",
            dependency="pandas",
        ),
        ExportCapability(
            "chimerax",
            CHIMERAX_AVAILABLE,
            "ChimeraX integration",
            dependency="chimerax",
        ),
        ExportCapability("transactional_write", True, "Atomic and transactional writes"),
        ExportCapability("manifest", True, "Manifest and provenance generation"),
        ExportCapability("migration", True, "Schema migration support"),
        ExportCapability("validation", True, "Strict and permissive validation"),
    ]


def inspect_export_object(
    value: Any,
    *,
    registry: SerializerRegistry = DEFAULT_SERIALIZER_REGISTRY,
) -> Dict[str, Any]:
    """Inspect serializer resolution and object export features."""
    entry = registry.resolve(value)
    attributes = {
        name: hasattr(value, name)
        for name in (
            "files",
            "metadata",
            "interactions",
            "scoring",
            "rank",
            "score",
            "total_score",
        )
    }
    return {
        "type": f"{type(value).__module__}.{type(value).__qualname__}",
        "serializer": entry.name if entry is not None else None,
        "is_dock_model_like": is_dock_model_like(value),
        "schema_version": str(detect_schema_version(value)),
        "attributes": attributes,
        "serializable": entry is not None or registry.allow_fallback,
    }


def introspect_export_system(
    *,
    registry: SerializerRegistry = DEFAULT_SERIALIZER_REGISTRY,
    include_public_names: bool = True,
) -> ExportIntrospectionReport:
    """Build a complete export subsystem report."""
    dependencies = {
        "numpy": NUMPY_AVAILABLE,
        "pandas": PANDAS_AVAILABLE,
        "openpyxl": OPENPYXL_AVAILABLE,
        "chimerax": CHIMERAX_AVAILABLE,
    }
    return ExportIntrospectionReport(
        formats=tuple(sorted(SUPPORTED_EXPORT_FORMATS)),
        serializers=serializer_registry_snapshot(registry),
        capabilities=export_capabilities(),
        public_names=tuple(sorted(set(__all__))) if include_public_names else (),
        optional_dependencies=dependencies,
        metadata={
            "serializer_registry": registry.name,
            "serializer_count": len(registry),
            "registry_frozen": registry.frozen,
        },
    )


def export_system_summary(
    *,
    registry: SerializerRegistry = DEFAULT_SERIALIZER_REGISTRY,
) -> str:
    """Return a concise human-readable subsystem summary."""
    report = introspect_export_system(registry=registry)
    lines = [
        f"{report.module} {report.module_version}",
        f"Formats: {', '.join(report.formats)}",
        f"Serializers: {len(report.serializers)}",
        f"Capabilities: {', '.join(report.available_capabilities)}",
    ]
    unavailable = [item.name for item in report.capabilities if not item.available]
    if unavailable:
        lines.append(f"Unavailable: {', '.join(unavailable)}")
    return "\n".join(lines)


__all__.extend([
    "INTROSPECTION_SCHEMA_VERSION",
    "ExportCapability",
    "ExportIntrospectionReport",
    "serializer_registry_snapshot",
    "export_capabilities",
    "inspect_export_object",
    "introspect_export_system",
    "export_system_summary",
])
# -----------------------------------------------------------------------------
# 30. Self-tests
# 30.1. Infrastructure
# -----------------------------------------------------------------------------

import time as _self_test_time

SELF_TEST_SCHEMA_VERSION: Final[str] = "1.1.0"
SELF_TEST_CODE_FAILURE: Final[str] = "code_failure"
SELF_TEST_TEST_FAILURE: Final[str] = "test_failure"
SELF_TEST_ENVIRONMENTAL_LIMITATION: Final[str] = "environmental_limitation"
SELF_TEST_FAILURE_CATEGORIES: Final[FrozenSet[str]] = frozenset(
    {
        SELF_TEST_CODE_FAILURE,
        SELF_TEST_TEST_FAILURE,
        SELF_TEST_ENVIRONMENTAL_LIMITATION,
    }
)


class SelfTestFailure(AssertionError):
    """Raised when an internal self-test fails."""


@dataclass(slots=True)
class SelfTestCaseResult:
    """Store one self-test result."""

    name: str
    passed: bool
    duration_seconds: float = 0.0
    message: str = ""
    exception_type: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    failure_category: Optional[str] = None

    def __post_init__(self) -> None:
        if self.failure_category is not None:
            category = str(self.failure_category).strip().lower()
            if category not in SELF_TEST_FAILURE_CATEGORIES:
                raise ValueError(f"Unknown self-test failure category: {category!r}")
            self.failure_category = category
        if self.passed:
            self.failure_category = None

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable test result."""
        return asdict(self)


@dataclass(slots=True)
class SelfTestReport:
    """Aggregate self-test results."""

    schema_version: str = SELF_TEST_SCHEMA_VERSION
    name: str = "export_self_tests"
    cases: List[SelfTestCaseResult] = field(default_factory=list)
    started_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    finished_at: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> int:
        """Return the number of passing cases."""
        return sum(case.passed for case in self.cases)

    @property
    def failed(self) -> int:
        """Return the number of failing cases."""
        return len(self.cases) - self.passed

    @property
    def success(self) -> bool:
        """Return whether all cases passed."""
        return bool(self.cases) and self.failed == 0

    @property
    def failure_counts(self) -> Dict[str, int]:
        """Return failure totals grouped by category."""
        counts = {category: 0 for category in sorted(SELF_TEST_FAILURE_CATEGORIES)}
        for case in self.cases:
            if not case.passed and case.failure_category in counts:
                counts[case.failure_category] += 1
        return counts

    def add(self, result: SelfTestCaseResult) -> SelfTestCaseResult:
        """Append one case result."""
        self.cases.append(result)
        return result

    def finish(self) -> "SelfTestReport":
        """Mark the report as finished."""
        self.finished_at = datetime.now(timezone.utc)
        return self

    def raise_for_failures(self) -> None:
        """Raise when at least one case failed."""
        if self.failed:
            messages = [case.message or case.name for case in self.cases if not case.passed]
            raise SelfTestFailure("; ".join(messages))

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable report."""
        return {
            "schema_version": self.schema_version,
            "name": self.name,
            "success": self.success,
            "passed": self.passed,
            "failed": self.failed,
            "failure_counts": self.failure_counts,
            "started_at": self.started_at.isoformat(),
            "finished_at": self.finished_at.isoformat() if self.finished_at else None,
            "cases": [case.to_dict() for case in self.cases],
            "metadata": dict(self.metadata),
        }


def self_test_assert(condition: Any, message: str = "Self-test assertion failed") -> None:
    """Assert a self-test condition."""
    if not condition:
        raise SelfTestFailure(message)


def self_test_equal(actual: Any, expected: Any, message: Optional[str] = None) -> None:
    """Assert equality in a self-test."""
    if actual != expected:
        raise SelfTestFailure(message or f"Expected {expected!r}, received {actual!r}")


def self_test_raises(
    exception_type: type[BaseException],
    function: Callable[..., Any],
    *args: Any,
    **kwargs: Any,
) -> BaseException:
    """Assert that a callable raises the expected exception."""
    try:
        function(*args, **kwargs)
    except exception_type as exc:
        return exc
    except BaseException as exc:
        raise SelfTestFailure(
            f"Expected {exception_type.__name__}, received {type(exc).__name__}"
        ) from exc
    raise SelfTestFailure(f"Expected {exception_type.__name__} to be raised")


def _classify_self_test_failure(exc: BaseException) -> str:
    """Classify an uncaught self-test exception."""
    chain: List[BaseException] = []
    current: Optional[BaseException] = exc
    seen: Set[int] = set()
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        chain.append(current)
        current = current.__cause__ or current.__context__

    environmental_types = (
        ImportError,
        ModuleNotFoundError,
        ExportDependencyError,
    )
    if any(isinstance(item, environmental_types) for item in chain):
        return SELF_TEST_ENVIRONMENTAL_LIMITATION
    if isinstance(exc, SelfTestFailure):
        return SELF_TEST_CODE_FAILURE

    traceback = exc.__traceback__
    deepest_name = ""
    while traceback is not None:
        deepest_name = traceback.tb_frame.f_code.co_name
        traceback = traceback.tb_next
    if deepest_name.startswith(("_test_", "self_test_", "run_self_test_")):
        return SELF_TEST_TEST_FAILURE
    return SELF_TEST_CODE_FAILURE


def run_self_test_case(
    name: str,
    function: Callable[[], Any],
    *,
    raise_on_error: bool = False,
    metadata: Optional[Mapping[str, Any]] = None,
) -> SelfTestCaseResult:
    """Execute one self-test case and classify failures."""
    start = _self_test_time.perf_counter()
    try:
        function()
    except BaseException as exc:
        result = SelfTestCaseResult(
            name=name,
            passed=False,
            duration_seconds=_self_test_time.perf_counter() - start,
            message=str(exc),
            exception_type=type(exc).__name__,
            failure_category=_classify_self_test_failure(exc),
            metadata=dict(metadata or {}),
        )
        if raise_on_error:
            raise
        return result
    return SelfTestCaseResult(
        name=name,
        passed=True,
        duration_seconds=_self_test_time.perf_counter() - start,
        metadata=dict(metadata or {}),
    )


@contextmanager
def self_test_workspace(prefix: str = "dockanalyzer_export_") -> Iterator[Path]:
    """Create a temporary self-test workspace."""
    with TemporaryDirectory(prefix=prefix) as directory:
        yield Path(directory)


@dataclass(slots=True)
class MockExportObject:
    """Minimal serializable object used by self-tests."""

    identifier: str = "pose_1"
    score: float = -7.5
    tags: List[str] = field(default_factory=lambda: ["test", "mock"])
    metadata: Dict[str, Any] = field(default_factory=lambda: {"source": "self-test"})


@dataclass(slots=True)
class MockDockModel:
    """Minimal DockModel-like object used by self-tests."""

    name: str = "pose_1"
    total_score: float = -7.5
    files: Any = field(default_factory=list)
    interactions: List[Any] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


def self_test_payload() -> Dict[str, Any]:
    """Return a deterministic serialization fixture."""
    return {
        "schema_version": CURRENT_EXPORT_SCHEMA_VERSION,
        "identifier": "pose_1",
        "score": -7.5,
        "active": True,
        "tags": ["alpha", "beta"],
        "metadata": {"replicate": 1, "source": "self-test"},
    }


__all__.extend([
    "SELF_TEST_SCHEMA_VERSION",
    "SELF_TEST_CODE_FAILURE",
    "SELF_TEST_TEST_FAILURE",
    "SELF_TEST_ENVIRONMENTAL_LIMITATION",
    "SELF_TEST_FAILURE_CATEGORIES",
    "SelfTestFailure",
    "SelfTestCaseResult",
    "SelfTestReport",
    "self_test_assert",
    "self_test_equal",
    "self_test_raises",
    "run_self_test_case",
    "self_test_workspace",
    "MockExportObject",
    "MockDockModel",
    "self_test_payload",
])
# -----------------------------------------------------------------------------
# 30.2. Path and serialization tests
# -----------------------------------------------------------------------------


def _test_path_normalization() -> None:
    with self_test_workspace() as workspace:
        path = absolute_path("nested/result.json", base_dir=workspace)
        self_test_equal(path, workspace / "nested" / "result.json")
        self_test_equal(path.suffix, ".json")


def _test_output_directory_creation() -> None:
    with self_test_workspace() as workspace:
        directory = resolve_output_directory(workspace / "exports", create=True)
        self_test_assert(directory.is_dir(), "Output directory was not created")


def _test_filename_construction() -> None:
    filename = build_export_filename(
        "dock result",
        format_name=EXPORT_FORMAT_JSON,
        table="scores",
        pose=3,
    )
    self_test_assert(filename.endswith(".json"), "JSON extension was not applied")
    self_test_assert("scores" in filename, "Table label was not included")
    self_test_assert("3" in filename, "Pose label was not included")


def _test_overwrite_modes() -> None:
    with self_test_workspace() as workspace:
        target = workspace / "result.json"
        target.write_text("{}", encoding=DEFAULT_ENCODING)
        versioned = resolve_overwrite_path(target, OverwriteMode.UNIQUE)
        self_test_assert(versioned != target, "Unique mode reused the existing path")
        self_test_assert(not versioned.exists(), "Versioned path already exists")
        self_test_raises(
            ExportPathError,
            resolve_overwrite_path,
            target,
            OverwriteMode.ERROR,
        )


def _test_json_roundtrip() -> None:
    payload = self_test_payload()
    with self_test_workspace() as workspace:
        exported = write_json(payload, output_dir=workspace, basename="payload")
        restored = read_json(exported.path)
        self_test_equal(restored, payload)
        self_test_assert(exported.path.exists(), "JSON file was not created")


def _test_json_lines_roundtrip() -> None:
    records = [self_test_payload(), {"identifier": "pose_2", "score": -6.9}]
    with self_test_workspace() as workspace:
        exported = write_json_lines(records, output_dir=workspace, basename="records")
        restored = read_json_lines(exported.path)
        self_test_equal(restored, records)


def _test_generic_serialization() -> None:
    value = MockExportObject()
    serialized = to_serializable(value)
    self_test_assert(isinstance(serialized, Mapping), "Dataclass was not serialized")
    self_test_equal(serialized["identifier"], value.identifier)
    self_test_equal(serialized["score"], value.score)


def _test_json_mapping_preservation() -> None:
    payload = {
        "record_type": "dock_model",
        "dock_model_id": "pose_1",
        "total_score": -7.5,
        "scoring": {"total_score": -7.5},
        "interactions": {"hbond": []},
    }
    restored = json.loads(json_dumps(payload))
    self_test_equal(restored, payload)


def _test_registry_resolution() -> None:
    value = MockDockModel()
    entry = DEFAULT_SERIALIZER_REGISTRY.resolve(value)
    self_test_assert(entry is not None, "No serializer resolved for DockModel-like object")
    self_test_equal(entry.name, "dock_model")


def _test_table_conversion() -> None:
    records = [self_test_payload(), {"identifier": "pose_2", "score": -6.9}]
    table = build_table(records, name="poses")
    self_test_equal(table.name, "poses")
    self_test_equal(len(table.rows), 2)
    self_test_assert("identifier" in table.columns, "Identifier column was not discovered")


def _test_schema_detection() -> None:
    payload = self_test_payload()
    version = detect_schema_version(payload)
    self_test_equal(version, SchemaVersion.parse(CURRENT_EXPORT_SCHEMA_VERSION))


def path_serialization_self_test_cases() -> Dict[str, Callable[[], None]]:
    """Return Section 30.2 self-test cases."""
    return {
        "path_normalization": _test_path_normalization,
        "output_directory_creation": _test_output_directory_creation,
        "filename_construction": _test_filename_construction,
        "overwrite_modes": _test_overwrite_modes,
        "json_roundtrip": _test_json_roundtrip,
        "json_lines_roundtrip": _test_json_lines_roundtrip,
        "generic_serialization": _test_generic_serialization,
        "json_mapping_preservation": _test_json_mapping_preservation,
        "registry_resolution": _test_registry_resolution,
        "table_conversion": _test_table_conversion,
        "schema_detection": _test_schema_detection,
    }


def run_path_serialization_self_tests(
    *,
    raise_on_error: bool = False,
) -> SelfTestReport:
    """Run path and serialization self-tests."""
    report = SelfTestReport(name="export_path_serialization_self_tests")
    for name, function in path_serialization_self_test_cases().items():
        report.add(run_self_test_case(name, function, raise_on_error=raise_on_error))
    report.metadata.update({
        "section": "30.2",
        "case_count": len(report.cases),
        "python": platform.python_version(),
    })
    report.finish()
    if raise_on_error:
        report.raise_for_failures()
    return report


__all__.extend([
    "path_serialization_self_test_cases",
    "run_path_serialization_self_tests",
])
# -----------------------------------------------------------------------------
# 30.3. Object and interaction tests
# -----------------------------------------------------------------------------


def _self_test_atom() -> Dict[str, Any]:
    return {
        "name": "N1",
        "element": "N",
        "serial": 1,
        "coord": (1.0, 2.0, 3.0),
        "residue": {"name": "LIG", "number": 1, "chain_id": "A"},
    }


def _self_test_residue() -> Dict[str, Any]:
    return {
        "name": "ASP",
        "number": 42,
        "chain_id": "B",
        "atoms": [_self_test_atom()],
    }


def _self_test_interactions() -> List[Dict[str, Any]]:
    return [
        {
            "interaction_id": "hbond_1",
            "family": "hbond",
            "interaction_type": "hydrogen_bond",
            "atom1": _self_test_atom(),
            "atom2": {"name": "O2", "element": "O", "serial": 2},
            "residue1": {"name": "LIG", "number": 1, "chain_id": "A"},
            "residue2": {"name": "ASP", "number": 42, "chain_id": "B"},
            "distance": 2.8,
            "angle": 164.0,
            "score": 1.2,
            "strength": "strong",
            "pose": "pose_1",
        },
        {
            "interaction_id": "contact_1",
            "family": "contact",
            "interaction_type": "van_der_waals",
            "atom1": {"name": "C1", "element": "C", "serial": 3},
            "atom2": {"name": "C2", "element": "C", "serial": 4},
            "distance": 3.6,
            "score": 0.4,
            "pose": "pose_1",
        },
    ]


def _test_atom_serialization() -> None:
    record = atom_to_record(_self_test_atom())
    self_test_equal(record.get("name"), "N1")
    self_test_equal(record.get("element"), "N")
    self_test_assert("coordinates" in record or "coord" in record, "Atom coordinates were not serialized")


def _test_residue_serialization() -> None:
    record = residue_to_record(_self_test_residue())
    self_test_equal(record.get("name"), "ASP")
    self_test_assert(
        record.get("number") == 42 or record.get("residue_number") == 42,
        "Residue number was not serialized",
    )


def _test_interaction_inference() -> None:
    interaction = _self_test_interactions()[0]
    self_test_equal(infer_interaction_family(interaction), "hbond")
    self_test_equal(infer_interaction_type(interaction), "hydrogen_bond")
    self_test_assert(is_interaction_like(interaction), "Interaction was not recognized")


def _test_interaction_serialization() -> None:
    record = interaction_to_record(_self_test_interactions()[0])
    self_test_equal(record.get("family"), "hbond")
    self_test_equal(record.get("interaction_type"), "hydrogen_bond")
    geometry = record.get("geometry", {})
    distance = geometry.get("distance") if isinstance(geometry, Mapping) else None
    if distance is None and isinstance(record.get("value"), Mapping):
        distance = record["value"].get("distance")
    self_test_equal(distance, 2.8)


def _test_interaction_collection_summary() -> None:
    result = interactions_to_records(_self_test_interactions(), include_summary=True)
    self_test_equal(result["summary"]["total"], 2)
    self_test_equal(result["summary"]["by_family"]["hbond"], 1)
    self_test_equal(result["summary"]["by_family"]["contact"], 1)


def _test_interaction_deduplication() -> None:
    interaction = _self_test_interactions()[0]
    duplicate = dict(interaction)
    duplicate["interaction_id"] = "hbond_duplicate"
    deduplicated = deduplicate_interactions([interaction, duplicate])
    self_test_equal(len(deduplicated), 1)


def _test_interaction_grouping() -> None:
    grouped = group_interactions_by_family(_self_test_interactions())
    self_test_equal(len(grouped["hbond"]), 1)
    self_test_equal(len(grouped["contact"]), 1)


def _test_interaction_table() -> None:
    table = interactions_table(_self_test_interactions(), name="interactions")
    self_test_equal(table.name, "interactions")
    self_test_equal(len(table.rows), 2)
    self_test_assert("family" in table.columns, "Interaction family column is missing")


def _test_interaction_cache_isolation() -> None:
    shared_atom = _self_test_atom()
    interactions = [
        {
            "family": "contact",
            "interaction_type": "contact",
            "atom1": shared_atom,
            "atom2": {"name": f"C{index}", "element": "C", "serial": index + 2},
            "distance": 3.5,
        }
        for index in range(2)
    ]
    records = interactions_to_records(interactions)
    first_atom = records[0]["participants"][0]["atom"]
    second_atom = records[1]["participants"][0]["atom"]
    self_test_equal(first_atom, second_atom)
    self_test_assert(first_atom is not second_atom, "Cached atom records were shared")
    first_coordinates = first_atom.get("coordinates")
    second_coordinates = second_atom.get("coordinates")
    if isinstance(first_coordinates, list) and isinstance(second_coordinates, list):
        self_test_assert(
            first_coordinates is not second_coordinates,
            "Cached coordinate lists were shared",
        )


def object_interaction_self_test_cases() -> Dict[str, Callable[[], None]]:
    """Return Section 30.3 self-test cases."""
    return {
        "atom_serialization": _test_atom_serialization,
        "residue_serialization": _test_residue_serialization,
        "interaction_inference": _test_interaction_inference,
        "interaction_serialization": _test_interaction_serialization,
        "interaction_collection_summary": _test_interaction_collection_summary,
        "interaction_deduplication": _test_interaction_deduplication,
        "interaction_grouping": _test_interaction_grouping,
        "interaction_table": _test_interaction_table,
        "interaction_cache_isolation": _test_interaction_cache_isolation,
    }


def run_object_interaction_self_tests(*, raise_on_error: bool = False) -> SelfTestReport:
    """Run object and interaction self-tests."""
    report = SelfTestReport(name="export_object_interaction_self_tests")
    for name, function in object_interaction_self_test_cases().items():
        report.add(run_self_test_case(name, function, raise_on_error=raise_on_error))
    report.metadata.update({"section": "30.3", "case_count": len(report.cases)})
    report.finish()
    if raise_on_error:
        report.raise_for_failures()
    return report


__all__.extend([
    "object_interaction_self_test_cases",
    "run_object_interaction_self_tests",
])
# -----------------------------------------------------------------------------
# 30.4. File format tests
# -----------------------------------------------------------------------------


def _test_csv_roundtrip() -> None:
    records = [self_test_payload(), {"identifier": "pose_2", "score": -6.8}]
    with self_test_workspace() as workspace:
        exported = write_csv(records, output_dir=workspace, basename="records")
        restored = read_delimited(exported.path)
        self_test_equal(len(restored.rows), 2)
        self_test_equal(restored.rows[0]["identifier"], "pose_1")


def _test_tsv_roundtrip() -> None:
    records = [self_test_payload(), {"identifier": "pose_2", "score": -6.8}]
    with self_test_workspace() as workspace:
        exported = write_tsv(records, output_dir=workspace, basename="records")
        restored = read_delimited(exported.path)
        self_test_equal(len(restored.rows), 2)
        self_test_equal(restored.rows[1]["identifier"], "pose_2")


def _test_text_roundtrip() -> None:
    with self_test_workspace() as workspace:
        exported = write_text(self_test_payload(), output_dir=workspace, basename="summary")
        restored = read_text(exported.path)
        self_test_assert("pose_1" in restored, "Text export omitted the identifier")
        self_test_assert(exported.path.suffix == ".txt", "Unexpected text extension")


def _test_excel_roundtrip() -> None:
    if not OPENPYXL_AVAILABLE:
        return
    records = [self_test_payload(), {"identifier": "pose_2", "score": -6.8}]
    with self_test_workspace() as workspace:
        exported = write_excel(records, output_dir=workspace, basename="records")
        restored = read_excel(exported.path)
        self_test_assert(bool(restored.tables), "Excel workbook contains no tables")
        first = next(iter(restored.tables.values()))
        self_test_equal(len(first.rows), 2)


def _test_format_extensions() -> None:
    for format_name, extension in FORMAT_EXTENSIONS.items():
        self_test_equal(normalize_export_format(extension), format_name)
        self_test_equal(canonical_extension(format_name), extension)


def _test_available_formats() -> None:
    formats = available_export_formats()
    self_test_assert(EXPORT_FORMAT_JSON in formats, "JSON format is unavailable")
    self_test_assert(EXPORT_FORMAT_CSV in formats, "CSV format is unavailable")
    if OPENPYXL_AVAILABLE:
        self_test_assert(EXPORT_FORMAT_EXCEL in formats, "Excel format was not exposed")


def _test_manifest_integrity() -> None:
    with self_test_workspace() as workspace:
        exported = write_json(self_test_payload(), output_dir=workspace, basename="payload")
        manifest = build_manifest([exported], source_name="self_test", include_hashes=True)
        manifest_file = write_manifest(manifest, output_dir=workspace, basename="manifest")
        restored = read_manifest(manifest_file.path)
        verification = verify_manifest(restored, root=workspace)
        self_test_assert(bool(verification.get("valid")), "Manifest verification failed")


def _test_load_export_dispatch() -> None:
    with self_test_workspace() as workspace:
        json_file = write_json(self_test_payload(), output_dir=workspace, basename="payload")
        self_test_equal(load_export(json_file.path), self_test_payload())


def file_format_self_test_cases() -> Dict[str, Callable[[], None]]:
    """Return Section 30.4 self-test cases."""
    return {
        "csv_roundtrip": _test_csv_roundtrip,
        "tsv_roundtrip": _test_tsv_roundtrip,
        "text_roundtrip": _test_text_roundtrip,
        "excel_roundtrip": _test_excel_roundtrip,
        "format_extensions": _test_format_extensions,
        "available_formats": _test_available_formats,
        "manifest_integrity": _test_manifest_integrity,
        "load_export_dispatch": _test_load_export_dispatch,
    }


def run_file_format_self_tests(*, raise_on_error: bool = False) -> SelfTestReport:
    """Run file format self-tests."""
    report = SelfTestReport(name="export_file_format_self_tests")
    for name, function in file_format_self_test_cases().items():
        report.add(run_self_test_case(name, function, raise_on_error=raise_on_error))
    report.metadata.update({
        "section": "30.4",
        "case_count": len(report.cases),
        "openpyxl_available": OPENPYXL_AVAILABLE,
    })
    report.finish()
    if raise_on_error:
        report.raise_for_failures()
    return report


__all__.extend([
    "file_format_self_test_cases",
    "run_file_format_self_tests",
])
# -----------------------------------------------------------------------------
# 30.5. DockModel and multipose tests
# -----------------------------------------------------------------------------


def _self_test_dock_models() -> List[MockDockModel]:
    first = MockDockModel(name="pose_1", total_score=-7.5)
    second = MockDockModel(name="pose_2", total_score=-6.8)
    first.interactions = _self_test_interactions()
    second.interactions = [_self_test_interactions()[1]]
    first.metadata = {"rank": 1}
    second.metadata = {"rank": 2}
    return [first, second]


def _test_dock_model_recognition() -> None:
    model = _self_test_dock_models()[0]
    self_test_assert(is_dock_model_like(model), "Mock DockModel was not recognized")
    self_test_equal(dock_model_identifier(model), "pose_1")


def _test_dock_model_record() -> None:
    model = _self_test_dock_models()[0]
    record = dock_model_to_record(model)
    self_test_equal(record.get("name"), "pose_1")
    self_test_equal(record.get("total_score"), -7.5)
    self_test_assert("interactions" in record, "DockModel interactions were omitted")


def _test_dock_model_collection() -> None:
    records = dock_models_to_records(_self_test_dock_models())
    self_test_equal(len(records), 2)
    summary = summarize_dock_model_records(records)
    self_test_equal(summary.count, 2)
    self_test_equal(summary.scored, 2)


def _test_dock_model_ranking() -> None:
    records = dock_models_to_records(_self_test_dock_models())
    ranked = rank_dock_model_records(records, direction=ScoreDirection.LOWER_IS_BETTER)
    self_test_equal(ranked[0].get("name"), "pose_1")
    self_test_equal(ranked[0].get("rank"), 1)


def _test_dock_model_files_update() -> None:
    model = MockDockModel()
    with self_test_workspace() as workspace:
        exported = write_json(self_test_payload(), output_dir=workspace, basename="pose")
        report = update_dock_model_files(model, [exported])
        self_test_assert(report.changed, "DockModel.files was not updated")
        self_test_assert(bool(model.files), "DockModel.files remained empty")


def _test_single_pose_export() -> None:
    model = _self_test_dock_models()[0]
    with self_test_workspace() as workspace:
        options = ExportOptions(
            output_dir=workspace,
            basename="single_pose",
            formats=(EXPORT_FORMAT_JSON,),
            include_manifest=False,
            include_provenance=False,
            update_model_files=False,
            overwrite=OverwriteMode.OVERWRITE.value,
        )
        result = export_single_pose(model, options=options)
        self_test_assert(result.succeeded, "Single-pose export failed")
        self_test_assert(any(item.path.suffix == ".json" for item in result.files), "JSON file was not created")


def _test_multipose_preparation() -> None:
    models = _self_test_dock_models()
    context = prepare_multipose_export(
        models,
        options=MultiposeExportOptions(
            export=ExportOptions(include_manifest=False, update_model_files=False)
        ),
        basename="multipose",
    )
    self_test_equal(len(context.records), 2)
    self_test_assert("poses" in context.tables.tables, "Poses table is missing")
    self_test_assert("ranking" in context.tables.tables, "Ranking table is missing")


def _test_multipose_export() -> None:
    models = _self_test_dock_models()
    with self_test_workspace() as workspace:
        options = MultiposeExportOptions(
            export=ExportOptions(
                output_dir=workspace,
                basename="multipose",
                formats=(EXPORT_FORMAT_JSON, EXPORT_FORMAT_CSV),
                include_manifest=False,
                include_provenance=False,
                update_model_files=False,
                overwrite=OverwriteMode.OVERWRITE.value,
            ),
            include_ranking=True,
        )
        result = export_multiple_poses(models, options=options)
        self_test_assert(result.succeeded, "Multipose export failed")
        suffixes = {item.path.suffix for item in result.files}
        self_test_assert(".json" in suffixes, "Multipose JSON file is missing")
        self_test_assert(".csv" in suffixes, "Multipose CSV files are missing")


def _test_batch_dynamic_fields() -> None:
    @dataclass
    class BatchSource:
        name: str
        group: str
        order: int

    items = prepare_batch_items(
        [BatchSource("pose_b", "group_1", 2), BatchSource("pose_a", "group_2", 1)],
        options=BatchExportOptions(group_key="group", sort_key="order"),
    )
    self_test_equal([item.name for item in items], ["pose_a", "pose_b"])
    self_test_equal([item.group for item in items], ["group_2", "group_1"])


def _test_extended_dock_model_interactions() -> None:
    contact = {"interaction_id": "contact_1", "family": "contact", "distance": 3.2}
    hbond = {"interaction_id": "hbond_1", "family": "hbond", "distance": 2.8}
    salt = {"interaction_id": "salt_1", "family": "saltbridge", "distance": 3.4}
    clash = {"interaction_id": "clash_1", "family": "clash", "distance": 1.8}

    @dataclass
    class ExtendedDockModel:
        name: str = "pose_extended"
        contacts: List[Any] = field(default_factory=lambda: [contact])
        hbonds: List[Any] = field(default_factory=lambda: [hbond])
        hydrophobic: List[Any] = field(default_factory=list)
        pi: List[Any] = field(default_factory=list)
        saltbridges: List[Any] = field(default_factory=lambda: [salt])
        clashes: List[Any] = field(default_factory=lambda: [clash])

        @property
        def interactions(self) -> Tuple[Any, ...]:
            return tuple(self.contacts + self.hbonds + self.saltbridges + self.clashes)

    grouped = extract_dock_model_interactions(ExtendedDockModel())
    self_test_equal(set(grouped), {"contact", "hbond", "saltbridge", "clash"})
    identifiers = [
        str(_interaction_get(item, _INTERACTION_ID_FIELDS))
        for values in grouped.values()
        for item in values
    ]
    self_test_equal(len(identifiers), len(set(identifiers)))


def _test_mapping_model_files_update() -> None:
    model = MockDockModel(files={"existing_json": "existing.json"})
    with self_test_workspace() as workspace:
        options = ExportOptions(
            output_dir=workspace,
            basename="mapping_files",
            formats=(EXPORT_FORMAT_JSON,),
            include_manifest=False,
            include_provenance=False,
            update_model_files=True,
            preserve_previous_files=True,
            overwrite=OverwriteMode.OVERWRITE.value,
        )
        result = export_single_pose(model, options=options)
        self_test_assert(result.succeeded, "Mapping-layout export failed")
        self_test_assert(isinstance(model.files, Mapping), "DockModel.files layout changed")
        self_test_assert("existing_json" in model.files, "Existing file entry was removed")
        self_test_assert(len(model.files) >= 2, "Generated file was not registered")


def _test_multipose_model_files_update() -> None:
    models = _self_test_dock_models()
    with self_test_workspace() as workspace:
        options = MultiposeExportOptions(
            export=ExportOptions(
                output_dir=workspace,
                basename="multipose_files",
                formats=(EXPORT_FORMAT_JSON,),
                include_manifest=False,
                include_provenance=False,
                update_model_files=True,
                overwrite=OverwriteMode.OVERWRITE.value,
            )
        )
        result = export_multiple_poses(models, options=options)
        self_test_assert(result.succeeded, "Multipose files export failed")
        for model in models:
            self_test_assert(bool(model.files), "Multipose files were not attached")


def _test_ranking_missing_scores_last() -> None:
    records = [
        {"name": "missing", "total_score": None},
        {"name": "low", "total_score": 1.0},
        {"name": "high", "total_score": 3.0},
    ]
    higher = rank_dock_model_records(
        records,
        direction=ScoreDirection.HIGHER_IS_BETTER,
    )
    lower = rank_dock_model_records(
        records,
        direction=ScoreDirection.LOWER_IS_BETTER,
    )
    self_test_equal([item["name"] for item in higher], ["high", "low", "missing"])
    self_test_equal([item["name"] for item in lower], ["low", "high", "missing"])


def _test_empty_and_invalid_pose_inputs() -> None:
    context = prepare_multipose_export(
        [],
        options=MultiposeExportOptions(
            export=ExportOptions(include_manifest=False, update_model_files=False)
        ),
    )
    self_test_equal(context.records, [])
    self_test_raises(ExportInputError, prepare_single_pose_export, None)
    self_test_raises(ExportInputError, prepare_multipose_export, None)
    self_test_raises(ExportInputError, prepare_batch_items, None)


def dock_model_multipose_self_test_cases() -> Dict[str, Callable[[], None]]:
    """Return Section 30.5 self-test cases."""
    return {
        "dock_model_recognition": _test_dock_model_recognition,
        "dock_model_record": _test_dock_model_record,
        "dock_model_collection": _test_dock_model_collection,
        "dock_model_ranking": _test_dock_model_ranking,
        "dock_model_files_update": _test_dock_model_files_update,
        "single_pose_export": _test_single_pose_export,
        "multipose_preparation": _test_multipose_preparation,
        "multipose_export": _test_multipose_export,
        "batch_dynamic_fields": _test_batch_dynamic_fields,
        "extended_dock_model_interactions": _test_extended_dock_model_interactions,
        "mapping_model_files_update": _test_mapping_model_files_update,
        "multipose_model_files_update": _test_multipose_model_files_update,
        "ranking_missing_scores_last": _test_ranking_missing_scores_last,
        "empty_and_invalid_pose_inputs": _test_empty_and_invalid_pose_inputs,
    }


def run_dock_model_multipose_self_tests(*, raise_on_error: bool = False) -> SelfTestReport:
    """Run DockModel and multipose self-tests."""
    report = SelfTestReport(name="export_dock_model_multipose_self_tests")
    for name, function in dock_model_multipose_self_test_cases().items():
        report.add(run_self_test_case(name, function, raise_on_error=raise_on_error))
    report.metadata.update({"section": "30.5", "case_count": len(report.cases)})
    report.finish()
    if raise_on_error:
        report.raise_for_failures()
    return report


__all__.extend([
    "dock_model_multipose_self_test_cases",
    "run_dock_model_multipose_self_tests",
])
# -----------------------------------------------------------------------------
# 30.6. Manifest and validation tests
# -----------------------------------------------------------------------------


def _self_test_exported_file(path: Path, *, include_hash: bool = True) -> ExportedFile:
    exported = ExportedFile(path=path, format=path.suffix.lstrip("."), record_count=1)
    exported.refresh_size()
    if include_hash:
        enrich_exported_file(exported, include_hash=True)
    return exported


def _test_file_hash_stability() -> None:
    with self_test_workspace() as workspace:
        path = workspace / "payload.txt"
        path.write_text("dockanalyzer\n", encoding=DEFAULT_ENCODING)
        first = compute_file_hash(path)
        second = compute_file_hash(path)
        self_test_equal(first, second)
        self_test_assert(bool(first), "File hash is empty")


def _test_manifest_roundtrip() -> None:
    with self_test_workspace() as workspace:
        payload = workspace / "payload.json"
        payload.write_text('{"value": 1}\n', encoding=DEFAULT_ENCODING)
        manifest = build_manifest(
            [_self_test_exported_file(payload)],
            source_name="self-test",
            metadata={"section": "30.6"},
        )
        exported = write_manifest(
            manifest,
            output_dir=workspace,
            basename="manifest",
            overwrite=OverwriteMode.OVERWRITE.value,
        )
        restored = read_manifest(exported.path)
        self_test_equal(restored.source_name, "self-test")
        self_test_equal(len(restored.files), 1)
        self_test_equal(restored.files[0].hash_value, manifest.files[0].hash_value)


def _test_manifest_verification_success() -> None:
    with self_test_workspace() as workspace:
        path = workspace / "data.txt"
        path.write_text("verified", encoding=DEFAULT_ENCODING)
        manifest = build_manifest([_self_test_exported_file(path)])
        result = verify_manifest(manifest)
        self_test_assert(result["valid"], "Valid manifest was rejected")
        self_test_equal(result["file_count"], 1)


def _test_manifest_detects_tampering() -> None:
    with self_test_workspace() as workspace:
        path = workspace / "data.txt"
        path.write_text("original", encoding=DEFAULT_ENCODING)
        manifest = build_manifest([_self_test_exported_file(path)])
        path.write_text("modified", encoding=DEFAULT_ENCODING)
        result = verify_manifest(manifest)
        self_test_assert(not result["valid"], "Modified file was not detected")
        self_test_equal(result["files"][0]["hash_valid"], False)


def _test_export_options_validation() -> None:
    with self_test_workspace() as workspace:
        valid_options = ExportOptions(
            output_dir=workspace,
            formats=(EXPORT_FORMAT_JSON,),
            include_manifest=False,
        )
        valid_report = validate_export_options(valid_options)
        self_test_assert(valid_report.valid, "Valid export options were rejected")
        invalid_options = ExportOptions(
            output_dir=workspace,
            formats=(EXPORT_FORMAT_JSON,),
            include_manifest=False,
        )
        invalid_options.formats = ()
        invalid_report = validate_export_options(invalid_options)
        self_test_assert(not invalid_report.valid, "Empty format list was accepted")


def _test_exported_file_validation() -> None:
    with self_test_workspace() as workspace:
        path = workspace / "result.json"
        path.write_text("{}\n", encoding=DEFAULT_ENCODING)
        exported = _self_test_exported_file(path)
        report = validate_exported_file(
            exported,
            validation=ExportValidationOptions(check_hashes=True),
        )
        self_test_assert(report.valid, "Existing exported file was rejected")
        path.unlink()
        missing_report = validate_exported_file(exported)
        self_test_assert(not missing_report.valid, "Missing file was accepted")


def _test_dock_model_validation() -> None:
    valid_report = validate_dock_model_for_export(MockDockModel())
    self_test_assert(valid_report.valid, "MockDockModel was rejected")
    invalid_report = validate_dock_model_for_export(None)
    self_test_assert(not invalid_report.valid, "None DockModel was accepted")


def _test_validation_raise_for_errors() -> None:
    report = ValidationReport(subject="self-test")
    report.add("self_test.error", "Expected validation error")
    self_test_raises(ExportValidationError, report.raise_for_errors)


def _test_public_api_contract() -> None:
    required = (
        "available_export_formats",
        "normalize_export_format",
        "export_data",
        "write_json",
        "write_json_lines",
        "write_csv",
        "write_tsv",
        "write_excel",
        "write_text",
        "build_table",
        "TableCollection",
        "ExportedFile",
        "ExportResult",
    )
    self_test_equal(len(__all__), len(set(__all__)))
    invalid = [name for name in __all__ if not isinstance(name, str) or not name]
    missing = [name for name in __all__ if isinstance(name, str) and name not in globals()]
    missing_required = [name for name in required if name not in globals() or name not in __all__]
    self_test_assert(not invalid, f"Invalid __all__ entries: {invalid!r}")
    self_test_assert(not missing, f"Missing public names: {missing!r}")
    self_test_assert(
        not missing_required,
        f"Report integration API is incomplete: {missing_required!r}",
    )


def _test_self_test_failure_classification() -> None:
    def _test_broken_fixture() -> None:
        raise NameError("missing self-test fixture")

    code_result = run_self_test_case(
        "deliberate_code_failure",
        lambda: self_test_assert(False, "deliberate mismatch"),
    )
    test_result = run_self_test_case(
        "deliberate_test_failure",
        _test_broken_fixture,
    )
    environment_result = run_self_test_case(
        "deliberate_environmental_limitation",
        lambda: (_ for _ in ()).throw(ModuleNotFoundError("optional dependency")),
    )
    self_test_equal(code_result.failure_category, SELF_TEST_CODE_FAILURE)
    self_test_equal(test_result.failure_category, SELF_TEST_TEST_FAILURE)
    self_test_equal(
        environment_result.failure_category,
        SELF_TEST_ENVIRONMENTAL_LIMITATION,
    )


def _test_non_serializable_validation() -> None:
    result = ExportResult(
        status=ExportStatus.SUCCESS.value,
        source_name="non_serializable",
        payload=object(),
    ).finish()
    report = validate_export_result(
        result,
        validation=ExportValidationOptions(
            check_existing_files=False,
            require_files_for_success=False,
            check_serializable=True,
        ),
    )
    self_test_assert(not report.valid, "Non-serializable payload was accepted")
    self_test_assert(
        any(issue.code == "result.not_serializable" for issue in report.issues),
        "Non-serializable payload was not reported",
    )


def manifest_validation_self_test_cases() -> Dict[str, Callable[[], None]]:
    """Return Section 30.6 self-test cases."""
    return {
        "file_hash_stability": _test_file_hash_stability,
        "manifest_roundtrip": _test_manifest_roundtrip,
        "manifest_verification_success": _test_manifest_verification_success,
        "manifest_detects_tampering": _test_manifest_detects_tampering,
        "export_options_validation": _test_export_options_validation,
        "exported_file_validation": _test_exported_file_validation,
        "dock_model_validation": _test_dock_model_validation,
        "validation_raise_for_errors": _test_validation_raise_for_errors,
        "public_api_contract": _test_public_api_contract,
        "failure_classification": _test_self_test_failure_classification,
        "non_serializable_validation": _test_non_serializable_validation,
    }


def run_manifest_validation_self_tests(*, raise_on_error: bool = False) -> SelfTestReport:
    """Run manifest and validation self-tests."""
    report = SelfTestReport(name="export_manifest_validation_self_tests")
    for name, function in manifest_validation_self_test_cases().items():
        report.add(run_self_test_case(name, function, raise_on_error=raise_on_error))
    report.metadata.update({"section": "30.6", "case_count": len(report.cases)})
    report.finish()
    if raise_on_error:
        report.raise_for_failures()
    return report


__all__.extend([
    "manifest_validation_self_test_cases",
    "run_manifest_validation_self_tests",
])
# -----------------------------------------------------------------------------
# 30.7. ChimeraX compatibility tests
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class _MockChimeraXResidue:
    chain_id: str = "A"
    number: int = 42


@dataclass(slots=True)
class _MockChimeraXAtom:
    name: str = "CA"
    residue: Any = field(default_factory=_MockChimeraXResidue)
    atomspec: Optional[str] = None


@dataclass(slots=True)
class _MockChimeraXModel:
    id: Tuple[int, ...] = (1, 2)
    id_string: Optional[str] = None


class _MockChimeraXLogger:
    def __init__(self) -> None:
        self.messages: Dict[str, List[str]] = {"info": [], "warning": [], "error": []}

    def info(self, message: str) -> None:
        self.messages["info"].append(message)

    def warning(self, message: str) -> None:
        self.messages["warning"].append(message)

    def error(self, message: str) -> None:
        self.messages["error"].append(message)


class _MockChimeraXSession:
    def __init__(self) -> None:
        self.logger = _MockChimeraXLogger()
        self.models: List[Any] = []


def _test_chimerax_session_recognition() -> None:
    self_test_assert(is_chimerax_session(_MockChimeraXSession()))
    self_test_assert(not is_chimerax_session(object()))


def _test_chimerax_atom_spec_explicit() -> None:
    atom = _MockChimeraXAtom(atomspec="#1/A:42@CA")
    self_test_equal(chimerax_atom_spec(atom), "#1/A:42@CA")


def _test_chimerax_atom_spec_inferred() -> None:
    atom = _MockChimeraXAtom()
    self_test_equal(chimerax_atom_spec(atom), "/A:42@CA")


def _test_chimerax_model_spec() -> None:
    self_test_equal(chimerax_model_spec(_MockChimeraXModel()), "#1.2")
    self_test_equal(chimerax_model_spec(_MockChimeraXModel(id_string="3")), "#3")


def _test_chimerax_export_commands() -> None:
    result = ExportResult(source_name="self-test")
    result.files.extend([
        ExportedFile(path=Path("pose.pdb"), format="pdb"),
        ExportedFile(path=Path("table.csv"), format="csv"),
        ExportedFile(path=Path("ligand.sdf"), format="sdf"),
    ])
    commands = chimerax_export_result_commands(result)
    self_test_equal(len(commands), 2)
    self_test_assert(any("pose.pdb" in command for command in commands))
    self_test_assert(any("ligand.sdf" in command for command in commands))


def _test_chimerax_notification_logging() -> None:
    session = _MockChimeraXSession()
    result = ExportResult(source_name="self-test")
    result.warnings.append("test warning")
    result.errors.append("test error")
    notified = notify_chimerax_export(
        session,
        result,
        options=ChimeraXExportOptions(log_results=True, open_exported=False),
    )
    self_test_assert(notified is result)
    self_test_equal(len(session.logger.messages["info"]), 1)
    self_test_equal(session.logger.messages["warning"], ["test warning"])
    self_test_equal(session.logger.messages["error"], ["test error"])


def _test_chimerax_invalid_session() -> None:
    result = ExportResult(source_name="self-test")
    self_test_raises(
        ExportDependencyError,
        notify_chimerax_export,
        object(),
        result,
    )
    self_test_raises(
        ExportDependencyError,
        run_chimerax_command,
        object(),
        "info",
    )


def chimerax_self_test_cases() -> Dict[str, Callable[[], None]]:
    """Return Section 30.7 self-test cases."""
    return {
        "session_recognition": _test_chimerax_session_recognition,
        "atom_spec_explicit": _test_chimerax_atom_spec_explicit,
        "atom_spec_inferred": _test_chimerax_atom_spec_inferred,
        "model_spec": _test_chimerax_model_spec,
        "export_commands": _test_chimerax_export_commands,
        "notification_logging": _test_chimerax_notification_logging,
        "invalid_session": _test_chimerax_invalid_session,
    }


def run_chimerax_self_tests(*, raise_on_error: bool = False) -> SelfTestReport:
    """Run ChimeraX compatibility self-tests."""
    report = SelfTestReport(name="export_chimerax_self_tests")
    for name, function in chimerax_self_test_cases().items():
        report.add(run_self_test_case(name, function, raise_on_error=raise_on_error))
    report.metadata.update({
        "section": "30.7",
        "case_count": len(report.cases),
        "chimerax_available": chimerax_available(),
    })
    report.finish()
    if raise_on_error:
        report.raise_for_failures()
    return report


__all__.extend([
    "chimerax_self_test_cases",
    "run_chimerax_self_tests",
])
# -----------------------------------------------------------------------------
# 30.8. Final runner
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class ExportSelfTestSuiteReport:
    """Aggregate report for all export self-test sections."""

    name: str = "dockanalyzer_export_self_tests"
    reports: List[SelfTestReport] = field(default_factory=list)
    started_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    finished_at: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> int:
        return sum(report.passed for report in self.reports)

    @property
    def failed(self) -> int:
        return sum(report.failed for report in self.reports)

    @property
    def success(self) -> bool:
        return self.failed == 0 and bool(self.reports)

    @property
    def failure_counts(self) -> Dict[str, int]:
        counts = {category: 0 for category in sorted(SELF_TEST_FAILURE_CATEGORIES)}
        for report in self.reports:
            for category, count in report.failure_counts.items():
                counts[category] += count
        return counts

    def finish(self) -> "ExportSelfTestSuiteReport":
        self.finished_at = datetime.now(timezone.utc)
        return self

    def raise_for_failures(self) -> None:
        if not self.success:
            failed_sections = [report.name for report in self.reports if not report.success]
            raise SelfTestFailure(
                "Export self-tests failed: " + ", ".join(failed_sections)
            )

    def to_dict(self) -> Dict[str, Any]:
        return {
            "schema_version": SELF_TEST_SCHEMA_VERSION,
            "name": self.name,
            "success": self.success,
            "passed": self.passed,
            "failed": self.failed,
            "failure_counts": self.failure_counts,
            "started_at": self.started_at.isoformat(),
            "finished_at": self.finished_at.isoformat() if self.finished_at else None,
            "reports": [report.to_dict() for report in self.reports],
            "metadata": dict(self.metadata),
        }

    def format_text(self) -> str:
        lines = [
            f"{self.name}: {'PASS' if self.success else 'FAIL'}",
            f"Passed: {self.passed}",
            f"Failed: {self.failed}",
            (
                "Failure categories: "
                f"code={self.failure_counts[SELF_TEST_CODE_FAILURE]}, "
                f"test={self.failure_counts[SELF_TEST_TEST_FAILURE]}, "
                "environment="
                f"{self.failure_counts[SELF_TEST_ENVIRONMENTAL_LIMITATION]}"
            ),
        ]
        for report in self.reports:
            lines.append(
                f"- {report.metadata.get('section', report.name)}: "
                f"{report.passed} passed, {report.failed} failed"
            )
        return "\n".join(lines)


def export_self_test_runners() -> Tuple[Callable[..., SelfTestReport], ...]:
    """Return all export self-test section runners."""
    return (
        run_path_serialization_self_tests,
        run_object_interaction_self_tests,
        run_file_format_self_tests,
        run_dock_model_multipose_self_tests,
        run_manifest_validation_self_tests,
        run_chimerax_self_tests,
    )


def run_self_tests(
    *,
    raise_on_error: bool = False,
    verbose: bool = False,
) -> ExportSelfTestSuiteReport:
    """Run the complete export.py self-test suite."""
    suite = ExportSelfTestSuiteReport()
    for runner in export_self_test_runners():
        report = runner(raise_on_error=False)
        suite.reports.append(report)
        if verbose:
            print(
                f"[{report.metadata.get('section', report.name)}] "
                f"passed={report.passed} failed={report.failed}"
            )
        if raise_on_error and not report.success:
            report.raise_for_failures()
    suite.metadata.update({
        "section": "30.8",
        "section_count": len(suite.reports),
        "python": platform.python_version(),
        "platform": platform.platform(),
        "chimerax_available": chimerax_available(),
    })
    suite.finish()
    if raise_on_error:
        suite.raise_for_failures()
    if verbose:
        print(suite.format_text())
    return suite


__all__.extend([
    "ExportSelfTestSuiteReport",
    "export_self_test_runners",
    "run_self_tests",
])


if __name__ == "__main__":
    _self_test_suite = run_self_tests(verbose=True)
    raise SystemExit(0 if _self_test_suite.success else 1)
