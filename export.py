# =============================================================================
# DockAnalyzer — Data export and serialization
# Section 1 — Header, imports, aliases and public interface
# =============================================================================

"""Serialization and file export utilities for DockAnalyzer.

This module converts DockAnalyzer results into portable representations and
writes them to JSON, JSON Lines, CSV, TSV, Excel and plain-text files.

The export layer consumes existing analysis and scoring results. It does not
redetect molecular interactions, recalculate scores or generate scientific
reports. Specialized modules remain responsible for their own chemistry,
geometry and scoring logic.

The implementation is designed to work:

1. inside UCSF ChimeraX with native molecular objects;
2. outside ChimeraX with ordinary Python objects;
3. with dictionaries loaded from previous exports;
4. with synthetic objects used by DockAnalyzer self-tests.

Optional dependencies are loaded defensively. JSON, JSON Lines, CSV, TSV and
text export remain available without pandas or openpyxl.
"""

from __future__ import annotations

# -----------------------------------------------------------------------------
# 1.1. Standard-library imports
# -----------------------------------------------------------------------------

from collections import defaultdict
from collections.abc import Iterable, Iterator, Mapping, MutableMapping, Sequence
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field, fields, is_dataclass, replace
from datetime import date, datetime, time, timedelta, timezone
from enum import Enum
from hashlib import sha256
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile, TemporaryDirectory
from types import MappingProxyType
from typing import (
    IO,
    Any,
    BinaryIO,
    Callable,
    ClassVar,
    Dict,
    Final,
    FrozenSet,
    Generic,
    Hashable,
    List,
    Literal,
    NamedTuple,
    Optional,
    Protocol,
    Set,
    TextIO,
    Tuple,
    TypeAlias,
    TypeVar,
    Union,
    runtime_checkable,
)
import csv
import gzip
import json
import logging
import math
import os
import platform
import re
import shutil
import sys
import warnings

# -----------------------------------------------------------------------------
# 1.2. Optional NumPy support
# -----------------------------------------------------------------------------

try:
    import numpy as np
    from numpy.typing import NDArray

    NUMPY_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - environment dependent
    np = None  # type: ignore[assignment]
    NDArray = Any  # type: ignore[misc,assignment]
    NUMPY_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.3. Optional pandas support
# -----------------------------------------------------------------------------

try:
    import pandas as pd

    PANDAS_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - environment dependent
    pd = None  # type: ignore[assignment]
    PANDAS_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.4. Optional openpyxl support
# -----------------------------------------------------------------------------

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - environment dependent
    openpyxl = None  # type: ignore[assignment]
    Workbook = Any  # type: ignore[misc,assignment]
    Alignment = Any  # type: ignore[misc,assignment]
    Font = Any  # type: ignore[misc,assignment]
    PatternFill = Any  # type: ignore[misc,assignment]
    Table = Any  # type: ignore[misc,assignment]
    TableStyleInfo = Any  # type: ignore[misc,assignment]
    get_column_letter = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.5. Optional ChimeraX support
# -----------------------------------------------------------------------------

try:
    from chimerax.atomic import Atom as ChimeraXAtom
    from chimerax.atomic import AtomicStructure as ChimeraXAtomicStructure
    from chimerax.atomic import Residue as ChimeraXResidue

    CHIMERAX_AVAILABLE: Final[bool] = True
except ImportError:  # pragma: no cover - expected outside ChimeraX
    ChimeraXAtom = Any  # type: ignore[misc,assignment]
    ChimeraXAtomicStructure = Any  # type: ignore[misc,assignment]
    ChimeraXResidue = Any  # type: ignore[misc,assignment]
    CHIMERAX_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1.6. Internal DockAnalyzer imports
# -----------------------------------------------------------------------------

try:
    from . import config
    from .utils import DockLogger, DockModel
except ImportError:
    import config
    from utils import DockLogger, DockModel

# -----------------------------------------------------------------------------
# 1.7. Module metadata
# -----------------------------------------------------------------------------

__author__: Final[str] = "Leonardo Bastos and DockAnalyzer contributors"
__version__: Final[str] = "0.1.0"
__license__: Final[str] = "MIT"
__status__: Final[str] = "Development"

_MODULE_NAME: Final[str] = "export"
_MODULE_DESCRIPTION: Final[str] = (
    "Serialization and multi-format export utilities for DockAnalyzer."
)

_LOGGER: Final[DockLogger] = DockLogger(_MODULE_NAME)

# -----------------------------------------------------------------------------
# 1.8. Public interface
# -----------------------------------------------------------------------------

# Public names are registered after their defining sections.
__all__: List[str] = []

# -----------------------------------------------------------------------------
# 1.9. Generic type variables
# -----------------------------------------------------------------------------

T = TypeVar("T")
K = TypeVar("K", bound=Hashable)
V = TypeVar("V")
RecordT = TypeVar("RecordT", bound=Mapping[str, Any])
ExportObjectT = TypeVar("ExportObjectT")

# -----------------------------------------------------------------------------
# 1.10. General aliases
# -----------------------------------------------------------------------------

PathLike: TypeAlias = Union[str, os.PathLike[str], Path]
FileLike: TypeAlias = Union[TextIO, BinaryIO]

Number: TypeAlias = Union[int, float]
if NUMPY_AVAILABLE:
    Number = Union[int, float, np.integer, np.floating]

JSONPrimitive: TypeAlias = Union[str, int, float, bool, None]
JSONValue: TypeAlias = Union[
    JSONPrimitive,
    List["JSONValue"],
    Dict[str, "JSONValue"],
]
JSONObject: TypeAlias = Dict[str, JSONValue]
JSONArray: TypeAlias = List[JSONValue]

Record: TypeAlias = Dict[str, Any]
RecordSequence: TypeAlias = Sequence[Mapping[str, Any]]
RecordList: TypeAlias = List[Record]
TableMapping: TypeAlias = Mapping[str, RecordSequence]
MutableTableMapping: TypeAlias = MutableMapping[str, RecordList]

Metadata: TypeAlias = Mapping[str, Any]
MutableMetadata: TypeAlias = MutableMapping[str, Any]

ColumnName: TypeAlias = str
SheetName: TypeAlias = str
TableName: TypeAlias = str
FormatName: TypeAlias = str
SchemaVersion: TypeAlias = str

# -----------------------------------------------------------------------------
# 1.11. Molecular and DockAnalyzer aliases
# -----------------------------------------------------------------------------

AtomLike: TypeAlias = Any
ResidueLike: TypeAlias = Any
StructureLike: TypeAlias = Any
InteractionLike: TypeAlias = Any
AnalysisResultLike: TypeAlias = Any
ScoringResultLike: TypeAlias = Any
DockModelLike: TypeAlias = Union[DockModel, Any]

AtomCollection: TypeAlias = Iterable[AtomLike]
ResidueCollection: TypeAlias = Iterable[ResidueLike]
InteractionCollection: TypeAlias = Iterable[InteractionLike]
DockModelCollection: TypeAlias = Iterable[DockModelLike]

# -----------------------------------------------------------------------------
# 1.12. Callable aliases
# -----------------------------------------------------------------------------

Serializer: TypeAlias = Callable[[Any], Any]
ObjectSerializer: TypeAlias = Callable[[Any, "ExportContextLike"], Any]
RecordBuilder: TypeAlias = Callable[[Any], Record]
PathResolver: TypeAlias = Callable[[PathLike], Path]
ErrorHandler: TypeAlias = Callable[[Exception], None]

# -----------------------------------------------------------------------------
# 1.13. Structural protocols
# -----------------------------------------------------------------------------

@runtime_checkable
class SupportsToDict(Protocol):
    """Object exposing ``to_dict``."""

    def to_dict(self) -> Mapping[str, Any]: ...


@runtime_checkable
class SupportsAsDict(Protocol):
    """Object exposing ``as_dict``."""

    def as_dict(self) -> Mapping[str, Any]: ...


@runtime_checkable
class SupportsJSON(Protocol):
    """Object exposing ``to_json``."""

    def to_json(self, *args: Any, **kwargs: Any) -> str: ...


@runtime_checkable
class SupportsFileRecord(Protocol):
    """Object exposing a file-record mapping."""

    files: MutableMapping[str, Any]


@runtime_checkable
class ExportContextLike(Protocol):
    """Minimum serializer-context interface."""

    include_metadata: bool
    include_private: bool
    strict: bool


# -----------------------------------------------------------------------------
# 1.14. Section registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "NUMPY_AVAILABLE",
        "PANDAS_AVAILABLE",
        "OPENPYXL_AVAILABLE",
        "CHIMERAX_AVAILABLE",
        "PathLike",
        "FileLike",
        "Number",
        "JSONPrimitive",
        "JSONValue",
        "JSONObject",
        "JSONArray",
        "Record",
        "RecordSequence",
        "RecordList",
        "TableMapping",
        "Metadata",
        "AtomLike",
        "ResidueLike",
        "StructureLike",
        "InteractionLike",
        "AnalysisResultLike",
        "ScoringResultLike",
        "DockModelLike",
        "Serializer",
        "ObjectSerializer",
        "RecordBuilder",
        "SupportsToDict",
        "SupportsAsDict",
        "SupportsJSON",
        "SupportsFileRecord",
        "ExportContextLike",
    ]
)

# =============================================================================
# End of Section 1
# =============================================================================


# =============================================================================
# Section 2 — Constants, formats and export conventions
# =============================================================================

# -----------------------------------------------------------------------------
# 2.1. Schema and encoding
# -----------------------------------------------------------------------------

EXPORT_SCHEMA_NAME: Final[str] = "dockanalyzer-export"
EXPORT_SCHEMA_VERSION: Final[str] = "1.0"
DEFAULT_ENCODING: Final[str] = "utf-8"
DEFAULT_NEWLINE: Final[str] = "\n"
DEFAULT_FLOAT_PRECISION: Final[int] = 6
DEFAULT_JSON_INDENT: Final[int] = 2
DEFAULT_CSV_DELIMITER: Final[str] = ","
DEFAULT_TSV_DELIMITER: Final[str] = "\t"
DEFAULT_NULL_TEXT: Final[str] = ""
DEFAULT_BOOL_TRUE: Final[str] = "true"
DEFAULT_BOOL_FALSE: Final[str] = "false"

MIN_FLOAT_PRECISION: Final[int] = 0
MAX_FLOAT_PRECISION: Final[int] = 15
MAX_JSON_DEPTH: Final[int] = 128
MAX_EXCEL_ROWS: Final[int] = 1_048_576
MAX_EXCEL_COLUMNS: Final[int] = 16_384
MAX_EXCEL_SHEET_NAME: Final[int] = 31

# -----------------------------------------------------------------------------
# 2.2. Export formats
# -----------------------------------------------------------------------------

class ExportFormat(str, Enum):
    """Supported output formats."""

    JSON = "json"
    JSONL = "jsonl"
    CSV = "csv"
    TSV = "tsv"
    EXCEL = "xlsx"
    TEXT = "txt"


EXPORT_FORMAT_JSON: Final[str] = ExportFormat.JSON.value
EXPORT_FORMAT_JSONL: Final[str] = ExportFormat.JSONL.value
EXPORT_FORMAT_CSV: Final[str] = ExportFormat.CSV.value
EXPORT_FORMAT_TSV: Final[str] = ExportFormat.TSV.value
EXPORT_FORMAT_EXCEL: Final[str] = ExportFormat.EXCEL.value
EXPORT_FORMAT_TEXT: Final[str] = ExportFormat.TEXT.value

SUPPORTED_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    item.value for item in ExportFormat
)
TABULAR_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    {EXPORT_FORMAT_CSV, EXPORT_FORMAT_TSV, EXPORT_FORMAT_EXCEL}
)
TEXT_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    {
        EXPORT_FORMAT_JSON,
        EXPORT_FORMAT_JSONL,
        EXPORT_FORMAT_CSV,
        EXPORT_FORMAT_TSV,
        EXPORT_FORMAT_TEXT,
    }
)
BINARY_EXPORT_FORMATS: Final[FrozenSet[str]] = frozenset(
    {EXPORT_FORMAT_EXCEL}
)

# -----------------------------------------------------------------------------
# 2.3. Extensions and aliases
# -----------------------------------------------------------------------------

FORMAT_EXTENSIONS: Final[Mapping[str, str]] = MappingProxyType(
    {
        EXPORT_FORMAT_JSON: ".json",
        EXPORT_FORMAT_JSONL: ".jsonl",
        EXPORT_FORMAT_CSV: ".csv",
        EXPORT_FORMAT_TSV: ".tsv",
        EXPORT_FORMAT_EXCEL: ".xlsx",
        EXPORT_FORMAT_TEXT: ".txt",
    }
)

EXTENSION_FORMATS: Final[Mapping[str, str]] = MappingProxyType(
    {extension: name for name, extension in FORMAT_EXTENSIONS.items()}
)

FORMAT_ALIASES: Final[Mapping[str, str]] = MappingProxyType(
    {
        "json": EXPORT_FORMAT_JSON,
        ".json": EXPORT_FORMAT_JSON,
        "application/json": EXPORT_FORMAT_JSON,
        "jsonl": EXPORT_FORMAT_JSONL,
        ".jsonl": EXPORT_FORMAT_JSONL,
        "ndjson": EXPORT_FORMAT_JSONL,
        ".ndjson": EXPORT_FORMAT_JSONL,
        "application/x-ndjson": EXPORT_FORMAT_JSONL,
        "csv": EXPORT_FORMAT_CSV,
        ".csv": EXPORT_FORMAT_CSV,
        "text/csv": EXPORT_FORMAT_CSV,
        "tsv": EXPORT_FORMAT_TSV,
        ".tsv": EXPORT_FORMAT_TSV,
        "tab": EXPORT_FORMAT_TSV,
        "text/tab-separated-values": EXPORT_FORMAT_TSV,
        "xlsx": EXPORT_FORMAT_EXCEL,
        ".xlsx": EXPORT_FORMAT_EXCEL,
        "excel": EXPORT_FORMAT_EXCEL,
        "spreadsheet": EXPORT_FORMAT_EXCEL,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": (
            EXPORT_FORMAT_EXCEL
        ),
        "txt": EXPORT_FORMAT_TEXT,
        ".txt": EXPORT_FORMAT_TEXT,
        "text": EXPORT_FORMAT_TEXT,
        "plain": EXPORT_FORMAT_TEXT,
        "text/plain": EXPORT_FORMAT_TEXT,
    }
)

# -----------------------------------------------------------------------------
# 2.4. Output modes
# -----------------------------------------------------------------------------

class ExportLayout(str, Enum):
    """File layout for multi-table exports."""

    SINGLE_FILE = "single_file"
    MULTIPLE_FILES = "multiple_files"
    PER_POSE = "per_pose"
    CONSOLIDATED = "consolidated"


class ExportDetail(str, Enum):
    """Payload detail level."""

    MINIMAL = "minimal"
    COMPACT = "compact"
    STANDARD = "standard"
    COMPLETE = "complete"


class OverwriteMode(str, Enum):
    """Existing-file behavior."""

    ERROR = "error"
    OVERWRITE = "overwrite"
    UNIQUE = "unique"
    BACKUP = "backup"


class ErrorMode(str, Enum):
    """Export error behavior."""

    RAISE = "raise"
    WARN = "warn"
    COLLECT = "collect"
    SKIP = "skip"


DEFAULT_EXPORT_LAYOUT: Final[str] = ExportLayout.SINGLE_FILE.value
DEFAULT_EXPORT_DETAIL: Final[str] = ExportDetail.STANDARD.value
DEFAULT_OVERWRITE_MODE: Final[str] = OverwriteMode.UNIQUE.value
DEFAULT_ERROR_MODE: Final[str] = ErrorMode.RAISE.value

SUPPORTED_EXPORT_LAYOUTS: Final[FrozenSet[str]] = frozenset(
    item.value for item in ExportLayout
)
SUPPORTED_EXPORT_DETAILS: Final[FrozenSet[str]] = frozenset(
    item.value for item in ExportDetail
)
SUPPORTED_OVERWRITE_MODES: Final[FrozenSet[str]] = frozenset(
    item.value for item in OverwriteMode
)
SUPPORTED_ERROR_MODES: Final[FrozenSet[str]] = frozenset(
    item.value for item in ErrorMode
)

# -----------------------------------------------------------------------------
# 2.5. Standard table names
# -----------------------------------------------------------------------------

TABLE_SUMMARY: Final[str] = "summary"
TABLE_POSES: Final[str] = "poses"
TABLE_INTERACTIONS: Final[str] = "interactions"
TABLE_CONTACTS: Final[str] = "contacts"
TABLE_HBONDS: Final[str] = "hydrogen_bonds"
TABLE_HYDROPHOBIC: Final[str] = "hydrophobic"
TABLE_PI: Final[str] = "pi_interactions"
TABLE_SALT_BRIDGES: Final[str] = "salt_bridges"
TABLE_RESIDUES: Final[str] = "residues"
TABLE_SCORES: Final[str] = "scores"
TABLE_SCORE_COMPONENTS: Final[str] = "score_components"
TABLE_RANKING: Final[str] = "ranking"
TABLE_CONSENSUS: Final[str] = "consensus"
TABLE_PERSISTENCE: Final[str] = "persistence"
TABLE_EXTERNAL_SCORES: Final[str] = "external_scores"
TABLE_EXPLAINABILITY: Final[str] = "explainability"
TABLE_METADATA: Final[str] = "metadata"
TABLE_FILES: Final[str] = "files"
TABLE_WARNINGS: Final[str] = "warnings"
TABLE_ERRORS: Final[str] = "errors"

STANDARD_TABLE_NAMES: Final[Tuple[str, ...]] = (
    TABLE_SUMMARY,
    TABLE_POSES,
    TABLE_INTERACTIONS,
    TABLE_CONTACTS,
    TABLE_HBONDS,
    TABLE_HYDROPHOBIC,
    TABLE_PI,
    TABLE_SALT_BRIDGES,
    TABLE_RESIDUES,
    TABLE_SCORES,
    TABLE_SCORE_COMPONENTS,
    TABLE_RANKING,
    TABLE_CONSENSUS,
    TABLE_PERSISTENCE,
    TABLE_EXTERNAL_SCORES,
    TABLE_EXPLAINABILITY,
    TABLE_METADATA,
    TABLE_FILES,
    TABLE_WARNINGS,
    TABLE_ERRORS,
)

# -----------------------------------------------------------------------------
# 2.6. Excel sheet names
# -----------------------------------------------------------------------------

EXCEL_SHEET_NAMES: Final[Mapping[str, str]] = MappingProxyType(
    {
        TABLE_SUMMARY: "Summary",
        TABLE_POSES: "Poses",
        TABLE_INTERACTIONS: "Interactions",
        TABLE_CONTACTS: "Contacts",
        TABLE_HBONDS: "Hydrogen Bonds",
        TABLE_HYDROPHOBIC: "Hydrophobic",
        TABLE_PI: "Pi Interactions",
        TABLE_SALT_BRIDGES: "Salt Bridges",
        TABLE_RESIDUES: "Residues",
        TABLE_SCORES: "Scores",
        TABLE_SCORE_COMPONENTS: "Score Components",
        TABLE_RANKING: "Ranking",
        TABLE_CONSENSUS: "Consensus",
        TABLE_PERSISTENCE: "Persistence",
        TABLE_EXTERNAL_SCORES: "External Scores",
        TABLE_EXPLAINABILITY: "Explainability",
        TABLE_METADATA: "Metadata",
        TABLE_FILES: "Files",
        TABLE_WARNINGS: "Warnings",
        TABLE_ERRORS: "Errors",
    }
)

DEFAULT_EXCEL_TABLE_STYLE: Final[str] = "TableStyleMedium2"
DEFAULT_EXCEL_FREEZE_PANES: Final[str] = "A2"
DEFAULT_EXCEL_HEADER_ROW: Final[int] = 1
DEFAULT_EXCEL_COLUMN_WIDTH: Final[float] = 12.0
MAX_AUTO_COLUMN_WIDTH: Final[float] = 60.0

# -----------------------------------------------------------------------------
# 2.7. Schema keys
# -----------------------------------------------------------------------------

KEY_SCHEMA_NAME: Final[str] = "schema_name"
KEY_SCHEMA_VERSION: Final[str] = "schema_version"
KEY_DOCKANALYZER_VERSION: Final[str] = "dockanalyzer_version"
KEY_GENERATED_AT: Final[str] = "generated_at"
KEY_PROVENANCE: Final[str] = "provenance"
KEY_METADATA: Final[str] = "metadata"
KEY_FILES: Final[str] = "files"
KEY_WARNINGS: Final[str] = "warnings"
KEY_ERRORS: Final[str] = "errors"
KEY_MODELS: Final[str] = "models"
KEY_INTERACTIONS: Final[str] = "interactions"
KEY_STATISTICS: Final[str] = "statistics"
KEY_SCORING: Final[str] = "scoring"
KEY_TABLES: Final[str] = "tables"
KEY_MANIFEST: Final[str] = "manifest"

REQUIRED_SCHEMA_KEYS: Final[Tuple[str, ...]] = (
    KEY_SCHEMA_NAME,
    KEY_SCHEMA_VERSION,
    KEY_GENERATED_AT,
)

# -----------------------------------------------------------------------------
# 2.8. Interaction families and aliases
# -----------------------------------------------------------------------------

INTERACTION_CONTACT: Final[str] = "contact"
INTERACTION_HBOND: Final[str] = "hbond"
INTERACTION_HYDROPHOBIC: Final[str] = "hydrophobic"
INTERACTION_PI: Final[str] = "pi"
INTERACTION_SALT_BRIDGE: Final[str] = "salt_bridge"
INTERACTION_CLASH: Final[str] = "clash"
INTERACTION_UNKNOWN: Final[str] = "unknown"

INTERACTION_FAMILIES: Final[FrozenSet[str]] = frozenset(
    {
        INTERACTION_CONTACT,
        INTERACTION_HBOND,
        INTERACTION_HYDROPHOBIC,
        INTERACTION_PI,
        INTERACTION_SALT_BRIDGE,
        INTERACTION_CLASH,
        INTERACTION_UNKNOWN,
    }
)

INTERACTION_FAMILY_ALIASES: Final[Mapping[str, str]] = MappingProxyType(
    {
        "contact": INTERACTION_CONTACT,
        "contacts": INTERACTION_CONTACT,
        "hbond": INTERACTION_HBOND,
        "hbonds": INTERACTION_HBOND,
        "hydrogen_bond": INTERACTION_HBOND,
        "hydrogen_bonds": INTERACTION_HBOND,
        "hydrophobic": INTERACTION_HYDROPHOBIC,
        "hydrophobics": INTERACTION_HYDROPHOBIC,
        "pi": INTERACTION_PI,
        "pi_interaction": INTERACTION_PI,
        "pi_interactions": INTERACTION_PI,
        "stacking": INTERACTION_PI,
        "saltbridge": INTERACTION_SALT_BRIDGE,
        "saltbridges": INTERACTION_SALT_BRIDGE,
        "salt_bridge": INTERACTION_SALT_BRIDGE,
        "salt_bridges": INTERACTION_SALT_BRIDGE,
        "ionic": INTERACTION_SALT_BRIDGE,
        "clash": INTERACTION_CLASH,
        "clashes": INTERACTION_CLASH,
    }
)

# -----------------------------------------------------------------------------
# 2.9. File naming
# -----------------------------------------------------------------------------

DEFAULT_EXPORT_BASENAME: Final[str] = "dockanalyzer_export"
DEFAULT_MANIFEST_BASENAME: Final[str] = "manifest"
DEFAULT_SUMMARY_BASENAME: Final[str] = "summary"
DEFAULT_POSE_PREFIX: Final[str] = "pose"
DEFAULT_BACKUP_SUFFIX: Final[str] = ".bak"
DEFAULT_TEMP_SUFFIX: Final[str] = ".tmp"

INVALID_FILENAME_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"[<>:\"/\\|?*\x00-\x1F]"
)
INVALID_SHEET_NAME_PATTERN: Final[re.Pattern[str]] = re.compile(r"[\\/*?:\[\]]")
WHITESPACE_PATTERN: Final[re.Pattern[str]] = re.compile(r"\s+")
MULTIPLE_UNDERSCORES_PATTERN: Final[re.Pattern[str]] = re.compile(r"_+")
WINDOWS_RESERVED_NAMES: Final[FrozenSet[str]] = frozenset(
    {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        *(f"COM{index}" for index in range(1, 10)),
        *(f"LPT{index}" for index in range(1, 10)),
    }
)

# -----------------------------------------------------------------------------
# 2.10. Delimited-file conventions
# -----------------------------------------------------------------------------

DELIMITER_BY_FORMAT: Final[Mapping[str, str]] = MappingProxyType(
    {
        EXPORT_FORMAT_CSV: DEFAULT_CSV_DELIMITER,
        EXPORT_FORMAT_TSV: DEFAULT_TSV_DELIMITER,
    }
)

DEFAULT_CSV_QUOTING: Final[int] = csv.QUOTE_MINIMAL
DEFAULT_CSV_QUOTECHAR: Final[str] = '"'
DEFAULT_CSV_ESCAPECHAR: Final[Optional[str]] = None
DEFAULT_CSV_DOUBLEQUOTE: Final[bool] = True
DEFAULT_CSV_LINETERMINATOR: Final[str] = DEFAULT_NEWLINE

# -----------------------------------------------------------------------------
# 2.11. JSON conventions
# -----------------------------------------------------------------------------

DEFAULT_JSON_ENSURE_ASCII: Final[bool] = False
DEFAULT_JSON_SORT_KEYS: Final[bool] = False
DEFAULT_JSON_ALLOW_NAN: Final[bool] = False
DEFAULT_JSON_COMPACT_SEPARATORS: Final[Tuple[str, str]] = (",", ":")
DEFAULT_JSON_PRETTY_SEPARATORS: Final[Tuple[str, str]] = (",", ": ")
DEFAULT_JSONL_INDENT: Final[None] = None

# -----------------------------------------------------------------------------
# 2.12. Compression and hashing
# -----------------------------------------------------------------------------

class CompressionFormat(str, Enum):
    """Supported stream compression."""

    NONE = "none"
    GZIP = "gzip"


DEFAULT_COMPRESSION: Final[str] = CompressionFormat.NONE.value
SUPPORTED_COMPRESSION_FORMATS: Final[FrozenSet[str]] = frozenset(
    item.value for item in CompressionFormat
)
COMPRESSED_SUFFIXES: Final[Mapping[str, str]] = MappingProxyType(
    {CompressionFormat.GZIP.value: ".gz"}
)
DEFAULT_HASH_ALGORITHM: Final[str] = "sha256"
DEFAULT_HASH_BLOCK_SIZE: Final[int] = 65_536

# -----------------------------------------------------------------------------
# 2.13. Export status
# -----------------------------------------------------------------------------

class ExportStatus(str, Enum):
    """Export operation status."""

    PENDING = "pending"
    SUCCESS = "success"
    PARTIAL = "partial"
    FAILED = "failed"
    SKIPPED = "skipped"


SUCCESS_EXPORT_STATUSES: Final[FrozenSet[str]] = frozenset(
    {ExportStatus.SUCCESS.value, ExportStatus.PARTIAL.value}
)
TERMINAL_EXPORT_STATUSES: Final[FrozenSet[str]] = frozenset(
    {
        ExportStatus.SUCCESS.value,
        ExportStatus.PARTIAL.value,
        ExportStatus.FAILED.value,
        ExportStatus.SKIPPED.value,
    }
)

# -----------------------------------------------------------------------------
# 2.14. Configuration lookup keys
# -----------------------------------------------------------------------------

CONFIG_DIRECTORY_KEYS: Final[Mapping[str, Tuple[str, ...]]] = MappingProxyType(
    {
        EXPORT_FORMAT_JSON: ("JSON_DIR", "JSON_OUTPUT_DIR", "OUTPUT_JSON_DIR"),
        EXPORT_FORMAT_JSONL: ("JSON_DIR", "JSON_OUTPUT_DIR", "OUTPUT_JSON_DIR"),
        EXPORT_FORMAT_CSV: ("CSV_DIR", "CSV_OUTPUT_DIR", "OUTPUT_CSV_DIR"),
        EXPORT_FORMAT_TSV: ("CSV_DIR", "CSV_OUTPUT_DIR", "OUTPUT_CSV_DIR"),
        EXPORT_FORMAT_EXCEL: (
            "EXCEL_DIR",
            "EXCEL_OUTPUT_DIR",
            "OUTPUT_EXCEL_DIR",
        ),
        EXPORT_FORMAT_TEXT: (
            "REPORT_DIR",
            "REPORTS_DIR",
            "TEXT_OUTPUT_DIR",
        ),
    }
)

CONFIG_ENABLE_KEYS: Final[Mapping[str, Tuple[str, ...]]] = MappingProxyType(
    {
        EXPORT_FORMAT_JSON: ("EXPORT_JSON", "ENABLE_JSON_EXPORT"),
        EXPORT_FORMAT_JSONL: ("EXPORT_JSON", "ENABLE_JSON_EXPORT"),
        EXPORT_FORMAT_CSV: ("EXPORT_CSV", "ENABLE_CSV_EXPORT"),
        EXPORT_FORMAT_TSV: ("EXPORT_CSV", "ENABLE_CSV_EXPORT"),
        EXPORT_FORMAT_EXCEL: ("EXPORT_EXCEL", "ENABLE_EXCEL_EXPORT"),
        EXPORT_FORMAT_TEXT: ("EXPORT_REPORT", "ENABLE_TEXT_EXPORT"),
    }
)

# -----------------------------------------------------------------------------
# 2.15. Default table order by detail level
# -----------------------------------------------------------------------------

TABLES_BY_DETAIL: Final[Mapping[str, Tuple[str, ...]]] = MappingProxyType(
    {
        ExportDetail.MINIMAL.value: (
            TABLE_SUMMARY,
            TABLE_POSES,
            TABLE_SCORES,
        ),
        ExportDetail.COMPACT.value: (
            TABLE_SUMMARY,
            TABLE_POSES,
            TABLE_INTERACTIONS,
            TABLE_RESIDUES,
            TABLE_SCORES,
            TABLE_RANKING,
        ),
        ExportDetail.STANDARD.value: (
            TABLE_SUMMARY,
            TABLE_POSES,
            TABLE_INTERACTIONS,
            TABLE_CONTACTS,
            TABLE_HBONDS,
            TABLE_HYDROPHOBIC,
            TABLE_PI,
            TABLE_SALT_BRIDGES,
            TABLE_RESIDUES,
            TABLE_SCORES,
            TABLE_SCORE_COMPONENTS,
            TABLE_RANKING,
            TABLE_CONSENSUS,
            TABLE_PERSISTENCE,
            TABLE_EXTERNAL_SCORES,
            TABLE_METADATA,
        ),
        ExportDetail.COMPLETE.value: STANDARD_TABLE_NAMES,
    }
)

# -----------------------------------------------------------------------------
# 2.16. Public registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "EXPORT_SCHEMA_NAME",
        "EXPORT_SCHEMA_VERSION",
        "DEFAULT_ENCODING",
        "DEFAULT_NEWLINE",
        "DEFAULT_FLOAT_PRECISION",
        "DEFAULT_JSON_INDENT",
        "ExportFormat",
        "ExportLayout",
        "ExportDetail",
        "OverwriteMode",
        "ErrorMode",
        "CompressionFormat",
        "ExportStatus",
        "SUPPORTED_EXPORT_FORMATS",
        "FORMAT_EXTENSIONS",
        "FORMAT_ALIASES",
        "STANDARD_TABLE_NAMES",
        "EXCEL_SHEET_NAMES",
        "INTERACTION_FAMILIES",
        "INTERACTION_FAMILY_ALIASES",
        "TABLES_BY_DETAIL",
    ]
)

# =============================================================================
# End of Section 2
# =============================================================================

# =============================================================================
# Section 3 — Export exceptions and warnings
# =============================================================================

# -----------------------------------------------------------------------------
# 3.1. Base exception
# -----------------------------------------------------------------------------


class ExportError(Exception):
    """Base exception for export failures."""

    default_code: ClassVar[str] = "export_error"

    def __init__(
        self,
        message: str,
        *,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        clean_message = str(message).strip() or self.__class__.__name__
        self.message = clean_message
        self.code = str(code or self.default_code)
        self.context: Dict[str, Any] = dict(context or {})
        self.cause = cause
        super().__init__(clean_message)

    def __str__(self) -> str:
        text = self.message
        if self.code:
            text = f"[{self.code}] {text}"
        if self.context:
            details = ", ".join(
                f"{key}={value!r}" for key, value in sorted(self.context.items())
            )
            text = f"{text} ({details})"
        return text

    def with_context(self, **context: Any) -> "ExportError":
        """Add context and return this exception."""
        self.context.update(context)
        return self

    def to_dict(self) -> Dict[str, Any]:
        """Return a serializable error record."""
        record: Dict[str, Any] = {
            "type": self.__class__.__name__,
            "code": self.code,
            "message": self.message,
            "context": dict(self.context),
        }
        if self.cause is not None:
            record["cause"] = {
                "type": self.cause.__class__.__name__,
                "message": str(self.cause),
            }
        return record


# -----------------------------------------------------------------------------
# 3.2. Input and configuration errors
# -----------------------------------------------------------------------------


class ExportInputError(ExportError):
    """Raised for unsupported or malformed input objects."""

    default_code = "invalid_input"


class ExportConfigurationError(ExportError):
    """Raised for invalid export options."""

    default_code = "invalid_configuration"


class ExportFormatError(ExportConfigurationError):
    """Raised for unsupported formats or format options."""

    default_code = "unsupported_format"


class ExportDependencyError(ExportConfigurationError):
    """Raised when an optional dependency is required."""

    default_code = "missing_dependency"

    def __init__(
        self,
        dependency: str,
        *,
        feature: Optional[str] = None,
        message: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        dependency_name = str(dependency).strip() or "unknown"
        details = dict(context or {})
        details.setdefault("dependency", dependency_name)
        if feature:
            details.setdefault("feature", str(feature))
        if message is None:
            message = f"Optional dependency {dependency_name!r} is required"
            if feature:
                message += f" for {feature}"
        super().__init__(
            message,
            context=details,
            cause=cause,
        )
        self.dependency = dependency_name
        self.feature = feature


# -----------------------------------------------------------------------------
# 3.3. Path and write errors
# -----------------------------------------------------------------------------


class ExportPathError(ExportError):
    """Raised for invalid or inaccessible output paths."""

    default_code = "invalid_path"

    def __init__(
        self,
        message: str,
        *,
        path: Optional[PathLike] = None,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        details = dict(context or {})
        if path is not None:
            details.setdefault("path", str(path))
        super().__init__(
            message,
            code=code,
            context=details,
            cause=cause,
        )
        self.path = Path(path) if path is not None else None


class ExportWriteError(ExportPathError):
    """Raised when writing or replacing an output file fails."""

    default_code = "write_failed"


# -----------------------------------------------------------------------------
# 3.4. Serialization and validation errors
# -----------------------------------------------------------------------------


class ExportSerializationError(ExportError):
    """Raised when an object cannot be serialized safely."""

    default_code = "serialization_failed"

    def __init__(
        self,
        message: str,
        *,
        object_type: Optional[Union[str, type]] = None,
        location: Optional[str] = None,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        details = dict(context or {})
        if object_type is not None:
            type_name = (
                object_type.__qualname__
                if isinstance(object_type, type)
                else str(object_type)
            )
            details.setdefault("object_type", type_name)
        if location:
            details.setdefault("location", str(location))
        super().__init__(
            message,
            code=code,
            context=details,
            cause=cause,
        )
        self.object_type = object_type
        self.location = location


class ExportValidationError(ExportError):
    """Raised when export data fail validation."""

    default_code = "validation_failed"

    def __init__(
        self,
        message: str,
        *,
        field_name: Optional[str] = None,
        value: Any = None,
        include_value: bool = False,
        code: Optional[str] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        details = dict(context or {})
        if field_name:
            details.setdefault("field", str(field_name))
        if include_value:
            details.setdefault("value", value)
        super().__init__(
            message,
            code=code,
            context=details,
            cause=cause,
        )
        self.field_name = field_name
        self.value = value


# -----------------------------------------------------------------------------
# 3.5. Batch errors
# -----------------------------------------------------------------------------


class ExportBatchError(ExportError):
    """Raised when one or more batch exports fail."""

    default_code = "batch_failed"

    def __init__(
        self,
        message: str,
        *,
        failures: Optional[Sequence[BaseException]] = None,
        completed: int = 0,
        total: Optional[int] = None,
        context: Optional[Mapping[str, Any]] = None,
        cause: Optional[BaseException] = None,
    ) -> None:
        self.failures: Tuple[BaseException, ...] = tuple(failures or ())
        self.completed = max(0, int(completed))
        self.total = None if total is None else max(0, int(total))
        details = dict(context or {})
        details.setdefault("completed", self.completed)
        details.setdefault("failure_count", len(self.failures))
        if self.total is not None:
            details.setdefault("total", self.total)
        super().__init__(message, context=details, cause=cause)

    def to_dict(self) -> Dict[str, Any]:
        """Return the batch error and nested failures."""
        record = super().to_dict()
        record["failures"] = [
            failure.to_dict()
            if isinstance(failure, ExportError)
            else {
                "type": failure.__class__.__name__,
                "message": str(failure),
            }
            for failure in self.failures
        ]
        return record


# -----------------------------------------------------------------------------
# 3.6. Warnings
# -----------------------------------------------------------------------------


class ExportWarning(UserWarning):
    """Base warning for recoverable export issues."""


class ExportLossyConversionWarning(ExportWarning):
    """Warn that conversion discarded or simplified data."""


class ExportDependencyWarning(ExportWarning):
    """Warn that an optional feature is unavailable."""


class ExportCompatibilityWarning(ExportWarning):
    """Warn about legacy or partially supported data."""


class ExportOverwriteWarning(ExportWarning):
    """Warn that an existing output may be replaced."""


# -----------------------------------------------------------------------------
# 3.7. Public registration
# -----------------------------------------------------------------------------

__all__.extend(
    [
        "ExportError",
        "ExportInputError",
        "ExportConfigurationError",
        "ExportFormatError",
        "ExportDependencyError",
        "ExportPathError",
        "ExportWriteError",
        "ExportSerializationError",
        "ExportValidationError",
        "ExportBatchError",
        "ExportWarning",
        "ExportLossyConversionWarning",
        "ExportDependencyWarning",
        "ExportCompatibilityWarning",
        "ExportOverwriteWarning",
    ]
)

# =============================================================================
# End of Section 3
# =============================================================================





